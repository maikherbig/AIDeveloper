# -*- coding: utf-8 -*-
"""
AIDeveloper
---------
@author: maikherbig
"""
VERSION = "0.4.0" #Python 3.9.9 Version

import os,sys,gc
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '3'

if not sys.platform.startswith("win"):
    from multiprocessing import freeze_support
    freeze_support()
# Make sure to get the right icon file on win,linux and mac
if sys.platform=="darwin":
    icon_suff = ".icns"
else:
    icon_suff = ".ico"

import pyqtgraph as pg
from pyqtgraph.Qt import QtCore, QtWidgets, QtGui
from pyqtgraph import Qt

import aid_start
dir_root = os.path.dirname(aid_start.__file__)#ask the module for its origin
dir_settings = os.path.join(dir_root,"aid_settings.json")#dir to settings
Default_dict = aid_start.get_default_dict(dir_settings) 

#try:
#    splashapp = QtWidgets.QApplication(sys.argv)
#    #splashapp.setWindowIcon(QtGui.QIcon("."+os.sep+"art"+os.sep+Default_dict["Icon theme"]+os.sep+"main_icon_simple_04_256.ico"))
#    # Create and display the splash screen
#    splash_pix = os.path.join(dir_root,"art",Default_dict["Icon theme"],"main_icon_simple_04_256"+icon_suff)
#    splash_pix = QtGui.QPixmap(splash_pix)
#    #splash_pix = QtGui.QPixmap("."+os.sep+"art"+os.sep+Default_dict["Icon theme"]+os.sep+"main_icon_simple_04_256"+icon_suff)
#    splash = QtWidgets.QSplashScreen(splash_pix, QtCore.Qt.WindowStaysOnTopHint)
#    splash.setMask(splash_pix.mask())
#    splash.show()
#except:
#    pass

#BEFORE importing tensorflow or anything from keras: make sure the keras.json has
#certain properties
keras_json_path = os.path.expanduser('~')+os.sep+'.keras'+os.sep+'keras.json'
if not os.path.isdir(os.path.expanduser('~')+os.sep+'.keras'):
    os.mkdir(os.path.expanduser('~')+os.sep+'.keras')

aid_start.banner() #show a fancy banner in console
aid_start.keras_json_check(keras_json_path)

import traceback,shutil,re,ast,io,platform
import h5py,json,time,copy,urllib,datetime
from stat import S_IREAD,S_IRGRP,S_IROTH,S_IWRITE,S_IWGRP,S_IWOTH
import tensorflow as tf
tf.compat.v1.experimental.output_all_intermediates(True)
# from tensorboard import program
# from tensorboard import default
# from tensorboard import assets

from tensorflow.python.client import device_lib
devices = device_lib.list_local_devices()
device_types = [devices[i].device_type for i in range(len(devices))]

#Get the number  of CPU cores and GPUs
cpu_nr = os.cpu_count()
gpu_nr = device_types.count("GPU")
print("Nr. of CPUs detected: "+str(cpu_nr))
print("Nr. of GPUs detected: "+str(gpu_nr))

print("List of device(s):")
print("------------------------")
for i in range(len(devices)):
    print("Device "+str(i)+": "+devices[i].name)
    print("Device type: "+devices[i].device_type)
    print("Device description: "+devices[i].physical_device_desc)
    print("------------------------")

#Split CPU and GPU into two lists of devices
devices_cpu = []
devices_gpu = []
for dev in devices:
    if dev.device_type=="CPU":
        devices_cpu.append(dev)
    elif dev.device_type=="GPU":
        devices_gpu.append(dev)
    else:
        print("Unknown device type:"+str(dev)+"\n")

import numpy as np
rand_state = np.random.RandomState(117) #to get the same random number on diff. PCs
#from scipy import ndimage,misc
from sklearn import metrics,preprocessing
import PIL
import cv2
import pandas as pd
import openpyxl,xlrd 
import psutil

from tensorflow.keras.models import model_from_json,model_from_config,load_model,clone_model
from tensorflow.keras import backend as K
# if 'GPU' in device_types:
#     keras_gpu_avail = tf.config.list_physical_devices()
#     if len(keras_gpu_avail)>0:
#         print("Following Devices are available:")
#         print(keras_gpu_avail)
#         print("------------------------")
#     else:
#         print("TensorFlow detected GPU, but Keras didn't")
#         print("------------------------")

from tensorflow.keras.preprocessing.image import load_img
from tensorflow.keras.utils import to_categorical

import model_zoo 
import tf2onnx
from onnx import save_model as save_onnx

import aid_img, aid_dl, aid_bin
import aid_frontend
from partial_trainability import partial_trainability
import aid_imports

model_zoo_version = model_zoo.__version__()
print("AIDeveloper Version: "+VERSION)
print("model_zoo.py Version: "+model_zoo.__version__())

try:
    _fromUtf8 = QtCore.QString.fromUtf8
except AttributeError:
    def _fromUtf8(s):
        return s

try:
    _encoding = QtWidgets.QApplication.UnicodeUTF8
    def _translate(context, text, disambig):
        return QtWidgets.QApplication.translate(context, text, disambig, _encoding)
except AttributeError:
    def _translate(context, text, disambig):
        return QtWidgets.QApplication.translate(context, text, disambig)

tooltips = aid_start.get_tooltips()

class MyPopup(QtWidgets.QWidget):
    def __init__(self):
        QtWidgets.QWidget.__init__(self)

class WorkerSignals(QtCore.QObject):
    '''
    Code inspired from here: https://www.learnpyqt.com/courses/concurrent-execution/multithreading-pyqt-applications-qthreadpool/
    
    Defines the signals available from a running worker thread.
    Supported signals are:
    finished
        No data
    error
        `tuple` (exctype, value, traceback.format_exc() )
    result
        `object` data returned from processing, anything
    progress
        `int` indicating % progress
    history
        `dict` containing keras model history.history resulting from .fit
    '''
    finished = QtCore.pyqtSignal()
    error = QtCore.pyqtSignal(tuple)
    result = QtCore.pyqtSignal(object)
    progress = QtCore.pyqtSignal(int)
    history = QtCore.pyqtSignal(dict)

class Worker(QtCore.QRunnable):
    '''
    Code inspired/copied from: https://www.learnpyqt.com/courses/concurrent-execution/multithreading-pyqt-applications-qthreadpool/
    Worker thread
    Inherits from QRunnable to handler worker thread setup, signals and wrap-up.
    :param callback: The function callback to run on this worker thread. Supplied args and 
                     kwargs will be passed through to the runner.
    :type callback: function
    :param args: Arguments to pass to the callback function
    :param kwargs: Keywords to pass to the callback function
    '''
    def __init__(self, fn, *args, **kwargs):
        super(Worker, self).__init__()

        # Store constructor arguments (re-used for processing)
        self.fn = fn
        self.args = args
        self.kwargs = kwargs
        self.signals = WorkerSignals()

        # Add the callback to our kwargs
        self.kwargs['progress_callback'] = self.signals.progress
        self.kwargs['history_callback'] = self.signals.history

    @QtCore.pyqtSlot()
    def run(self):
        '''
        Initialise the runner function with passed args, kwargs.
        '''

        # Retrieve args/kwargs here; and fire processing using them
        try:
            result = self.fn(*self.args, **self.kwargs)
        except:
            traceback.print_exc()
            exctype, value = sys.exc_info()[:2]
            self.signals.error.emit((exctype, value, traceback.format_exc()))
        else:
            self.signals.result.emit(result)  # Return the result of the processing
        finally:
            self.signals.finished.emit()  # Done


class MainWindow(QtWidgets.QMainWindow):
    def __init__(self, *args, **kwargs):
        super(MainWindow, self).__init__(*args, **kwargs)
        self.setupUi()
    def add_app(self,app):
        self.app = app
        
    def setupUi(self):
        aid_frontend.setup_main_ui(self,gpu_nr)
    
    def retranslateUi(self):
        aid_frontend.retranslate_main_ui(self,gpu_nr,VERSION)

    def dataDropped(self, l):
        #If there is data stored on ram tell user that RAM needs to be refreshed!
        if len(self.ram)>0:
            self.statusbar.showMessage("Newly added data is not yet in RAM. Only RAM data will be used. Use ->'File'->'Data to RAM now' to update RAM",5000)
        #l is a list of some filenames (.rtdc) or folders (containing .jpg, jpeg, .png)
        
        #Iterate over l and check if it is a folder or a file (directory)    
        isfile = [os.path.isfile(str(url)) for url in l]
        isfolder = [os.path.isdir(str(url)) for url in l]


        #####################For folders with images:##########################            
        #where are folders?
        ind_true = np.where(np.array(isfolder)==True)[0]
        foldernames = list(np.array(l)[ind_true]) #select the indices that are valid
        #On mac, there is a trailing / in case of folders; remove them
        foldernames = [os.path.normpath(url) for url in foldernames]

        basename = [os.path.basename(f) for f in foldernames]
        #Look quickly inside the folders and ask the user if he wants to convert
        #to .rtdc (might take a while!)
        if len(foldernames)>0: #User dropped (also) folders (which may contain images)
#            filecounts = []
#            for i in range(len(foldernames)):
#                url = foldernames[i]
#                files = os.listdir(url)
#                files_full = [os.path.join(url,files[i]) for i in range(len(files))]
#                filecounts.append(len([f for f in files_full if os.path.isfile(f)]))
#            Text = []
#            for b,n in zip(basename,filecounts):
#                Text.append(b+": "+str(n)+" images")
#            Text = "\n".join(Text)

            Text = "Images from single folders are read and saved to individual \
            .rtdc files with the same name like the corresponding folder.<b>If \
            you have RGB images you can either save the full RGB information, \
            or do a conversion to Grayscale (saves some diskspace but information \
            about color is lost). RGB is recommended since AID will automatically\
            do the conversion to grayscale later if required.<b>If you have \
            Grayscale images, a conversion to RGB will just copy the info to all \
            channels, which allows you to use RGB-mode and Grayscale-mode lateron."

            Text = Text+"\nImages from following folders will be converted:\n"+"\n".join(basename)
            #Show the user a summary with all the found folders and how many files are
            #contained. Ask if he want to convert
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Question)
            text = "<html><head/><body><p>Should the images of the chosen folder(s)\
            be converted to .rtdc using <b>RGB</b> or <b>Grayscale</b> format? <b>\
            (RGB is recommended!)</b>  Either option might take some time. You can \
            reuse the .rtdc file next time.</p></body></html>"
            msg.setText(text)
            msg.setDetailedText(Text)
            msg.setWindowTitle("Format for conversion to .rtdc (RGB/Grayscale)")
            msg.addButton(QtGui.QPushButton('Convert to Grayscale'), QtGui.QMessageBox.YesRole)
            msg.addButton(QtGui.QPushButton('Convert to RGB'), QtGui.QMessageBox.NoRole)
            msg.addButton(QtGui.QPushButton('Cancel'), QtGui.QMessageBox.RejectRole)
            retval = msg.exec_()

            #Conversion of images in folders is (almost) independent from what 
            #is going to be fitted (So I leave the option menu still!)
            #In options: Color Mode one can still use RGB mode and export here as
            #Grayscale (but this would actually not work since RGB information is lost).
            #The other way around works. Therefore it is recommended to export RGB!
            if retval==0: 
                color_mode = "Grayscale"
                channels = 1
            elif retval==1:
                color_mode = "RGB"
                channels = 3
            else:
                return
            self.statusbar.showMessage("Color mode' "+color_mode+"' is used",5000)
            url_converted = []
            for i in range(len(foldernames)):
                url = foldernames[i]
                print("Start converting images in\n"+url)
                #try:
                #get a list of files inside this directory:
                images,pos_x,pos_y = [],[],[]
                for root, dirs, files in os.walk(url):
                    for file in files:
                        try:
                            path = os.path.join(root, file)
                            img = load_img(path,color_mode=color_mode.lower()) #This uses PIL and supports many many formats!
                            images.append(np.array(img)) #append nice numpy array to list
                            #create pos_x and pos_y
                            pos_x.append( int(np.round(img.width/2.0,0)) ) 
                            pos_y.append( int(np.round(img.height/2.0,0)) )  
                        except:
                            pass
                
                #Thanks to andko76 for pointing that unequal image sizes cause an error:
                #https://github.com/maikherbig/AIDeveloper/issues/1
                #Check that all images have the same size
#                img_shape_errors = 0
#                text_error = "Images have unequal dimensions:"
#                img_h = [a.shape[0] for a in images]
#                img_h_uni = len(np.unique(img_h))
#                if img_h_uni!=1:
#                    text_error += "\n- found unequal heights"
#                    img_shape_errors=1
#                img_w = [a.shape[1] for a in images]
#                img_w_uni = len(np.unique(img_w))
#                if img_w_uni!=1:
#                    text_error += "\n- found unequal widths"
#                    img_shape_errors=1
#                img_c = [len(a.shape) for a in images]
#                img_c_uni = len(np.unique(img_c))
#                if img_c_uni!=1:
#                    text_error += "\n- found unequal numbers of channels"
#                    img_shape_errors=1
#                #If there were issues detected, show error message
#                if img_shape_errors==1:
#                    msg = QtWidgets.QMessageBox()
#                    msg.setIcon(QtWidgets.QMessageBox.Warning)       
#                    msg.setText(str(text_error))
#                    msg.setWindowTitle("Error: Unequal image shapes")
#                    msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
#                    msg.exec_()
#                    return

                #Get a list of occuring image dimensions (width and height)
                img_shape = [a.shape[0] for a in images] + [a.shape[1] for a in images]
                dims = np.unique(img_shape)
                #Get a list of occurences of image shapes
                img_shape = [str(a.shape[0])+" x "+str(a.shape[1]) for a in images]
                occurences = np.unique(img_shape,return_counts=True)
                #inform user if there is more than one img shape 
                if len(occurences[0])>1 or len(dims)>1:
                    text_detail = "Path: "+url
                    text_detail += "\nFollowing image shapes are present"
                    for i in range(len(occurences[0])):
                        text_detail+="\n- "+str(occurences[1][i])+" times: "+str(occurences[0][i])
                    
                    self.popup_imgRes = QtGui.QDialog()
                    self.popup_imgRes_ui = aid_frontend.popup_imageLoadResize()
                    self.popup_imgRes_ui.setupUi(self.popup_imgRes) #open a popup to show options for image resizing (make image equally sized)
                    #self.popup_imgRes.setWindowModality(QtCore.Qt.WindowModal)
                    self.popup_imgRes.setWindowModality(QtCore.Qt.ApplicationModal)
                    #Insert information into textBrowser
                    self.popup_imgRes_ui.textBrowser_imgResize_occurences.setText(text_detail)
                    Image_import_dimension = Default_dict["Image_import_dimension"]
                    self.popup_imgRes_ui.spinBox_ingResize_h_1.setValue(Image_import_dimension)
                    self.popup_imgRes_ui.spinBox_ingResize_h_2.setValue(Image_import_dimension)
                    self.popup_imgRes_ui.spinBox_ingResize_w_1.setValue(Image_import_dimension)
                    self.popup_imgRes_ui.spinBox_ingResize_w_2.setValue(Image_import_dimension)
                    Image_import_interpol_method = Default_dict["Image_import_interpol_method"]
                    index = self.popup_imgRes_ui.comboBox_resizeMethod.findText(Image_import_interpol_method, QtCore.Qt.MatchFixedString)
                    if index >= 0:
                         self.popup_imgRes_ui.comboBox_resizeMethod.setCurrentIndex(index)
                    #Define function for the OK button:
                    def popup_imgRes_ok(images,channels,pos_x,pos_y):
                        print("Start resizing operation")
                        #Get info from GUI
                        final_h = int(self.popup_imgRes_ui.spinBox_ingResize_h_1.value())
                        print("Height:"+str(final_h))
                        final_w = int(self.popup_imgRes_ui.spinBox_ingResize_w_1.value())
                        print("Width:"+str(final_w))
                        Default_dict["Image_import_dimension"] = final_h
                        
                        pix = 1
                        if self.popup_imgRes_ui.radioButton_imgResize_cropPad.isChecked():#cropping and padding method
                            images = aid_img.image_crop_pad_cv2(images,pos_x,pos_y,pix,final_h,final_w,padding_mode="cv2.BORDER_CONSTANT")
                        elif self.popup_imgRes_ui.radioButton_imgResize_interpolate.isChecked():
                            interpolation_method = str(self.popup_imgRes_ui.comboBox_resizeMethod.currentText())
                            Default_dict["Image_import_interpol_method"] = interpolation_method
                            images = aid_img.image_resize_scale(images,pos_x,pos_y,final_h,final_w,channels,interpolation_method,verbose=False)
                        else:
                            print("Invalid image resize method!")
                        #Save the Default_dict
                        aid_bin.save_aid_settings(Default_dict) 
                        self.popup_imgRes.accept()
                        return images
                    
                    #Define function for the Cancel button:                    
                    def popup_imgRes_cancel():
                        self.popup_imgRes.close()
                        return

                    self.popup_imgRes_ui.pushButton_imgResize_ok.clicked.connect(lambda: popup_imgRes_ok(images,channels,pos_x,pos_y))
                    self.popup_imgRes_ui.pushButton_imgResize_cancel.clicked.connect(popup_imgRes_cancel)
                    
                    retval = self.popup_imgRes.exec_()
                    #retval is 0 if the user clicked cancel or just closed the window; in this case just exist the function
                    if retval==0:
                        return

                #get new pos_x, pos_y (after cropping, the pixel value for the middle of the image is different!)
                pos_x = [int(np.round(img.shape[1]/2.0,0)) for img in images]
                pos_y = [int(np.round(img.shape[0]/2.0,0)) for img in images]
                
                #Now, all images are of identical shape and can be converted to a numpy array
                images = np.array((images), dtype="uint8")
                pos_x = np.array((pos_x), dtype="uint8")
                pos_y = np.array((pos_y), dtype="uint8")
                
                #Save as foldername.rtdc
                fname = url+".rtdc"
                if os.path.isfile(fname):
                    #ask user if file can be overwritten
                    msg = QtWidgets.QMessageBox()
                    msg.setIcon(QtWidgets.QMessageBox.Question)
                    text = "<html><head/><body><p>File:"+fname+" already exists. Should it be overwritten?</p></body></html>"
                    msg.setText(text)
                    msg.setWindowTitle("Overwrite file?")
                    msg.addButton(QtGui.QPushButton('Yes'), QtGui.QMessageBox.YesRole)
                    msg.addButton(QtGui.QPushButton('No'), QtGui.QMessageBox.NoRole)
                    msg.addButton(QtGui.QPushButton('Cancel'), QtGui.QMessageBox.RejectRole)
                    retval = msg.exec_()
        
                    if retval==0:
                        try:
                            os.remove(fname)
                            aid_img.imgs_2_rtdc(fname,images,pos_x,pos_y)
                            url_converted.append(fname)
                        except Exception as e:
                            msg = QtWidgets.QMessageBox()
                            msg.setIcon(QtWidgets.QMessageBox.Critical)
                            msg.setText(str(e))
                            msg.setWindowTitle("Error")
                            retval = msg.exec_()
                    elif retval==1:
                        pass
                    else:
                        pass
                else:#file does not yet exist. Create it
                    aid_img.imgs_2_rtdc(fname,images,pos_x,pos_y)
                    url_converted.append(fname)

            print("Finished converting! Final dimension of image tensor is:"+str(images.shape))
            #Now load the created files directly to drag/drop-region!
            self.dataDropped(url_converted)

        #####################For .rtdc files:##################################            
        #where are files?
        ind_true = np.where(np.array(isfile)==True)[0]
        filenames = list(np.array(l)[ind_true]) #select the indices that are valid
        #check if the file can be opened and get some information
        fileinfo = []
        for i in range(len(filenames)):
            rtdc_path = filenames[i]
            
            try:
                failed,rtdc_ds = aid_bin.load_rtdc(rtdc_path)
                if failed:
                    msg = QtWidgets.QMessageBox()
                    msg.setIcon(QtWidgets.QMessageBox.Critical)       
                    msg.setText(str(rtdc_ds))
                    msg.setWindowTitle("Error occurred during loading file")
                    msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                    msg.exec_()
                    return
                
                features = list(rtdc_ds["events"].keys())
                #Make sure that there is "images", "pos_x" and "pos_y" available
                if "image" in features and "pos_x" in features and "pos_y" in features:
                    nr_images = rtdc_ds["events"]["image"].len()
                    pix = rtdc_ds.attrs["imaging:pixel size"]
                    xtra_in_available = len(rtdc_ds.keys())>2 #Is True, only if there are more than 2 elements. 
                    fileinfo.append({"rtdc_ds":rtdc_ds,"rtdc_path":rtdc_path,"features":features,"nr_images":nr_images,"pix":pix,"xtra_in":xtra_in_available})
                else:
                    missing = []
                    for feat in ["image","pos_x","pos_y"]:
                        if feat not in features:
                            missing.append(feat)    
                    msg = QtWidgets.QMessageBox()
                    msg.setIcon(QtWidgets.QMessageBox.Information)       
                    msg.setText("Essential feature(s) are missing in data-set")
                    msg.setDetailedText("Data-set: "+rtdc_path+"\nis missing "+str(missing))
                    msg.setWindowTitle("Missing essential features")
                    msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                    msg.exec_()                      
                                
            except Exception as e:
                print(e)
        
        #Add the stuff to the combobox on Plot/Peak Tab
        url_list = [fileinfo[iterator]["rtdc_path"] for iterator in range(len(fileinfo))]
        self.comboBox_chooseRtdcFile.addItems(url_list)
        self.comboBox_selectData.addItems(url_list)
        if len(url_list)==0: #This fixes the issue that the prog. crashes if accidentially a tableitem is dragged and "dropped" on the table
            return
        width=self.comboBox_selectData.fontMetrics().boundingRect(max(url_list, key=len)).width()
        self.comboBox_selectData.view().setFixedWidth(width+10)             
        
        for rowNumber in range(len(fileinfo)):#for url in l:
            url = fileinfo[rowNumber]["rtdc_path"]
            #add to table
            rowPosition = self.table_dragdrop.rowCount()
            self.table_dragdrop.insertRow(rowPosition)
            
            columnPosition = 0
            line = QtWidgets.QTableWidgetItem()
            line.setText(url)
            line.setFlags( QtCore.Qt.ItemIsSelectable |  QtCore.Qt.ItemIsEnabled )
            #line.setDisabled(True)
            #line.setAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
            self.table_dragdrop.setItem(rowPosition, columnPosition, line)            
            
#            item = QtWidgets.QTableWidgetItem(url) 
#            item.setFlags( QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable )
#            print(item.textAlignment())
#            item.setTextAlignment(QtCore.Qt.AlignRight) # change the alignment
#            #item.setTextAlignment(QtCore.Qt.AlignLeading | QtCore.Qt.AnchorRight) # change the alignment
#            self.table_dragdrop.setItem(rowPosition , columnPosition, item ) #

            columnPosition = 1
            spinb = QtWidgets.QSpinBox(self.table_dragdrop)
            spinb.valueChanged.connect(self.dataOverviewOn)
            self.table_dragdrop.setCellWidget(rowPosition, columnPosition, spinb)            

            for columnPosition in range(2,4):
                #for each item, also create 2 checkboxes (train/valid)
                item = QtWidgets.QTableWidgetItem()#("item {0} {1}".format(rowNumber, columnNumber))
                item.setFlags( QtCore.Qt.ItemIsUserCheckable | QtCore.Qt.ItemIsEnabled  )
                item.setCheckState(QtCore.Qt.Unchecked)
                self.table_dragdrop.setItem(rowPosition, columnPosition, item)
            
            columnPosition = 4
            #Place a button which allows to show a plot (scatter, histo...lets see)
            btn = QtWidgets.QPushButton(self.table_dragdrop)
            btn.setMinimumSize(QtCore.QSize(50, 30))
            btn.setMaximumSize(QtCore.QSize(50, 30))
            btn.clicked.connect(self.button_hist)
            btn.setText('Plot')
            self.table_dragdrop.setCellWidget(rowPosition, columnPosition, btn)            
            self.table_dragdrop.resizeRowsToContents()

#            columnPosition = 5
#            #Place a combobox with the available features
#            cb = QtWidgets.QComboBox(self.table_dragdrop)
#            cb.addItems(fileinfo[rowNumber]["features"])
#            cb.setMinimumSize(QtCore.QSize(70, 30))
#            cb.setMaximumSize(QtCore.QSize(70, 30))            
#            width=cb.fontMetrics().boundingRect(max(fileinfo[rowNumber]["features"], key=len)).width()
#            cb.view().setFixedWidth(width+30)             
#            self.table_dragdrop.setCellWidget(rowPosition, columnPosition, cb)            
          

            columnPosition = 5
            #Place a combobox with the available features
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.DisplayRole, fileinfo[rowNumber]["nr_images"])
            item.setFlags(item.flags() &~QtCore.Qt.ItemIsEnabled &~ QtCore.Qt.ItemIsSelectable )
            self.table_dragdrop.setItem(rowPosition, columnPosition, item)

            columnPosition = 6
            #Field to user-define nr. of cells/epoch
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.EditRole,100)
            #item.cellChanged.connect(self.dataOverviewOn)
            self.table_dragdrop.setItem(rowPosition, columnPosition, item)

            columnPosition = 7
            #Pixel size
            item = QtWidgets.QTableWidgetItem()
            pix = float(fileinfo[rowNumber]["pix"])
            #print(pix)
            item.setData(QtCore.Qt.EditRole,pix)
            self.table_dragdrop.setItem(rowPosition, columnPosition, item)

            columnPosition = 8           
            #Should data be shuffled (random?)
            item = QtWidgets.QTableWidgetItem()#("item {0} {1}".format(rowNumber, columnNumber))
            item.setFlags( QtCore.Qt.ItemIsUserCheckable | QtCore.Qt.ItemIsEnabled  )
            item.setCheckState(QtCore.Qt.Checked)
            self.table_dragdrop.setItem(rowPosition, columnPosition, item)

            columnPosition = 9
            #Zooming factor
            item = QtWidgets.QTableWidgetItem()
            zoom = 1.0
            item.setData(QtCore.Qt.EditRole,zoom)
            self.table_dragdrop.setItem(rowPosition, columnPosition, item)

            columnPosition = 10           
            #Should xtra_data be used?
            item = QtWidgets.QTableWidgetItem()#("item {0} {1}".format(rowNumber, columnNumber))
            xtra_in_available = fileinfo[rowNumber]["xtra_in"]
            if xtra_in_available:
                item.setFlags( QtCore.Qt.ItemIsUserCheckable | QtCore.Qt.ItemIsEnabled  )
            else:
                item.setFlags( QtCore.Qt.ItemIsUserCheckable )
            item.setCheckState(QtCore.Qt.Unchecked)

            self.table_dragdrop.setItem(rowPosition, columnPosition, item)


    #Functions for Keras augmentation checkboxes
    def keras_changed_rotation(self,on_or_off):
        if on_or_off==0:
            self.lineEdit_Rotation.setText(str(0))
            self.lineEdit_Rotation.setEnabled(False)
        elif on_or_off==2:
            self.lineEdit_Rotation.setText(str(Default_dict ["rotation"]))
            self.lineEdit_Rotation.setEnabled(True)
        else:
            return
    def keras_changed_width_shift(self,on_or_off):
        if on_or_off==0:
            self.lineEdit_widthShift.setText(str(0))
            self.lineEdit_widthShift.setEnabled(False)
        elif on_or_off==2:
            self.lineEdit_widthShift.setText(str(Default_dict ["width_shift"]))
            self.lineEdit_widthShift.setEnabled(True)
        else:
            return
    def keras_changed_height_shift(self,on_or_off):
        if on_or_off==0:
            self.lineEdit_heightShift.setText(str(0))
            self.lineEdit_heightShift.setEnabled(False)
        elif on_or_off==2:
            self.lineEdit_heightShift.setText(str(Default_dict ["height_shift"]))
            self.lineEdit_heightShift.setEnabled(True)
        else:
            return
    def keras_changed_zoom(self,on_or_off):
        if on_or_off==0:
            self.lineEdit_zoomRange.setText(str(0))
            self.lineEdit_zoomRange.setEnabled(False)
        elif on_or_off==2:
            self.lineEdit_zoomRange.setText(str(Default_dict ["zoom"]))
            self.lineEdit_zoomRange.setEnabled(True)
        else:
            return
    def keras_changed_shear(self,on_or_off):
        if on_or_off==0:
            self.lineEdit_shearRange.setText(str(0))
            self.lineEdit_shearRange.setEnabled(False)
        elif on_or_off==2:
            self.lineEdit_shearRange.setText(str(Default_dict ["shear"]))
            self.lineEdit_shearRange.setEnabled(True)
        else:
            return
    def keras_changed_brightplus(self,on_or_off):
        if on_or_off==0:
            self.spinBox_PlusLower.setValue(0)
            self.spinBox_PlusLower.setEnabled(False)
            self.spinBox_PlusUpper.setValue(0)
            self.spinBox_PlusUpper.setEnabled(False)
        elif on_or_off==2:
            self.spinBox_PlusLower.setValue(Default_dict ["Brightness add. lower"])
            self.spinBox_PlusLower.setEnabled(True)
            self.spinBox_PlusUpper.setValue(Default_dict ["Brightness add. upper"])
            self.spinBox_PlusUpper.setEnabled(True)
        else:
            return
    def keras_changed_brightmult(self,on_or_off):
        if on_or_off==0:
            self.doubleSpinBox_MultLower.setValue(1.0)
            self.doubleSpinBox_MultLower.setEnabled(False)
            self.doubleSpinBox_MultUpper.setValue(1.0)
            self.doubleSpinBox_MultUpper.setEnabled(False)
        elif on_or_off==2:
            self.doubleSpinBox_MultLower.setValue(Default_dict ["Brightness mult. lower"])
            self.doubleSpinBox_MultLower.setEnabled(True)
            self.doubleSpinBox_MultUpper.setValue(Default_dict ["Brightness mult. upper"])
            self.doubleSpinBox_MultUpper.setEnabled(True)
        else:
            return
    def keras_changed_noiseMean(self,on_or_off):
        if on_or_off==0:
            self.doubleSpinBox_GaussianNoiseMean.setValue(0.0)
            self.doubleSpinBox_GaussianNoiseMean.setEnabled(False)
        elif on_or_off==2:
            self.doubleSpinBox_GaussianNoiseMean.setValue(Default_dict ["Gaussnoise Mean"])
            self.doubleSpinBox_GaussianNoiseMean.setEnabled(True)
        else:
            return
    def keras_changed_noiseScale(self,on_or_off):
        if on_or_off==0:
            self.doubleSpinBox_GaussianNoiseScale.setValue(0.0)
            self.doubleSpinBox_GaussianNoiseScale.setEnabled(False)
        elif on_or_off==2:
            self.doubleSpinBox_GaussianNoiseScale.setValue(Default_dict ["Gaussnoise Scale"])
            self.doubleSpinBox_GaussianNoiseScale.setEnabled(True)
        else:
            return
    def keras_changed_contrast(self,on_or_off):
        if on_or_off==0:
            self.doubleSpinBox_contrastLower.setEnabled(False)
            self.doubleSpinBox_contrastHigher.setEnabled(False)

        elif on_or_off==2:
            self.doubleSpinBox_contrastLower.setEnabled(True)
            self.doubleSpinBox_contrastHigher.setEnabled(True)
        else:
            return
    def keras_changed_saturation(self,on_or_off):
        if on_or_off==0:
            self.doubleSpinBox_saturationLower.setEnabled(False)
            self.doubleSpinBox_saturationHigher.setEnabled(False)
        elif on_or_off==2:
            self.doubleSpinBox_saturationLower.setEnabled(True)
            self.doubleSpinBox_saturationHigher.setEnabled(True)
        else:
            return
    def keras_changed_hue(self,on_or_off):
        if on_or_off==0:
            self.doubleSpinBox_hueDelta.setEnabled(False)
        elif on_or_off==2:
            self.doubleSpinBox_hueDelta.setEnabled(True)
        else:
            return

    def expert_mode_off(self,on_or_off):
        """
        Reset all values on the expert tab to the default values, excluding the metrics
        metrics are defined only once when starting fitting and should not be changed
        """
        if on_or_off==0: #switch off
            self.spinBox_batchSize.setValue(Default_dict["spinBox_batchSize"])
            self.spinBox_epochs.setValue(1)
            self.checkBox_expt_loss.setChecked(False)
            self.expert_loss_off(0)
            self.groupBox_learningRate.setChecked(False)        
            self.expert_learningrate_off(0)
            self.checkBox_optimizer.setChecked(False)
            self.expert_optimizer_off(0)
            
    def expert_loss_off(self,on_or_off):
        if on_or_off==0: #switch off
            #switch back to categorical_crossentropy 
            index = self.comboBox_expt_loss.findText("categorical_crossentropy", QtCore.Qt.MatchFixedString)
            if index >= 0:
                self.comboBox_expt_loss.setCurrentIndex(index)
    
    def expert_learningrate_off(self,on_or_off):
        if on_or_off==0: #switch off
            #which optimizer is used? (there are different default learning-rates
            #for each optimizer!)
            optimizer = str(self.comboBox_optimizer.currentText())
            self.doubleSpinBox_learningRate.setValue(Default_dict["doubleSpinBox_learningRate_"+optimizer])
            self.radioButton_LrCycl.setChecked(False)
            self.radioButton_LrExpo.setChecked(False)
            self.radioButton_LrConst.setChecked(True)
    
    def expert_optimizer_off(self,on_or_off):
        if on_or_off==0: #switch off, set back to categorical_crossentropy
            optimizer = "Adam"
            index = self.comboBox_optimizer.findText(optimizer, QtCore.Qt.MatchFixedString)
            if index >= 0:
                self.comboBox_optimizer.setCurrentIndex(index)
                #also reset the learning rate to the default
                self.doubleSpinBox_learningRate.setValue(Default_dict["doubleSpinBox_learningRate_"+optimizer])
    

    def expert_optimizer_changed(self,optimizer_text,listindex):
#        print("optimizer_text: "+str(optimizer_text))
#        print("listindex: "+str(listindex))
        
        if optimizer_text=="":
            return
        if listindex==-1:
            item_ui = self
        else:
            item_ui = self.fittingpopups_ui[listindex]
        #set the learning rate to the default for this optimizer
        value_current = float(item_ui.doubleSpinBox_learningRate.value())
        value_wanted = Default_dict["doubleSpinBox_learningRate_"+optimizer_text]
        
        #insert the current value in the optimizer_settings:
        item_ui.optimizer_settings["doubleSpinBox_lr_"+optimizer_text.lower()] = value_current    
        item_ui.optimizer_settings["comboBox_optimizer"] = optimizer_text    

        try: #only works on the fitting-popup
            text = str(item_ui.textBrowser_FittingInfo.toPlainText())
        except:
            text = "Epoch"
#        print("text: "+str(text))

        if value_current!=value_wanted and "Epoch" in text:#avoid that the message pops up when window is created
            item_ui.doubleSpinBox_learningRate.setValue(value_wanted)
            item_ui.doubleSpinBox_expDecInitLr.setValue(value_wanted)

            #Inform user
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setWindowTitle("Learning rate to default")
            msg.setText("Learning rate was set to the default for "+optimizer_text)
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return

    def expert_lr_changed(self,value,optimizer_text,listindex):

        if listindex==-1:
            item_ui = self
        else:
            item_ui = self.fittingpopups_ui[listindex]
        item_ui.optimizer_settings["doubleSpinBox_lr_"+optimizer_text.lower()] = value
        
    def update_hist1(self):
        feature = str(self.comboBox_feat1.currentText())
        feature_values = self.rtdc_ds["events"][feature]
        #if len(feature_values)==len(self.rtdc_ds['area_cvx']):
#        self.histogram = pg.GraphicsWindow()        
        #plt1 = self.histogram.addPlot()
        y,x = np.histogram(feature_values, bins='auto')
        self.plt1.plot(x, y, stepMode=True, fillLevel=0, brush=(0,0,255,150),clear=True)
#        self.gridLayout_w2.addWidget(self.histogram,1, 0, 1, 1)        
#        self.w.show()
    def update_hist2(self):
        feature = str(self.comboBox_feat2.currentText())
        feature_values = self.rtdc_ds["events"][feature]
        #if len(feature_values)==len(self.rtdc_ds['area_cvx']):
        #self.histogram = pg.GraphicsWindow()        
        #plt1 = self.histogram.addPlot()
        y,x = np.histogram(feature_values, bins='auto')
        self.plt1.plot(x, y, stepMode=True, fillLevel=0, brush=(0,0,255,150),clear=True)
#        self.gridLayout_w2.addWidget(self.histogram,1, 0, 1, 1)        
#        self.w.show()
    def update_scatter(self):
        feature_x = str(self.comboBox_feat1.currentText())
        feature_x_values = self.rtdc_ds["events"][feature_x]
        feature_y = str(self.comboBox_feat2.currentText())
        feature_y_values = self.rtdc_ds["events"][feature_y]
        if len(feature_x_values)==len(feature_y_values):
            #self.histogram = pg.GraphicsWindow()        
            #plt1 = self.histogram.addPlot()
            #y,x = np.histogram(feature_values, bins='auto')
            self.plt1.plot(feature_x_values, feature_y_values,pen=None,symbol='o',clear=True)
#            self.gridLayout_w2.addWidget(self.histogram,1, 0, 1, 1)        
#            self.w.show()

    def button_hist(self,item):
        buttonClicked = self.sender()
        index = self.table_dragdrop.indexAt(buttonClicked.pos())
        rowPosition = index.row()
        rtdc_path = self.table_dragdrop.item(rowPosition, 0).text()
        rtdc_path = str(rtdc_path)

        failed,rtdc_ds = aid_bin.load_rtdc(rtdc_path)
        if failed:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Critical)       
            msg.setText(str(rtdc_ds))
            msg.setWindowTitle("Error occurred during loading file")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return
                    
        self.rtdc_ds = rtdc_ds
#        feature_values = rtdc_ds[feature]
        #Init a popup window
        self.w = MyPopup()
        self.w.setWindowTitle(rtdc_path)
        self.w.setObjectName(_fromUtf8("w"))
        self.gridLayout_w2 = QtWidgets.QGridLayout(self.w)
        self.gridLayout_w2.setContentsMargins(0, 0, 0, 0)

        self.gridLayout_w2.setObjectName(_fromUtf8("gridLayout_w2"))
        self.widget = QtWidgets.QWidget(self.w)
        self.widget.setMinimumSize(QtCore.QSize(0, 65))
        self.widget.setMaximumSize(QtCore.QSize(16777215, 65))
        self.widget.setObjectName(_fromUtf8("widget"))
        self.horizontalLayout_w3 = QtWidgets.QHBoxLayout(self.widget)        
        self.horizontalLayout_w3.setContentsMargins(0, 0, 0, 0)
        self.horizontalLayout_w3.setObjectName(_fromUtf8("horizontalLayout_w3"))
        self.verticalLayout_w = QtWidgets.QVBoxLayout()
        self.verticalLayout_w.setObjectName(_fromUtf8("verticalLayout_w"))
        self.horizontalLayout_w = QtWidgets.QHBoxLayout()
        self.horizontalLayout_w.setObjectName(_fromUtf8("horizontalLayout_w"))
        self.comboBox_feat1 = QtWidgets.QComboBox(self.widget)
        self.comboBox_feat1.setObjectName(_fromUtf8("comboBox_feat1"))
        features = list(self.rtdc_ds["events"].keys())
        self.comboBox_feat1.addItems(features)
        self.horizontalLayout_w.addWidget(self.comboBox_feat1)
        self.comboBox_feat2 = QtWidgets.QComboBox(self.widget)
        self.comboBox_feat2.setObjectName(_fromUtf8("comboBox_feat2"))
        self.comboBox_feat2.addItems(features)
        self.horizontalLayout_w.addWidget(self.comboBox_feat2)
        self.verticalLayout_w.addLayout(self.horizontalLayout_w)
        self.horizontalLayout_w2 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_w2.setObjectName(_fromUtf8("horizontalLayout_w2"))
        self.pushButton_Hist1 = QtWidgets.QPushButton(self.widget)
        self.pushButton_Hist1.setObjectName(_fromUtf8("pushButton_Hist1"))
        self.horizontalLayout_w2.addWidget(self.pushButton_Hist1)
        self.pushButton_Hist2 = QtWidgets.QPushButton(self.widget)
        self.pushButton_Hist2.setObjectName(_fromUtf8("pushButton_Hist2"))
        self.horizontalLayout_w2.addWidget(self.pushButton_Hist2)
        self.verticalLayout_w.addLayout(self.horizontalLayout_w2)
        self.horizontalLayout_w3.addLayout(self.verticalLayout_w)
        self.verticalLayout_w2 = QtWidgets.QVBoxLayout()
        self.verticalLayout_w2.setObjectName(_fromUtf8("verticalLayout_w2"))
        self.pushButton_Scatter = QtWidgets.QPushButton(self.widget)
        self.pushButton_Scatter.setObjectName(_fromUtf8("pushButton_Scatter"))
        self.verticalLayout_w2.addWidget(self.pushButton_Scatter)
        self.checkBox_ScalePix = QtWidgets.QCheckBox(self.widget)
        self.checkBox_ScalePix.setObjectName(_fromUtf8("checkBox_ScalePix"))
        self.verticalLayout_w2.addWidget(self.checkBox_ScalePix)
        self.horizontalLayout_w3.addLayout(self.verticalLayout_w2)
        self.gridLayout_w2.addWidget(self.widget, 0, 0, 1, 1)
      
        self.pushButton_Hist1.setText("Hist")
        self.pushButton_Hist1.clicked.connect(self.update_hist1)
        self.pushButton_Hist2.setText("Hist")
        self.pushButton_Hist2.clicked.connect(self.update_hist2)
        self.pushButton_Scatter.setText("Scatter")
        self.pushButton_Scatter.clicked.connect(self.update_scatter)

        self.checkBox_ScalePix.setText("Scale by pix")

        self.histogram = pg.GraphicsWindow()        
        self.plt1 = self.histogram.addPlot()
#        y,x = np.histogram(feature_values, bins='auto')
#        plt1.plot(x, y, stepMode=True, fillLevel=0, brush=(0,0,255,150))
        self.gridLayout_w2.addWidget(self.histogram,1, 0, 1, 1)        
        self.w.show()

    def update_historyplot_pop(self,listindex):
        #listindex = self.popupcounter-1 #len(self.fittingpopups_ui)-1
        #After the first epoch there are checkboxes available. Check, if user checked some:
        colcount = int(self.fittingpopups_ui[listindex].tableWidget_HistoryInfo_pop.columnCount())
        #Collect items that are checked
        selected_items,Colors = [],[]
        for colposition in range(colcount):  
            #is it checked for train?
            cb = self.fittingpopups_ui[listindex].tableWidget_HistoryInfo_pop.item(0, colposition)
            if not cb==None:
                if cb.checkState() == QtCore.Qt.Checked:
                    selected_items.append(str(cb.text()))
                    Colors.append(cb.background())
        self.Colors = Colors
        Histories = self.fittingpopups_ui[listindex].Histories
        DF1 = [[ h[h_i][-1] for h_i in h] for h in Histories] #if nb_epoch in .fit() is >1, only save the last history item, beacuse this would a model that could be saved
        DF1 = np.r_[DF1]
        DF1 = pd.DataFrame( DF1,columns=Histories[0].keys() )
#        if len(DF1)>0:
#            DF1 = pd.concat(DF1)
#        else:
#            return
        self.fittingpopups_ui[listindex].widget_pop.clear()
        
        #Create fresh plot
        plt1 = self.fittingpopups_ui[listindex].widget_pop.addPlot()
        plt1.showGrid(x=True,y=True)
        plt1.addLegend()
        plt1.setLabel('bottom', 'Epoch', units='')
        #Create a dict that stores plots for each metric (for real time plotting)
        self.fittingpopups_ui[listindex].historyscatters = dict()
        for i in range(len(selected_items)):
            key = selected_items[i]
            df = DF1[key]
            color = self.Colors[i]
            pen_rollmedi = list(color.color().getRgb())
            pen_rollmedi = pg.mkColor(pen_rollmedi)
            pen_rollmedi = pg.mkPen(color=pen_rollmedi,width=6)
            color = list(color.color().getRgb())
            color[-1] = int(0.6*color[-1])
            color = tuple(color)                
            pencolor = pg.mkColor(color)
            brush = pg.mkBrush(color=pencolor)
            #print(df)
            
            historyscatter = plt1.plot(range(len(df)), df.values, pen=None,symbol='o',symbolPen=None,symbolBrush=brush,name=key,clear=False)
            #self.fittingpopups_ui[listindex].historyscatters.append(historyscatter)
            self.fittingpopups_ui[listindex].historyscatters[key]=historyscatter


    def stop_fitting_pop(self,listindex):
        #listindex = len(self.fittingpopups_ui)-1
        epochs = self.fittingpopups_ui[listindex].epoch_counter                            
        #Stop button on the fititng popup
        #Should stop the fitting process and save the metafile
        #1. Change the nr. requested epochs to a smaller number
        self.fittingpopups_ui[listindex].spinBox_NrEpochs.setValue(epochs-1)
        #2. Check the box which will cause that the new parameters are applied at next epoch
        self.fittingpopups_ui[listindex].checkBox_ApplyNextEpoch.setChecked(True)


    def pause_fitting_pop(self,listindex):
        #Just change the text on the button
        if str(self.fittingpopups_ui[listindex].pushButton_Pause_pop.text())==" ":
            #If the the text on the button was Pause, change it to Continue
            self.fittingpopups_ui[listindex].pushButton_Pause_pop.setText("")
            self.fittingpopups_ui[listindex].pushButton_Pause_pop.setStyleSheet("background-color: green")
            self.fittingpopups_ui[listindex].pushButton_Pause_pop.setIcon(QtGui.QIcon(os.path.join(dir_root,"art",Default_dict["Icon theme"],"continue.png")))

        elif str(self.fittingpopups_ui[listindex].pushButton_Pause_pop.text())=="":
            #If the the text on the button was Continue, change it to Pause
            self.fittingpopups_ui[listindex].pushButton_Pause_pop.setText(" ")
            self.fittingpopups_ui[listindex].pushButton_Pause_pop.setIcon(QtGui.QIcon(os.path.join(dir_root,"art",Default_dict["Icon theme"],"pause.png")))
            self.fittingpopups_ui[listindex].pushButton_Pause_pop.setStyleSheet("")



    def saveTextWindow_pop(self,listindex):
        #Get the entire content of textBrowser_FittingInfo
        text = str(self.fittingpopups_ui[listindex].textBrowser_FittingInfo.toPlainText())
        #Ask the user where to save the stuff
        filename = QtWidgets.QFileDialog.getSaveFileName(self, 'Fitting info', Default_dict["Path of last model"]," (*.txt)")
        filename = filename[0]
        #Save to this filename
        if len(filename)>0:
            f = open(filename,'w')
            f.write(text)
            f.close()                

    def clearTextWindow_pop(self,listindex):
        self.fittingpopups_ui[listindex].textBrowser_FittingInfo.clear()
        
    def showModelSumm_pop(self,listindex):
        text5 = "Model summary:\n"
        summary = []
        self.model_keras.summary(print_fn=summary.append)
        summary = "\n".join(summary)
        text = text5+summary
        self.fittingpopups_ui[listindex].textBrowser_FittingInfo.append(text)

    def saveModelSumm_pop(self,listindex):
        text5 = "Model summary:\n"
        summary = []
        self.model_keras.summary(print_fn=summary.append)
        summary = "\n".join(summary)
        text = text5+summary
        #Ask the user where to save the stuff
        filename = QtWidgets.QFileDialog.getSaveFileName(self, 'Model summary', Default_dict["Path of last model"]," (*.txt)")
        filename = filename[0]
        #Save to this filename
        f = open(filename,'w')
        f.write(text)
        f.close()                

    #class_weight = self.get_class_weight(self.fittingpopups_ui[listindex].SelectedFiles,lossW_expert) #
    def get_class_weight(self,SelectedFiles,lossW_expert,custom_check_classes=False):
        t1 = time.perf_counter()
        print("Getting dictionary for class_weight")
        if lossW_expert=="None":
            return None
        elif lossW_expert=="":
            return None
        elif lossW_expert=="Balanced":
            #Which are training files?
            ind = [selectedfile["TrainOrValid"] == "Train" for selectedfile in SelectedFiles]
            ind = np.where(np.array(ind)==True)[0]
            SelectedFiles_train = list(np.array(SelectedFiles)[ind])
            classes = [int(selectedfile["class"]) for selectedfile in SelectedFiles_train]
            nr_events_epoch = [int(selectedfile["nr_events_epoch"]) for selectedfile in SelectedFiles_train]
            classes_uni = np.unique(classes)
            counter = {}
            for class_ in classes_uni:
                ind = np.where(np.array(classes)==class_)[0]
                nr_events_epoch_class = np.array(nr_events_epoch)[ind]
                counter[class_] = np.sum(nr_events_epoch_class)
            max_val = float(max(counter.values()))
            return {class_id : max_val/num_images for class_id, num_images in counter.items()}
            
        elif lossW_expert.startswith("{"):#Custom loss weights
            class_weights = eval(lossW_expert)
            if custom_check_classes:#Check that each element in classes_uni is contained in class_weights.keys()
                ind = [selectedfile["TrainOrValid"] == "Train" for selectedfile in SelectedFiles]
                ind = np.where(np.array(ind)==True)[0]
                SelectedFiles_train = list(np.array(SelectedFiles)[ind])
                classes = [int(selectedfile["class"]) for selectedfile in SelectedFiles_train]
                classes_uni = np.unique(classes)
                classes_uni = np.sort(classes_uni)
                class_weights_keys = np.sort([int(a) for a in class_weights.keys()])
                #each element in classes_uni has to be equal to class_weights_keys
                equal = np.array_equal(classes_uni,class_weights_keys)
                if equal == True:
                    return class_weights
                else:    
                    #If the equal is false I'm really in trouble...
                    #run the function again, but request 'Balanced' weights. I'm not sure if this should be the default...
                    class_weights = self.get_class_weight(SelectedFiles,"Balanced")
                    return ["Balanced",class_weights]
            else:
                return class_weights
        t2 = time.perf_counter()
        dt = np.round(t2-t1,2)
        print("Comp. time = "+str(dt))


    def accept_lr_range(self):
        lr_start = str(self.popup_lrfinder_ui.lineEdit_LrMin.text())
        lr_stop = str(self.popup_lrfinder_ui.lineEdit_LrMax.text())
        if len(lr_start)>0 and len(lr_stop)>0:
            self.lineEdit_cycLrMin.setText(lr_start)
            self.lineEdit_cycLrMax.setText(lr_stop)
        else:
            print("Found no values for LR range")
        
    def accept_lr_value(self):
        single_lr = self.popup_lrfinder_ui.lineEdit_singleLr.text()
        if len(single_lr)>0:
            lr_value = float(single_lr)
            self.doubleSpinBox_learningRate.setValue(lr_value)
            self.doubleSpinBox_expDecInitLr.setValue(lr_value)
        else:
            print("Found no value for single LR!")
    
    def reset_lr_settings(self):
        self.popup_lrfinder_ui.lineEdit_startLr.setText(_translate("Form_LrFinder", "1e-10", None))
        self.popup_lrfinder_ui.lineEdit_stopLr.setText(_translate("Form_LrFinder", "0.1", None))
        self.popup_lrfinder_ui.doubleSpinBox_percDataT.setProperty("value", 100.0)
        self.popup_lrfinder_ui.doubleSpinBox_percDataV.setProperty("value", 100.0)
        self.popup_lrfinder_ui.spinBox_batchSize.setValue(Default_dict["spinBox_batchSize"])       
        self.popup_lrfinder_ui.spinBox_lineWidth.setProperty("value", 6)
        self.popup_lrfinder_ui.spinBox_epochs.setProperty("value", 5)
        
    def reset_lr_value(self):
        self.popup_lrfinder_ui.lineEdit_singleLr.setText("")
        #Uncheck and Check the groupbox to refresh the line
        self.popup_lrfinder_ui.groupBox_singleLr.setChecked(False)
        self.popup_lrfinder_ui.groupBox_singleLr.setChecked(True)

    def reset_lr_range(self):
        self.popup_lrfinder_ui.lineEdit_LrMin.setText("")
        self.popup_lrfinder_ui.lineEdit_LrMax.setText("")
        #Uncheck and Check the groupbox to refresh the range
        self.popup_lrfinder_ui.groupBox_LrRange.setChecked(False)
        self.popup_lrfinder_ui.groupBox_LrRange.setChecked(True)

          
    def popup_lr_finder(self):
        SelectedFiles = self.items_clicked()
            
        self.popup_lrfinder = MyPopup()
        self.popup_lrfinder_ui = aid_frontend.popup_lrfinder()
        self.popup_lrfinder_ui.setupUi(self.popup_lrfinder) #open a popup for lr finder

        #Get information about the model
        #check, which radiobutton is clicked and just copy paste the text from there
        if self.radioButton_NewModel.isChecked():
            modelname = str(self.comboBox_ModelSelection.currentText())
            if modelname==None:
                msg = QtWidgets.QMessageBox()
                msg.setIcon(QtWidgets.QMessageBox.Information)       
                msg.setText("No model specified!")
                msg.setWindowTitle("No model specified!")
                msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                msg.exec_()
                return       
        elif self.radioButton_LoadContinueModel.isChecked():
            modelname = str(self.lineEdit_LoadModelPath.text())
        elif self.radioButton_LoadRestartModel.isChecked():
            modelname = str(self.lineEdit_LoadModelPath.text())
        else:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("Please specify a model using the radiobuttons on the 'Define Model' -tab")
            msg.setWindowTitle("No model specified!")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return 
        
        in_dim = int(self.spinBox_imagecrop.value())

        #Put information onto UI
        self.popup_lrfinder_ui.lineEdit_loadModel.setText(modelname)
        self.popup_lrfinder_ui.spinBox_Crop_inpImgSize.setValue(in_dim)
        color_mode = self.get_color_mode()
        self.popup_lrfinder_ui.comboBox_colorMode.addItem(color_mode)
        loss_str = str(self.comboBox_expt_loss.currentText())
        self.popup_lrfinder_ui.comboBox_expt_loss.addItem(loss_str)
        optimizer_str = str(self.comboBox_optimizer.currentText())
        self.popup_lrfinder_ui.comboBox_optimizer.addItem(optimizer_str)

        batch_size = self.spinBox_batchSize.value()
        self.popup_lrfinder_ui.spinBox_batchSize.setValue(batch_size)

        #Connect action_lr_finder function to button
        self.popup_lrfinder_ui.pushButton_LrFindRun.clicked.connect(lambda: self.action_initialize_model(duties="initialize_lrfind"))
        self.popup_lrfinder_ui.pushButton_rangeAccept.clicked.connect(self.accept_lr_range)
        self.popup_lrfinder_ui.pushButton_singleAccept.clicked.connect(self.accept_lr_value)
        self.popup_lrfinder_ui.pushButton_LrReset.clicked.connect(self.reset_lr_settings)
        self.popup_lrfinder_ui.pushButton_singleReset.clicked.connect(self.reset_lr_value)
        self.popup_lrfinder_ui.pushButton_rangeReset.clicked.connect(self.reset_lr_range)
        #Update the plot when any plotting option is changed
        self.popup_lrfinder_ui.comboBox_metric.currentIndexChanged.connect(self.update_lrfind_plot)
        self.popup_lrfinder_ui.spinBox_lineWidth.valueChanged.connect(self.update_lrfind_plot)
        self.popup_lrfinder_ui.checkBox_smooth.toggled.connect(self.update_lrfind_plot)
        
        #LR single value when groupbox is toggled
        self.popup_lrfinder_ui.groupBox_singleLr.toggled.connect(self.get_lr_single)
        #LR range when groupbox is toggled
        self.popup_lrfinder_ui.groupBox_LrRange.toggled.connect(self.get_lr_range)

        #compute the number of steps/epoch
        ind = [selectedfile["TrainOrValid"] == "Train" for selectedfile in SelectedFiles]
        ind = np.where(np.array(ind)==True)[0]
        SelectedFiles_train = np.array(SelectedFiles)[ind]
        SelectedFiles_train = list(SelectedFiles_train)
        nr_events_train_total = np.sum([int(selectedfile["nr_events_epoch"]) for selectedfile in SelectedFiles_train])

        def update_stepsPerEpoch():
            batch_size = self.popup_lrfinder_ui.spinBox_batchSize.value()
            perc_data = self.popup_lrfinder_ui.doubleSpinBox_percDataT.value()
            nr_events = (perc_data/100)*nr_events_train_total
            stepsPerEpoch = np.ceil(nr_events / float(batch_size))
            self.popup_lrfinder_ui.spinBox_stepsPerEpoch.setValue(stepsPerEpoch)

        update_stepsPerEpoch()
        self.popup_lrfinder_ui.spinBox_batchSize.valueChanged.connect(update_stepsPerEpoch)
        self.popup_lrfinder_ui.doubleSpinBox_percDataT.valueChanged.connect(update_stepsPerEpoch)

        self.popup_lrfinder.show()

    def popup_clr_settings(self,listindex):
        if listindex==-1:
            item_ui = self
        else:
            item_ui = self.fittingpopups_ui[listindex]
            
        item_ui.popup_clrsettings = MyPopup()
        item_ui.popup_clrsettings_ui = aid_frontend.Ui_Clr_settings()
        item_ui.popup_clrsettings_ui.setupUi(item_ui.popup_clrsettings) #open a popup for lr plotting

        ##Manual insertion##        
        item_ui.popup_clrsettings_ui.spinBox_stepSize.setProperty("value", item_ui.clr_settings["step_size"])
        item_ui.popup_clrsettings_ui.doubleSpinBox_gamma.setProperty("value", item_ui.clr_settings["gamma"])

        def clr_settings_ok():
            step_size = int(item_ui.popup_clrsettings_ui.spinBox_stepSize.value())
            gamma = float(item_ui.popup_clrsettings_ui.doubleSpinBox_gamma.value())
            item_ui.clr_settings["step_size"] = step_size #Number of epochs to fulfill half a cycle
            item_ui.clr_settings["gamma"] = gamma #gamma factor for Exponential decrease method (exp_range)
            print("Settings for cyclical learning rates were changed.")
            #close the popup
            item_ui.popup_clrsettings = None
            item_ui.popup_clrsettings_ui = None

        def clr_settings_cancel():#close the popup
            item_ui.popup_clrsettings = None
            item_ui.popup_clrsettings_ui = None

        item_ui.popup_clrsettings_ui.pushButton_ok.clicked.connect(clr_settings_ok)
        item_ui.popup_clrsettings_ui.pushButton_cancel.clicked.connect(clr_settings_cancel)
        
        item_ui.popup_clrsettings.show()
        
        
    def popup_lr_plot(self,listindex):
        if listindex==-1:
            item_ui = self
        else:
            item_ui = self.fittingpopups_ui[listindex]

        item_ui.popup_lrplot = MyPopup()
        item_ui.popup_lrplot_ui = aid_frontend.popup_lrplot()
        item_ui.popup_lrplot_ui.setupUi(item_ui.popup_lrplot) #open a popup for lr plotting
        
        #compute total number of epochs that will be fitted
        spinBox_NrEpochs = item_ui.spinBox_NrEpochs.value() #my own loop
        spinBox_epochs = item_ui.spinBox_epochs.value() #inside model.fit()
        nr_epochs = spinBox_NrEpochs*spinBox_epochs
        item_ui.popup_lrplot_ui.spinBox_totalEpochs.setValue(nr_epochs)
        
        #Get the number of training examples
        SelectedFiles = self.items_clicked()
        ind = [selectedfile["TrainOrValid"] == "Train" for selectedfile in SelectedFiles]
        ind = np.where(np.array(ind)==True)[0]
        SelectedFiles_train = np.array(SelectedFiles)[ind]
        SelectedFiles_train = list(SelectedFiles_train)
        nr_events_train_total = np.sum([int(selectedfile["nr_events_epoch"]) for selectedfile in SelectedFiles_train])
        if nr_events_train_total==0 and item_ui.radioButton_LrConst.isChecked()==False:
            #for Cyclical learning rates and Exponential learning rates, the 
            #number of training images is needed
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("There is no training data. Nr. of training images is required for this plot.")
            msg.setWindowTitle("Nr. of training images = 0")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return 
        
        text_info = ""
        if item_ui.radioButton_LrConst.isChecked():
            text_info+="Constant learning rate\n"
            epochs_plot = np.array(range(nr_epochs))
            const_lr = float(self.doubleSpinBox_learningRate.value())
            learningrates = np.repeat(const_lr,nr_epochs)
            
        elif item_ui.radioButton_LrCycl.isChecked():
            text_info+="Cyclical learning rates\n"
            base_lr = float(item_ui.lineEdit_cycLrMin.text())
            max_lr = float(item_ui.lineEdit_cycLrMax.text())
            batch_size = int(item_ui.spinBox_batchSize.value())
            step_size = item_ui.clr_settings["step_size"] #batch updates in a half cycle
            step_size_ = step_size*int(np.round(nr_events_train_total / batch_size))#number of steps in one epoch
            mode = str(item_ui.comboBox_cycLrMethod.currentText())
            clr_iterations = nr_epochs*int(np.round(nr_events_train_total / batch_size))#number of cycles
            nr_cycles = (clr_iterations/step_size_)/2.0#number of cycles
            gamma = item_ui.clr_settings["gamma"] #gamma factor for the exp_range
            
            #Generate text to diplay the settings used
            text_info+="Nr. of training images: "+str(nr_events_train_total)+"\n"
            text_info+="base_lr: "+str(base_lr)+"\n"
            text_info+="max_lr: "+str(max_lr)+"\n"
            text_info+="batch_size: "+str(batch_size)+"\n"
            text_info+="mode: "+str(mode)+"\n"
            text_info+="gamma: "+str(gamma)+"\n"

            text_info+="Nr. of epochs to fulfill one cycle: "+str(2*step_size)+"\n"
            #text_info+="Total nr. of lr adjustmend: "+str(step_size_)+"\n"
            text_info+="Total nr. of lr adjustments: "+str(clr_iterations)+"\n"
            text_info+="Total nr. of cycles: "+str(nr_cycles)+"\n"
            
            #Request the learning rates from the class cyclicLR
            clr_iterations = np.arange(clr_iterations) 
            clr_1 = aid_dl.cyclicLR(base_lr=base_lr,max_lr=max_lr,step_size=step_size_,mode=mode,gamma=gamma)
            clr_1.clr_iterations=clr_iterations#pass the number of clr iterations to the class
            
            learningrates = clr_1.clr() #compute the learning rates for each iteration
            #convert clr_iterations back to "epochs"
            epochs_plot = clr_iterations/int(np.round(nr_events_train_total / batch_size))
          
        
        elif item_ui.radioButton_LrExpo.isChecked():
            text_info+="Exponentially decreased learning rates\n"
            initial_lr = float(item_ui.doubleSpinBox_expDecInitLr.value())
            decay_steps = int(item_ui.spinBox_expDecSteps.value())
            decay_rate = float(item_ui.doubleSpinBox_expDecRate.value())
            batch_size = int(item_ui.spinBox_batchSize.value())
            text_info+="Nr. of training images: "+str(nr_events_train_total)+"\n"
            text_info+="initial_lr: "+str(initial_lr)+"\n"
            text_info+="decay_steps: "+str(decay_steps)+"\n"
            text_info+="decay_rate: "+str(decay_rate)+"\n"
        
            #epochs_plot = np.array(range(nr_epochs))
            epochs_plot = nr_epochs * int(np.round(nr_events_train_total / batch_size))
            epochs_plot = np.arange(epochs_plot)
            exp_decay = aid_dl.exponentialDecay(initial_lr=initial_lr, decay_steps=decay_steps, decay_rate=decay_rate)
            exp_decay.iterations=epochs_plot#pass the number of clr iterations to the class
            learningrates = exp_decay.exp_decay()
            epochs_plot = epochs_plot/int(np.round(nr_events_train_total / batch_size))
            #learningrates = aid_dl.exponentialDecay(epochs_plot,initial_lr=initial_lr, decay_steps=decay_steps, decay_rate=decay_rate)
        
        
        def refreshPlot():
            try: # try to empty the plot
                item_ui.popup_lrplot_ui.lr_plot.removeItem(item_ui.lr_line2)   
            except:
                pass
            #Get design settings
            color = item_ui.popup_lrplot_ui.pushButton_color.palette().button().color()
            width = int(item_ui.popup_lrplot_ui.spinBox_lineWidth.value())
            color = list(color.getRgb())
            color = tuple(color)                
            pencolor=pg.mkPen(color, width=width)
            #define curve and add to plot  
            item_ui.lr_line2 = pg.PlotCurveItem(x=epochs_plot, y=learningrates,pen=pencolor)
            item_ui.popup_lrplot_ui.lr_plot.addItem(item_ui.lr_line2)            

        refreshPlot()
        item_ui.popup_lrplot_ui.pushButton_refreshPlot.clicked.connect(refreshPlot)
        item_ui.popup_lrplot_ui.textBrowser_lrSettings.setText(text_info)
        item_ui.popup_lrplot.show()
        
        
    def lossWeights_activated(self,on_or_off,listindex):
        if listindex==-1:
            item_ui = self
        else:
            item_ui = self.fittingpopups_ui[listindex]
        
        if on_or_off==False:#0 means switched OFF
            item_ui.lineEdit_lossW.setText("")
            item_ui.pushButton_lossW.setEnabled(False)

        #this happens when the user activated the expert option "loss weights"
        elif on_or_off==True:#2 means switched ON
            #Activate button
            item_ui.pushButton_lossW.setEnabled(True)
            self.lossWeights_popup(listindex)

        
    def lossWeights_popup(self,listindex):
        if listindex==-1:
            item_ui = self
            SelectedFiles = self.items_clicked()
        else:
            item_ui = self.fittingpopups_ui[listindex]
            SelectedFiles = item_ui.SelectedFiles
            
        item_ui.popup_lossW = MyPopup()
        item_ui.popup_lossW_ui = aid_frontend.popup_lossweights()
        item_ui.popup_lossW_ui.setupUi(item_ui.popup_lossW) #open a popup to show the numbers of events in each class in a table
        
        SelectedFiles = [x for x in SelectedFiles if x["TrainOrValid"]=="Train"]
        indices = [SelectedFiles[i]["class"] for i in range(len(SelectedFiles))]

        #Initiate the table with 4 columns : this will be ["Index","Nr of cells","Clr","Name"]
        item_ui.popup_lossW_ui.tableWidget_lossW.setColumnCount(5)
        nr_ind = len(set(indices)) #each index could occur for train and valid
        nr_rows = nr_ind
        item_ui.popup_lossW_ui.tableWidget_lossW.setRowCount(nr_rows)
        #Wich selected file has the most features?
        header_labels = ["Class", "Events tot." ,"Events/Epoch", "Events/Epoch[%]", "Loss weight"]
        item_ui.popup_lossW_ui.tableWidget_lossW.setHorizontalHeaderLabels(header_labels) 
        header = item_ui.popup_lossW_ui.tableWidget_lossW.horizontalHeader()
        for i in range(len(header_labels)):
            header.setResizeMode(i, QtWidgets.QHeaderView.ResizeToContents)        

        #Fill the table 
        rowPosition = 0      
        #Training info
        ind = [selectedfile["TrainOrValid"] == "Train" for selectedfile in SelectedFiles]
        ind = np.where(np.array(ind)==True)[0]
        SelectedFiles_train = np.array(SelectedFiles)[ind]
        SelectedFiles_train = list(SelectedFiles_train)
        indices_train = [selectedfile["class"] for selectedfile in SelectedFiles_train]
        nr_events_train_total = np.sum([int(selectedfile["nr_events_epoch"]) for selectedfile in SelectedFiles_train])

        #Total nr of cells for each index
        for index in np.unique(indices_train):
            colPos = 0 #"Class" #put the index (class!) in column nr. 0
            item = QtWidgets.QTableWidgetItem()
            item.setFlags(item.flags() &~QtCore.Qt.ItemIsEnabled &~ QtCore.Qt.ItemIsSelectable )
            item.setData(QtCore.Qt.EditRole,str(index))
            item_ui.popup_lossW_ui.tableWidget_lossW.setItem(rowPosition, colPos, item)
            
            #Get the training files of that index
            ind = np.where(indices_train==index)[0]
            SelectedFiles_train_index = np.array(SelectedFiles_train)[ind]
    
            colPos = 1 #"Events tot."
            nr_events = [int(selectedfile["nr_events"]) for selectedfile in SelectedFiles_train_index]
            item = QtWidgets.QTableWidgetItem()
            item.setFlags(item.flags() &~QtCore.Qt.ItemIsEnabled &~ QtCore.Qt.ItemIsSelectable )
            item.setData(QtCore.Qt.EditRole, str(np.sum(nr_events)))
            item_ui.popup_lossW_ui.tableWidget_lossW.setItem(rowPosition, colPos, item)

            colPos = 2 #"Events/Epoch"
            nr_events_epoch = [int(selectedfile["nr_events_epoch"]) for selectedfile in SelectedFiles_train_index]
            item = QtWidgets.QTableWidgetItem()
            item.setFlags(item.flags() &~QtCore.Qt.ItemIsEnabled &~ QtCore.Qt.ItemIsSelectable )
            item.setData(QtCore.Qt.EditRole, str(np.sum(nr_events_epoch)))
            item_ui.popup_lossW_ui.tableWidget_lossW.setItem(rowPosition, colPos, item)
            
            colPos = 3 #"Events/Epoch[%]"
            item = QtWidgets.QTableWidgetItem()
            item.setFlags(item.flags() &~QtCore.Qt.ItemIsEnabled &~ QtCore.Qt.ItemIsSelectable )
            item.setData(QtCore.Qt.EditRole, str(np.round(np.sum(nr_events_epoch)/float(nr_events_train_total),2)))
            item_ui.popup_lossW_ui.tableWidget_lossW.setItem(rowPosition, colPos, item)

            colPos = 4 #"Loss weights"
            #for each item create a spinbopx (trainability)
            spinb = QtWidgets.QDoubleSpinBox(item_ui.popup_lossW_ui.tableWidget_lossW)
            spinb.setEnabled(False)
            spinb.setMinimum(-99999)
            spinb.setMaximum(99999)
            spinb.setSingleStep(0.1)
            spinb.setValue(1.0) #Default in Keras is "None", which means class_weight=1.0
            item_ui.popup_lossW_ui.tableWidget_lossW.setCellWidget(rowPosition, colPos, spinb)            
            
            rowPosition += 1

        item_ui.popup_lossW_ui.tableWidget_lossW.resizeColumnsToContents()            
        item_ui.popup_lossW_ui.tableWidget_lossW.resizeRowsToContents()

        item_ui.popup_lossW.show()
                
        item_ui.popup_lossW_ui.pushButton_pop_lossW_cancel.clicked.connect(lambda: self.lossW_cancel(listindex))
        item_ui.popup_lossW_ui.pushButton_pop_lossW_ok.clicked.connect(lambda: self.lossW_ok(listindex))
        item_ui.popup_lossW_ui.comboBox_lossW.currentIndexChanged.connect(lambda on_or_off: self.lossW_comboB(on_or_off,listindex))


    def optimizer_change_settings_popup(self,listindex):
        if listindex==-1:
            item_ui = self
        else:
            item_ui = self.fittingpopups_ui[listindex]

        item_ui.popup_optim = MyPopup()
        item_ui.popup_optim_ui = aid_frontend.Ui_Form_expt_optim()
        item_ui.popup_optim_ui.setupUi(item_ui.popup_optim) #open a popup to show advances settings for optimizer
        
        ##Manual insertion##
        optimizer_name = item_ui.optimizer_settings["comboBox_optimizer"].lower()
        if optimizer_name=='sgd':
            item_ui.popup_optim_ui.radioButton_sgd.setChecked(True)        
        elif optimizer_name=='rmsprop':
            item_ui.popup_optim_ui.radioButton_rms.setChecked(True)        
        elif optimizer_name=='adagrad':
            item_ui.popup_optim_ui.radioButton_adagrad.setChecked(True)        
        elif optimizer_name=='adadelta':
            item_ui.popup_optim_ui.radioButton_adadelta.setChecked(True)        
        elif optimizer_name=='adam':
            item_ui.popup_optim_ui.radioButton_adam.setChecked(True)        
        elif optimizer_name=='adamax':
            item_ui.popup_optim_ui.radioButton_adamax.setChecked(True)        
        elif optimizer_name=='nadam':
            item_ui.popup_optim_ui.radioButton_nadam.setChecked(True)        

        item_ui.popup_optim_ui.doubleSpinBox_lr_sgd.setValue(item_ui.optimizer_settings["doubleSpinBox_lr_sgd"])        
        item_ui.popup_optim_ui.doubleSpinBox_sgd_momentum.setValue(item_ui.optimizer_settings["doubleSpinBox_sgd_momentum"])  
        item_ui.popup_optim_ui.checkBox_sgd_nesterov.setChecked(item_ui.optimizer_settings["checkBox_sgd_nesterov"])  

        item_ui.popup_optim_ui.doubleSpinBox_lr_rmsprop.setValue(item_ui.optimizer_settings["doubleSpinBox_lr_rmsprop"])  
        item_ui.popup_optim_ui.doubleSpinBox_rms_rho.setValue(item_ui.optimizer_settings["doubleSpinBox_rms_rho"])  

        item_ui.popup_optim_ui.doubleSpinBox_lr_adam.setValue(item_ui.optimizer_settings["doubleSpinBox_lr_adam"])  
        item_ui.popup_optim_ui.doubleSpinBox_adam_beta1.setValue(item_ui.optimizer_settings["doubleSpinBox_adam_beta1"])  
        item_ui.popup_optim_ui.doubleSpinBox_adam_beta2.setValue(item_ui.optimizer_settings["doubleSpinBox_adam_beta2"])  
        item_ui.popup_optim_ui.checkBox_adam_amsgrad.setChecked(item_ui.optimizer_settings["checkBox_adam_amsgrad"])  

        item_ui.popup_optim_ui.doubleSpinBox_lr_nadam.setValue(item_ui.optimizer_settings["doubleSpinBox_lr_nadam"])          
        item_ui.popup_optim_ui.doubleSpinBox_nadam_beta1.setValue(item_ui.optimizer_settings["doubleSpinBox_nadam_beta1"])  
        item_ui.popup_optim_ui.doubleSpinBox_nadam_beta2.setValue(item_ui.optimizer_settings["doubleSpinBox_nadam_beta2"])  

        item_ui.popup_optim_ui.doubleSpinBox_lr_adadelta.setValue(item_ui.optimizer_settings["doubleSpinBox_lr_adadelta"])  
        item_ui.popup_optim_ui.doubleSpinBox_adadelta_rho.setValue(item_ui.optimizer_settings["doubleSpinBox_adadelta_rho"])  

        item_ui.popup_optim_ui.doubleSpinBox_lr_adagrad.setValue(item_ui.optimizer_settings["doubleSpinBox_lr_adagrad"])  

        item_ui.popup_optim_ui.doubleSpinBox_lr_adamax.setValue(item_ui.optimizer_settings["doubleSpinBox_lr_adamax"])  
        item_ui.popup_optim_ui.doubleSpinBox_adamax_beta1.setValue(item_ui.optimizer_settings["doubleSpinBox_adamax_beta1"])  
        item_ui.popup_optim_ui.doubleSpinBox_adamax_beta2.setValue(item_ui.optimizer_settings["doubleSpinBox_adamax_beta2"])

        def change_lr(lr):
            item_ui.doubleSpinBox_learningRate.setValue(lr)
            item_ui.doubleSpinBox_expDecInitLr.setValue(lr)

        item_ui.popup_optim_ui.doubleSpinBox_lr_adam.valueChanged.connect(change_lr)
        item_ui.popup_optim_ui.doubleSpinBox_lr_sgd.valueChanged.connect(change_lr)
        item_ui.popup_optim_ui.doubleSpinBox_lr_rmsprop.valueChanged.connect(change_lr)
        item_ui.popup_optim_ui.doubleSpinBox_lr_adagrad.valueChanged.connect(change_lr)
        item_ui.popup_optim_ui.doubleSpinBox_lr_adadelta.valueChanged.connect(change_lr)
        item_ui.popup_optim_ui.doubleSpinBox_lr_adamax.valueChanged.connect(change_lr)
        item_ui.popup_optim_ui.doubleSpinBox_lr_nadam.valueChanged.connect(change_lr)

        def change_optimizer(optimizer_name):
            index = item_ui.comboBox_optimizer.findText(optimizer_name, QtCore.Qt.MatchFixedString)
            if index >= 0:
                item_ui.comboBox_optimizer.setCurrentIndex(index)
            #get the learning rate for that optimizer
            lr = item_ui.optimizer_settings["doubleSpinBox_lr_"+optimizer_name.lower()]
            change_lr(lr)
            
        item_ui.popup_optim_ui.radioButton_adam.toggled.connect(lambda: change_optimizer("Adam"))
        item_ui.popup_optim_ui.radioButton_sgd.toggled.connect(lambda: change_optimizer("SGD"))
        item_ui.popup_optim_ui.radioButton_rms.toggled.connect(lambda: change_optimizer("RMSprop"))
        item_ui.popup_optim_ui.radioButton_adagrad.toggled.connect(lambda: change_optimizer("Adagrad"))
        item_ui.popup_optim_ui.radioButton_adadelta.toggled.connect(lambda: change_optimizer("Adadelta"))
        item_ui.popup_optim_ui.radioButton_adamax.toggled.connect(lambda: change_optimizer("Adamax"))
        item_ui.popup_optim_ui.radioButton_nadam.toggled.connect(lambda: change_optimizer("Nadam"))

        def ok():
            doubleSpinBox_lr_sgd = float(item_ui.popup_optim_ui.doubleSpinBox_lr_sgd.value())
            doubleSpinBox_sgd_momentum = float(item_ui.popup_optim_ui.doubleSpinBox_sgd_momentum.value())
            checkBox_sgd_nesterov = bool(item_ui.popup_optim_ui.checkBox_sgd_nesterov.isChecked())

            doubleSpinBox_lr_rmsprop = float(item_ui.popup_optim_ui.doubleSpinBox_lr_rmsprop.value())            
            doubleSpinBox_rms_rho = float(item_ui.popup_optim_ui.doubleSpinBox_rms_rho.value())

            doubleSpinBox_lr_adam = float(item_ui.popup_optim_ui.doubleSpinBox_lr_adam.value())
            doubleSpinBox_adam_beta1 = float(item_ui.popup_optim_ui.doubleSpinBox_adam_beta1.value())
            doubleSpinBox_adam_beta2 = float(item_ui.popup_optim_ui.doubleSpinBox_adam_beta2.value())
            checkBox_adam_amsgrad = bool(item_ui.popup_optim_ui.checkBox_adam_amsgrad.isChecked())
            
            doubleSpinBox_lr_adadelta = float(item_ui.popup_optim_ui.doubleSpinBox_lr_adadelta.value())
            doubleSpinBox_adadelta_rho = float(item_ui.popup_optim_ui.doubleSpinBox_adadelta_rho.value())

            doubleSpinBox_lr_nadam = float(item_ui.popup_optim_ui.doubleSpinBox_lr_nadam.value())
            doubleSpinBox_nadam_beta1 = float(item_ui.popup_optim_ui.doubleSpinBox_nadam_beta1.value())
            doubleSpinBox_nadam_beta2 = float(item_ui.popup_optim_ui.doubleSpinBox_nadam_beta2.value())

            doubleSpinBox_lr_adagrad = float(item_ui.popup_optim_ui.doubleSpinBox_lr_adagrad.value())

            doubleSpinBox_lr_adamax = float(item_ui.popup_optim_ui.doubleSpinBox_lr_adamax.value())
            doubleSpinBox_adamax_beta2 = float(item_ui.popup_optim_ui.doubleSpinBox_adamax_beta2.value())
            doubleSpinBox_adamax_beta1 = float(item_ui.popup_optim_ui.doubleSpinBox_adamax_beta1.value())

            item_ui.optimizer_settings["doubleSpinBox_lr_sgd"] = doubleSpinBox_lr_sgd
            item_ui.optimizer_settings["doubleSpinBox_sgd_momentum"] = doubleSpinBox_sgd_momentum
            item_ui.optimizer_settings["checkBox_sgd_nesterov"] = checkBox_sgd_nesterov

            item_ui.optimizer_settings["doubleSpinBox_lr_rmsprop"] = doubleSpinBox_lr_rmsprop
            item_ui.optimizer_settings["doubleSpinBox_rms_rho"] = doubleSpinBox_rms_rho

            item_ui.optimizer_settings["doubleSpinBox_lr_adam"] = doubleSpinBox_lr_adam
            item_ui.optimizer_settings["doubleSpinBox_adam_beta1"] = doubleSpinBox_adam_beta1
            item_ui.optimizer_settings["doubleSpinBox_adam_beta2"] = doubleSpinBox_adam_beta2
            item_ui.optimizer_settings["checkBox_adam_amsgrad"] = checkBox_adam_amsgrad

            item_ui.optimizer_settings["doubleSpinBox_lr_adadelta"] = doubleSpinBox_lr_adadelta
            item_ui.optimizer_settings["doubleSpinBox_adadelta_rho"] = doubleSpinBox_adadelta_rho

            item_ui.optimizer_settings["doubleSpinBox_lr_nadam"] = doubleSpinBox_lr_nadam
            item_ui.optimizer_settings["doubleSpinBox_nadam_beta1"] = doubleSpinBox_nadam_beta1
            item_ui.optimizer_settings["doubleSpinBox_nadam_beta2"] = doubleSpinBox_nadam_beta2

            item_ui.optimizer_settings["doubleSpinBox_lr_adagrad"] = doubleSpinBox_lr_adagrad

            item_ui.optimizer_settings["doubleSpinBox_lr_adamax"] = doubleSpinBox_lr_adamax
            item_ui.optimizer_settings["doubleSpinBox_adamax_beta1"] = doubleSpinBox_adamax_beta1
            item_ui.optimizer_settings["doubleSpinBox_adamax_beta2"] = doubleSpinBox_adamax_beta2
            
            #close the popup
            item_ui.popup_optim = None
            item_ui.popup_optim_ui = None
            print("Advanced settings for optimizer were changed.")

        def cancel():#close the popup
            item_ui.popup_optim = None
            item_ui.popup_optim_ui = None
        def reset():
            print("Reset optimizer settings (in UI). To accept, click OK")
            optimizer_default = aid_dl.get_optimizer_settings()
            item_ui.popup_optim_ui.doubleSpinBox_lr_sgd.setValue(optimizer_default["doubleSpinBox_lr_sgd"])        
            item_ui.popup_optim_ui.doubleSpinBox_sgd_momentum.setValue(optimizer_default["doubleSpinBox_sgd_momentum"])  
            item_ui.popup_optim_ui.checkBox_sgd_nesterov.setChecked(optimizer_default["checkBox_sgd_nesterov"])  
    
            item_ui.popup_optim_ui.doubleSpinBox_lr_rmsprop.setValue(optimizer_default["doubleSpinBox_lr_rmsprop"])  
            item_ui.popup_optim_ui.doubleSpinBox_rms_rho.setValue(optimizer_default["doubleSpinBox_rms_rho"])  
    
            item_ui.popup_optim_ui.doubleSpinBox_lr_adam.setValue(optimizer_default["doubleSpinBox_lr_adam"])  
            item_ui.popup_optim_ui.doubleSpinBox_adam_beta1.setValue(optimizer_default["doubleSpinBox_adam_beta1"])  
            item_ui.popup_optim_ui.doubleSpinBox_adam_beta2.setValue(optimizer_default["doubleSpinBox_adam_beta2"])  
            item_ui.popup_optim_ui.checkBox_adam_amsgrad.setChecked(optimizer_default["checkBox_adam_amsgrad"])  
    
            item_ui.popup_optim_ui.doubleSpinBox_lr_nadam.setValue(optimizer_default["doubleSpinBox_lr_nadam"])          
            item_ui.popup_optim_ui.doubleSpinBox_nadam_beta1.setValue(optimizer_default["doubleSpinBox_nadam_beta1"])  
            item_ui.popup_optim_ui.doubleSpinBox_nadam_beta2.setValue(optimizer_default["doubleSpinBox_nadam_beta2"])  
    
            item_ui.popup_optim_ui.doubleSpinBox_lr_adadelta.setValue(optimizer_default["doubleSpinBox_lr_adadelta"])  
            item_ui.popup_optim_ui.doubleSpinBox_adadelta_rho.setValue(optimizer_default["doubleSpinBox_adadelta_rho"])  
    
            item_ui.popup_optim_ui.doubleSpinBox_lr_adagrad.setValue(optimizer_default["doubleSpinBox_lr_adagrad"])  
    
            item_ui.popup_optim_ui.doubleSpinBox_lr_adamax.setValue(optimizer_default["doubleSpinBox_lr_adamax"])  
            item_ui.popup_optim_ui.doubleSpinBox_adamax_beta1.setValue(optimizer_default["doubleSpinBox_adamax_beta1"])  
            item_ui.popup_optim_ui.doubleSpinBox_adamax_beta2.setValue(optimizer_default["doubleSpinBox_adamax_beta2"])


        item_ui.popup_optim_ui.pushButton_ok.clicked.connect(ok)
        item_ui.popup_optim_ui.pushButton_cancel.clicked.connect(cancel)
        item_ui.popup_optim_ui.pushButton_reset.clicked.connect(reset)
        
        item_ui.popup_optim.show()


    def onLayoutChange(self):
        #Get the text of the triggered layout
        layout_trig = (self.sender().text()).split(" layout")[0]
        layout_current = Default_dict["Layout"]

        if layout_trig == layout_current:
            self.statusbar.showMessage(layout_current+" layout is already in use",2000)
            return
        
        elif layout_trig == "Normal":
            #Change Layout in Defaultdict to "Normal", such that next start will use Normal layout
            Default_dict["Layout"] = "Normal"
            self.app.setStyleSheet("")
            #Standard is with tooltip
            self.actionTooltipOnOff.setChecked(True)

        elif layout_trig == "Dark":
            #Change Layout in Defaultdict to "Dark", such that next start will use Dark layout
            Default_dict["Layout"] = "Dark"
            dir_layout = os.path.join(dir_root,"layout_dark.txt")#dir to settings
            f = open(dir_layout, "r") #I obtained the layout file from: https://github.com/ColinDuquesnoy/QDarkStyleSheet/blob/master/qdarkstyle/style.qss
            f = f.read()
            self.app.setStyleSheet(f)
            #Standard is with tooltip
            self.actionTooltipOnOff.setChecked(True)
        
        elif layout_trig == "DarkOrange":
            #Change Layout in Defaultdict to "Dark", such that next start will use Dark layout
            Default_dict["Layout"] = "DarkOrange"
            dir_layout = os.path.join(dir_root,"layout_darkorange.txt")#dir to settings
            f = open(dir_layout, "r") #I obtained the layout file from: https://github.com/nphase/qt-ping-grapher/blob/master/resources/darkorange.stylesheet
            f = f.read()
            self.app.setStyleSheet(f)
            #Standard is with tooltip
            self.actionTooltipOnOff.setChecked(True)

        #Save the layout to Default_dict
        with open(dir_settings, 'w') as f:
            json.dump(Default_dict,f)
        
    def onTooltipOnOff(self):
        #what is the current layout?
        if bool(self.actionLayout_Normal.isChecked())==True: #use normal layout
            if bool(self.actionTooltipOnOff.isChecked())==True: #with tooltips
                self.app.setStyleSheet("")
            elif bool(self.actionTooltipOnOff.isChecked())==False: #no tooltips
                self.app.setStyleSheet("""QToolTip {
                                         opacity: 0
                                           }""")

        elif bool(self.actionLayout_Dark.isChecked())==True: #use dark layout
            if bool(self.actionTooltipOnOff.isChecked())==True: #with tooltips
                dir_layout = os.path.join(dir_root,"layout_dark.txt")#dir to settings
                f = open(dir_layout, "r") #I obtained the layout file from: https://github.com/ColinDuquesnoy/QDarkStyleSheet/blob/master/qdarkstyle/style.qss
                f = f.read()
                self.app.setStyleSheet(f)

            elif bool(self.actionTooltipOnOff.isChecked())==False: #no tooltips
                dir_layout = os.path.join(dir_root,"layout_dark_notooltip.txt")#dir to settings
                f = open(dir_layout, "r")#I obtained the layout file from: https://github.com/ColinDuquesnoy/QDarkStyleSheet/blob/master/qdarkstyle/style.qss
                f = f.read()
                self.app.setStyleSheet(f)

        elif bool(self.actionLayout_DarkOrange.isChecked())==True: #use darkorange layout
            if bool(self.actionTooltipOnOff.isChecked())==True: #with tooltips
                dir_layout = os.path.join(dir_root,"layout_darkorange.txt")#dir to settings
                f = open(dir_layout, "r") #I obtained the layout file from: https://github.com/nphase/qt-ping-grapher/blob/master/resources/darkorange.stylesheet
                f = f.read()
                self.app.setStyleSheet(f)

            elif bool(self.actionTooltipOnOff.isChecked())==False: #no tooltips
                dir_layout = os.path.join(dir_root,"layout_darkorange_notooltip.txt")#dir to settings
                f = open(dir_layout, "r")
                f = f.read()
                self.app.setStyleSheet(f)

    def onIconThemeChange(self):
        #Get the text of the triggered icon theme
        icontheme_trig = self.sender().text()
        icontheme_currenent = Default_dict["Icon theme"]

        if icontheme_trig == icontheme_currenent:
            self.statusbar.showMessage(icontheme_currenent+" is already in use",2000)
            return
        
        elif icontheme_trig == "Icon theme 1":
            Default_dict["Icon theme"] = "Icon theme 1"
            self.statusbar.showMessage("Icon theme 1 will be used after restart",2000)

        elif icontheme_trig == "Icon theme 2":
            Default_dict["Icon theme"] = "Icon theme 2"
            self.statusbar.showMessage("Icon theme 2 will be used after restart",2000)
        
        #Save the layout to Default_dict
        with open(dir_settings, 'w') as f:
            json.dump(Default_dict,f)


    def items_clicked(self):
        #This function checks, which data has been checked on table_dragdrop and returns the necessary data
        rowCount = self.table_dragdrop.rowCount()
        #Collect urls to files that are checked
        SelectedFiles = []
        for rowPosition in range(rowCount):  
            #get the filename/path
            rtdc_path = str(self.table_dragdrop.item(rowPosition, 0).text())
            #get the index (celltype) of it
            index = int(self.table_dragdrop.cellWidget(rowPosition, 1).value())
            #is it checked for train?
            cb_t = self.table_dragdrop.item(rowPosition, 2)
            #How many Events contains dataset in total?
            nr_events = int(self.table_dragdrop.item(rowPosition, 5).text())
            #how many cells/epoch during training or validation?
            nr_events_epoch = int(self.table_dragdrop.item(rowPosition, 6).text())            
            #should the dataset be randomized (shuffled?)            
            shuffle = bool(self.table_dragdrop.item(rowPosition, 8).checkState())           
            #should the images be zoomed in/out by a factor?
            zoom_factor = float(self.table_dragdrop.item(rowPosition, 9).text())            
            #should xtra_data be used for training?
            xtra_in = bool(self.table_dragdrop.item(rowPosition, 10).checkState())           
            
            if cb_t.checkState() == QtCore.Qt.Checked and nr_events_epoch>0: #add to training files if the user wants more than 0 images per epoch
                failed,rtdc_ds = aid_bin.load_rtdc(rtdc_path)
                if failed:
                    msg = QtWidgets.QMessageBox()
                    msg.setIcon(QtWidgets.QMessageBox.Critical)       
                    msg.setText(str(rtdc_ds))
                    msg.setWindowTitle("Error occurred during loading file")
                    msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                    msg.exec_()
                    return
                    
                hash_ = aid_bin.hashfunction(rtdc_path)#rtdc_ds.hash
                features = list(rtdc_ds["events"].keys())
                nr_images = rtdc_ds["events"]["image"].len()
                SelectedFiles.append({"rtdc_ds":rtdc_ds,"rtdc_path":rtdc_path,"features":features,"nr_images":nr_images,"class":index,"TrainOrValid":"Train","nr_events":nr_events,"nr_events_epoch":nr_events_epoch,"shuffle":shuffle,"zoom_factor":zoom_factor,"hash":hash_,"xtra_in":xtra_in})
            
            cb_v = self.table_dragdrop.item(rowPosition, 3)
            if cb_v.checkState() == QtCore.Qt.Checked and nr_events_epoch>0:
                failed,rtdc_ds = aid_bin.load_rtdc(rtdc_path)
                if failed:
                    msg = QtWidgets.QMessageBox()
                    msg.setIcon(QtWidgets.QMessageBox.Critical)       
                    msg.setText(str(rtdc_ds))
                    msg.setWindowTitle("Error occurred during loading file")
                    msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                    msg.exec_()
                    return
                hash_ = aid_bin.hashfunction(rtdc_path)
                features = list(rtdc_ds["events"].keys())
                nr_images = rtdc_ds["events"]["image"].len()
                SelectedFiles.append({"rtdc_ds":rtdc_ds,"rtdc_path":rtdc_path,"features":features,"nr_images":nr_images,"class":index,"TrainOrValid":"Valid","nr_events":nr_events,"nr_events_epoch":nr_events_epoch,"shuffle":shuffle,"zoom_factor":zoom_factor,"hash":hash_,"xtra_in":xtra_in})
        return SelectedFiles


    def items_available(self):
        """
        Function grabs all information from table_dragdrop. Checked and Unchecked
        Does not load rtdc_ds (save time)
        """
        rowCount = self.table_dragdrop.rowCount()
        #Collect urls to files that are checked
        SelectedFiles = []
        for rowPosition in range(rowCount):  
            #get the filename/path
            rtdc_path = str(self.table_dragdrop.item(rowPosition, 0).text())
            #get the index (celltype) of it
            index = int(self.table_dragdrop.cellWidget(rowPosition, 1).value())
            #How many Events contains dataset in total?
            nr_events = int(self.table_dragdrop.item(rowPosition, 5).text())
            #how many cells/epoch during training or validation?
            nr_events_epoch = int(self.table_dragdrop.item(rowPosition, 6).text())            
            #should the dataset be randomized (shuffled?)            
            shuffle = bool(self.table_dragdrop.item(rowPosition, 8).checkState())           
            #should the images be zoomed in/out by a factor?
            zoom_factor = float(self.table_dragdrop.item(rowPosition, 9).text())            
            #should xtra_data be used for training?
            xtra_in = bool(self.table_dragdrop.item(rowPosition, 10).checkState())           

            SelectedFiles.append({"rtdc_path":rtdc_path,"class":index,"TrainOrValid":"NotSpecified","nr_events":nr_events,"nr_events_epoch":nr_events_epoch,"shuffle":shuffle,"zoom_factor":zoom_factor,"xtra_in":xtra_in})
            
        return SelectedFiles


    def items_clicked_no_rtdc_ds(self):
        #This function checks, which data has been checked on table_dragdrop and returns the necessary data
        rowCount = self.table_dragdrop.rowCount()
        #Collect urls to files that are checked
        SelectedFiles = []
        for rowPosition in range(rowCount):  
            #get the filename/path
            rtdc_path = str(self.table_dragdrop.item(rowPosition, 0).text())
            #get the index (celltype) of it
            index = int(self.table_dragdrop.cellWidget(rowPosition, 1).value())
            #How many Events contains dataset in total?
            nr_events = int(self.table_dragdrop.item(rowPosition, 5).text())
            #how many cells/epoch during training or validation?
            nr_events_epoch = int(self.table_dragdrop.item(rowPosition, 6).text())            
            #should the dataset be randomized (shuffled?)            
            shuffle = bool(self.table_dragdrop.item(rowPosition, 8).checkState())           
            #should the images be zoomed in/out by a factor?
            zoom_factor = float(self.table_dragdrop.item(rowPosition, 9).text())            
            #should xtra_data be used for training?
            xtra_in = bool(self.table_dragdrop.item(rowPosition, 10).checkState())           

            #is it checked for train?
            cb_t = self.table_dragdrop.item(rowPosition, 2)
            if cb_t.checkState() == QtCore.Qt.Checked and nr_events_epoch>0: #add to training files if the user wants more than 0 images per epoch
                #SelectedFiles.append({"nr_images":nr_events,"class":index,"TrainOrValid":"Train","nr_events":nr_events,"nr_events_epoch":nr_events_epoch})
                SelectedFiles.append({"rtdc_path":rtdc_path,"class":index,"TrainOrValid":"Train","nr_events":nr_events,"nr_events_epoch":nr_events_epoch,"shuffle":shuffle,"zoom_factor":zoom_factor,"xtra_in":xtra_in})

            cb_v = self.table_dragdrop.item(rowPosition, 3)
            if cb_v.checkState() == QtCore.Qt.Checked and nr_events_epoch>0:
                #SelectedFiles.append({"nr_images":nr_events,"class":index,"TrainOrValid":"Valid","nr_events":nr_events,"nr_events_epoch":nr_events_epoch})
                SelectedFiles.append({"rtdc_path":rtdc_path,"class":index,"TrainOrValid":"Valid","nr_events":nr_events,"nr_events_epoch":nr_events_epoch,"shuffle":shuffle,"zoom_factor":zoom_factor,"xtra_in":xtra_in})

        return SelectedFiles


    def uncheck_if_zero(self,item):
        #If the Nr. of epochs is changed to zero:
        #uncheck the dataset for train/valid
        row = item.row()
        col = item.column()
        #if the user changed Nr. of cells per epoch to zero
        if col==6 and int(item.text())==0:
            #get the checkstate of the coresponding T/V
            cb_t = self.table_dragdrop.item(row, 2)
            if cb_t.checkState() == QtCore.Qt.Checked:
                cb_t.setCheckState(False)
            cb_v = self.table_dragdrop.item(row, 3)
            if cb_v.checkState() == QtCore.Qt.Checked:
                cb_v.setCheckState(False)

      
    def item_click(self,item): 
        colPosition = item.column()
        rowPosition = item.row()
        #if Shuffle was clicked (col=8), check if this checkbox is not deactivated
        if colPosition==8:
            if bool(self.table_dragdrop.item(rowPosition, 8).checkState())==False:
                rtdc_path = self.table_dragdrop.item(rowPosition, 0).text()
                rtdc_path = str(rtdc_path)

                failed,rtdc_ds = aid_bin.load_rtdc(rtdc_path)
                if failed:
                    msg = QtWidgets.QMessageBox()
                    msg.setIcon(QtWidgets.QMessageBox.Critical)       
                    msg.setText(str(rtdc_ds))
                    msg.setWindowTitle("Error occurred during loading file")
                    msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                    msg.exec_()
                    return
                nr_images = rtdc_ds["events"]["image"].len()
        
                columnPosition = 6
                item = QtWidgets.QTableWidgetItem()
                item.setData(QtCore.Qt.DisplayRole, nr_images)
                item.setFlags(item.flags() &~QtCore.Qt.ItemIsEnabled &~ QtCore.Qt.ItemIsSelectable )
                self.table_dragdrop.setItem(rowPosition, columnPosition, item)
            if bool(self.table_dragdrop.item(rowPosition, 8).checkState())==True:
                #Inspect this table item. If shuffle was checked before, it will be grayed out. Invert normal cell then
                item = self.table_dragdrop.item(rowPosition, 6)
                item.setFlags(item.flags() |QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable )

        if len(self.ram)>0:
            self.statusbar.showMessage("Make sure to update RAM (->Edit->Data to RAM now) after changing Data-set",2000)
            self.ram = dict() #clear the ram, since the data was changed
            
        self.dataOverviewOn()

        #When data is clicked, always reset the validation set (only important for 'Assess Model'-tab)
        self.ValidationSet = None
        self.Metrics = dict() #Also reset the metrics

    def dataOverviewOn(self):
        if self.groupBox_DataOverview.isChecked()==True:
            if self.threadpool_single_queue == 0:
                SelectedFiles = self.items_clicked_no_rtdc_ds()
                self.update_data_overview(SelectedFiles)
                self.update_data_overview_2(SelectedFiles)

    def dataOverviewOn_OnChange(self,item):
        #When a value is entered in Events/Epoch and enter is hit
        #there is no update of the table called
        if self.groupBox_DataOverview.isChecked()==True:
            if self.threadpool_single_queue == 0:
                rowPosition = item.row()
                colPosition = item.column()
                if colPosition==6:#one when using the spinbox (Class),or when entering a new number in "Events/Epoch", the table is not updated. 
                    #get the new value
                    nr_cells = self.table_dragdrop.cellWidget(rowPosition, colPosition)
                    if nr_cells==None:
                        return
                    else:
                        SelectedFiles = self.items_clicked_no_rtdc_ds()
                        self.update_data_overview(SelectedFiles)
                        self.update_data_overview_2(SelectedFiles)
                            
    def update_data_overview(self,SelectedFiles):
        #Check if there are custom class names (determined by user)
        rows = self.tableWidget_Info.rowCount()
        self.classes_custom = [] #by default assume there are no custom classes
        classes_custom_bool = False
        if rows>0:#if >0, then there is already a table existing
            classes,self.classes_custom = [],[]
            for row in range(rows):
                try:
                    class_ = self.tableWidget_Info.item(row,0).text()
                    if class_.isdigit():    
                        classes.append(class_)#get the classes
                except:
                    pass
                try:
                    self.classes_custom.append(self.tableWidget_Info.item(row,3).text())#get the classes
                except:
                    pass
            classes = np.unique(classes)
            if len(classes)==len(self.classes_custom):#equal in length
                same = [i for i, j in zip(classes, self.classes_custom) if i == j] #which items are identical?
                if len(same)==0:
                    #apparently there are custom classes! Save them
                    classes_custom_bool = True                            
            
        if len(SelectedFiles)==0:#reset the table
            #Table1
            #Prepare a table in tableWidget_Info
            self.tableWidget_Info.setColumnCount(0)
            self.tableWidget_Info.setRowCount(0)
            self.tableWidget_Info.setColumnCount(4)
            header = self.tableWidget_Info.horizontalHeader()
            header_labels = ["Class","Events tot.","Events/Epoch","Name"]
            self.tableWidget_Info.setHorizontalHeaderLabels(header_labels) 
            header = self.tableWidget_Info.horizontalHeader()
            for i in range(4):
                header.setResizeMode(i, QtWidgets.QHeaderView.ResizeToContents)
            return
        #Prepare a table in tableWidget_Info
        self.tableWidget_Info.setColumnCount(0)
        self.tableWidget_Info.setRowCount(0)

        indices = [SelectedFiles[i]["class"] for i in range(len(SelectedFiles))]
        self.tableWidget_Info.setColumnCount(4)
        header = self.tableWidget_Info.horizontalHeader()

        nr_ind = len(set(indices)) #each index could occur for train and valid
        nr_rows = 2*nr_ind+2 #add two rows for intermediate headers (Train/Valid)
        self.tableWidget_Info.setRowCount(nr_rows)
        #Wich selected file has the most features?
        header_labels = ["Class","Events tot.","Events/Epoch","Name"]
        self.tableWidget_Info.setHorizontalHeaderLabels(header_labels) 
        #self.tableWidget_Info.resizeColumnsToContents()            
        header = self.tableWidget_Info.horizontalHeader()
        for i in range(4):
            header.setResizeMode(i, QtWidgets.QHeaderView.ResizeToContents)        
        
        #Training info
        rowPosition = 0
        self.tableWidget_Info.setSpan(rowPosition, 0, 1, 2) 
        item = QtWidgets.QTableWidgetItem("Train. data") 
        item.setFlags(QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable)
        self.tableWidget_Info.setItem(rowPosition, 0, item)            
        rowPosition += 1
        ind = [selectedfile["TrainOrValid"] == "Train" for selectedfile in SelectedFiles]
        ind = np.where(np.array(ind)==True)[0]
        SelectedFiles_train = np.array(SelectedFiles)[ind]
        SelectedFiles_train = list(SelectedFiles_train)
        indices_train = [selectedfile["class"] for selectedfile in SelectedFiles_train]

        classes = np.unique(indices_train)
        if len(classes)==len(self.classes_custom):
            classes_custom_bool = True
        else:
            classes_custom_bool = False
 
       #display information for each individual class            
        for index_ in range(len(classes)):
        #for index in np.unique(indices_train):
            index = classes[index_]
            #put the index in column nr. 0
            item = QtWidgets.QTableWidgetItem()
            item.setFlags(QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable)
            item.setData(QtCore.Qt.EditRole,str(index))
            self.tableWidget_Info.setItem(rowPosition, 0, item)
            #Get the training files of that index
            ind = np.where(indices_train==index)[0]
            SelectedFiles_train_index = np.array(SelectedFiles_train)[ind]
            #Total nr of cells for each class
            nr_events = [int(selectedfile["nr_events"]) for selectedfile in SelectedFiles_train_index]
            item = QtWidgets.QTableWidgetItem()
            item.setFlags(QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable)
            item.setData(QtCore.Qt.EditRole, str(np.sum(nr_events)))
            self.tableWidget_Info.setItem(rowPosition, 1, item)
            nr_events_epoch = [int(selectedfile["nr_events_epoch"]) for selectedfile in SelectedFiles_train_index]
            item = QtWidgets.QTableWidgetItem()
            item.setFlags(QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable)
            item.setData(QtCore.Qt.EditRole, str(np.sum(nr_events_epoch)))
            self.tableWidget_Info.setItem(rowPosition, 2, item)    

            item = QtWidgets.QTableWidgetItem()
            if classes_custom_bool==False:
                item.setData(QtCore.Qt.EditRole,str(index))
            else:
                item.setData(QtCore.Qt.EditRole,self.classes_custom[index_])
            self.tableWidget_Info.setItem(rowPosition, 3, item)                

            rowPosition += 1
        
        #Validation info
        self.tableWidget_Info.setSpan(rowPosition, 0, 1, 2) 
        item = QtWidgets.QTableWidgetItem("Val. data")  
        item.setFlags(QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable)
        self.tableWidget_Info.setItem(rowPosition, 0, item)            
        rowPosition += 1
        ind = [selectedfile["TrainOrValid"] == "Valid" for selectedfile in SelectedFiles]
        ind = np.where(np.array(ind)==True)[0]
        SelectedFiles_valid = np.array(SelectedFiles)[ind]
        SelectedFiles_valid = list(SelectedFiles_valid)
        indices_valid = [selectedfile["class"] for selectedfile in SelectedFiles_valid]
        #Total nr of cells for each index
        for index in np.unique(indices_valid):
            #put the index in column nr. 0
            item = QtWidgets.QTableWidgetItem()
            item.setFlags(QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable)
            item.setData(QtCore.Qt.EditRole,str(index))
            self.tableWidget_Info.setItem(rowPosition, 0, item)
            #Get the validation files of that index
            ind = np.where(indices_valid==index)[0]
            SelectedFiles_valid_index = np.array(SelectedFiles_valid)[ind]
            nr_events = [int(selectedfile["nr_events"]) for selectedfile in SelectedFiles_valid_index]
            item = QtWidgets.QTableWidgetItem()
            item.setFlags(QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable)
            item.setData(QtCore.Qt.EditRole, str(np.sum(nr_events)))
            self.tableWidget_Info.setItem(rowPosition, 1, item)
            nr_events_epoch = [int(selectedfile["nr_events_epoch"]) for selectedfile in SelectedFiles_valid_index]
            item = QtWidgets.QTableWidgetItem()
            item.setFlags(QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable)
            item.setData(QtCore.Qt.EditRole, str(np.sum(nr_events_epoch)))
            self.tableWidget_Info.setItem(rowPosition, 2, item)
            rowPosition += 1
        self.tableWidget_Info.resizeColumnsToContents()            
        self.tableWidget_Info.resizeRowsToContents()


    def update_data_overview_2(self,SelectedFiles):
        if len(SelectedFiles)==0:
            #Table2
            self.tableWidget_Info_2.setColumnCount(0)
            self.tableWidget_Info_2.setRowCount(0)               
            #In case user specified X_valid and y_valid before, delete it again:
            self.ValidationSet = None
            self.Metrics = dict() #Also reset the metrics
            #Initiate the table with 4 columns : this will be ["Index","Nr of cells","Clr","Name"]
            self.tableWidget_Info_2.setColumnCount(4)
            header_labels = ["Class","Nr of cells","Clr","Name"]
            self.tableWidget_Info_2.setHorizontalHeaderLabels(header_labels) 
            header = self.tableWidget_Info_2.horizontalHeader()
            for i in range(4):
                header.setResizeMode(i, QtWidgets.QHeaderView.ResizeToContents)  
            return

        #Prepare a table in tableWidget_Info
        self.tableWidget_Info_2.setColumnCount(0)
        self.tableWidget_Info_2.setRowCount(0)
        
        #In case user specified X_valid and y_valid before, delete it again:
        self.ValidationSet = None
        self.Metrics = dict() #Also reset the metrics

        indices = [SelectedFiles[i]["class"] for i in range(len(SelectedFiles))]
        #Initiate the table with 4 columns : this will be ["Index","Nr of cells","Clr","Name"]
        self.tableWidget_Info_2.setColumnCount(4)
        nr_ind = len(set(indices)) #each index could occur for train and valid
        nr_rows = nr_ind
        self.tableWidget_Info_2.setRowCount(nr_rows)
        #Wich selected file has the most features?
        header_labels = ["Class","Nr of cells","Clr","Name"]
        self.tableWidget_Info_2.setHorizontalHeaderLabels(header_labels) 
        header = self.tableWidget_Info_2.horizontalHeader()
        for i in range(4):
            header.setResizeMode(i, QtWidgets.QHeaderView.ResizeToContents)        
        
        rowPosition = 0      
        #Validation info
        ind = [selectedfile["TrainOrValid"] == "Valid" for selectedfile in SelectedFiles]
        ind = np.where(np.array(ind)==True)[0]
        SelectedFiles_valid = np.array(SelectedFiles)[ind]
        SelectedFiles_valid = list(SelectedFiles_valid)
        indices_valid = [selectedfile["class"] for selectedfile in SelectedFiles_valid]
        #Total nr of cells for each index
        for index in np.unique(indices_valid):
            #put the index in column nr. 0
            item = QtWidgets.QTableWidgetItem()
            item.setFlags(item.flags() &~QtCore.Qt.ItemIsEnabled &~ QtCore.Qt.ItemIsSelectable )
            item.setData(QtCore.Qt.EditRole,str(index))
            self.tableWidget_Info_2.setItem(rowPosition, 0, item)
            #Get the validation files of that index
            ind = np.where(indices_valid==index)[0]
            SelectedFiles_valid_index = np.array(SelectedFiles_valid)[ind]
            nr_events_epoch = [int(selectedfile["nr_events_epoch"]) for selectedfile in SelectedFiles_valid_index]
            item = QtWidgets.QTableWidgetItem()
            item.setFlags(item.flags() &~QtCore.Qt.ItemIsEnabled &~ QtCore.Qt.ItemIsSelectable )
            #item.setFlags(QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable)
            item.setData(QtCore.Qt.EditRole, str(np.sum(nr_events_epoch)))
            self.tableWidget_Info_2.setItem(rowPosition, 1, item)
            
            #Column for color
            item = QtWidgets.QTableWidgetItem()
            item.setFlags(item.flags() &~QtCore.Qt.ItemIsEnabled &~ QtCore.Qt.ItemIsSelectable )
            #item.setFlags(QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable)
            item.setData(QtCore.Qt.EditRole, "")
            item.setBackground(QtGui.QColor(self.colorsQt[index]))            
            self.tableWidget_Info_2.setItem(rowPosition, 2, item)

            #Column for User specified name
            item = QtWidgets.QTableWidgetItem()
            #item.setFlags(QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable)
            item.setData(QtCore.Qt.EditRole,str(index))
            self.tableWidget_Info_2.setItem(rowPosition, 3, item)
            
            rowPosition += 1
        self.tableWidget_Info_2.resizeColumnsToContents()            
        self.tableWidget_Info_2.resizeRowsToContents()


    def tableWidget_Info_2_click(self,item):
        if item is not None:
            if item.column()==2:
                tableitem = self.tableWidget_Info_2.item(item.row(), item.column())
                color = QtGui.QColorDialog.getColor()
                if color.getRgb()==(0, 0, 0, 255):#no black!
                    return
                else:
                    tableitem.setBackground(color)

    def tableWidget_HistoryItems_dclick(self,item):
        if item is not None:
            tableitem = self.tableWidget_HistoryItems.item(item.row(), item.column())
            if str(tableitem.text())!="Show saved only":
                color = QtGui.QColorDialog.getColor()
                if color.getRgb()==(0, 0, 0, 255):#no black!
                    return
                else:
                    tableitem.setBackground(color)
                    self.update_historyplot()
 
    def select_all(self,col):
        """
        Check/Uncheck items on table_dragdrop
        """
        apply_at_col = [2,3,8,10]
        if col not in apply_at_col:
            return
        #otherwiese continue
        rows = range(self.table_dragdrop.rowCount()) #Number of rows of the table
        
        tableitems = [self.table_dragdrop.item(row, col) for row in rows]
        checkStates = [tableitem.checkState() for tableitem in tableitems]
        #Checked?
        checked = [state==QtCore.Qt.Checked for state in checkStates]
        if set(checked)=={True}:#all are checked!
            #Uncheck all!
            for tableitem in tableitems:
                tableitem.setCheckState(QtCore.Qt.Unchecked)
        else:#otherwise check all   
            for tableitem in tableitems:
                tableitem.setCheckState(QtCore.Qt.Checked)
                
        #If shuffle column was clicked do some extra
        if col==8:
            for rowPosition in rows:
                if bool(self.table_dragdrop.item(rowPosition, 8).checkState())==False:
                    rtdc_path = self.table_dragdrop.item(rowPosition, 0).text()
                    rtdc_path = str(rtdc_path)
    
                    failed,rtdc_ds = aid_bin.load_rtdc(rtdc_path)
                    if failed:
                        msg = QtWidgets.QMessageBox()
                        msg.setIcon(QtWidgets.QMessageBox.Critical)       
                        msg.setText(str(rtdc_ds))
                        msg.setWindowTitle("Error occurred during loading file")
                        msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                        msg.exec_()
                        return
                    nr_images = rtdc_ds["events"]["image"].len()
            
                    columnPosition = 6
                    item = QtWidgets.QTableWidgetItem()
                    item.setData(QtCore.Qt.DisplayRole, nr_images)
                    item.setFlags(item.flags() &~QtCore.Qt.ItemIsEnabled &~ QtCore.Qt.ItemIsSelectable )
                    self.table_dragdrop.setItem(rowPosition, columnPosition, item)
                if bool(self.table_dragdrop.item(rowPosition, 8).checkState())==True:
                    #Inspect this table item. If shuffle was checked before, it will be grayed out. Invert normal cell then
                    item = self.table_dragdrop.item(rowPosition, 6)
                    item.setFlags(item.flags() |QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable )

        #Finally, update the Data-Overview-Box
        self.dataOverviewOn()#update the overview box

    def item_dclick(self, item):
        #Check/Uncheck if item is from column 2 or 3
        tableitem = self.table_dragdrop.item(item.row(), item.column())
        if item.column() in [2,3]:
            #If the item is unchecked ->check it!
            if tableitem.checkState() == QtCore.Qt.Unchecked:
                tableitem.setCheckState(QtCore.Qt.Checked)
            #else, the other way around
            elif tableitem.checkState() == QtCore.Qt.Checked:
                tableitem.setCheckState(QtCore.Qt.Unchecked)  
        
        #Show example image if item on column 0 was dclicked
        if item.column() == 0: 
            #rtdc_path = str(item.text())
            #rtdc_path = tableitem.text()
            rtdc_path = self.table_dragdrop.item(item.row(), item.column()).text()

            failed,rtdc_ds = aid_bin.load_rtdc(rtdc_path)
            if failed:
                msg = QtWidgets.QMessageBox()
                msg.setIcon(QtWidgets.QMessageBox.Critical)       
                msg.setText(str(rtdc_ds))
                msg.setWindowTitle("Error occurred during loading file")
                msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                msg.exec_()
                    
            nr_images = rtdc_ds["events"]["image"].len()
            ind = np.random.randint(0,nr_images)
            img = rtdc_ds["events"]["image"][ind]
            if len(img.shape)==2:
                height, width = img.shape
                channels = 1
            elif len(img.shape)==3:
                height, width, channels = img.shape
            else:
                print("Invalid image format: "+str(img.shape))
                return
            self.w = MyPopup()
            self.gridLayout_w = QtWidgets.QGridLayout(self.w)
            self.label_image = QtWidgets.QLabel(self.w)
            self.label_cropimage = QtWidgets.QLabel(self.w)

            #zoom image such that longest side is 512
            zoom_factor = np.round(float(512.0/np.max(img.shape)),0)
            #Get the order, specified in Image processing->Zoom Order
            zoom_order = int(self.comboBox_zoomOrder.currentIndex()) #the combobox-index is already the zoom order
            #Convert to corresponding cv2 zooming method
            zoom_interpol_method = aid_img.zoom_arguments_scipy2cv(zoom_factor,zoom_order)

            img_zoomed = cv2.resize(img, dsize=None,fx=zoom_factor, fy=zoom_factor, interpolation=eval(zoom_interpol_method))
            
            if channels==1:
                height, width = img_zoomed.shape
            if channels==3:
                height, width, _ = img_zoomed.shape
            
            if channels==1:
                qi=QtGui.QImage(img_zoomed.data, width, height,width, QtGui.QImage.Format_Indexed8)
            if channels==3:
                qi = QtGui.QImage(img_zoomed.data,img_zoomed.shape[1], img_zoomed.shape[0], QtGui.QImage.Format_RGB888)
                
            self.label_image.setPixmap(QtGui.QPixmap.fromImage(qi))
            self.gridLayout_w.addWidget(self.label_image, 1,1)
              
            #get the location of the cell
            rowPosition = item.row()
            pix = float(self.table_dragdrop.item(rowPosition, 7).text())
            #pix = rtdc_ds.config["imaging"]["pixel size"]
            PIX = pix
            
            pos_x,pos_y = rtdc_ds["events"]["pos_x"][ind]/PIX,rtdc_ds["events"]["pos_y"][ind]/PIX
            cropsize = self.spinBox_imagecrop.value()
            y1 = int(round(pos_y))-cropsize/2                
            x1 = int(round(pos_x))-cropsize/2 
            y2 = y1+cropsize                
            x2 = x1+cropsize
            
            #Crop the image
            img_crop = img[int(y1):int(y2),int(x1):int(x2)]
            #zoom image such that the height gets the same as for non-cropped img
            zoom_factor = float(img_zoomed.shape[0])/img_crop.shape[0]
            
            if zoom_factor == np.inf:
                factor = 1
                if self.actionVerbose.isChecked()==True:
                    print("Set resize factor to 1. Before, it was: "+str(factor))     
            #Get the order, specified in Image processing->Zoom Order
            zoom_order = str(self.comboBox_zoomOrder.currentText()) #
            zoom_interpol_method = aid_img.zoom_arguments_scipy2cv(zoom_factor,zoom_order)
            img_crop = cv2.resize(img_crop, dsize=None,fx=zoom_factor, fy=zoom_factor, interpolation=eval(zoom_interpol_method))
            
            if channels==1:
                height, width = img_crop.shape
                qi=QtGui.QImage(img_crop.data, width, height,width, QtGui.QImage.Format_Indexed8)

            if channels==3:
                height, width, _ = img_crop.shape
                qi = QtGui.QImage(img_crop.data,width, height, QtGui.QImage.Format_RGB888)
            
            self.label_cropimage.setPixmap(QtGui.QPixmap.fromImage(qi))
            self.gridLayout_w.addWidget(self.label_cropimage, 1,2)
            self.w.show()

    def get_norm_from_modelparafile(self):
        #Get the normalization method from a modelparafile
        filename = QtWidgets.QFileDialog.getOpenFileName(self, 'Open meta-data', Default_dict["Path of last model"],"AIDeveloper Meta file (*meta.xlsx)")
        filename = filename[0]
        if len(str(filename))==0:
            return
        norm = pd.read_excel(filename,sheet_name='Parameters',engine="openpyxl")["Normalization"]
        norm = str(norm[0])
        index = self.comboBox_Normalization.findText(norm, QtCore.Qt.MatchFixedString)
        if index >= 0:
            self.comboBox_Normalization.setCurrentIndex(index)
            self.w.close()
        else:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("Invalid normalization method was specified.\
            Likely this version of AIDeveloper does not support that normalization method\
            Please define a valid normalization method")
            msg.setDetailedText("Supported normalization methods are: "+"\n".join(self.norm_methods))
            msg.setWindowTitle("Invalid Normalization method")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return
            #raise ValueError("Invalid Normalization method")

    def update_plottingTab(self):           
        #Get current text of combobox (url to data set)
        url = str(self.comboBox_chooseRtdcFile.currentText())
        if len(url)==0:
            return

        failed,rtdc_ds = aid_bin.load_rtdc(url)
        if failed:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Critical)       
            msg.setText(str(rtdc_ds))
            msg.setWindowTitle("Error occurred during loading file")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return

        keys = list(rtdc_ds["events"].keys())
        #find keys of image_channels
        keys_0d,keys_1d,keys_2d = [],[],[]
        for key in keys:
            if type(rtdc_ds["events"][key])==h5py._hl.dataset.Dataset:
                shape = rtdc_ds["events"][key].shape
                if len(shape)==1: #zero-dimensional info (single number per cell)
                    keys_0d.append(key)
                elif len(shape)==2: #one-dimensional info (multiple numbers per cell)
                    keys_1d.append(key)
                elif len(shape)==3: #two-dimensional info (images)
                    keys_2d.append(key)
                elif len(shape)==4: #two-dimensional RBG info (images)
                    keys_2d.append(key)

        #add the traces to the 1d features
        if "trace" in keys:
            for key_trace in list(rtdc_ds["events"]["trace"].keys()):
                keys_1d.append(key_trace+" (RTFDC)")


        #Sort keys_2d: "image" first; "mask" last 
        keys_2d.insert(0, keys_2d.pop(keys_2d.index("image")))
        if "mask" in keys_2d:
            keys_2d.insert(len(keys_2d), keys_2d.pop(keys_2d.index("mask")))        
        
        #Fill those feautues in the comboboxes at the scatterplot
        self.comboBox_featurex.addItems(keys_0d)
        self.comboBox_featurey.addItems(keys_0d)

        #check if masks or contours are available
        cont_available = "mask" in keys or "contour" in keys
        self.checkBox_contour.setEnabled(cont_available)
        self.checkBox_contour.setChecked(cont_available)
        #Centroid is always available (prerequisite for AIDeveloper)
        self.checkBox_centroid.setEnabled(True)
        self.checkBox_centroid.setChecked(True)

        #Intialize option menus 
        self.contour_options_nr = 0
        self.centroid_options_nr = 0
        self.show_1d_options_nr = 0
        self.show_2d_options_nr = 0
        self.init_contour_options(keys_2d)
        self.init_centroid_options(keys_1d)
        self.init_2d_options(keys_2d)
        self.init_1d_options(keys_1d)
            
    def init_contour_options(self,keys_2d):
        print("Work in progress")
        # self.popup_layercontrols = MyPopup()
        # self.popup_layercontrols_ui = frontend.Ui_LayerControl()
        # self.popup_layercontrols_ui.setupUi(self.popup_layercontrols,keys_2d) #open a popup
    def init_centroid_options(self,keys_image):
        print("Work in progress")
        # self.popup_centroid_options = MyPopup()
        # self.popup_centroid_options_ui = aid_frontend.Ui_centroid_options()
        # self.popup_centroid_options_ui.setupUi(self.popup_centroid_options,keys_image) #open a popup

    def init_2d_options(self,keys_2d):
        #Initialize 2d Option Menu. Range values are saved and manipulated here
        self.popup_2dOptions = MyPopup()
        self.popup_2dOptions_ui = aid_frontend.Ui_2dOptions()
        self.popup_2dOptions_ui.setupUi(self.popup_2dOptions,keys_2d) #open a popup

    def init_1d_options(self,keys_1d):
        self.popup_1dOptions = MyPopup()
        self.popup_1dOptions_ui = aid_frontend.Ui_1dOptions()
        self.popup_1dOptions_ui.setupUi(self.popup_1dOptions,keys_1d) #open a popup

    def show_contour_options(self):
        self.contour_options_nr += 1
        print("Work in progress")

    def show_centroid_options(self):
        print("Work in progress")
        self.centroid_options_nr += 1
        #self.popup_layercontrols_ui.pushButton_close.clicked.connect(self.visualization_settings)
        if self.centroid_options_nr==1:
            for iterator in range(len(self.popup_layercontrols_ui.spinBox_minChX)):
                print(1)

    def show_2d_options(self):
        self.show_2d_options_nr += 1
        #self.popup_layercontrols_ui.pushButton_close.clicked.connect(self.visualization_settings)
        if self.show_2d_options_nr==1:
            for iterator in range(len(self.popup_2dOptions_ui.spinBox_minChX)):
                slider = self.popup_2dOptions_ui.horizontalSlider_chX[iterator]
                slider.startValueChanged.connect(lambda _, b=None: self.put_image(ind=b))
                slider.endValueChanged.connect(lambda _, b=None: self.put_image(ind=b))
                checkBox = self.popup_2dOptions_ui.checkBox_show_chX[iterator]
                checkBox.stateChanged.connect(lambda _, b=None: self.put_image(ind=b))
                comboBox = self.popup_2dOptions_ui.comboBox_cmap_chX[iterator]
                comboBox.currentIndexChanged.connect(lambda _, b=None: self.put_image(ind=b))
                checkBox = self.popup_2dOptions_ui.checkBox_auto_chX[iterator]
                checkBox.stateChanged.connect(lambda _, b=None: self.put_image(ind=b))
        self.popup_2dOptions.show()


    def show_1d_options(self):
        self.show_1d_options_nr += 1
        #self.popup_layercontrols_ui.pushButton_close.clicked.connect(self.visualization_settings)
        if self.show_1d_options_nr==1:
            for iterator in range(len(self.popup_1dOptions_ui.checkBox_show_chX)):
                checkBox = self.popup_1dOptions_ui.checkBox_show_chX[iterator]
                checkBox.stateChanged.connect(lambda _, b=None: self.put_line(index=b))
                comboBox = self.popup_1dOptions_ui.comboBox_cmap_chX[iterator]
                comboBox.clicked.connect(lambda _, b=None: self.put_line(index=b))
        self.popup_1dOptions.show()


    def activate_deactivate_spinbox(self,newstate):
        #get the checkstate of the Input model crop 
        if newstate==2:
            #activate the spinbox
            self.spinBox_imagecrop.setEnabled(True)
        elif newstate==0:
            self.spinBox_imagecrop.setEnabled(False)


    def gray_or_rgb_augmentation(self,index):
        #When Color-Mode is changed:
        #Get the new colormode:
        new_colormode = self.colorModes[index]
        #when the new Color Mode is Grayscale, disable saturation and hue augmentation
        if new_colormode=="Grayscale":

            self.checkBox_contrast.setEnabled(True)
            self.checkBox_contrast.setChecked(True)
            self.doubleSpinBox_contrastLower.setEnabled(True)
            self.doubleSpinBox_contrastHigher.setEnabled(True)

            self.checkBox_saturation.setEnabled(False)
            self.checkBox_saturation.setChecked(False)
            self.doubleSpinBox_saturationLower.setEnabled(False)
            self.doubleSpinBox_saturationHigher.setEnabled(False)

            self.checkBox_hue.setEnabled(False)
            self.checkBox_hue.setChecked(False)
            self.doubleSpinBox_hueDelta.setEnabled(False)

        elif new_colormode=="RGB":
            self.checkBox_contrast.setEnabled(True)
            self.checkBox_contrast.setChecked(True)
            self.doubleSpinBox_contrastLower.setEnabled(True)
            self.doubleSpinBox_contrastHigher.setEnabled(True)

            self.checkBox_saturation.setEnabled(True)
            self.checkBox_saturation.setChecked(True)
            self.doubleSpinBox_saturationLower.setEnabled(True)
            self.doubleSpinBox_saturationHigher.setEnabled(True)

            self.checkBox_hue.setEnabled(True)
            self.checkBox_hue.setChecked(True)
            self.doubleSpinBox_hueDelta.setEnabled(True)
        else:
            print("Invalid Color Mode")

           
    def onClick(self,points,pointermethod):
        #delete the last item if the user selected already one:
        try:
            self.scatter_xy.removeItem(self.point_clicked)
        except:
            pass
        
        if pointermethod=="point":
            points = points[0]
            p = points.pos()
            clicked_x, clicked_y = p.x(), p.y()
            a1 = (clicked_x)/float(np.max(self.feature_x))            
            a2 = (clicked_y)/float(np.max(self.feature_y))
            #Which is the closest scatter point?
            dist = np.sqrt(( a1-self.scatter_x_norm )**2 + ( a2-self.scatter_y_norm )**2)
            index =  np.argmin(dist)

        elif pointermethod=="index":
            index = points
            
        clicked_x = self.feature_x[index]
        clicked_y = self.feature_y[index]
        
        self.point_clicked = pg.ScatterPlotItem()
        self.point_clicked.setData([clicked_x], [clicked_y],brush="r",symbol='o',symbolPen="w",size=15)
        self.scatter_xy.addItem(self.point_clicked)
        #self.scatter_xy.plot([clicked_x], [clicked_y],pen=None,symbol='o',symbolPen='w',clear=False)
        self.point_was_selected_before = True

        #I dont care if the user click or used the slider->always adjust spinbox and slider without running the onChange functions 
        self.changedbyuser = False
        self.spinBox_cellInd.setValue(index)
        self.horizontalSlider_cellInd.setValue(index)
        self.changedbyuser = True
        
        self.put_image(index)
        self.put_line(index)

    
    def put_image(self,ind):
        #check that the user is looking at the plotting tab
        curr_ind = self.tabWidget_Modelbuilder.currentIndex()
        if curr_ind!=3:
            return
        try:
            self.widget_showCell.removeItem(self.dot)
        except:
            pass
        try:
            self.widget_showCell.removeItem(self.plot_contour)
        except:
            pass
        
        if ind==None:
            index = int(self.spinBox_cellInd.value())
        else:
            index = ind
            
        rtdc_ds = self.rtdc_ds
        
        #which channel shouldbe displayed                
        channels = len(self.popup_2dOptions_ui.spinBox_minChX)
        keys_2d = [self.popup_2dOptions_ui.label_layername_chX[i].text() for i in range(channels)]

        #Define variable on self that carries all image information
        if channels==1: 
            img = np.expand_dims(rtdc_ds["events"]["image"][index],-1)
        elif channels>1:
            img = np.stack( [rtdc_ds["events"][key][index] for key in keys_2d] ,axis=-1)            

        if len(img.shape)==2:
            channels = 1
        elif len(img.shape)==3:
            height, width, channels = img.shape
        else:
            print("Invalid image format: "+str(img.shape))
            return

        color_mode = str(self.comboBox_GrayOrRGB_2.currentText())
        
        if color_mode=="Grayscale": #Slider allows to show individual layers: each is shown as grayscale
            img = img
            
        elif color_mode == "RGB":#User can define, which layers are shown in R,G,and B
            #Retrieve the setting from self.popup_layercontrols_ui
            ui_item = self.popup_2dOptions_ui
            layer_names = [obj.text() for obj in ui_item.label_layername_chX]
            layer_active = [obj.isChecked() for obj in ui_item.checkBox_show_chX]
            layer_range = [obj.getRange() for obj in ui_item.horizontalSlider_chX]
            layer_auto = [obj.isChecked() for obj in ui_item.checkBox_auto_chX]
            layer_cmap = [obj.currentText() for obj in ui_item.comboBox_cmap_chX]
    
            #Assemble the image according to the settings in self.popup_layercontrols_ui
            #Find activated layers for each color:
            ind_active_r,ind_active_g,ind_active_b = [],[],[]
            for ch in range(len(layer_cmap)):
            #for color,active in zip(layer_cmap,layer_active):
                if layer_cmap[ch]=="Red" and layer_active[ch]==True:
                    ind_active_r.append(ch)
                if layer_cmap[ch]=="Green" and layer_active[ch]==True:
                    ind_active_g.append(ch)
                if layer_cmap[ch]=="Blue" and layer_active[ch]==True:
                    ind_active_b.append(ch)
            
            if len(ind_active_r)>0:
                img_ch = img[:,:,np.array(ind_active_r)]
                layer_range_ch = np.array(layer_range)[np.array(ind_active_r)] #Range of all red channels 
                layer_auto_ch = np.array(layer_auto)[np.array(ind_active_r)] #Automatic range
                #Scale each red channel according to layer_range
                for layer in range(img_ch.shape[-1]):
                    limits,auto = layer_range_ch[layer],layer_auto_ch[layer]
                    img_ch[:,:,layer] = aid_img.clip_contrast(img=img_ch[:,:,layer],low=limits[0],high=limits[1],auto=auto)
                img_r = np.mean(img_ch,axis=-1).astype(np.uint8)
            else:
                img_r = np.zeros(shape=(img.shape[0],img.shape[1]),dtype=np.uint8)
                
            if len(ind_active_g)>0:
                img_ch = img[:,:,np.array(ind_active_g)]
                layer_range_ch = np.array(layer_range)[np.array(ind_active_g)] #Range of all red channels 
                layer_auto_ch = np.array(layer_auto)[np.array(ind_active_g)] #Automatic range
                #Scale each red channel according to layer_range
                for layer in range(img_ch.shape[-1]):
                    limits,auto = layer_range_ch[layer],layer_auto_ch[layer]
                    img_ch[:,:,layer] = aid_img.clip_contrast(img=img_ch[:,:,layer],low=limits[0],high=limits[1],auto=auto)
                img_g = np.mean(img_ch,axis=-1).astype(np.uint8)
            else:
                img_g = np.zeros(shape=(img.shape[0],img.shape[1]),dtype=np.uint8)
    
            if len(ind_active_b)>0:
                img_ch = img[:,:,np.array(ind_active_b)]
                layer_range_ch = np.array(layer_range)[np.array(ind_active_b)] #Range of all red channels 
                layer_auto_ch = np.array(layer_auto)[np.array(ind_active_b)] #Automatic range
                #Scale each red channel according to layer_range
                for layer in range(img_ch.shape[-1]):
                    limits,auto = layer_range_ch[layer],layer_auto_ch[layer]
                    img_ch[:,:,layer] = aid_img.clip_contrast(img=img_ch[:,:,layer],low=limits[0],high=limits[1],auto=auto)
                img_b = np.mean(img_ch,axis=-1).astype(np.uint8)
            else:
                img_b = np.zeros(shape=(img.shape[0],img.shape[1]),dtype=np.uint8)
            
            #Assemble image by stacking all layers
            img = np.stack([img_r,img_g,img_b],axis=-1)        

        
        #Get the levels of the previous frame
        levels_init = self.widget_showCell.getLevels()
        if levels_init==(0,1.0):
            levels_init = (0,255)
        
        #Get the layer index of the previous frame
        index_ = self.widget_showCell.currentIndex
            
        if color_mode=="Grayscale":
            self.widget_showCell.setImage(img.T,autoRange=False,levels=levels_init,levelMode="mono")
            
            self.widget_showCell.setCurrentIndex(index_)
        elif color_mode=="RGB":
            self.widget_showCell.setImage(np.swapaxes(img,0,1))

        pix = rtdc_ds.attrs["imaging:pixel size"]
        pos_x = rtdc_ds["events"]["pos_x"][index]/pix
        pos_y = rtdc_ds["events"]["pos_y"][index]/pix

        #Indicate the centroid of the cell        
        if self.checkBox_centroid.isChecked():
            self.dot = pg.CircleROI(pos=(pos_x-2, pos_y-2), size=4, pen=QtGui.QPen(QtCore.Qt.red, 0.1), movable=False)
            self.widget_showCell.getView().addItem(self.dot)
            self.widget_showCell.show()
        
        if self.checkBox_contour.isChecked():
            #get the contour based on the mask
            contour,_ = cv2.findContours(rtdc_ds["events"]["mask"][index], cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_NONE)
            contour = contour[0][:,0,:]
            self.plot_contour = pg.PlotCurveItem(contour[:,0],contour[:,1],width=6,pen="r")
            self.widget_showCell.getView().addItem(self.plot_contour)

    def put_line(self,index):
        curr_ind = self.tabWidget_Modelbuilder.currentIndex()
        if curr_ind!=3:
            return
        
        #Fluorescence traces: clear first
        try:
            self.plot_fl_trace_.clear() #clear the plot
            self.plot_fl_trace.clear() #clear the plot
        except:
            pass

        if index==None:
            index = int(self.spinBox_cellInd.value())

        rtdc_ds = self.rtdc_ds
        feature_keys = list(rtdc_ds.keys())
        
        #which features shouldbe displayed                
        features_nr = len(self.popup_1dOptions_ui.checkBox_show_chX)
        keys_1d = [self.popup_1dOptions_ui.checkBox_show_chX[i].text() for i in range(features_nr)]
        keys_1d_on = [self.popup_1dOptions_ui.checkBox_show_chX[i].isChecked() for i in range(features_nr)]
        colors = [self.popup_1dOptions_ui.comboBox_cmap_chX[i].palette().button().color() for i in range(features_nr)]
        colors = [list(c.getRgb()) for c in colors]
        colors = [tuple(c) for c in colors]
        ind = np.where(np.array(keys_1d_on)==True)[0]
        keys_1d = list(np.array(keys_1d)[ind])
        colors = list(np.array(colors)[ind])
                
        for key_1d,color in zip(keys_1d,colors):
            if key_1d.endswith(" (RTFDC)"):
                key_1d = key_1d.split(" (RTFDC)")[0]
                trace_flx = rtdc_ds["events"]["trace"][key_1d][index]
                pencolor = pg.mkPen(color, width=2)
                self.plot_fl_trace_ = self.plot_fl_trace.plot(range(len(trace_flx)),trace_flx,width=6,pen=pencolor,clear=False)
                # if "fl1_max" in feature_keys and "fl1_pos" in feature_keys: #if also the maxima and position of the max are available: use it to put the region accordingly
                #     fl1_max,fl1_pos = rtdc_ds["events"]["fl1_max"][index],rtdc_ds["events"]["fl1_pos"][index]
            else:
                values = rtdc_ds["events"][key_1d][index]
                pencolor = pg.mkPen(color, width=2)
                self.plot_fl_trace_ = self.plot_fl_trace.plot(range(len(trace_flx)),trace_flx,width=6,pen=pencolor,clear=False)

                #get the maximum of [fl1_max,fl2_max,fl3_max] and put the region to the corresponding fl-position
                # ind = np.argmax(np.array([fl1_max,fl2_max,fl3_max]))
                # region_pos = np.array([fl1_pos,fl2_pos,fl3_pos])[ind] #this region is already given in us. translate this back to range
                # peak_height = np.array([fl1_max,fl2_max,fl3_max])[ind]
                # sample_rate = rtdc_ds.attrs["fluorescence:sample rate"]
                # fl_pos_ind = float((sample_rate*region_pos))/1E6 #
                # #Indicate the used flx_max and flx_pos by a scatter dot
                # self.peak_dot = self.plot_fl_trace.plot([float(fl_pos_ind)], [float(peak_height)],pen=None,symbol='o',symbolPen='w',clear=False)

    def onScatterClick(self,event, points):
        pointermethod = 'point'
        if self.changedbyuser:
            self.onClick(points,pointermethod)
    def onIndexChange(self,index):
        pointermethod = 'index'
        if self.changedbyuser:
            self.onClick(index,pointermethod)
        #Set self.changedbyuser to False and change the spinbox and slider. changedbyuser=False prevents onClick function
        self.changedbyuser = False
        self.spinBox_cellInd.setValue(index)
        self.horizontalSlider_cellInd.setValue(index)
        self.changedbyuser = True

    def updateScatterPlot(self):
        #If the Plot is updated, delete the dot in the cell-image
        try:
            self.widget_showCell.removeItem(self.dot)
        except:
            pass
        
        try:
            self.scatter_xy.removeItem(self.point_clicked)
        except:
            pass

        self.point_was_selected_before = False
        #read url from current comboBox_chooseRtdcFile
        url = str(self.comboBox_chooseRtdcFile.currentText())
        if len(url)==0:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("Please use the 'Build' tab to load files first")
            msg.setWindowTitle("No file selected")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return

        failed,rtdc_ds = aid_bin.load_rtdc(url)
        if failed:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Critical)       
            msg.setText(str(rtdc_ds))
            msg.setWindowTitle("Error occurred during loading file")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return

        self.rtdc_ds = rtdc_ds

        feature_x_name = str(self.comboBox_featurex.currentText())
        feature_y_name = str(self.comboBox_featurey.currentText())
        
        features = list(self.rtdc_ds["events"].keys())
        if feature_x_name in features:
            self.feature_x = self.rtdc_ds["events"][feature_x_name]
        else:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("Feature on x axis is not contained in data set")
            msg.setWindowTitle("Invalid x feature")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return
        if feature_y_name in features:    
            self.feature_y = self.rtdc_ds["events"][feature_y_name]
        else:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("Feature on y axis is not contained in data set")
            msg.setWindowTitle("Invalid y feature")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return            

        self.changedbyuser = True #variable used to prevent plotting if spinbox or slider is changed programmatically
        
        #density estimation
        kde = self.comboBox_kde.currentText()
        if kde=="None":
            brush = "b"
        
        elif kde=="2d Histogram" or kde=="Gauss":
            if kde=="2d Histogram":
                density = aid_bin.kde_histogram(np.array(self.feature_x), np.array(self.feature_y))
            elif kde=="Gauss":
                density = aid_bin.kde_gauss(np.array(self.feature_x), np.array(self.feature_y))
            
            density_min,density_max = np.min(density),np.max(density)
            density = (density-density_min)/density_max
            
            # define colormap
            brush = []
            from pyqtgraph.graphicsItems.GradientEditorItem import Gradients
            cmap = pg.ColorMap(*zip(*Gradients["viridis"]["ticks"]))
            for k in density:
                brush.append(cmap.mapToQColor(k))

        #Add plot  
        #self.scatter = self.scatter_xy.plot(np.array(self.feature_x), np.array(self.feature_y),symbolPen=None,pen=None,symbol='o',brush=brush[100],clear=True)
        #try to remove existing scatterplot
        try:
            self.scatter_xy.removeItem(self.scatter)
        except:
            print("Not cleared")

        self.scatter = pg.ScatterPlotItem()
        self.scatter.setData(np.array(self.feature_x), np.array(self.feature_y),brush=brush,symbolPen=None,pen=None,symbol='o',size=10)
        self.scatter_xy.addItem(self.scatter)
        #pen=None,symbol='o',symbolPen=None,symbolBrush=density,clear=True)   
        
        self.scatter.sigClicked.connect(self.onScatterClick) #When scatterplot is clicked, show the desired cell

        #Fill histogram for x-axis; widget_histx
        y,x = np.histogram(self.feature_x, bins='auto')
        self.hist_x.plot(x, y, stepMode=True, fillLevel=0, brush=(0,0,255,150),clear=True)
        #Manually clear y hist first. Only clear=True did not do the job
        self.hist_y.clear()
        #Fill histogram for y-axis; widget_histy
        y,x = np.histogram(self.feature_y, bins='auto')
        curve = pg.PlotCurveItem(-1.*x, y, stepMode=True, fillLevel=0, brush=(0, 0, 255, 150),clear=True)
        curve.rotate(-90)
        self.hist_y.addItem(curve)
        
        self.scatter_x_norm = (np.array(self.feature_x).astype(np.float32))/float(np.max(self.feature_x))
        self.scatter_y_norm = (np.array(self.feature_y).astype(np.float32))/float(np.max(self.feature_y))

        #Adjust the horizontalSlider_cellInd and spinBox_cellInd
        self.horizontalSlider_cellInd.setSingleStep(1)
        self.horizontalSlider_cellInd.setMinimum(0)
        self.horizontalSlider_cellInd.setMaximum(len(self.feature_x)-1)
        self.spinBox_cellInd.setMinimum(0)
        self.spinBox_cellInd.setMaximum(len(self.feature_x)-1)                  

        
    def selectPeakPos(self):
        #Check if self.region exists
        #If not, show a message and return:
        if not hasattr(self, 'region'):
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("There is no region defined yet")
            msg.setWindowTitle("No region defined")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return   
        #Try to get the user defined peak position
        if not hasattr(self, 'new_peak'):
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("There is no peak defined yet")
            msg.setWindowTitle("No peak defined")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return
                
        #how much rows are already in table?
        rowcount = self.tableWidget_showSelectedPeaks.rowCount()
        self.tableWidget_showSelectedPeaks.setRowCount(rowcount+1)
        rowPosition = rowcount        
        item = QtWidgets.QTableWidgetItem()
        item.setData(QtCore.Qt.EditRole, float(self.new_peak["fl_max"]))
        self.tableWidget_showSelectedPeaks.setItem(rowPosition, 0, item)
        item = QtWidgets.QTableWidgetItem()
        fl_pos_us = float(float(self.new_peak["fl_pos"])*float(1E6))/float(self.rtdc_ds.attrs["fluorescence:sample rate"])
        item.setData(QtCore.Qt.EditRole,fl_pos_us)
        self.tableWidget_showSelectedPeaks.setItem(rowPosition, 1, item)
        item = QtWidgets.QTableWidgetItem()
        pos_x_um = float(self.new_peak["pos_x"])*float(self.rtdc_ds.attrs["imaging:pixel size"])
        item.setData(QtCore.Qt.EditRole,pos_x_um)
        self.tableWidget_showSelectedPeaks.setItem(rowPosition, 2, item)
        item = QtWidgets.QTableWidgetItem()
        item.setData(QtCore.Qt.EditRole, float(self.new_peak["fl_pos"]))
        self.tableWidget_showSelectedPeaks.setItem(rowPosition, 3, item)
        item = QtWidgets.QTableWidgetItem()
        item.setData(QtCore.Qt.EditRole, float(self.new_peak["pos_x"]))
        self.tableWidget_showSelectedPeaks.setItem(rowPosition, 4, item)
        self.tableWidget_showSelectedPeaks.resizeColumnsToContents()            
        self.tableWidget_showSelectedPeaks.resizeRowsToContents()

        #Update the widget_showSelectedPeaks
        self.update_peak_plot()

    def selectPeakRange(self):
        new_region = self.region.getRegion()
        region_width = np.max(new_region) - np.min(new_region) #in [samples]
        sample_rate = self.rtdc_ds.attrs["fluorescence:sample rate"]
        region_width = (float(region_width)/float(sample_rate))*1E6 #range[samples]*(1/sample_rate[1/s]) = range[s]; div by 1E6 to conver to us
        self.region_width = region_width
        #put this in the table
        item = QtWidgets.QTableWidgetItem()
        item.setData(QtCore.Qt.EditRole, "Range [us]")
        self.tableWidget_peakModelParameters.setItem(1, 0, item)
        item = QtWidgets.QTableWidgetItem()
        item.setData(QtCore.Qt.EditRole, float(self.region_width))
        self.tableWidget_peakModelParameters.setItem(1, 1, item)
        item = QtWidgets.QTableWidgetItem()
        
    def onPeaksPlotClick(self,event, points):
        points = points[0]
        p = points.pos()
        clicked_x, clicked_y = p.x(), p.y()
        a1 = (clicked_x)/float(np.max(self.Pos_x))            
        a2 = (clicked_y)/float(np.max(self.Fl_pos))
        
        #Which is the closest scatter point?
        pos_x_norm = self.Pos_x/np.max(self.Pos_x)#normalized pos_x
        fl_pos_norm = self.Fl_pos/np.max(self.Fl_pos)#normalized fl_pos
        dist = np.sqrt(( a1-pos_x_norm )**2 + ( a2-fl_pos_norm )**2)
        index =  np.argmin(dist)
        #Highlight this row
        self.tableWidget_showSelectedPeaks.selectRow(index)
        #Delete the highlighted rows
#        try:
#            self.actionRemoveSelectedPeaks_function()
#        except:
#            pass
        
    def update_peak_plot(self):
        #This function reads tableWidget_showSelectedPeaks and 
        #fits a function and 
        #puts fitting parameters on tableWidget_peakModelParameters
        
        #read the data on tableWidget_showSelectedPeaks        
        rowcount = self.tableWidget_showSelectedPeaks.rowCount()
        Fl_pos,Pos_x = [],[]
        for row in range(rowcount):
            line = [float(self.tableWidget_showSelectedPeaks.item(row, col).text()) for col in [1,2]] #use the values for [us] and [um]
            Fl_pos.append(line[0])
            Pos_x.append(line[1])
            
        self.Fl_pos = np.array(Fl_pos)
        self.Pos_x = np.array(Pos_x)
        
        self.selectedPeaksPlotPlot = self.selectedPeaksPlot.plot(self.Pos_x, self.Fl_pos,pen=None,symbol='o',symbolPen=None,symbolBrush='b',clear=True)
        #if user clicks in the plot, show him the corresponding row in the table
        self.selectedPeaksPlotPlot.sigPointsClicked.connect(self.onPeaksPlotClick)

        if not hasattr(self, 'region_width'): #if there was no region_width defined yet...
            #to get a reasonable initial range, use 20% of the nr. of availeble samples
            samples_per_event = self.rtdc_ds.attrs["fluorescence:samples per event"]
            self.region_width = 0.2*samples_per_event #width of the region in samples
            #Convert to SI unit:
            sample_rate = self.rtdc_ds.attrs["fluorescence:sample rate"] 
            self.region_width = (float(self.region_width)/float(sample_rate))*1E6 #range[samples]*(1/sample_rate[1/s]) = range[s]; div by 1E6 to convert to us

        #which model should be used?
        if str(self.comboBox_peakDetModel.currentText()) == "Linear dependency and max in range" and len(Pos_x)>1:
            slope,intercept = np.polyfit(Pos_x, Fl_pos,deg=1) #Linear FIT, y=mx+n; y=FL_pos[us] x=Pos_x[um]
            xlin = np.round(np.linspace(np.min(Pos_x),np.max(Pos_x),25),1)
            ylin = intercept + slope*xlin
            self.selectedPeaksPlot.plot(xlin, ylin,width=6,pen='b',clear=False)
            
            #Put info to tableWidget_peakModelParameters
            self.tableWidget_peakModelParameters.setColumnCount(2)
            self.tableWidget_peakModelParameters.setRowCount(5)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.EditRole, "Model")
            self.tableWidget_peakModelParameters.setItem(0, 0, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.EditRole, "Linear dependency and max in range")
            self.tableWidget_peakModelParameters.setItem(0, 1, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.EditRole, "Range [us]")
            self.tableWidget_peakModelParameters.setItem(1, 0, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.EditRole, float(self.region_width))
            self.tableWidget_peakModelParameters.setItem(1, 1, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.EditRole, "Intercept [us]")
            self.tableWidget_peakModelParameters.setItem(2, 0, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.EditRole, float(intercept))
            self.tableWidget_peakModelParameters.setItem(2, 1, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.EditRole, "Slope [us/um]")
            self.tableWidget_peakModelParameters.setItem(3, 0, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.EditRole, float(slope))
            self.tableWidget_peakModelParameters.setItem(3, 1, item)
            #Calculate velocity
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.EditRole, "Velocity[m/s]")
            self.tableWidget_peakModelParameters.setItem(4, 0, item)
            velocity = float(1.0/float(slope))
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.EditRole, float(velocity))
            self.tableWidget_peakModelParameters.setItem(4, 1, item)

            
    def addHighestXPctPeaks(self):
        #how many x%?
        x_pct = float(self.doubleSpinBox_highestXPercent.value())
        #Get the flourescence traces and maxima/positions of maxima
        #->it could be that the user did not yet load the dataset:
        if not hasattr(self,"rtdc_ds"):
            #run the function updateScatterPlot()
            self.updateScatterPlot()
            
        trace = self.rtdc_ds["events"]["trace"]
        fl_keys = list(trace.keys())
        fl1_max,fl1_pos,fl2_max,fl2_pos,fl3_max,fl3_pos,pos_x = [],[],[],[],[],[],[]
        for i in range(len(fl_keys)):
            if "fl1_median" in fl_keys[i] and self.checkBox_fl1.isChecked():
                for index in range(len(trace[fl_keys[i]])):
                    trace_flx = trace[fl_keys[i]][index]
                    ind = np.argmax(trace_flx)
                    fl1_max.append(trace_flx[ind])
                    fl1_pos.append(ind)
                #Get the x% maxima
                fl1_max = np.array(fl1_max)
                fl1_pos = np.array(fl1_pos)
                sorter = np.argsort(fl1_max)[::-1]
                sorter = sorter[0:int(x_pct/100.0*len(fl1_max))]
                fl1_max = fl1_max[sorter]
                fl1_pos = fl1_pos[sorter]
                pos_x.append(self.rtdc_ds["events"]["pos_x"][sorter])
                
            elif "fl2_median" in fl_keys[i] and self.checkBox_fl2.isChecked():
                for index in range(len(trace[fl_keys[i]])):
                    trace_flx = trace[fl_keys[i]][index]
                    ind = np.argmax(trace_flx)
                    fl2_max.append(trace_flx[ind])
                    fl2_pos.append(ind)
                #Get the x% maxima
                fl2_max = np.array(fl2_max)
                fl2_pos = np.array(fl2_pos)
                sorter = np.argsort(fl2_max)[::-1]
                sorter = sorter[0:int(x_pct/100.0*len(fl2_max))]
                fl2_max = fl2_max[sorter]
                fl2_pos = fl2_pos[sorter]
                pos_x.append(self.rtdc_ds["events"]["pos_x"][sorter])

            elif "fl3_median" in fl_keys[i] and self.checkBox_fl3.isChecked():
                for index in range(len(trace[fl_keys[i]])):
                    trace_flx = trace[fl_keys[i]][index]
                    ind = np.argmax(trace_flx)
                    fl3_max.append(trace_flx[ind])
                    fl3_pos.append(ind)
                #Get the x% maxima
                fl3_max = np.array(fl3_max)
                fl3_pos = np.array(fl3_pos)
                sorter = np.argsort(fl3_max)[::-1]
                sorter = sorter[0:int(x_pct/100.0*len(fl3_max))]
                fl3_max = fl3_max[sorter]
                fl3_pos = fl3_pos[sorter]
                pos_x.append(self.rtdc_ds["events"]["pos_x"][sorter])

        #Add fl1 fl2 and fl3 information
        flx_max = np.array(list(fl1_max)+list(fl2_max)+list(fl3_max))
        flx_pos = np.array(list(fl1_pos)+list(fl2_pos)+list(fl3_pos))
        pos_x_um = np.concatenate(np.atleast_2d(np.array(pos_x)))
        pix = self.rtdc_ds.attrs["imaging:pixel size"]
        pos_x = pos_x_um/pix #convert from um to pix

        rowcount = self.tableWidget_showSelectedPeaks.rowCount()
        self.tableWidget_showSelectedPeaks.setRowCount(rowcount+len(flx_max))
        
        for i in range(len(flx_max)):
            rowPosition = rowcount+i
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.EditRole, float(flx_max[i]))
            self.tableWidget_showSelectedPeaks.setItem(rowPosition, 0, item)
            item = QtWidgets.QTableWidgetItem()
            fl_pos_us = float(float(flx_pos[i])*float(1E6))/float(self.rtdc_ds.attrs["fluorescence:sample rate"] )
            item.setData(QtCore.Qt.EditRole,fl_pos_us)
            self.tableWidget_showSelectedPeaks.setItem(rowPosition, 1, item)
            item = QtWidgets.QTableWidgetItem()
            #pos_x_um = float(pos_x[i])*float(self.rtdc_ds.config["imaging"]["pixel size"])
            item.setData(QtCore.Qt.EditRole,float(pos_x_um[i]))
            self.tableWidget_showSelectedPeaks.setItem(rowPosition, 2, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.EditRole, float(flx_pos[i]))
            self.tableWidget_showSelectedPeaks.setItem(rowPosition, 3, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.EditRole, float(pos_x[i]))
            self.tableWidget_showSelectedPeaks.setItem(rowPosition, 4, item)
        #Update the widget_showSelectedPeaks
        self.update_peak_plot()



    def savePeakDetModel(self):
        #Get tableWidget_peakModelParameters and write it to excel file
        #Get filename from user:
        filename = QtWidgets.QFileDialog.getSaveFileName(self, 'Save peak fitting model', Default_dict["Path of last model"],"Excel file (*.xlsx)")
        filename = filename[0]
        if len(filename)==0:
            return
        #add the suffix .csv
        if not filename.endswith(".xlsx"):
            filename = filename +".xlsx"               

        table = self.tableWidget_peakModelParameters
        cols = table.columnCount()
        header = range(cols)
        rows = table.rowCount()
        model_df = pd.DataFrame(columns=header,index=range(rows)) 
        for i in range(rows):
            for j in range(cols):
                try:
                    model_df.iloc[i, j] = table.item(i, j).text()
                except:
                    model_df.iloc[i, j] = np.nan

        table = self.tableWidget_showSelectedPeaks
        cols = table.columnCount()
        header = [table.horizontalHeaderItem(col).text() for col in range(cols)]
        rows = table.rowCount()
        peaks_df = pd.DataFrame(columns=header,index=range(rows)) 
        for i in range(rows):
            for j in range(cols):
                try:
                    peaks_df.iloc[i, j] = table.item(i, j).text()
                except:
                    peaks_df.iloc[i, j] = np.nan

        writer = pd.ExcelWriter(filename, engine='openpyxl')
        #Used files go to a separate sheet on the MetaFile.xlsx
        pd.DataFrame().to_excel(writer,sheet_name='Model') #initialize empty Sheet
        model_df.to_excel(writer,sheet_name='Model') #initialize empty Sheet
        pd.DataFrame().to_excel(writer,sheet_name='Peaks') #initialize empty Sheet
        peaks_df.to_excel(writer,sheet_name='Peaks')
        writer.save()
        writer.close()

    def loadPeakDetModel(self):
        filename = QtWidgets.QFileDialog.getOpenFileName(self, 'Open peak fitting model', Default_dict["Path of last model"],"Excel file (*.xlsx)")
        filename = filename[0]
        if len(str(filename))==0:
            return
        peak_model_df = pd.read_excel(filename,sheet_name='Model',engine="openpyxl")
        model = peak_model_df.iloc[0,1]
        if model=="Linear dependency and max in range":
            #set the combobox accordingly
            index = self.comboBox_peakDetModel.findText(model, QtCore.Qt.MatchFixedString)
            if index >= 0:
                self.comboBox_peakDetModel.setCurrentIndex(index)
            else:
                msg = QtWidgets.QMessageBox()
                msg.setIcon(QtWidgets.QMessageBox.Information)       
                msg.setText("Could not find a valid model in the chosen file. Did you accidentially load a session or history file?!")
                msg.setWindowTitle("No valid model found")
                msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                msg.exec_()
                return                

            range_ = float(peak_model_df.iloc[1,1])
            intercept = float(peak_model_df.iloc[2,1])
            slope = float(peak_model_df.iloc[3,1])
            velocity = float(peak_model_df.iloc[4,1])

            #put the information in the table
            xlin = np.round(np.linspace(np.min(0),np.max(100),25),1)
            ylin = intercept + slope*xlin
            self.selectedPeaksPlot.plot(xlin, ylin,width=6,pen='b',clear=False)
            
            #Put info to tableWidget_peakModelParameters
            self.tableWidget_peakModelParameters.setColumnCount(2)
            self.tableWidget_peakModelParameters.setRowCount(5)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.EditRole, "Model")
            self.tableWidget_peakModelParameters.setItem(0, 0, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.EditRole, "Linear dependency and max in range")
            self.tableWidget_peakModelParameters.setItem(0, 1, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.EditRole, "Range [us]")
            self.tableWidget_peakModelParameters.setItem(1, 0, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.EditRole, float(range_))
            self.tableWidget_peakModelParameters.setItem(1, 1, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.EditRole, "Intercept [us]")
            self.tableWidget_peakModelParameters.setItem(2, 0, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.EditRole, float(intercept))
            self.tableWidget_peakModelParameters.setItem(2, 1, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.EditRole, "Slope [us/um]")
            self.tableWidget_peakModelParameters.setItem(3, 0, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.EditRole, float(slope))
            self.tableWidget_peakModelParameters.setItem(3, 1, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.EditRole, "Velocity[m/s]")
            self.tableWidget_peakModelParameters.setItem(4, 0, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.EditRole, float(velocity))
            self.tableWidget_peakModelParameters.setItem(4, 1, item)


    def applyPeakModel_and_export(self):
        print("Deprecated. Please use earlier version if AIDeveloper (<0.4.0)")



    def partialtrainability_activated(self,on_or_off):
        if on_or_off==False:#0 means switched OFF 
            
            self.lineEdit_partialTrainability.setText("")
            #self.lineEdit_partialTrainability.setEnabled(False)#enables the lineEdit which shows the trainability status of each layer.
            self.pushButton_partialTrainability.setEnabled(False)
            #Also, remove the model from self!
            self.model_keras = None

            self.radioButton_NewModel.setChecked(False)
            self.radioButton_LoadRestartModel.setChecked(False)
            self.radioButton_LoadContinueModel.setChecked(False)
            self.lineEdit_LoadModelPath.setText("")#put the filename in the lineedit
            
        #this happens when the user activated the expert option "partial trainability"
        elif on_or_off==True:#2 means switched ON
            #Has the user already chosen a model?
            if self.model_keras == None: #if there is no model yet chosen
                self.action_initialize_model(duties="initialize")
            #If there is still no model...
            if self.model_keras == None:# or self.model_keras_path==None: #if there is no model yet chosen
                #Tell the user to initiate a model first!
                msg = QtWidgets.QMessageBox()
                msg.setIcon(QtWidgets.QMessageBox.Information)       
                msg.setText("<html><head/><body><p>To use this option please first select and load a model. To do that choose/load a model in 'Define Model'-Tab and hit the button 'Initialize/Fit Model'. Choose to only initialize the model.</p></body></html>")
                msg.setWindowTitle("Please load a model first")
                msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                msg.exec_()
                
                #Switch off
                self.lineEdit_partialTrainability.setText("")
                self.radioButton_NewModel.setChecked(False)
                self.radioButton_LoadRestartModel.setChecked(False)
                self.radioButton_LoadContinueModel.setChecked(False)
                self.lineEdit_LoadModelPath.setText("")
                #self.lineEdit_partialTrainability.setEnabled(False)#enables the lineEdit which shows the trainability status of each layer.
                self.pushButton_partialTrainability.setEnabled(False)
                self.checkBox_partialTrainability.setChecked(False)
                return

            #Otherwise, there is a model on self and we can continue :)
            
            #Collections are not supported
            if type(self.model_keras)==tuple:
                msg = QtWidgets.QMessageBox()
                msg.setIcon(QtWidgets.QMessageBox.Information)       
                msg.setText("<html><head/><body><p>Partial trainability is not available for collections of models. Please specify a single model.</p></body></html>")
                msg.setWindowTitle("Collections of models not supported for collections of models")
                msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                msg.exec_()
                return
            
            #Switch on lineedit and the button
            #self.lineEdit_partialTrainability.setEnabled(True)#enables the lineEdit which shows the trainability status of each layer.
            self.pushButton_partialTrainability.setEnabled(True)#enables the lineEdit which shows the trainability status of each layer.

            #Load trainability states of the model           
            Layer_types = [self.model_keras.layers[i].__class__.__name__ for i in range(len(self.model_keras.layers))]
            #Count Dense and Conv layers
            is_dense_or_conv = [layer_type in ["Dense","Conv2D"] for layer_type in Layer_types]  
            index = np.where(np.array(is_dense_or_conv)==True)[0]
            Layer_train_status = [self.model_keras.layers[layerindex].trainable for layerindex in index]
            self.lineEdit_partialTrainability.setText(str(Layer_train_status))#enables the lineEdit which shows the trainability status of each layer.



    def partialTrainability(self):
        self.popup_trainability = MyPopup()
        self.popup_trainability_ui = aid_frontend.popup_trainability()
        self.popup_trainability_ui.setupUi(self.popup_trainability) #open a popup to show the layers in a table

        #One can only activate this function when there was a model loaded already!
        #self.model_keras has to exist!!!

        if self.model_keras == None: #if there is no model yet chosen
            self.action_initialize_model(duties="initialize")

        if self.model_keras == None: #if there is still no model...            
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("<html><head/><body><p>To use this option please first select and load a model. To do that choose/load a model in 'Define Model'-Tab and hit the button 'Initialize/Fit Model'. Choose to only initialize the model.</p></body></html>")
            msg.setWindowTitle("Please load a model first")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            #Switch this On in the final version
            self.lineEdit_partialTrainability.setText("")
            self.lineEdit_partialTrainability.setEnabled(False)#enables the lineEdit which shows the trainability status of each layer.
            self.pushButton_partialTrainability.setEnabled(False)
            return

        #Fill information about the model 
        if self.radioButton_NewModel.isChecked():#a new model is loaded          
            self.popup_trainability_ui.lineEdit_pop_pTr_modelPath.setText("New model")
        elif self.radioButton_LoadRestartModel.isChecked():#a new model is loaded          
            load_model_path =  str(self.lineEdit_LoadModelPath.text())
            self.popup_trainability_ui.lineEdit_pop_pTr_modelPath.setText("Restart model: "+load_model_path)
        elif self.radioButton_LoadContinueModel.isChecked():#a new model is loaded          
            load_model_path =  str(self.lineEdit_LoadModelPath.text())
            self.popup_trainability_ui.lineEdit_pop_pTr_modelPath.setText("Continue model: "+load_model_path)

        in_dim = self.model_keras.input_shape
        #Retrieve the color_mode from the model (nr. of channels in last in_dim)
        channels = in_dim[-1] #TensorFlow: channels in last dimension
        out_dim = self.model_keras.output_shape[-1]
        
        self.popup_trainability_ui.spinBox_pop_pTr_inpSize.setValue(int(in_dim[1]))
        self.popup_trainability_ui.spinBox_pop_pTr_outpSize.setValue(int(out_dim))
        if channels==1:
            self.popup_trainability_ui.comboBox_pop_pTr_colorMode.addItem("Grayscale")
        elif channels==3:
            self.popup_trainability_ui.comboBox_pop_pTr_colorMode.addItem("RGB")
            
        #Model summary to textBrowser_pop_pTr_modelSummary
        summary = []
        self.model_keras.summary(print_fn=summary.append)
        summary = "\n".join(summary)
        self.popup_trainability_ui.textBrowser_pop_pTr_modelSummary.setText(summary)
                
        #Work on the tableWidget_pop_pTr_layersTable
        Layer_types = [self.model_keras.layers[i].__class__.__name__ for i in range(len(self.model_keras.layers))]
        #Count Dense and Conv layers
        is_dense_or_conv = [layer_type in ["Dense","Conv2D"] for layer_type in Layer_types]  
        index = np.where(np.array(is_dense_or_conv)==True)[0]
        nr_layers = len(index) #total nr. of dense and conv layers with parameters

        for rowNumber in range(nr_layers):
            layerindex = index[rowNumber]
            columnPosition = 0
            layer = self.model_keras.layers[layerindex]
            rowPosition = self.popup_trainability_ui.tableWidget_pop_pTr_layersTable.rowCount()
            self.popup_trainability_ui.tableWidget_pop_pTr_layersTable.insertRow(rowPosition)
            Name = layer.name
            item = QtWidgets.QTableWidgetItem(Name) 
            item.setFlags( QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable )
            item.setTextAlignment(QtCore.Qt.AlignCenter) # change the alignment
            self.popup_trainability_ui.tableWidget_pop_pTr_layersTable.setItem(rowPosition , columnPosition, item ) #

            columnPosition = 1
            layer_type = layer.__class__.__name__
            item = QtWidgets.QTableWidgetItem(layer_type)             
            item.setFlags( QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable )
            item.setTextAlignment(QtCore.Qt.AlignCenter) # change the alignment
            self.popup_trainability_ui.tableWidget_pop_pTr_layersTable.setItem(rowPosition , columnPosition, item ) #

            columnPosition = 2
            Params = layer.count_params()
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.DisplayRole, Params)
            item.setFlags(item.flags() &~QtCore.Qt.ItemIsEnabled &~ QtCore.Qt.ItemIsSelectable )
            self.popup_trainability_ui.tableWidget_pop_pTr_layersTable.setItem(rowPosition, columnPosition, item)

            columnPosition = 3
            if layer_type == "Dense":
                split_property = "units" #'units' are the number of nodes in dense layers
            elif layer_type == "Conv2D":
                split_property = "filters"
            else:
                print("other splitprop!")
                return
            layer_config = layer.get_config()
            nr_units = layer_config[split_property] #units are either nodes or filters for dense and convolutional layer, respectively
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.DisplayRole, int(nr_units))
            item.setFlags(item.flags() &~QtCore.Qt.ItemIsEnabled &~ QtCore.Qt.ItemIsSelectable )
            self.popup_trainability_ui.tableWidget_pop_pTr_layersTable.setItem(rowPosition, columnPosition, item)

            columnPosition = 4
            #for each item create a spinbopx (trainability)
            spinb = QtWidgets.QDoubleSpinBox(self.popup_trainability_ui.tableWidget_pop_pTr_layersTable)
            spinb.setMinimum(0)
            spinb.setMaximum(1)
            spinb.setSingleStep(0.1)
            trainability = int(layer.trainable) #.trainable actually returns True or False. Make it integer
            spinb.setValue(trainability) #this should be always 1
            self.popup_trainability_ui.tableWidget_pop_pTr_layersTable.setCellWidget(rowPosition, columnPosition, spinb)            

        self.popup_trainability.show()
        
        #self.popup_trainability_ui.pushButton_pop_pTr_reset.clicked.connect(self.pop_pTr_reset)
        self.popup_trainability_ui.pushButton_pop_pTr_update.clicked.connect(self.pop_pTr_update_2)
        self.popup_trainability_ui.pushButton_pop_pTr_ok.clicked.connect(self.pop_pTr_ok)






    ###############Functions for the partial trainability popup################

    def pop_pTr_reset(self):
        #Reset the model to initial state, with partial trainability
        print("Not implemented yet")
        msg = QtWidgets.QMessageBox()
        msg.setIcon(QtWidgets.QMessageBox.Information)       
        msg.setText("<html><head/><body><p>Not implemented yet.</p></body></html>")
        msg.setWindowTitle("Not implemented")
        msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
        msg.exec_()

    def pop_pTr_update_1(self):#main worker function
        #Apply the requested changes and display updated model in table
        pTr_table = self.popup_trainability_ui.tableWidget_pop_pTr_layersTable

        #Read the table: 
        Layer_names,Layer_trainabilities = [],[]
        rowCount = pTr_table.rowCount()
        for row in range(rowCount):
            #Layer_indices.append(str(pTr_table.item(row, 0).text()))
            Layer_names.append(str(pTr_table.item(row, 0).text()))
            Layer_trainabilities.append(float(pTr_table.cellWidget(row, 4).value()))
        Layer_trainabilities = np.array(Layer_trainabilities)
        
        #What are the current trainability statuses of the model
        Layer_trainabilities_orig = np.array([self.model_keras.get_layer(l_name).trainable for l_name in Layer_names])
        diff = abs( Layer_trainabilities - Layer_trainabilities_orig )
        ind = np.where( diff>0 )[0]
        
        #Where do we have a trainability between 0 and 1
        #ind = np.where( (Layer_trainabilities>0) & (Layer_trainabilities<1) )[0]
        if len(ind)>0:
            Layer_trainabilities = list(Layer_trainabilities[ind])
            Layer_names = list(np.array(Layer_names)[ind])
            #Update the model using user-specified trainabilities
            self.model_keras = partial_trainability(self.model_keras,Layer_names,Layer_trainabilities)

            #Update lineEdit_partialTrainability
            Layer_types = [self.model_keras.layers[i].__class__.__name__ for i in range(len(self.model_keras.layers))]
            #Count Dense and Conv layers
            is_dense_or_conv = [layer_type in ["Dense","Conv2D"] for layer_type in Layer_types]  
            index = np.where(np.array(is_dense_or_conv)==True)[0]
            Layer_train_status = [self.model_keras.layers[layerindex].trainable for layerindex in index]
            self.lineEdit_partialTrainability.setText(str(Layer_train_status))#enables the lineEdit which shows the trainability status of each layer.
        else:
            print("Nothing to do. All trainabilities are either 0 or 1")

    def pop_pTr_update_2(self):#call pop_pTr_update_1 to do the work and then update the window
        try:
            self.pop_pTr_update_1()#Change the model on self.model_keras according to the table
            self.partialTrainability()#Update the popup window by calling the partialTrainability function
        except Exception as e: 
            #There is an issue building the model!
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Critical)       
            msg.setText(str(e))
            msg.setWindowTitle("Error occured when building model:")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return       
        
    def pop_pTr_ok(self):
        self.pop_pTr_update_1()#Change the model on self.model_keras according to the table; If 'Update' was used before, there will not be done work again, but the model is used as it is
        #To make the model accessible, it has to be saved to a new .model file
        filename = QtWidgets.QFileDialog.getSaveFileName(self, 'Save model', Default_dict["Path of last model"],"AIDeveloper model file (*.model)")
        filename = filename[0]
        path, fname = os.path.split(filename)
        if len(fname)==0:
            return
        #add the suffix _session.xlsx
        if not fname.endswith(".model"):
            fname = fname +".model"
        filename = os.path.join(path,fname)
        
        self.model_keras.save(filename,save_format='h5')
        #Activate 'load and restart' and put this file
        #Avoid the automatic popup
        self.radioButton_NewModel.setChecked(False)
        self.radioButton_LoadRestartModel.setChecked(False)
        self.radioButton_LoadContinueModel.setChecked(True)

        self.lineEdit_LoadModelPath.setText(filename)#put the filename in the lineedit
        #Destroy the window
        self.popup_trainability = None

        msg = QtWidgets.QMessageBox()
        msg.setIcon(QtWidgets.QMessageBox.Information)       
        msg.setText(tooltips["modelsaved_success"])
        msg.setWindowTitle("Sucessfully created and selected model")
        msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
        msg.exec_()
        


    def lossW_comboB(self,state_nr,listindex):
        if listindex==-1:
            ui_item = self.popup_lossW_ui
        else:
            ui_item = self.fittingpopups_ui[listindex].popup_lossW_ui

        state_str = ui_item.comboBox_lossW.itemText(int(state_nr))
        rows_nr = int(ui_item.tableWidget_lossW.rowCount())
        if rows_nr==0:
            state_str = "None"
            
        if state_str=="None":
            for rowPos in range(rows_nr):
                colPos = 4 #"Loss weights"
                ui_item.tableWidget_lossW.cellWidget(rowPos,colPos).setEnabled(False)
                ui_item.tableWidget_lossW.cellWidget(rowPos,colPos).setValue(1.0)
            
        elif state_str=="Custom":
            for rowPos in range(rows_nr):
                colPos = 4 #"Loss weights"
                ui_item.tableWidget_lossW.cellWidget(rowPos,colPos).setEnabled(True)

        elif state_str=="Balanced":
            #How many cells in total per epoch
            events_epoch = [int(ui_item.tableWidget_lossW.item(rowPos,2).text()) for rowPos in range(rows_nr)]
            classes = [int(ui_item.tableWidget_lossW.item(rowPos,0).text()) for rowPos in range(rows_nr)]
            counter = {}
            for i in range(len(classes)):
                counter[classes[i]]=events_epoch[i]
                
            max_val = float(max(counter.values()))       
            class_weights = {class_id : max_val/num_images for class_id, num_images in counter.items()}                     
            class_weights = list(class_weights.values())
            for rowPos in range(rows_nr):
                colPos = 4 #"Loss weights"
                ui_item.tableWidget_lossW.cellWidget(rowPos,colPos).setEnabled(False)
                ui_item.tableWidget_lossW.cellWidget(rowPos,colPos).setValue(class_weights[rowPos])

    def lossW_ok(self,listindex):
        #This happens when the user presses the OK button on the popup for 
        #custom loss weights
        if listindex==-1:
            ui_item = self
        else:
            ui_item = self.fittingpopups_ui[listindex]
        
        #Which option was used on comboBox_lossW?
        state_str = ui_item.popup_lossW_ui.comboBox_lossW.currentText()
        if state_str=="None":#User left None. This actually means its off
            ui_item.lineEdit_lossW.setText("")
            ui_item.pushButton_lossW.setEnabled(False)
            ui_item.checkBox_lossW.setChecked(False)

        elif state_str=="Custom":#User left None. This actually means its off
            #There are custom values         
            #Read the loss values on the table
            rows_nr = int(ui_item.popup_lossW_ui.tableWidget_lossW.rowCount())
            classes = [int(ui_item.popup_lossW_ui.tableWidget_lossW.item(rowPos,0).text()) for rowPos in range(rows_nr)]
            loss_weights = [float(ui_item.popup_lossW_ui.tableWidget_lossW.cellWidget(rowPos,4).value()) for rowPos in range(rows_nr)]
            counter = {}
            for i in range(len(classes)):
                counter[classes[i]]=loss_weights[i]
            #Put counter (its a dictionary) to lineedit
            ui_item.lineEdit_lossW.setText(str(counter))
        
        elif state_str=="Balanced":#Balanced, the values are computed later fresh, even when user changes the cell-numbers again
            ui_item.lineEdit_lossW.setText("Balanced")

        #Destroy the window        
        ui_item.popup_lossW = None
        
    def lossW_cancel(self,listindex):
        #This happens when the user presses the Cancel button on the popup for 
        #custom loss weights       
        if listindex==-1:
            ui_item = self
        else:
            ui_item = self.fittingpopups_ui[listindex]

        if ui_item.lineEdit_lossW.text()=="":
        #if state_str=="None":#User left None. This actually means its off
            ui_item.lineEdit_lossW.setText("")
            ui_item.pushButton_lossW.setEnabled(False)
            ui_item.checkBox_lossW.setChecked(False)
            ui_item.popup_lossW = None
            return
        #Destroy the window
        ui_item.popup_lossW = None



    def get_norm_from_manualselection(self):
        norm = self.comboBox_w.currentText()
        index = self.comboBox_Normalization.findText(norm, QtCore.Qt.MatchFixedString)
        if index >= 0:
            self.comboBox_Normalization.setCurrentIndex(index)
            self.w.close()
    
    def popup_normalization(self):
            self.w = MyPopup()
            self.gridLayout_w = QtWidgets.QGridLayout(self.w)
            self.gridLayout_w.setObjectName(_fromUtf8("gridLayout"))
            self.verticalLayout_w = QtWidgets.QVBoxLayout()
            self.verticalLayout_w.setObjectName(_fromUtf8("verticalLayout"))
            self.label_w = QtWidgets.QLabel(self.w)
            self.label_w.setAlignment(QtCore.Qt.AlignCenter)
            self.label_w.setObjectName(_fromUtf8("label_w"))
            self.verticalLayout_w.addWidget(self.label_w)
            self.horizontalLayout_2_w = QtWidgets.QHBoxLayout()
            self.horizontalLayout_2_w.setObjectName(_fromUtf8("horizontalLayout_2"))
            self.pushButton_w = QtWidgets.QPushButton(self.w)
            self.pushButton_w.setObjectName(_fromUtf8("pushButton"))
            self.horizontalLayout_2_w.addWidget(self.pushButton_w)
            self.horizontalLayout_w = QtWidgets.QHBoxLayout()
            self.horizontalLayout_w.setObjectName(_fromUtf8("horizontalLayout"))
            self.label_2_w = QtWidgets.QLabel(self.w)
            self.label_2_w.setAlignment(QtCore.Qt.AlignRight|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
            self.label_2_w.setObjectName(_fromUtf8("label_2_w"))
            self.horizontalLayout_w.addWidget(self.label_2_w)
            self.comboBox_w = QtWidgets.QComboBox(self.w)
            self.comboBox_w.setObjectName(_fromUtf8("comboBox"))
            self.comboBox_w.addItems(["Select"]+self.norm_methods)
            self.comboBox_w.setMinimumSize(QtCore.QSize(200,22))
            self.comboBox_w.setMaximumSize(QtCore.QSize(200, 22))
            width=self.comboBox_w.fontMetrics().boundingRect(max(self.norm_methods, key=len)).width()
            self.comboBox_w.view().setFixedWidth(width+10)             
            self.comboBox_w.currentIndexChanged.connect(self.get_norm_from_manualselection)
            self.horizontalLayout_w.addWidget(self.comboBox_w)
            self.horizontalLayout_2_w.addLayout(self.horizontalLayout_w)
            self.verticalLayout_w.addLayout(self.horizontalLayout_2_w)
            self.gridLayout_w.addLayout(self.verticalLayout_w, 0, 0, 1, 1)

            self.w.setWindowTitle("Select normalization method")
            self.label_w.setText("You are about to continue training a pretrained model\n"
    "Please select the meta file of that model to load the normalization method\n"
    "or choose the normalization method manually")
            self.pushButton_w.setText("Load meta file")
            self.label_2_w.setText("Manual \n"
    "selection")

            #one button that allows to load a meta file containing the norm-method
            self.pushButton_w.clicked.connect(self.get_norm_from_modelparafile)
            self.w.show()

    def action_preview_model(self,enabled):#function runs when radioButton_LoadRestartModel or radioButton_LoadContinueModel was clicked
        if enabled: 
            #if the "Load and restart" radiobutton was clicked:
            if self.radioButton_LoadRestartModel.isChecked():
                modelname = QtWidgets.QFileDialog.getOpenFileName(self, 'Open model architecture', Default_dict["Path of last model"],"Architecture or model (*.arch *.model)")
                modelname = modelname[0]
                #modelname_for_dict = modelname
            #if the "Load and continue" radiobutton was clicked:
            elif self.radioButton_LoadContinueModel.isChecked():
                modelname = QtWidgets.QFileDialog.getOpenFileName(self, 'Open model with all parameters', Default_dict["Path of last model"],"Keras model (*.model)")
                modelname = modelname[0]
                #modelname_for_dict = modelname
            self.lineEdit_LoadModelPath.setText(modelname) #Put the filename to the line edit

            #Remember the location for next time
            if len(str(modelname))>0:
                Default_dict["Path of last model"] = os.path.split(modelname)[0]
                aid_bin.save_aid_settings(Default_dict)
            #If user wants to load and restart a model
            if self.radioButton_LoadRestartModel.isChecked():
                #load the model and print summary
                if modelname.endswith(".arch"):
                    json_file = open(modelname, 'r')
                    model_config = json_file.read()
                    json_file.close()
                    model_config = json.loads(model_config)
                    #cut the .json off
                    modelname = modelname.split(".arch")[0]
                                    
                #Or a .model (FULL model with trained weights) , but for display only load the architecture        
                elif modelname.endswith(".model"):
                    #Load the model config (this is the architecture)                  
                    model_full_h5 = h5py.File(modelname, 'r')
                    model_config = model_full_h5.attrs['model_config']
                    model_full_h5.close() #close the hdf5
                    model_config = json.loads(model_config)
                    #model = model_from_config(model_config)
                    modelname = modelname.split(".model")[0]
                else:
                    msg = QtWidgets.QMessageBox()
                    msg.setIcon(QtWidgets.QMessageBox.Information)       
                    msg.setText("No valid file was chosen. Please specify a file that was created using AIDeveloper or Keras")
                    msg.setWindowTitle("No valid file was chosen")
                    msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                    msg.exec_()
                    return
                    #raise ValueError("No valid file was chosen")
                    
                text1 = "Architecture: loaded from .arch\nWeights: will be randomly initialized'\n"                    
                    
                #Try to find the corresponding .meta
                #All models have a number:
                metaname = modelname.rsplit('_',1)[0]+"_meta.xlsx"
                if os.path.isfile(metaname):
                    #open the metafile
                    meta = pd.read_excel(metaname,sheet_name="Parameters",engine="openpyxl")
                    if "Chosen Model" in list(meta.keys()):
                        chosen_model = meta["Chosen Model"].iloc[-1]
                    else:
                        #Try to get the model architecture and adjust the combobox
                        try:
                            ismlp,chosen_model = model_zoo.mlpconfig_to_str(model_config)
                        except:#No model could be identified
                            chosen_model = "None"
                else:
                    #Try to get the model architecture and adjust the combobox
                    try:
                        ismlp,chosen_model = model_zoo.mlpconfig_to_str(model_config)
                    except:#No model could be identified
                        chosen_model = "None"

                if chosen_model is not None:
                    #chosen_model is a string that should be contained in comboBox_ModelSelection
                    index = self.comboBox_ModelSelection.findText(chosen_model, QtCore.Qt.MatchFixedString)
                    if index >= 0:
                        self.comboBox_ModelSelection.setCurrentIndex(index)
                else:
                    index = self.comboBox_ModelSelection.findText('None', QtCore.Qt.MatchFixedString)
                    if index >= 0:
                        self.comboBox_ModelSelection.setCurrentIndex(index)
                    

            #Otherwise, user wants to load and continue training a model
            elif self.radioButton_LoadContinueModel.isChecked():
                #User can only choose a .model (FULL model with trained weights) , but for display only load the architecture
                if modelname.endswith(".model"):
                    #Load the model config (this is the architecture)                  
                    model_full_h5 = h5py.File(modelname, 'r')
                    model_config = model_full_h5.attrs['model_config']
                    model_full_h5.close() #close the hdf5
                    model_config = json.loads(model_config)
                    #model = model_from_config(model_config)
                    modelname = modelname.split(".model")[0]

                    #Try to find the corresponding .meta
                    #All models have a number:
                    metaname = modelname.rsplit('_',1)[0]+"_meta.xlsx"
                    if os.path.isfile(metaname):
                        #open the metafile
                        meta = pd.read_excel(metaname,sheet_name="Parameters",engine="openpyxl")
                        if "Chosen Model" in list(meta.keys()):
                            chosen_model = meta["Chosen Model"].iloc[-1]
                        else:
                            #Try to get the model architecture and adjust the combobox
                            try:
                                ismlp,chosen_model = model_zoo.mlpconfig_to_str(model_config)
                            except:#No model could be identified
                                chosen_model = "None"
                    else:
                        #Try to get the model architecture and adjust the combobox
                        try:
                            ismlp,chosen_model = model_zoo.mlpconfig_to_str(model_config)
                        except:#No model could be identified
                            chosen_model = "None"

                    if chosen_model is not None:
                        #chosen_model is a string that should be contained in comboBox_ModelSelection
                        index = self.comboBox_ModelSelection.findText(chosen_model, QtCore.Qt.MatchFixedString)
                        if index >= 0:
                            self.comboBox_ModelSelection.setCurrentIndex(index)
                    else:
                        index = self.comboBox_ModelSelection.findText('None', QtCore.Qt.MatchFixedString)
                        if index >= 0:
                            self.comboBox_ModelSelection.setCurrentIndex(index)
                    text1 = "Architecture: loaded from .model\nWeights: pretrained weights will be loaded and used when hitting button 'Initialize model!'\n"
                else:
                    msg = QtWidgets.QMessageBox()
                    msg.setIcon(QtWidgets.QMessageBox.Information)       
                    msg.setText("No valid file was chosen. Please specify a file that was created using AIDeveloper or Keras")
                    msg.setWindowTitle("No valid file was chosen")
                    msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                    msg.exec_()
                    return
                    #raise ValueError("No valid file was chosen")

            #In both cases (restart or continue) the input dimensions have to fit
            #The number of output classes should also fit but this is not essential  
            #but most users certainly want the same number of classes (output)->Give Info
            in_dim, out_dim = aid_dl.model_in_out_dim(model_config,"config")

            #Retrieve the color_mode from the model (nr. of channels in last in_dim)
            channels = in_dim[-1] #TensorFlow: channels in last dimension
            if channels==1:
                channel_text = "1 channel (Grayscale)"
                if self.get_color_mode()!="Grayscale":
                    #when model needs Grayscale, set the color mode in comboBox_GrayOrRGB to that
                    index = self.comboBox_GrayOrRGB.findText("Grayscale", QtCore.Qt.MatchFixedString)
                    if index >= 0:
                        self.comboBox_GrayOrRGB.setCurrentIndex(index)                                        
                    self.statusbar.showMessage("Color Mode set to Grayscale",5000)

            elif channels==3:
                channel_text = "3 channels (RGB)"
                if self.get_color_mode()!="RGB":
                    #when model needs RGB, set the color mode in the ui to that
                    index = self.comboBox_GrayOrRGB.findText("RGB", QtCore.Qt.MatchFixedString)
                    if index >= 0:
                        self.comboBox_GrayOrRGB.setCurrentIndex(index)
                    self.statusbar.showMessage("Color Mode set to RGB",5000)

            text2 = "Model Input: loaded Model takes: "+str(in_dim[-3])+" x "+str(in_dim[-2]) + " pixel images and "+channel_text+"\n"
            if int(self.spinBox_imagecrop.value())!=int(in_dim[-2]):
                self.spinBox_imagecrop.setValue(in_dim[-2])
                text2 = text2+ "'Input image size'  in GUI was changed accordingly\n"
            
            #check that the nr. of classes are equal to the model out put
            SelectedFiles = self.items_clicked_no_rtdc_ds()
            indices = [s["class"] for s in SelectedFiles]
            
            nr_classes = np.max(indices)+1

            if int(nr_classes)==int(out_dim):
                text3 = "Output: "+str(out_dim)+" classes\n"
            elif int(nr_classes)>int(out_dim):#Dataset has more classes than the model provides!
                text3 = "Loaded model has only "+(str(out_dim))+\
                " output nodes (classes) but your selected data has "+str(nr_classes)+\
                " classes. Therefore, the model will be adjusted before fitting, by customizing the final Dense layer.\n"                
                #aid_dl.model_add_classes(model_keras,nr_classes)#this function changes model_keras inplace
            elif int(nr_classes)<int(out_dim):#Dataset has less classes than the model provides!
                text3 = "Model output: The architecture you chose has "+(str(out_dim))+\
                " output nodes (classes) and your selected data has only "+str(nr_classes)+\
                " classes. This is fine. The model will essentially have some excess classes that are not used.\n"                

            text = text1+text2+text3
            self.textBrowser_Info.setText(text)

            if self.radioButton_LoadContinueModel.isChecked():
                #"Load the parameter file of the model that should be continued and apply the same normalization"
                #Make a popup: You are about to continue to train a pretrained model
                #Please select the parameter file of that model to load the normalization method
                #or choose the normalization method manually:
                #this is important
                self.popup_normalization()

    def get_metrics(self):
        Metrics =  []
        f1 = bool(self.checkBox_expertF1.isChecked())
        if f1==True:
            Metrics.append("auc")
        precision = bool(self.checkBox_expertPrecision.isChecked())
        if precision==True:
            Metrics.append("precision")
        recall = bool(self.checkBox_expertRecall.isChecked())
        if recall==True:
            Metrics.append("recall")
        metrics =  ['accuracy'] + Metrics
        #metrics = aid_dl.get_metrics_tensors(metrics,nr_classes)
        return metrics

    def action_set_modelpath_and_name(self):
        #Get the path and filename for the new model
        filename = QtWidgets.QFileDialog.getSaveFileName(self, 'Save model', Default_dict["Path of last model"],"Keras Model file (*.model)")
        filename = filename[0]
        if len(filename)==0:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("No valid filename was chosen.")
            msg.setWindowTitle("No valid filename was chosen")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return
            
        if filename.endswith(".arch"):
            filename = filename.split(".arch")[0]
        #add the suffix .model
        if not filename.endswith(".model"):
            filename = filename +".model"
        self.lineEdit_modelname.setText(filename)
        #Write to Default_dict
        Default_dict["Path of last model"] = os.path.split(filename)[0]
        aid_bin.save_aid_settings(Default_dict)
        
    def get_dataOverview(self):
        table = self.tableWidget_Info
        cols = table.columnCount()
        header = [table.horizontalHeaderItem(col).text() for col in range(cols)]
        rows = table.rowCount()
        tmp_df = pd.DataFrame(columns=header,index=range(rows)) 
        for i in range(rows):
            for j in range(cols):
                try:
                    tmp_df.iloc[i, j] = table.item(i, j).text()
                except:
                    tmp_df.iloc[i, j] = np.nan
        return tmp_df
        
                 
    def action_initialize_model(self,duties="initialize_train"):
        """
        duties: which tasks should be performed: "initialize", "initialize_train", "initialize_lrfind"
        """
        #print("duties: "+str(duties))
        
        #Create config (define which device to use)
        if self.radioButton_cpu.isChecked():
            deviceSelected = str(self.comboBox_cpu.currentText())
        elif self.radioButton_gpu.isChecked():
            deviceSelected = str(self.comboBox_gpu.currentText())
        gpu_memory = float(self.doubleSpinBox_memory.value())
        config_gpu = aid_dl.get_config(cpu_nr,gpu_nr,deviceSelected,gpu_memory)

#        try:
#            K.clear_session()
#        except:
#            print("Could not clear_session (7)")

        with tf.compat.v1.Session(graph = tf.Graph(), config=config_gpu) as sess:            
            sess.run(tf.compat.v1.global_variables_initializer())
            #Initialize the model
            #######################Load and restart model##########################
            if self.radioButton_LoadRestartModel.isChecked():
                
                load_modelname = str(self.lineEdit_LoadModelPath.text())
                text0 = "Loaded model: "+load_modelname
                #load the model and print summary
                if load_modelname.endswith(".arch"):
                    json_file = open(load_modelname, 'r')
                    model_config = json_file.read()
                    json_file.close()
                    model_keras = model_from_json(model_config)
                    model_config = json.loads(model_config)
                    text1 = "\nArchitecture: loaded from .arch\nWeights: randomly initialized\n"
    
                #Or a .model (FULL model with trained weights) , but for display only load the architecture        
                elif load_modelname.endswith(".model"):
                    #Load the model config (this is the architecture)                  
                    model_full_h5 = h5py.File(load_modelname, 'r')
                    model_config = model_full_h5.attrs['model_config']
                    model_full_h5.close() #close the hdf5                
                    model_config = json.loads(model_config)
                    model_keras = model_from_config(model_config)
                    text1 = "\nArchitecture: loaded from .model\nWeights: randomly initialized\n"
                else:
                    msg = QtWidgets.QMessageBox()
                    msg.setIcon(QtWidgets.QMessageBox.Information)       
                    msg.setText("No valid file was chosen. Please specify a file that was created using AIDeveloper or Keras")
                    msg.setWindowTitle("No valid file was chosen")
                    msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                    msg.exec_()
                    return
    
                try:
                    metaname = load_modelname.rsplit('_',1)[0]+"_meta.xlsx"
                    if os.path.isfile(metaname):
                        #open the metafile
                        meta = pd.read_excel(metaname,sheet_name="Parameters",engine="openpyxl")
                        if "Chosen Model" in list(meta.keys()):
                            chosen_model = meta["Chosen Model"].iloc[-1]
                except:
                    chosen_model = str(self.comboBox_ModelSelection.currentText())
                    
                #In both cases (restart or continue) the input dimensions have to fit
                #The number of output classes should also fit but this is not essential  
                #but most users certainly want the same number of classes (output)->Give Info
    
                in_dim, out_dim = aid_dl.model_in_out_dim(model_config,"config")  
                
                channels = in_dim[-1] #TensorFlow: channels in last dimension
    
                #Compile model (consider user-specific metrics)
                model_metrics = self.get_metrics()

                model_keras.compile(loss='categorical_crossentropy',optimizer='adam',metrics=aid_dl.get_metrics_tensors(model_metrics,out_dim))#dont specify loss and optimizer yet...expert stuff will follow and model will be recompiled

                if channels==1:
                    channel_text = "1 channel (Grayscale)"
                    if self.get_color_mode()!="Grayscale":
                        #when model needs Grayscale, set the color mode in comboBox_GrayOrRGB to that
                        index = self.comboBox_GrayOrRGB.findText("Grayscale", QtCore.Qt.MatchFixedString)
                        if index >= 0:
                            self.comboBox_GrayOrRGB.setCurrentIndex(index)                                        
                        self.statusbar.showMessage("Color Mode set to Grayscale",5000)
    
                elif channels==3:
                    channel_text = "3 channels (RGB)"
                    if self.get_color_mode()!="RGB":
                        #when model needs RGB, set the color mode in the ui to that
                        index = self.comboBox_GrayOrRGB.findText("RGB", QtCore.Qt.MatchFixedString)
                        if index >= 0:
                            self.comboBox_GrayOrRGB.setCurrentIndex(index)
                        self.statusbar.showMessage("Color Mode set to RGB",5000)
    
                text2 = "Model Input: "+str(in_dim[-3])+" x "+str(in_dim[-2]) + " pixel images and "+channel_text+"\n"
    
                if int(self.spinBox_imagecrop.value())!=int(in_dim[-2]):
                    self.spinBox_imagecrop.setValue(in_dim[-2])
                    text2 = text2+ "'Input image size'  in GUI was changed accordingly\n"
                
                #check that the nr. of classes are equal to the model out put
                SelectedFiles = self.items_clicked()
                indices = [s["class"] for s in SelectedFiles]
                nr_classes = np.max(indices)+1
                    
                if int(nr_classes)==int(out_dim):
                    text3 = "Output: "+str(out_dim)+" classes\n"
                elif int(nr_classes)>int(out_dim):#Dataset has more classes than the model provides!
                    text3 = "Loaded model has only "+(str(out_dim))+\
                    " output nodes (classes) but your selected data has "+str(nr_classes)+\
                    " classes. Therefore, the model will be adjusted before fitting, by customizing the final Dense layer.\n"                
                    aid_dl.model_add_classes(model_keras,nr_classes)#this function changes model_keras inplace
                elif int(nr_classes)<int(out_dim):#Dataset has less classes than the model provides!
                    text3 = "Model output: The architecture you chose has "+(str(out_dim))+\
                    " output nodes (classes) and your selected data has only "+str(nr_classes)+\
                    " classes. This is fine. The model will essentially have some excess classes that are not used.\n"                
    
            ###############Load and continue training the model####################
            elif self.radioButton_LoadContinueModel.isChecked():
                load_modelname = str(self.lineEdit_LoadModelPath.text())
                text0 = "Loaded model: "+load_modelname+"\n"

                #User can only choose a .model (FULL model with trained weights) , but for display only load the architecture
                if load_modelname.endswith(".model"):              
                    #Load the full model
                    try:
                        model_keras = load_model(load_modelname,custom_objects=aid_dl.get_custom_metrics())
                    except:
                        K.clear_session() #On linux It happened that there was an error, if another fitting run before
                        model_keras = load_model(load_modelname,custom_objects=aid_dl.get_custom_metrics())
                    #model_config = model_keras.config() #Load the model config (this is the architecture)
                    #load_modelname = load_modelname.split(".model")[0]
                    text1 = "Architecture: loaded from .model\nWeights: pretrained weights were loaded\n"
                else:
                    msg = QtWidgets.QMessageBox()
                    msg.setIcon(QtWidgets.QMessageBox.Information)       
                    msg.setText("No valid file was chosen. Please specify a file that was created using AIDeveloper or Keras")
                    msg.setWindowTitle("No valid file was chosen")
                    msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                    msg.exec_()
                    return
                    #raise ValueError("No valid file was chosen")
    
                try:
                    metaname = load_modelname.rsplit('_',1)[0]+"_meta.xlsx"
                    if os.path.isfile(metaname):
                        #open the metafile
                        meta = pd.read_excel(metaname,sheet_name="Parameters",engine="openpyxl")
                        if "Chosen Model" in list(meta.keys()):
                            chosen_model = meta["Chosen Model"].iloc[-1]
                    else:
                        chosen_model = str(self.comboBox_ModelSelection.currentText())
    
                except:
                    chosen_model = str(self.comboBox_ModelSelection.currentText())
    
    
                #Check input dimensions
                #The number of output classes should also fit but this is not essential  
                #but most users certainly want the same number of classes (output)->Give Info
    #            in_dim = model_config['config'][0]['config']['batch_input_shape']
    #            out_dim = model_config['config'][-2]['config']['units']
                in_dim = np.array(model_keras.input.shape[1:])
                out_dim = model_keras.output.shape[1]
                channels = in_dim[-1] #TensorFlow: channels in last dimension
    
                if channels==1:
                    channel_text = "1 channel (Grayscale)"
                    if self.get_color_mode()!="Grayscale":
                        #when model needs Grayscale, set the color mode in comboBox_GrayOrRGB to that
                        index = self.comboBox_GrayOrRGB.findText("Grayscale", QtCore.Qt.MatchFixedString)
                        if index >= 0:
                            self.comboBox_GrayOrRGB.setCurrentIndex(index)                                        
                        self.statusbar.showMessage("Color Mode set to Grayscale",5000)
    
                elif channels==3:
                    channel_text = "3 channels (RGB)"
                    if self.get_color_mode()!="RGB":
                        #when model needs RGB, set the color mode in the ui to that
                        index = self.comboBox_GrayOrRGB.findText("RGB", QtCore.Qt.MatchFixedString)
                        if index >= 0:
                            self.comboBox_GrayOrRGB.setCurrentIndex(index)
                        self.statusbar.showMessage("Color Mode set to RGB",5000)
    
                text2 = "Model Input: "+str(in_dim[-3])+" x "+str(in_dim[-2]) + " pixel images and "+channel_text+"\n"
                if int(self.spinBox_imagecrop.value())!=int(in_dim[-2]):
                    self.spinBox_imagecrop.setValue(in_dim[-2])
                    text2 = text2+ "'Input image size'  in GUI was changed accordingly\n"
                
                #check that the nr. of classes are equal to the model out put
                SelectedFiles = self.items_clicked()
                indices = [s["class"] for s in SelectedFiles]
                nr_classes = np.max(indices)+1

                if int(nr_classes)==int(out_dim):
                    text3 = "Output: "+str(out_dim)+" classes\n"
                elif int(nr_classes)>int(out_dim):#Dataset has more classes than the model provides!
                    text3 = "Loaded model has only "+(str(out_dim))+\
                    " output nodes (classes) but your selected data has "+str(nr_classes)+\
                    " classes. Therefore, the model will be adjusted before fitting, by customizing the final Dense layer.\n"                
                    aid_dl.model_add_classes(model_keras,nr_classes)#this function changes model_keras inplace
                elif int(nr_classes)<int(out_dim):#Dataset has less classes than the model provides!
                    text3 = "Model output: The architecture you chose has "+(str(out_dim))+\
                    " output nodes (classes) and your selected data has only "+str(nr_classes)+\
                    " classes. This is fine. The model will essentially have some excess classes that are not used.\n"                
    
            ###########################New model###################################
            elif self.radioButton_NewModel.isChecked():
                load_modelname = "" #No model is loaded
                text0 = load_modelname
                #Create a new model!
                #Get what the user wants from the dropdown menu!
                chosen_model = str(self.comboBox_ModelSelection.currentText())
                if chosen_model==None:
                    msg = QtWidgets.QMessageBox()
                    msg.setIcon(QtWidgets.QMessageBox.Information)       
                    msg.setText("No model specified!")
                    msg.setWindowTitle("No model specified!")
                    msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                    msg.exec_()
                    return       
                    
                in_dim = int(self.spinBox_imagecrop.value())
                SelectedFiles = self.items_clicked()
                #rtdc_ds = SelectedFiles[0]["rtdc_ds"]
    
                if str(self.comboBox_GrayOrRGB.currentText())=="Grayscale":
                    channels=1
                elif str(self.comboBox_GrayOrRGB.currentText())=="RGB":
                    channels=3
                    
                indices = [s["class"] for s in SelectedFiles]
                indices_unique = np.unique(np.array(indices))
                if len(indices_unique)<2:
                    msg = QtWidgets.QMessageBox()
                    msg.setIcon(QtWidgets.QMessageBox.Information)       
                    msg.setText("Need at least two classes to fit. Please specify .rtdc files and corresponding indeces")
                    msg.setWindowTitle("No valid file was chosen")
                    msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                    msg.exec_()
                    return  
                
                out_dim = np.max(indices)+1
                nr_classes = out_dim
                
                if chosen_model=="None":
                    msg = QtWidgets.QMessageBox()
                    msg.setIcon(QtWidgets.QMessageBox.Information)       
                    msg.setText("No model specified!")
                    msg.setWindowTitle("No model specified!")
                    msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                    msg.exec_()
                    return       



                try:                
                    model_keras = model_zoo.get_model(chosen_model,in_dim,channels,out_dim)
                except Exception as e: 
                    #There is an issue building the model!
                    msg = QtWidgets.QMessageBox()
                    msg.setIcon(QtWidgets.QMessageBox.Critical)       
                    msg.setText(str(e))
                    msg.setWindowTitle("Error occured when building model:")
                    msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                    msg.exec_()
                    return       
                
                text1 = "Architecture: created "+chosen_model+" design\nWeights: Initialized random weights\n"
                
                if self.get_color_mode()=="Grayscale":
                    channels = 1
                    channel_text = "1 channel (Grayscale)"
                elif self.get_color_mode()=="RGB":
                    channels = 3
                    channel_text = "3 channels (RGB)"
                        
                text2 = "Model Input: "+str(in_dim)+" x "+str(in_dim) + " pixel images and "+channel_text+"\n"
    
                if int(nr_classes)==int(out_dim):
                    text3 = "Output: "+str(out_dim)+" classes\n"
                elif int(nr_classes)>int(out_dim):#Dataset has more classes than the model provides!
                    text3 = "Loaded model has only "+(str(out_dim))+\
                    " output nodes (classes) but your selected data has "+str(nr_classes)+\
                    " classes. Therefore, the model will be adjusted before fitting, by customizing the final Dense layer.\n"                
                    aid_dl.model_add_classes(model_keras,nr_classes)#this function changes model_keras inplace
                elif int(nr_classes)<int(out_dim):#Dataset has less classes than the model provides!
                    text3 = "Model output: The architecture you chose has "+(str(out_dim))+\
                    " output nodes (classes) and your selected data has only "+str(nr_classes)+\
                    " classes. This is fine. The model will essentially have some excess classes that are not used.\n"                

            else:
                #No radio-button was chosen
                msg = QtWidgets.QMessageBox()
                msg.setIcon(QtWidgets.QMessageBox.Information)       
                msg.setText("Please use the radiobuttons to define the model")
                msg.setWindowTitle("No model defined")
                msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                msg.exec_()
                return
    

            #If expert mode is on, apply the requested options
            #This affects learning rate, trainability of layers and dropout rate
            expert_mode = bool(self.groupBox_expertMode.isChecked())
            learning_rate_const = float(self.doubleSpinBox_learningRate.value())
            learning_rate_expert_on = bool(self.groupBox_learningRate.isChecked())   
            train_last_layers = bool(self.checkBox_trainLastNOnly.isChecked())             
            train_last_layers_n = int(self.spinBox_trainLastNOnly.value())              
            train_dense_layers = bool(self.checkBox_trainDenseOnly.isChecked())             
            dropout_expert_on = bool(self.checkBox_dropout.isChecked())
            loss_expert_on = bool(self.checkBox_expt_loss.isChecked())
            loss_expert = str(self.comboBox_expt_loss.currentText()).lower()
            optimizer_expert_on = bool(self.checkBox_optimizer.isChecked())
            optimizer_expert = str(self.comboBox_optimizer.currentText()).lower()
            optimizer_settings = self.optimizer_settings.copy() #get the current optimizer settings
            paddingMode = str(self.comboBox_paddingMode.currentText())#.lower()
            model_metrics = self.get_metrics()    
            if "collection" in chosen_model.lower():
                for m in model_keras[1]: #in a collection, model_keras[0] are the names of the models and model_keras[1] is a list of all models
                    model_metrics_t = aid_dl.get_metrics_tensors(self.get_metrics(),nr_classes)
                    aid_dl.model_compile(m,loss_expert,optimizer_settings,learning_rate_const,model_metrics_t,nr_classes)

            if not "collection" in chosen_model.lower():
                model_metrics_t = aid_dl.get_metrics_tensors(self.get_metrics(),nr_classes)
                aid_dl.model_compile(model_keras,loss_expert,optimizer_settings,learning_rate_const,model_metrics_t,nr_classes)
                
            try:
                dropout_expert = str(self.lineEdit_dropout.text()) #due to the validator, there are no squ.brackets
                dropout_expert = "["+dropout_expert+"]"
                dropout_expert = ast.literal_eval(dropout_expert)        
            except:
                dropout_expert = []             
    
            if type(model_keras)==tuple:#when user chose a Collection of models, a tuple is returned by get_model
                collection = True
            else:
                collection = False

            if collection==False: #if there is a single model:
                #Original learning rate (before expert mode is switched on!)
                try:
                    self.learning_rate_original = model_keras.optimizer.get_config()["learning_rate"]
                except:
                    print("Session busy. Try again in fresh session...")
                    #tf.reset_default_graph() #Make sure to start with a fresh session
                    K.clear_session()
                    sess = tf.compat.v1.Session(graph = tf.Graph(), config=config_gpu)
                    #K.set_session(sess)
                    self.learning_rate_original = model_keras.optimizer.get_config()["learning_rate"]
                    
                #Get initial trainability states of model
                self.trainable_original, self.layer_names = aid_dl.model_get_trainable_list(model_keras)

                trainable_original, layer_names = self.trainable_original, self.layer_names

                self.do_list_original = aid_dl.get_dropout(model_keras)#Get a list of dropout values of the current model

                do_list_original = self.do_list_original

            if collection==True: #if there is a collection of models:
                #Original learning rate (before expert mode is switched on!)
                self.learning_rate_original = [model_keras[1][i].optimizer.get_config()["learning_rate"] for i in range(len(model_keras[1]))]
                #Get initial trainability states of model
                trainable_layerName = [aid_dl.model_get_trainable_list(model_keras[1][i]) for i in range(len(model_keras[1]))]
                self.trainable_original = [trainable_layerName[i][0] for i in range(len(trainable_layerName))]
                self.layer_names = [trainable_layerName[i][1] for i in range(len(trainable_layerName))]
                trainable_original, layer_names = self.trainable_original, self.layer_names
                self.do_list_original = [aid_dl.get_dropout(model_keras[1][i]) for i in range(len(model_keras[1]))]#Get a list of dropout values of the current model
                do_list_original = self.do_list_original
    
            #TODO add expert mode ability for collection of models. Maybe define self.model_keras as a list in general. So, fitting a single model is just a special case
    
            if expert_mode==True:
                #Apply the changes to trainable states:
                if train_last_layers==True:#Train only the last n layers
                    print("Train only the last "+str(train_last_layers_n)+ " layer(s)")
                    trainable_new = (len(trainable_original)-train_last_layers_n)*[False]+train_last_layers_n*[True]
                    aid_dl.model_change_trainability(model_keras,trainable_new,model_metrics,out_dim,loss_expert,optimizer_settings,learning_rate_const)
    
                if train_dense_layers==True:#Train only dense layers
                    print("Train only dense layers")
                    layer_dense_ind = ["Dense" in x for x in layer_names]
                    layer_dense_ind = np.where(np.array(layer_dense_ind)==True)[0] #at which indices are dropout layers?
                    #create a list of trainable states
                    trainable_new = len(trainable_original)*[False]
                    for index in layer_dense_ind:
                        trainable_new[index] = True
                    aid_dl.model_change_trainability(model_keras,trainable_new,model_metrics,out_dim,loss_expert,optimizer_settings,learning_rate_const)
                
                if dropout_expert_on==True:
                    #The user apparently want to change the dropout rates
                    do_list = aid_dl.get_dropout(model_keras)#Get a list of dropout values of the current model
                    #Compare the dropout values in the model to the dropout values requested by user
                    if len(dropout_expert)==1:#if the user gave a float
                        dropout_expert_list=len(do_list)*dropout_expert #convert to list
                    elif len(dropout_expert)>1:
                        dropout_expert_list = dropout_expert
                        if not len(dropout_expert_list)==len(do_list):
                            msg = QtWidgets.QMessageBox()
                            msg.setIcon(QtWidgets.QMessageBox.Information)       
                            msg.setText("Issue with dropout: you defined "+str(len(dropout_expert_list))+" dropout rates, but model has "+str(len(do_list))+" dropout layers")
                            msg.setWindowTitle("Issue with Expert->Dropout")
                            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                            msg.exec_()
                            dropout_expert_list = []
                            return
                    else:
                        msg = QtWidgets.QMessageBox()
                        msg.setIcon(QtWidgets.QMessageBox.Information)       
                        msg.setText("Could not understand user input at Expert->Dropout")
                        msg.setWindowTitle("Issue with Expert->Dropout")
                        msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                        msg.exec_()
                        
                        dropout_expert_list = []
                     
                    if len(dropout_expert_list)>0 and do_list!=dropout_expert_list:#if the dropout rates of the current model is not equal to the required do_list from user...
                        do_changed = aid_dl.change_dropout(model_keras,dropout_expert_list,model_metrics_t,nr_classes,loss_expert,optimizer_settings,learning_rate_const)
                        if do_changed==1:
                            text_do = "Dropout rate(s) in model was/were changed to: "+str(dropout_expert_list)
                        else:
                            text_do = "Dropout rate(s) in model was/were not changed"
                    else:
                        text_do = "Dropout rate(s) in model was/were not changed"
                    print(text_do)

            text_updates = ""
            #Learning Rate: Compare current lr and the lr on expert tab:
            if collection == False:
                lr_current = model_keras.optimizer.get_config()["learning_rate"]
            else:
                lr_current = model_keras[1][0].optimizer.get_config()["learning_rate"]
            lr_diff = learning_rate_const-lr_current
            if  abs(lr_diff) > 1e-6: #If there is a difference, change lr accordingly
                K.set_value(model_keras.optimizer.lr, learning_rate_const)
            text_updates += "Learning rate: "+str(lr_current)+"\n"

            recompile = False
            #Compare current optimizer and the optimizer on expert tab:
            if collection==False:
                optimizer_current = aid_dl.get_optimizer_name(model_keras).lower()#get the current optimizer of the model
            else:
                optimizer_current = aid_dl.get_optimizer_name(model_keras[1][0]).lower()#get the current optimizer of the model
    
            if optimizer_current!=optimizer_expert.lower():#if the current model has a different optimizer
                recompile = True
            text_updates+="Optimizer: "+optimizer_expert+"\n"

            #Loss function: Compare current loss function and the loss-function on expert tab:
            if collection==False:
                if model_keras.loss!=loss_expert:
                    recompile = True
            if collection==True:
                if model_keras[1][0].loss!=loss_expert:
                    recompile = True
            text_updates += "Loss function: "+loss_expert+"\n"

            if recompile==True:
                if collection==False:
                    print("Recompiling...")
                    model_metrics_t = aid_dl.get_metrics_tensors(self.get_metrics(),nr_classes)
                    aid_dl.model_compile(model_keras,loss_expert,optimizer_settings,learning_rate_const,model_metrics_t,nr_classes)
                if collection==True:
                    for m in model_keras[1]:
                        print("Recompiling...")
                        model_metrics_t = aid_dl.get_metrics_tensors(self.get_metrics(),nr_classes)
                        aid_dl.model_compile(m,loss_expert,optimizer_settings,learning_rate_const,model_metrics_t,nr_classes)
            self.model_keras = model_keras #overwrite the model in self
    
            if collection == False:
                #Get user-specified filename for the new model
                new_modelname = str(self.lineEdit_modelname.text())
                if len(new_modelname)>0:
                    text_new_modelname = "Model will be saved as: "+new_modelname+"\n"
                else:
                    text_new_modelname = "Please specify a model path (name for the model to be fitted)\n"
    
            if collection == True:
                new_modelname = str(self.lineEdit_modelname.text())
                if len(new_modelname)>0:
                    new_modelname = os.path.split(new_modelname)
                    text_new_modelname = "Collection of Models will be saved into: "+new_modelname[0]+"\n"
                else:
                    text_new_modelname = "Please specify a model path (name for the model to be fitted)\n"
    
    
            #Info about normalization method
            norm = str(self.comboBox_Normalization.currentText())
    
            text4 = "Input image normalization method: "+norm+"\n"
    
            #Check if there are dropout layers:
            #do_list = aid_dl.get_dropout(model_keras)#Get a list of dropout values of the current model
            if len(do_list_original)>0:
                text4 = text4+"Found "+str(len(do_list_original)) +" dropout layers with rates: "+str(do_list_original)+"\n"
            else:
                text4 = text4+"Found no dropout layers\n"

            if expert_mode==True:
                if dropout_expert_on:
                    text4 = text4+text_do+"\n"
    #            if learning_rate_expert_on==True:
    #                if K.eval(model_keras.optimizer.lr) != learning_rate_const: #if the learning rate in UI is NOT equal to the lr of the model...
    #                    text_lr = "Changed the learning rate to: "+ str(learning_rate_const)+"\n"
    #                    text4 = text4+text_lr
    
            text5 = "Model summary:\n"
            summary = []
            if collection==False:
                model_keras.summary(print_fn=summary.append)
                summary = "\n".join(summary)
                text = text_new_modelname+text0+text1+text2+text3+text4+text_updates+text5+summary
                self.textBrowser_Info.setText(text)
                    
                #Save the model architecture: serialize to JSON
                model_json = model_keras.to_json()
                with open(new_modelname.split(".model")[0]+".arch", "w") as json_file:
                    json_file.write(model_json)
    
            elif collection==True:
                if self.groupBox_expertMode.isChecked()==True:
                    self.groupBox_expertMode.setChecked(False)
                    print("Turned off expert mode. Not implemented yet for collections of models. This does not affect user-specified metrics (precision/recall/auc)")
                
                self.model_keras_arch_path = [new_modelname[0]+os.sep+new_modelname[1].split(".model")[0]+"_"+model_keras[0][i]+".arch" for i in range(len(model_keras[0]))]                
                for i in range(len(model_keras[1])):
                    model_keras[1][i].summary(print_fn=summary.append)
                        
                    #Save the model architecture: serialize to JSON
                    model_json = model_keras[1][i].to_json()
                    with open(self.model_keras_arch_path[i], "w") as json_file:
                        json_file.write(model_json)
    
                summary = "\n".join(summary)
                text = text_new_modelname+text0+text1+text2+text3+text4+text_updates+text5+summary
                self.textBrowser_Info.setText(text)
    
            #Save the model to a variable on self
            self.model_keras = model_keras

            #Get the user-defined cropping size
            crop = int(self.spinBox_imagecrop.value())          
            #Make the cropsize a bit larger since the images will later be rotated
            cropsize2 = np.sqrt(crop**2+crop**2)
            cropsize2 = np.ceil(cropsize2 / 2.) * 2 #round to the next even number
    
            #Estimate RAM needed
            nr_imgs = np.sum([np.array(list(SelectedFiles)[i]["nr_images"]) for i in range(len(list(SelectedFiles)))])
            ram_needed = np.round(nr_imgs * aid_bin.calc_ram_need(cropsize2),2)
     
            if duties=="initialize":#Stop here if the model just needs to be intialized (for expert mode->partial trainability)
                return
            
            elif duties=="initialize_train":
                #Tell the user if the data is stored and read from ram or not
                msg = QtWidgets.QMessageBox()
                msg.setIcon(QtWidgets.QMessageBox.Question)
                text = "<html><head/><body><p>Should the model only be initialized,\
                or do you want to start fitting right after? For fitting, data will\
                be loaded to RAM (since Edit->Data to RAM is enabled), which will\
                require "+str(ram_needed)+"MB of RAM.</p></body></html>"
                msg.setText(text) 
                msg.setWindowTitle("Initialize model or initialize and fit model?")
                msg.addButton(QtGui.QPushButton('Stop after model initialization'), QtGui.QMessageBox.RejectRole)
                msg.addButton(QtGui.QPushButton('Start fitting'), QtGui.QMessageBox.ApplyRole)
                retval = msg.exec_()
            
            elif duties=="initialize_lrfind":
                retval = 1
            
            else:
                print("Invalid duties: "+duties)
                return
            
            if retval==0: #yes role: Only initialize model
                print("Closing session")
                del model_keras
                sess.close()
                return
            
            elif retval == 1:
                if self.actionDataToRam.isChecked():
                    color_mode = self.get_color_mode()
                    zoom_factors = [selectedfile["zoom_factor"] for selectedfile in SelectedFiles]
                    #zoom_order = [self.actionOrder0.isChecked(),self.actionOrder1.isChecked(),self.actionOrder2.isChecked(),self.actionOrder3.isChecked(),self.actionOrder4.isChecked(),self.actionOrder5.isChecked()]
                    #zoom_order = int(np.where(np.array(zoom_order)==True)[0])
                    zoom_order = int(self.comboBox_zoomOrder.currentIndex()) #the combobox-index is already the zoom order
                    
                    #Check if there is data already available in RAM
                    if len(self.ram)==0:#if there is already data stored on ram
                        print("No data on RAM. I have to load")
                        dic = aid_img.crop_imgs_to_ram(list(SelectedFiles),cropsize2,zoom_factors=zoom_factors,zoom_order=zoom_order,color_mode=color_mode)
                        self.ram = dic 

                    else: 
                        print("There is already some data on RAM")
                        new_fileinfo = {"SelectedFiles":list(SelectedFiles),"cropsize2":cropsize2,"zoom_factors":zoom_factors,"zoom_order":zoom_order,"color_mode":color_mode}
                        identical = aid_bin.ram_compare_data(self.ram,new_fileinfo)
                        if not identical:
                            #Load the data
                            dic = aid_img.crop_imgs_to_ram(list(SelectedFiles),cropsize2,zoom_factors=zoom_factors,zoom_order=zoom_order,color_mode=color_mode)
                            self.ram = dic 
                        if identical:
                            msg = QtWidgets.QMessageBox()
                            msg.setIcon(QtWidgets.QMessageBox.Question)
                            text = "Data was loaded before! Should same data be reused? If not, click 'Reload data', e.g. if you altered the Data-table."
                            text = "<html><head/><body><p>"+text+"</p></body></html>"
                            msg.setText(text)
                            msg.setWindowTitle("Found data on RAM")
                            msg.addButton(QtGui.QPushButton('Reuse data'), QtGui.QMessageBox.YesRole)
                            msg.addButton(QtGui.QPushButton('Reload data'), QtGui.QMessageBox.NoRole)
                            retval = msg.exec_()
                
                            if retval==0: 
                                print("Re-use data")
                                #Re-use same data
                            elif retval==1:
                                print("Re-load data")
                                dic = aid_img.crop_imgs_to_ram(list(SelectedFiles),cropsize2,zoom_factors=zoom_factors,zoom_order=zoom_order,color_mode=color_mode)
                                self.ram = dic 

                #Finally, activate the 'Fit model' button again
                #self.pushButton_FitModel.setEnabled(True)
                if duties=="initialize_train":
                    self.action_fit_model()
                if duties=="initialize_lrfind":
                    self.action_lr_finder()
    
            del model_keras

    def action_fit_model_worker(self,progress_callback,history_callback):
        if self.radioButton_cpu.isChecked():
            gpu_used = False
            deviceSelected = str(self.comboBox_cpu.currentText())
        elif self.radioButton_gpu.isChecked():
            gpu_used = True
            deviceSelected = str(self.comboBox_gpu.currentText())
        gpu_memory = float(self.doubleSpinBox_memory.value())

        #Retrieve more Multi-GPU Options from Menubar:
        cpu_merge = bool(self.actioncpu_merge.isEnabled())
        cpu_relocation = bool(self.actioncpu_relocation.isEnabled())
        cpu_weight_merge = bool(self.actioncpu_weightmerge.isEnabled())    

        #Create config (define which device to use)
        config_gpu = aid_dl.get_config(cpu_nr,gpu_nr,deviceSelected,gpu_memory)

        with tf.compat.v1.Session(graph = tf.Graph(), config=config_gpu) as sess:                   
            sess.run(tf.compat.v1.global_variables_initializer())

            #get an index of the fitting popup
            listindex = self.popupcounter-1
            #Get user-specified filename for the new model
            new_modelname = str(self.lineEdit_modelname.text())
            model_keras_path = self.model_keras_path
            
            if type(model_keras_path)==list:
                collection = True
                #Take the initialized models
                model_keras_path = self.model_keras_path
                model_keras = [load_model(model_keras_path[i],custom_objects=aid_dl.get_custom_metrics()) for i in range(len(model_keras_path)) ]
                model_architecture_names = self.model_keras[0]
                print(model_architecture_names)    
                #self.model_keras = None
    
            else:
                collection = False
    
                if deviceSelected=="Multi-GPU" and cpu_weight_merge==True:
                    strategy = tf.distribute.MirroredStrategy()
                    with strategy.scope():
                        model_keras = load_model(model_keras_path,custom_objects=aid_dl.get_custom_metrics()) 
                else:
                    model_keras = load_model(model_keras_path,custom_objects=aid_dl.get_custom_metrics())

            #Initialize a variable for the parallel model
            model_keras_p = None
    
            #Multi-GPU
            if deviceSelected=="Multi-GPU":
                if collection==False:
                    print("Adjusting the model for Multi-GPU")
                    with tf.device("/cpu:0"):#I dont think this line is correct...CHECK!
                        model_keras_p = model_keras#multi_gpu_model(model_keras, gpus=gpu_nr, cpu_merge=cpu_merge, cpu_relocation=cpu_relocation)#indicate the numbers of gpus that you have
                    if self.radioButton_LoadContinueModel.isChecked():#calling multi_gpu_model resets the weights. Hence, they need to be put in place again
                        model_keras_p.layers[-2].set_weights(model_keras.get_weights())
                elif collection==True:
                    print("Collection & Multi-GPU is not supported yet")
                    return
    #                model_keras_p = []
    #                for m in model_keras_p:
    #                    print("Adjusting the model for Multi-GPU")
    #                    model_keras_p.append(multi_gpu_model(m, gpus=gpu_nr)) #indicate the numbers of gpus that you have
    
            ##############Main function after hitting FIT MODEL####################
            if self.radioButton_LoadRestartModel.isChecked():
                load_modelname = str(self.lineEdit_LoadModelPath.text())
            elif self.radioButton_LoadContinueModel.isChecked():
                load_modelname = str(self.lineEdit_LoadModelPath.text())
            elif self.radioButton_NewModel.isChecked():
                load_modelname = "" #No model is loaded
    
            if collection==False:    
                #model_config = model_keras.get_config()#["layers"] 
                nr_classes = int(model_keras.output.shape.dims[1])
            if collection==True:
                #model_config = model_keras.get_config()#["layers"] 
                nr_classes = int(model_keras[0].output.shape.dims[1])
            
            #Metrics to be displayed during fitting (real-time)
            model_metrics = self.get_metrics()
            model_metrics_t = aid_dl.get_metrics_tensors(model_metrics,nr_classes)

            #Compile model
            if collection==False and deviceSelected=="Single-GPU":
                model_keras.compile(loss='categorical_crossentropy',optimizer='adam',metrics=aid_dl.get_metrics_tensors(model_metrics,nr_classes))#dont specify loss and optimizer yet...expert stuff will follow and model will be recompiled
            elif collection==False and deviceSelected=="Multi-GPU":
                model_keras_p.compile(loss='categorical_crossentropy',optimizer='adam',metrics=aid_dl.get_metrics_tensors(model_metrics,nr_classes))#dont specify loss and optimizer yet...expert stuff will follow and model will be recompiled
            elif collection==True and deviceSelected=="Single-GPU":
                #Switch off the expert tab!
                self.fittingpopups_ui[listindex].groupBox_expertMode_pop.setChecked(False)
                self.fittingpopups_ui[listindex].groupBox_expertMode_pop.setEnabled(False)
                for m in model_keras:
                    model_metrics_ = self.get_metrics()
                    m.compile(loss='categorical_crossentropy',optimizer='adam',metrics=aid_dl.get_metrics_tensors(model_metrics_,nr_classes))#dont specify loss and optimizer yet...expert stuff will follow and model will be recompiled
            elif collection==True and deviceSelected=="Multi-GPU":
                print("Collection & Multi-GPU is not supported yet")
                return

            #Original learning rate:
            #learning_rate_original = self.learning_rate_original#K.eval(model_keras.optimizer.lr)
            #Original trainable states of layers with parameters
            trainable_original, layer_names = self.trainable_original, self.layer_names
            do_list_original = self.do_list_original
            
            #Collect all information about the fitting routine that was user
            #defined
            if self.actionVerbose.isChecked()==True:
                verbose = 1
            else:
                verbose = 0
    
            new_model = self.radioButton_NewModel.isChecked()
            chosen_model = str(self.comboBox_ModelSelection.currentText())
                
            crop = int(self.spinBox_imagecrop.value())      
            color_mode = str(self.comboBox_GrayOrRGB.currentText())
            
            loadrestart_model = self.radioButton_LoadRestartModel.isChecked()
            loadcontinue_model = self.radioButton_LoadContinueModel.isChecked()
    
            norm = str(self.comboBox_Normalization.currentText())
    
            nr_epochs = int(self.spinBox_NrEpochs.value())
            keras_refresh_nr_epochs = int(self.spinBox_RefreshAfterEpochs.value())
            h_flip = bool(self.checkBox_HorizFlip.isChecked())
            v_flip = bool(self.checkBox_VertFlip.isChecked())
            rotation = float(self.lineEdit_Rotation.text())
     
            width_shift = float(self.lineEdit_widthShift.text())
            height_shift = float(self.lineEdit_heightShift.text())
            zoom = float(self.lineEdit_zoomRange.text())
            shear = float(self.lineEdit_shearRange.text())
            
            brightness_refresh_nr_epochs = int(self.spinBox_RefreshAfterNrEpochs.value())
            brightness_add_lower = float(self.spinBox_PlusLower.value())
            brightness_add_upper = float(self.spinBox_PlusUpper.value())
            brightness_mult_lower = float(self.doubleSpinBox_MultLower.value())
            brightness_mult_upper = float(self.doubleSpinBox_MultUpper.value())
            gaussnoise_mean = float(self.doubleSpinBox_GaussianNoiseMean.value())
            gaussnoise_scale = float(self.doubleSpinBox_GaussianNoiseScale.value())
    
            contrast_on = bool(self.checkBox_contrast.isChecked())        
            contrast_lower = float(self.doubleSpinBox_contrastLower.value())
            contrast_higher = float(self.doubleSpinBox_contrastHigher.value())
            saturation_on = bool(self.checkBox_saturation.isChecked())        
            saturation_lower = float(self.doubleSpinBox_saturationLower.value())
            saturation_higher = float(self.doubleSpinBox_saturationHigher.value())
            hue_on = bool(self.checkBox_hue.isChecked())        
            hue_delta = float(self.doubleSpinBox_hueDelta.value())
    
            avgBlur_on = bool(self.checkBox_avgBlur.isChecked())        
            avgBlur_min = int(self.spinBox_avgBlurMin.value())
            avgBlur_max = int(self.spinBox_avgBlurMax.value())
            gaussBlur_on = bool(self.checkBox_gaussBlur.isChecked())        
            gaussBlur_min = int(self.spinBox_gaussBlurMin.value())
            gaussBlur_max = int(self.spinBox_gaussBlurMax.value())
            motionBlur_on = bool(self.checkBox_motionBlur.isChecked())        
            motionBlur_kernel = str(self.lineEdit_motionBlurKernel.text())
            motionBlur_angle = str(self.lineEdit_motionBlurAngle.text())
            motionBlur_kernel = tuple(ast.literal_eval(motionBlur_kernel)) #translate string in the lineEdits to a tuple
            motionBlur_angle = tuple(ast.literal_eval(motionBlur_angle)) #translate string in the lineEdits to a tuple
    
    
            if collection==False:
                expert_mode = bool(self.groupBox_expertMode.isChecked())
            elif collection==True:
                expert_mode = self.groupBox_expertMode.setChecked(False)
                print("Expert mode was switched off. Not implemented yet for collections")
                expert_mode = False
    
            batchSize_expert = int(self.spinBox_batchSize.value())
            epochs_expert = int(self.spinBox_epochs.value())
            
            learning_rate_expert_on = bool(self.groupBox_learningRate.isChecked()) 
            learning_rate_const_on = bool(self.radioButton_LrConst.isChecked()) 
            learning_rate_const = float(self.doubleSpinBox_learningRate.value())
            learning_rate_cycLR_on = bool(self.radioButton_LrCycl.isChecked())
            try:
                cycLrMin = float(self.lineEdit_cycLrMin.text())
                cycLrMax = float(self.lineEdit_cycLrMax.text())
            except:
                cycLrMin = []
                cycLrMax = []
            cycLrMethod = str(self.comboBox_cycLrMethod.currentText())
            #clr_settings = self.fittingpopups_ui[listindex].clr_settings.copy()
            cycLrGamma = self.clr_settings["gamma"]            
            SelectedFiles = self.items_clicked()#to compute cycLrStepSize, the number of training images is needed
            cycLrStepSize = aid_dl.get_cyclStepSize(SelectedFiles,self.clr_settings["step_size"],batchSize_expert)
            #put clr_settings onto fittingpopup,
            self.fittingpopups_ui[listindex].clr_settings = self.clr_settings.copy()#assign a copy. Otherwise values in both dicts are changed when manipulating one dict            
            #put optimizer_settings onto fittingpopup,
            self.fittingpopups_ui[listindex].optimizer_settings = self.optimizer_settings.copy()#assign a copy. Otherwise values in both dicts are changed when manipulating one dict            
            
            learning_rate_expo_on = bool(self.radioButton_LrExpo.isChecked()) 
            expDecInitLr = float(self.doubleSpinBox_expDecInitLr.value())
            expDecSteps = int(self.spinBox_expDecSteps.value())
            expDecRate = float(self.doubleSpinBox_expDecRate.value())

            loss_expert_on = bool(self.checkBox_expt_loss.isChecked())
            loss_expert = str(self.comboBox_expt_loss.currentText()).lower()
            optimizer_expert_on = bool(self.checkBox_optimizer.isChecked())
            optimizer_expert = str(self.comboBox_optimizer.currentText()).lower()
            optimizer_settings = self.fittingpopups_ui[listindex].optimizer_settings.copy()#make a copy to make sure that changes in the UI are not immediately used

            paddingMode = str(self.comboBox_paddingMode.currentText())#.lower()
    
            train_last_layers = bool(self.checkBox_trainLastNOnly.isChecked())             
            train_last_layers_n = int(self.spinBox_trainLastNOnly.value())              
            train_dense_layers = bool(self.checkBox_trainDenseOnly.isChecked())             
            dropout_expert_on = bool(self.checkBox_dropout.isChecked())             
            try:
                dropout_expert = str(self.lineEdit_dropout.text()) #due to the validator, there are no squ.brackets
                dropout_expert = "["+dropout_expert+"]"
                dropout_expert = ast.literal_eval(dropout_expert)        
            except:
                dropout_expert = []
            lossW_expert_on = bool(self.checkBox_lossW.isChecked())             
            lossW_expert = str(self.lineEdit_lossW.text())
    
            #To get the class weights (loss), the SelectedFiles are required 
            #SelectedFiles = self.items_clicked()
            #Check if xtra_data should be used for training
            xtra_in = [s["xtra_in"] for s in SelectedFiles]
            if len(set(xtra_in))==1:
                xtra_in = list(set(xtra_in))[0]
            elif len(set(xtra_in))>1:# False and True is present. Not supported
                print("Xtra data is used only for some files. Xtra data needs to be used either by all or by none!")
                return

            self.fittingpopups_ui[listindex].SelectedFiles = SelectedFiles #save to self. to make it accessible for popup showing loss weights
            #Get the class weights. This function runs now the first time in the fitting routine. 
            #It is possible that the user chose Custom weights and then changed the classes. Hence first check if 
            #there is a weight for each class available.
            class_weight = self.get_class_weight(self.fittingpopups_ui[listindex].SelectedFiles,lossW_expert,custom_check_classes=True)
            if type(class_weight)==list:
                #There has been a mismatch between the classes described in class_weight and the classes available in SelectedFiles!
                lossW_expert = class_weight[0] #overwrite 
                class_weight = class_weight[1]
                print("class_weight:" +str(class_weight))
                print("There has been a mismatch between the classes described in \
                      Loss weights and the classes available in the selected files! \
                      Hence, the Loss weights are set to Balanced")

            #Get callback for the learning rate scheduling
            callback_lr = aid_dl.get_lr_callback(learning_rate_const_on,learning_rate_const,
                                               learning_rate_cycLR_on,cycLrMin,cycLrMax,
                                               cycLrMethod,cycLrStepSize,
                                               learning_rate_expo_on,
                                               expDecInitLr,expDecSteps,expDecRate,cycLrGamma)
            #save a dictionary with initial values
            lr_dict_original = aid_dl.get_lr_dict(learning_rate_const_on,learning_rate_const,
                                               learning_rate_cycLR_on,cycLrMin,cycLrMax,
                                               cycLrMethod,cycLrStepSize,
                                               learning_rate_expo_on,
                                               expDecInitLr,expDecSteps,expDecRate,cycLrGamma)
            
            if collection==False:    
                #Create an excel file
                writer = pd.ExcelWriter(new_modelname.split(".model")[0]+'_meta.xlsx', engine='openpyxl')
                self.fittingpopups_ui[listindex].writer = writer
                #Used files go to a separate sheet on the MetaFile.xlsx
                SelectedFiles_df = pd.DataFrame(SelectedFiles)
                pd.DataFrame().to_excel(writer,sheet_name='UsedData') #initialize empty Sheet
                SelectedFiles_df.to_excel(writer,sheet_name='UsedData')
                DataOverview_df = self.get_dataOverview()
                DataOverview_df.to_excel(writer,sheet_name='DataOverview') #write data overview to separate sheet            
                pd.DataFrame().to_excel(writer,sheet_name='Parameters') #initialize empty Sheet
                pd.DataFrame().to_excel(writer,sheet_name='History') #initialize empty Sheet
                

            elif collection==True: 
                SelectedFiles_df = pd.DataFrame(SelectedFiles)
    
                Writers = []
                #Create excel files
                for i in range(len(model_keras_path)):
                    writer = pd.ExcelWriter(model_keras_path[i].split(".model")[0]+'_meta.xlsx', engine='openpyxl')
                    Writers.append(writer)
                for writer in Writers:
                    #Used files go to a separate sheet on the MetaFile.xlsx
                    pd.DataFrame().to_excel(writer,sheet_name='UsedData') #initialize empty Sheet
                    SelectedFiles_df.to_excel(writer,sheet_name='UsedData')
                    DataOverview_df = self.get_dataOverview()
                    DataOverview_df.to_excel(writer,sheet_name='DataOverview') #write data overview to separate sheet            
                    pd.DataFrame().to_excel(writer,sheet_name='Parameters') #initialize empty Sheet
                    pd.DataFrame().to_excel(writer,sheet_name='History') #initialize empty Sheet
            
            ###############################Expert Mode values##################
            expert_mode_before = False #There was no expert mode used before.
            if expert_mode==True:
                #activate groupBox_expertMode_pop
                self.fittingpopups_ui[listindex].groupBox_expertMode_pop.setChecked(True)
                expert_mode_before = True
                #Some settings only need to be changed once, after user clicked apply at next epoch
                        
                #Apply the changes to trainable states:
                if train_last_layers==True:#Train only the last n layers
                    print("Train only the last "+str(train_last_layers_n)+ " layer(s)")
                    trainable_new = (len(trainable_original)-train_last_layers_n)*[False]+train_last_layers_n*[True]
                    summary = aid_dl.model_change_trainability(model_keras,trainable_new,model_metrics_t,nr_classes,loss_expert,optimizer_settings,learning_rate_const)
                    if model_keras_p!=None:#if this is NOT None, there exists a parallel model, which also needs to be re-compiled
                        model_metrics_t = aid_dl.get_metrics_tensors(self.get_metrics(),nr_classes)
                        aid_dl.model_compile(model_keras_p,loss_expert,optimizer_settings,learning_rate_const,model_metrics_t,nr_classes)
                        print("Recompiled parallel model for train_last_layers==True")
                    text1 = "Expert mode: Request for custom trainability states: train only the last "+str(train_last_layers_n)+ " layer(s)\n"
                    #text2 = "\n--------------------\n"
                    self.fittingpopups_ui[listindex].textBrowser_FittingInfo.append(text1+summary)
                if train_dense_layers==True:#Train only dense layers
                    print("Train only dense layers")
                    layer_dense_ind = ["Dense" in x for x in layer_names]
                    layer_dense_ind = np.where(np.array(layer_dense_ind)==True)[0] #at which indices are dropout layers?
                    #create a list of trainable states
                    trainable_new = len(trainable_original)*[False]
                    for index in layer_dense_ind:
                        trainable_new[index] = True
                    summary = aid_dl.model_change_trainability(model_keras,trainable_new,model_metrics_t,nr_classes,loss_expert,optimizer_settings,learning_rate_const)                  
                    if model_keras_p!=None:#if this is NOT None, there exists a parallel model, which also needs to be re-compiled
                        model_metrics_t = aid_dl.get_metrics_tensors(self.get_metrics(),nr_classes)
                        aid_dl.model_compile(model_keras_p,loss_expert,optimizer_settings,learning_rate_const,model_metrics_t,nr_classes)
                        print("Recompiled parallel model for train_dense_layers==True")
                    text1 = "Expert mode: Request for custom trainability states: train only dense layer(s)\n"
                    #text2 = "\n--------------------\n"
                    self.fittingpopups_ui[listindex].textBrowser_FittingInfo.append(text1+summary)
    
                if dropout_expert_on==True:
                    #The user apparently want to change the dropout rates
                    do_list = aid_dl.get_dropout(model_keras)#Get a list of dropout values of the current model
                    #Compare the dropout values in the model to the dropout values requested by user
                    if len(dropout_expert)==1:#if the user gave a single float
                        dropout_expert_list = len(do_list)*dropout_expert #convert to list
                    elif len(dropout_expert)>1:
                        dropout_expert_list = dropout_expert
                        if not len(dropout_expert_list)==len(do_list):
                            text = "Issue with dropout: you defined "+str(len(dropout_expert_list))+" dropout rates, but model has "+str(len(do_list))+" dropout layers"
                            self.fittingpopups_ui[listindex].textBrowser_FittingInfo.append(text)
                    else:
                        text = "Could not understand user input at Expert->Dropout"
                        self.fittingpopups_ui[listindex].textBrowser_FittingInfo.append(text)
                        dropout_expert_list = []
                    if len(dropout_expert_list)>0 and do_list!=dropout_expert_list:#if the dropout rates of the current model is not equal to the required do_list from user...
                        do_changed = aid_dl.change_dropout(model_keras,dropout_expert_list,model_metrics_t,nr_classes,loss_expert,optimizer_settings,learning_rate_const)
                        if model_keras_p!=None:#if this is NOT None, there exists a parallel model, which also needs to be re-compiled
                            model_metrics_t = aid_dl.get_metrics_tensors(self.get_metrics(),nr_classes)
                            aid_dl.model_compile(model_keras_p,loss_expert,optimizer_settings,learning_rate_const,model_metrics_t,nr_classes)
                            print("Recompiled parallel model to change dropout. I'm not sure if this works already!")
                        if do_changed==1:
                            text_do = "Dropout rate(s) in model was/were changed to: "+str(dropout_expert_list)
                        else:
                            text_do = "Dropout rate(s) in model was/were not changed"
                    else:
                        text_do = "Dropout rate(s) in model was/were not changed"
                    print(text_do)
                    self.fittingpopups_ui[listindex].textBrowser_FittingInfo.append(text_do)
    
    
            text_updates = ""
            #Compare current lr and the lr on expert tab:
            if collection == False:
                lr_current = model_keras.optimizer.get_config()["learning_rate"]
            else:
                lr_current = model_keras[0].optimizer.get_config()["learning_rate"]
    
            lr_diff = learning_rate_const-lr_current
            if  abs(lr_diff) > 1e-6:
                if collection == False:
                    K.set_value(model_keras.optimizer.lr, learning_rate_const)
                if collection == True:
                    for m in model_keras:
                        K.set_value(m.optimizer.lr, learning_rate_const)
                text_updates +=  "Changed the learning rate to "+ str(learning_rate_const)+"\n"
            
            #Check if model has to be compiled again
            recompile = False #by default, dont recompile (happens for "Load and continue" training a model)
            if new_model==True:
                recompile = True
            
            #Compare current optimizer and the optimizer on expert tab:
            if collection==False:
                optimizer_current = aid_dl.get_optimizer_name(model_keras).lower()#get the current optimizer of the model
            if collection==True:
                optimizer_current = aid_dl.get_optimizer_name(model_keras[0]).lower()#get the current optimizer of the model
            if optimizer_current!=optimizer_expert.lower():#if the current model has a different optimizer
                recompile = True
                text_updates+="Changed the optimizer to "+optimizer_expert+"\n"
    
            #Compare current loss function and the loss-function on expert tab:
            if collection==False:
                if model_keras.loss!=loss_expert:
                    recompile = True
                    text_updates+="Changed the loss function to "+loss_expert+"\n"
            if collection==True:
                if model_keras[0].loss!=loss_expert:
                    recompile = True
                    text_updates+="Changed the loss function to "+loss_expert+"\n"
    
            if recompile==True:
                print("Recompiling...")
                if collection==False:
                    aid_dl.model_compile(model_keras,loss_expert,optimizer_settings,learning_rate_const,model_metrics_t,nr_classes)
                if collection==True:
                    for m in model_keras:
                        model_metrics_t = aid_dl.get_metrics_tensors(self.get_metrics(),nr_classes)
                        aid_dl.model_compile(m, loss_expert, optimizer_settings, learning_rate_const,model_metrics_t, nr_classes)
                if model_keras_p!=None:#if this is NOT None, there exists a parallel model, which also needs to be re-compiled
                    model_metrics_t = aid_dl.get_metrics_tensors(self.get_metrics(),nr_classes)
                    aid_dl.model_compile(model_keras_p,loss_expert,optimizer_settings,learning_rate_const,model_metrics_t,nr_classes)
                    print("Recompiled parallel model to adjust learning rate, loss, optimizer")
    
            self.fittingpopups_ui[listindex].textBrowser_FittingInfo.append(text_updates)
    
            #self.model_keras = model_keras #overwrite the model on self
    
            ######################Load the Training Data################################
            ind = [selectedfile["TrainOrValid"] == "Train" for selectedfile in SelectedFiles]
            ind = np.where(np.array(ind)==True)[0]
            SelectedFiles_train = np.array(SelectedFiles)[ind]
            SelectedFiles_train = list(SelectedFiles_train)
            indices_train = [selectedfile["class"] for selectedfile in SelectedFiles_train]
            nr_events_epoch_train = [selectedfile["nr_events_epoch"] for selectedfile in SelectedFiles_train]
            rtdc_path_train = [selectedfile["rtdc_path"] for selectedfile in SelectedFiles_train]
            zoom_factors_train = [selectedfile["zoom_factor"] for selectedfile in SelectedFiles_train]
            #zoom_order = [self.actionOrder0.isChecked(),self.actionOrder1.isChecked(),self.actionOrder2.isChecked(),self.actionOrder3.isChecked(),self.actionOrder4.isChecked(),self.actionOrder5.isChecked()]
            #zoom_order = int(np.where(np.array(zoom_order)==True)[0])
            zoom_order = int(self.comboBox_zoomOrder.currentIndex()) #the combobox-index is already the zoom order

            shuffle_train = [selectedfile["shuffle"] for selectedfile in SelectedFiles_train]
            xtra_in = set([selectedfile["xtra_in"] for selectedfile in SelectedFiles_train])   
            if len(xtra_in)>1:# False and True is present. Not supported
                print("Xtra data is used only for some files. Xtra data needs to be used either by all or by none!")
                return
            xtra_in = list(xtra_in)[0]#this is either True or False
            
            #read self.ram to new variable ; next clear ram. This is required for multitasking (training multiple models with maybe different data)
            DATA = self.ram
            if verbose==1:
                print("Length of DATA (in RAM) = "+str(len(DATA)))
            #clear the ram again if desired
            if not self.actionKeep_Data_in_RAM.isChecked():
                self.ram = dict()
                print("Removed data from self.ram. For further training sessions, data has to be reloaded.")
            #If the scaling method is "divide by mean and std of the whole training set":
            if norm == "StdScaling using mean and std of all training data":
                mean_trainingdata,std_trainingdata = [],[]
                for i in range(len(SelectedFiles_train)):
                    #if Data_to_RAM was not enabled:
                    #if not self.actionDataToRam.isChecked():
                    if len(DATA)==0: #Here, the entire training set needs to be used! Not only random images!
                        #Replace=true: means individual cells could occur several times
                        gen_train = aid_img.gen_crop_img(crop,rtdc_path_train[i],random_images=False,zoom_factor=zoom_factors_train[i],zoom_order=zoom_order,color_mode=self.get_color_mode(),padding_mode=paddingMode) 
    #                    else: #get a similar generator, using the ram-data
    #                        if len(DATA)==0:
    #                            gen_train = aid_img.gen_crop_img(crop,rtdc_path_train[i],random_images=False) #Replace true means that individual cells could occur several times
                    else:
                        gen_train = aid_img.gen_crop_img_ram(DATA,rtdc_path_train[i],random_images=False) #Replace true means that individual cells could occur several times
                        if self.actionVerbose.isChecked():
                            print("Loaded data from RAM")
                        
                    images = next(gen_train)[0]
                    mean_trainingdata.append(np.mean(images))
                    std_trainingdata.append(np.std(images))
                mean_trainingdata = np.mean(np.array(mean_trainingdata))
                std_trainingdata = np.mean(np.array(std_trainingdata))
                
                if np.allclose(std_trainingdata,0):
                    std_trainingdata = 0.0001
    
                    msg = QtWidgets.QMessageBox()
                    msg.setIcon(QtWidgets.QMessageBox.Information)       
                    text = "<html><head/><body><p>The standard deviation of your training data is zero! This would lead to division by zero. To avoid this, I will divide by 0.0001 instead.</p></body></html>"
                    msg.setText(text) 
                    msg.setWindowTitle("Std. is zero")
                    msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                    msg.exec_()
        
            Para_dict = pd.DataFrame()
            def update_para_dict():
                #Document changes in the meta-file
                Para_dict["AIDeveloper_Version"]=VERSION,
                Para_dict["model_zoo_version"]=model_zoo_version,
                try:
                    Para_dict["OS"]=platform.platform(),
                    Para_dict["CPU"]=platform.processor(),
                except:
                    Para_dict["OS"]="Unknown",
                    Para_dict["CPU"]="Unknown",
    
                Para_dict["Modelname"]=new_modelname,
                Para_dict["Chosen Model"]=chosen_model,
                Para_dict["new_model"]=new_model,
                Para_dict["loadrestart_model"]=loadrestart_model,
                Para_dict["loadcontinue_model"]=loadcontinue_model,
                Para_dict["Continued_Fitting_From"]=load_modelname,                        
                Para_dict["Input image size"]=crop,
                Para_dict["Color Mode"]=color_mode,
                Para_dict["Zoom order"]=zoom_order,                
                Para_dict["Device"]=deviceSelected,
                Para_dict["gpu_used"]=gpu_used,
                Para_dict["gpu_memory"]=gpu_memory,
                Para_dict["Output Nr. classes"]=nr_classes,
                Para_dict["Normalization"]=norm,
                Para_dict["Nr. epochs"]=nr_epochs,
                Para_dict["Keras refresh after nr. epochs"]=keras_refresh_nr_epochs,
                Para_dict["Horz. flip"]=h_flip,
                Para_dict["Vert. flip"]=v_flip,
                Para_dict["rotation"]=rotation,
                Para_dict["width_shift"]=width_shift,
                Para_dict["height_shift"]=height_shift,
                Para_dict["zoom"]=zoom,
                Para_dict["shear"]=shear,
                Para_dict["Brightness refresh after nr. epochs"]=brightness_refresh_nr_epochs,
                Para_dict["Brightness add. lower"]=brightness_add_lower,
                Para_dict["Brightness add. upper"]=brightness_add_upper,
                Para_dict["Brightness mult. lower"]=brightness_mult_lower,  
                Para_dict["Brightness mult. upper"]=brightness_mult_upper,
                Para_dict["Gaussnoise Mean"]=gaussnoise_mean,
                Para_dict["Gaussnoise Scale"]=gaussnoise_scale,
    
                Para_dict["Contrast on"]=contrast_on,                
                Para_dict["Contrast Lower"]=contrast_lower,
                Para_dict["Contrast Higher"]=contrast_higher,
                Para_dict["Saturation on"]=saturation_on,
                Para_dict["Saturation Lower"]=saturation_lower,
                Para_dict["Saturation Higher"]=saturation_higher,
                Para_dict["Hue on"]=hue_on,                
                Para_dict["Hue delta"]=hue_delta,                
    
                Para_dict["Average blur on"]=avgBlur_on,                
                Para_dict["Average blur Lower"]=avgBlur_min,
                Para_dict["Average blur Higher"]=avgBlur_max,
                Para_dict["Gauss blur on"]=gaussBlur_on,                
                Para_dict["Gauss blur Lower"]=gaussBlur_min,
                Para_dict["Gauss blur Higher"]=gaussBlur_max,
                Para_dict["Motion blur on"]=motionBlur_on,                
                Para_dict["Motion blur Kernel"]=motionBlur_kernel,                
                Para_dict["Motion blur Angle"]=motionBlur_angle,                
           
                Para_dict["Epoch_Started_Using_These_Settings"]=counter,
    
                Para_dict["expert_mode"]=expert_mode,
                Para_dict["batchSize_expert"]=batchSize_expert,
                Para_dict["epochs_expert"]=epochs_expert,

                Para_dict["learning_rate_expert_on"]=learning_rate_expert_on,
                Para_dict["learning_rate_const_on"]=learning_rate_const_on,
                Para_dict["learning_rate_const"]=learning_rate_const,
                Para_dict["learning_rate_cycLR_on"]=learning_rate_cycLR_on,
                Para_dict["cycLrMin"]=cycLrMin,
                Para_dict["cycLrMax"]=cycLrMax,
                Para_dict["cycLrMethod"] = cycLrMethod,
                Para_dict["clr_settings"] = self.fittingpopups_ui[listindex].clr_settings,
                
                Para_dict["learning_rate_expo_on"]=learning_rate_expo_on,
                Para_dict["expDecInitLr"]=expDecInitLr,
                Para_dict["expDecSteps"]=expDecSteps,
                Para_dict["expDecRate"]=expDecRate,
                
                Para_dict["loss_expert_on"]=loss_expert_on,
                Para_dict["loss_expert"]=loss_expert,
                Para_dict["optimizer_expert_on"]=optimizer_expert_on,
                Para_dict["optimizer_expert"]=optimizer_expert,                
                Para_dict["optimizer_settings"]=optimizer_settings,                
    
                Para_dict["paddingMode"]=paddingMode,                
                
                Para_dict["train_last_layers"]=train_last_layers,
                Para_dict["train_last_layers_n"]=train_last_layers_n,
                Para_dict["train_dense_layers"]=train_dense_layers,
                Para_dict["dropout_expert_on"]=dropout_expert_on,
                Para_dict["dropout_expert"]=dropout_expert,
                Para_dict["lossW_expert_on"]=lossW_expert_on,
                Para_dict["lossW_expert"]=lossW_expert,
                Para_dict["class_weight"]=class_weight,
                Para_dict["metrics"]=model_metrics,
                
                #training data cannot be changed during training
                if norm == "StdScaling using mean and std of all training data":                                
                    #This needs to be saved into Para_dict since it will be required for inference
                    Para_dict["Mean of training data used for scaling"]=mean_trainingdata,
                    Para_dict["Std of training data used for scaling"]=std_trainingdata,
    
                if collection==False:
                    if counter == 0:
                        Para_dict.to_excel(self.fittingpopups_ui[listindex].writer,sheet_name='Parameters')
                    else:
                        Para_dict.to_excel(self.fittingpopups_ui[listindex].writer,sheet_name='Parameters',startrow=self.fittingpopups_ui[listindex].writer.sheets['Parameters'].max_row,header= False)

                    if os.path.isfile(new_modelname.split(".model")[0]+'_meta.xlsx'):
                        os.chmod(new_modelname.split(".model")[0]+'_meta.xlsx', S_IREAD|S_IRGRP|S_IROTH|S_IWRITE|S_IWGRP|S_IWOTH)#change to read/write
                    try:
                        self.fittingpopups_ui[listindex].writer.save()
                    except:
                        pass
                    os.chmod(new_modelname.split(".model")[0]+'_meta.xlsx', S_IREAD|S_IRGRP|S_IROTH)#change to only readable
    
                if collection==True:
                    for i in range(len(Writers)):
                        Para_dict["Chosen Model"]=model_architecture_names[i],
                        writer = Writers[i]
                        if counter==0:
                            Para_dict.to_excel(Writers[i],sheet_name='Parameters')
                        else:
                            Para_dict.to_excel(writer,sheet_name='Parameters',startrow=writer.sheets['Parameters'].max_row,header= False)

                        if os.path.isfile(model_keras_path[i].split(".model")[0]+'_meta.xlsx'):
                            os.chmod(model_keras_path[i].split(".model")[0]+'_meta.xlsx', S_IREAD|S_IRGRP|S_IROTH|S_IWRITE|S_IWGRP|S_IWOTH) #read/write
                        try:
                            writer.save()
                        except:
                            pass
                        os.chmod(model_keras_path[i].split(".model")[0]+'_meta.xlsx', S_IREAD|S_IRGRP|S_IROTH) #read only
    
                
            ######################Load the Validation Data################################
            ind = [selectedfile["TrainOrValid"] == "Valid" for selectedfile in SelectedFiles]
            ind = np.where(np.array(ind)==True)[0]
            SelectedFiles_valid = np.array(SelectedFiles)[ind]
            SelectedFiles_valid = list(SelectedFiles_valid)
            indices_valid = [selectedfile["class"] for selectedfile in SelectedFiles_valid]
            nr_events_epoch_valid = [selectedfile["nr_events_epoch"] for selectedfile in SelectedFiles_valid]
            rtdc_path_valid = [selectedfile["rtdc_path"] for selectedfile in SelectedFiles_valid]
            zoom_factors_valid = [selectedfile["zoom_factor"] for selectedfile in SelectedFiles_valid]
            #zoom_order = [self.actionOrder0.isChecked(),self.actionOrder1.isChecked(),self.actionOrder2.isChecked(),self.actionOrder3.isChecked(),self.actionOrder4.isChecked(),self.actionOrder5.isChecked()]
            #zoom_order = int(np.where(np.array(zoom_order)==True)[0])
            zoom_order = int(self.comboBox_zoomOrder.currentIndex()) #the combobox-index is already the zoom order
            shuffle_valid = [selectedfile["shuffle"] for selectedfile in SelectedFiles_valid]
            xtra_in = set([selectedfile["xtra_in"] for selectedfile in SelectedFiles_valid])   
            if len(xtra_in)>1:# False and True is present. Not supported
                print("Xtra data is used only for some files. Xtra data needs to be used either by all or by none!")
                return
            xtra_in = list(xtra_in)[0]#this is either True or False

            ############Cropping#####################
            X_valid,y_valid,Indices,xtra_valid = [],[],[],[]
            for i in range(len(SelectedFiles_valid)):
                if not self.actionDataToRam.isChecked():
                    #Replace=true means individual cells could occur several times
                    gen_valid = aid_img.gen_crop_img(crop,rtdc_path_valid[i],nr_events_epoch_valid[i],random_images=shuffle_valid[i],replace=True,zoom_factor=zoom_factors_valid[i],zoom_order=zoom_order,color_mode=self.get_color_mode(),padding_mode=paddingMode,xtra_in=xtra_in)
                else: #get a similar generator, using the ram-data
                    if len(DATA)==0:
                        #Replace=true means individual cells could occur several times
                        gen_valid = aid_img.gen_crop_img(crop,rtdc_path_valid[i],nr_events_epoch_valid[i],random_images=shuffle_valid[i],replace=True,zoom_factor=zoom_factors_valid[i],zoom_order=zoom_order,color_mode=self.get_color_mode(),padding_mode=paddingMode,xtra_in=xtra_in)
                    else:
                        gen_valid = aid_img.gen_crop_img_ram(DATA,rtdc_path_valid[i],nr_events_epoch_valid[i],random_images=shuffle_valid[i],replace=True,xtra_in=xtra_in) #Replace true means that individual cells could occur several times
                        if self.actionVerbose.isChecked():
                            print("Loaded data from RAM")
                generator_cropped_out = next(gen_valid)
                X_valid.append(generator_cropped_out[0])
                #y_valid.append(np.repeat(indices_valid[i],nr_events_epoch_valid[i]))
                y_valid.append(np.repeat(indices_valid[i],X_valid[-1].shape[0]))
                Indices.append(generator_cropped_out[1])
                xtra_valid.append(generator_cropped_out[2])
                del generator_cropped_out
            #Save the validation set (BEFORE normalization!)
            #Write to.rtdc files
            if bool(self.actionExport_Original.isChecked())==True:
                print("Export original images")
                save_cropped = False
                aid_bin.write_rtdc(new_modelname.split(".model")[0]+'_Valid_Data.rtdc',rtdc_path_valid,X_valid,Indices,cropped=save_cropped,color_mode=self.get_color_mode(),xtra_in=xtra_valid)
    
            elif bool(self.actionExport_Cropped.isChecked())==True:
                print("Export cropped images")
                save_cropped = True
                aid_bin.write_rtdc(new_modelname.split(".model")[0]+'_Valid_Data.rtdc',rtdc_path_valid,X_valid,Indices,cropped=save_cropped,color_mode=self.get_color_mode(),xtra_in=xtra_valid)
    
            elif bool(self.actionExport_Off.isChecked())==True:
                print("Exporting is turned off")
    #                msg = QtWidgets.QMessageBox()
    #                msg.setIcon(QtWidgets.QMessageBox.Information)       
    #                msg.setText("Use a different Exporting option in ->Edit if you want to export the data")
    #                msg.setWindowTitle("Export is turned off!")
    #                msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
    #                msg.exec_()
                
            X_valid = np.concatenate(X_valid)
            y_valid = np.concatenate(y_valid)
            Y_valid = to_categorical(y_valid, nr_classes)# * 2 - 1
            xtra_valid = np.concatenate(xtra_valid)
            if not bool(self.actionExport_Off.isChecked())==True:
                #Save the labels
                np.savetxt(new_modelname.split(".model")[0]+'_Valid_Labels.txt',y_valid.astype(int),fmt='%i')            
    
            if len(X_valid.shape)==4:
                channels=3
            elif len(X_valid.shape)==3:
                channels=1
            else:
                print("Invalid data dimension:" +str(X_valid.shape))
            if channels==1:
                #Add the "channels" dimension
                X_valid = np.expand_dims(X_valid,3)
    
            #get it to theano image format (channels first)
            #X_valid = X_valid.swapaxes(-1,-2).swapaxes(-2,-3)
            if norm == "StdScaling using mean and std of all training data":
                X_valid = aid_img.image_normalization(X_valid,norm,mean_trainingdata,std_trainingdata)
            else:
                X_valid = aid_img.image_normalization(X_valid,norm)
    
            #Validation data can be cropped to final size already since no augmentation
            #will happen on this data set
            dim_val = X_valid.shape
            print("Current dim. of validation set (pixels x pixels) = "+str(dim_val[2]))
            if dim_val[2]!=crop:
                print("Change dim. (pixels x pixels) of validation set to = "+str(crop))
                remove = int(dim_val[2]/2.0 - crop/2.0)
                X_valid = X_valid[:,remove:remove+crop,remove:remove+crop,:] #crop to crop x crop pixels #TensorFlow
            
            if xtra_in==True:
                print("Add Xtra Data to X_valid")
                X_valid = [X_valid,xtra_valid]
    
    
            ####################Update the PopupFitting########################
            self.fittingpopups_ui[listindex].lineEdit_modelname_pop.setText(new_modelname) #set the progress bar to zero
            self.fittingpopups_ui[listindex].spinBox_imagecrop_pop.setValue(crop)
            self.fittingpopups_ui[listindex].spinBox_NrEpochs.setValue(nr_epochs)
            self.fittingpopups_ui[listindex].comboBox_ModelSelection_pop.addItems(self.predefined_models)
            chosen_model = str(self.comboBox_ModelSelection.currentText())
            index = self.fittingpopups_ui[listindex].comboBox_ModelSelection_pop.findText(chosen_model, QtCore.Qt.MatchFixedString)
            if index >= 0:
                self.fittingpopups_ui[listindex].comboBox_ModelSelection_pop.setCurrentIndex(index)
            self.fittingpopups_ui[listindex].comboBox_Normalization_pop.addItems(self.norm_methods)            
            index = self.fittingpopups_ui[listindex].comboBox_Normalization_pop.findText(norm, QtCore.Qt.MatchFixedString)
            if index >= 0:
                self.fittingpopups_ui[listindex].comboBox_Normalization_pop.setCurrentIndex(index)
            #padding
            index = self.fittingpopups_ui[listindex].comboBox_paddingMode_pop.findText(paddingMode, QtCore.Qt.MatchFixedString)
            if index >= 0:
                self.fittingpopups_ui[listindex].comboBox_paddingMode_pop.setCurrentIndex(index)
            #zoom_order
            self.fittingpopups_ui[listindex].comboBox_zoomOrder.setCurrentIndex(zoom_order)
            #CPU setting
            self.fittingpopups_ui[listindex].comboBox_cpu_pop.addItem("Default CPU")        
            if gpu_used==False:
                self.fittingpopups_ui[listindex].radioButton_cpu_pop.setChecked(True)
                self.fittingpopups_ui[listindex].doubleSpinBox_memory_pop.setValue(gpu_memory)        
            #GPU setting
            if gpu_used==True:
                self.fittingpopups_ui[listindex].radioButton_gpu_pop.setChecked(True)
                self.fittingpopups_ui[listindex].comboBox_gpu_pop.addItem(deviceSelected)        
                self.fittingpopups_ui[listindex].doubleSpinBox_memory_pop.setValue(gpu_memory)        
    
            self.fittingpopups_ui[listindex].spinBox_RefreshAfterEpochs_pop.setValue(keras_refresh_nr_epochs)
            self.fittingpopups_ui[listindex].checkBox_HorizFlip_pop.setChecked(h_flip)
            self.fittingpopups_ui[listindex].checkBox_VertFlip_pop.setChecked(v_flip)
            self.fittingpopups_ui[listindex].lineEdit_Rotation_pop.setText(str(rotation))
            self.fittingpopups_ui[listindex].lineEdit_widthShift_pop.setText(str(width_shift))
            self.fittingpopups_ui[listindex].lineEdit_heightShift_pop.setText(str(height_shift))
            self.fittingpopups_ui[listindex].lineEdit_zoomRange_pop.setText(str(zoom))
            self.fittingpopups_ui[listindex].lineEdit_shearRange_pop.setText(str(shear))
            self.fittingpopups_ui[listindex].spinBox_RefreshAfterNrEpochs_pop.setValue(brightness_refresh_nr_epochs)
            self.fittingpopups_ui[listindex].spinBox_PlusLower_pop.setValue(brightness_add_lower)
            self.fittingpopups_ui[listindex].spinBox_PlusUpper_pop.setValue(brightness_add_upper)
            self.fittingpopups_ui[listindex].doubleSpinBox_MultLower_pop.setValue(brightness_mult_lower)
            self.fittingpopups_ui[listindex].doubleSpinBox_MultUpper_pop.setValue(brightness_mult_upper)
            self.fittingpopups_ui[listindex].doubleSpinBox_GaussianNoiseMean_pop.setValue(gaussnoise_mean)
            self.fittingpopups_ui[listindex].doubleSpinBox_GaussianNoiseScale_pop.setValue(gaussnoise_scale) 
    
            self.fittingpopups_ui[listindex].checkBox_contrast_pop.setChecked(contrast_on) 
            self.fittingpopups_ui[listindex].doubleSpinBox_contrastLower_pop.setValue(contrast_lower) 
            self.fittingpopups_ui[listindex].doubleSpinBox_contrastHigher_pop.setValue(contrast_higher) 
            self.fittingpopups_ui[listindex].checkBox_saturation_pop.setChecked(saturation_on) 
            self.fittingpopups_ui[listindex].doubleSpinBox_saturationLower_pop.setValue(saturation_lower) 
            self.fittingpopups_ui[listindex].doubleSpinBox_saturationHigher_pop.setValue(saturation_higher) 
            self.fittingpopups_ui[listindex].checkBox_hue_pop.setChecked(hue_on) 
            self.fittingpopups_ui[listindex].doubleSpinBox_hueDelta_pop.setValue(hue_delta) 
            #Special for saturation and hue. Only enabled for RGB:
            saturation_enabled = bool(self.checkBox_saturation.isEnabled())
            self.fittingpopups_ui[listindex].checkBox_saturation_pop.setEnabled(saturation_enabled)
            self.fittingpopups_ui[listindex].doubleSpinBox_saturationLower_pop.setEnabled(saturation_enabled)
            self.fittingpopups_ui[listindex].doubleSpinBox_saturationHigher_pop.setEnabled(saturation_enabled)
                
            hue_enabled = bool(self.checkBox_hue.isEnabled())
            self.fittingpopups_ui[listindex].checkBox_hue_pop.setEnabled(hue_enabled) 
            self.fittingpopups_ui[listindex].doubleSpinBox_hueDelta_pop.setEnabled(hue_enabled)
    
    
            self.fittingpopups_ui[listindex].checkBox_avgBlur_pop.setChecked(avgBlur_on)
            self.fittingpopups_ui[listindex].spinBox_avgBlurMin_pop.setEnabled(avgBlur_on)
            self.fittingpopups_ui[listindex].label_avgBlurMin_pop.setEnabled(avgBlur_on)
            self.fittingpopups_ui[listindex].spinBox_avgBlurMin_pop.setValue(avgBlur_min) 
            self.fittingpopups_ui[listindex].spinBox_avgBlurMax_pop.setEnabled(avgBlur_on)
            self.fittingpopups_ui[listindex].label_avgBlurMax_pop.setEnabled(avgBlur_on)
            self.fittingpopups_ui[listindex].spinBox_avgBlurMax_pop.setValue(avgBlur_max) 
    
            self.fittingpopups_ui[listindex].checkBox_gaussBlur_pop.setChecked(gaussBlur_on)
            self.fittingpopups_ui[listindex].spinBox_gaussBlurMin_pop.setEnabled(gaussBlur_on)
            self.fittingpopups_ui[listindex].label_gaussBlurMin_pop.setEnabled(gaussBlur_on)
            self.fittingpopups_ui[listindex].spinBox_gaussBlurMin_pop.setValue(gaussBlur_min) 
            self.fittingpopups_ui[listindex].spinBox_gaussBlurMax_pop.setEnabled(gaussBlur_on)
            self.fittingpopups_ui[listindex].label_gaussBlurMax_pop.setEnabled(gaussBlur_on)
            self.fittingpopups_ui[listindex].spinBox_gaussBlurMax_pop.setValue(gaussBlur_max) 
    
            self.fittingpopups_ui[listindex].checkBox_motionBlur_pop.setChecked(motionBlur_on)
            self.fittingpopups_ui[listindex].label_motionBlurKernel_pop.setEnabled(motionBlur_on)
            self.fittingpopups_ui[listindex].lineEdit_motionBlurKernel_pop.setEnabled(motionBlur_on)
            self.fittingpopups_ui[listindex].label_motionBlurAngle_pop.setEnabled(motionBlur_on)
            self.fittingpopups_ui[listindex].lineEdit_motionBlurAngle_pop.setEnabled(motionBlur_on)
            if len(motionBlur_kernel)==1:
                self.fittingpopups_ui[listindex].lineEdit_motionBlurKernel_pop.setText(str(motionBlur_kernel[0]))
            if len(motionBlur_kernel)==2:
                self.fittingpopups_ui[listindex].lineEdit_motionBlurKernel_pop.setText(str(motionBlur_kernel[0])+","+str(motionBlur_kernel[1]))
            if len(motionBlur_angle)==1:
                self.fittingpopups_ui[listindex].lineEdit_motionBlurAngle_pop.setText(str(motionBlur_angle[0]))
            if len(motionBlur_kernel)==2:
                self.fittingpopups_ui[listindex].lineEdit_motionBlurAngle_pop.setText(str(motionBlur_angle[0])+","+str(motionBlur_angle[1]))
    
            self.fittingpopups_ui[listindex].groupBox_expertMode_pop.setChecked(expert_mode)
            self.fittingpopups_ui[listindex].spinBox_batchSize.setValue(batchSize_expert)
            self.fittingpopups_ui[listindex].spinBox_epochs.setValue(epochs_expert)
    
            self.fittingpopups_ui[listindex].groupBox_learningRate_pop.setChecked(learning_rate_expert_on)
            self.fittingpopups_ui[listindex].radioButton_LrConst.setChecked(learning_rate_const_on)
            self.fittingpopups_ui[listindex].doubleSpinBox_learningRate.setValue(learning_rate_const)
            self.fittingpopups_ui[listindex].radioButton_LrCycl.setChecked(learning_rate_cycLR_on)
            self.fittingpopups_ui[listindex].lineEdit_cycLrMin.setText(str(cycLrMin))
            self.fittingpopups_ui[listindex].lineEdit_cycLrMax.setText(str(cycLrMax))
            index = self.fittingpopups_ui[listindex].comboBox_cycLrMethod.findText(cycLrMethod, QtCore.Qt.MatchFixedString)
            if index >= 0:
                self.fittingpopups_ui[listindex].comboBox_cycLrMethod.setCurrentIndex(index)
            self.fittingpopups_ui[listindex].radioButton_LrExpo.setChecked(learning_rate_expo_on)
            self.fittingpopups_ui[listindex].doubleSpinBox_expDecInitLr.setValue(expDecInitLr)
            self.fittingpopups_ui[listindex].spinBox_expDecSteps.setValue(expDecSteps)
            self.fittingpopups_ui[listindex].doubleSpinBox_expDecRate.setValue(expDecRate) 

            self.fittingpopups_ui[listindex].checkBox_expt_loss_pop.setChecked(loss_expert_on)

            self.fittingpopups_ui[listindex].checkBox_expt_loss_pop.setChecked(loss_expert_on)
            index = self.fittingpopups_ui[listindex].comboBox_expt_loss_pop.findText(loss_expert, QtCore.Qt.MatchFixedString)
            if index >= 0:
                self.fittingpopups_ui[listindex].comboBox_expt_loss_pop.setCurrentIndex(index)
            self.fittingpopups_ui[listindex].checkBox_optimizer_pop.setChecked(optimizer_expert_on)
            index = self.fittingpopups_ui[listindex].comboBox_optimizer.findText(optimizer_expert, QtCore.Qt.MatchFixedString)
            if index >= 0:
                self.fittingpopups_ui[listindex].comboBox_optimizer.setCurrentIndex(index)
            self.fittingpopups_ui[listindex].doubleSpinBox_learningRate.setValue(learning_rate_const)
    
            index = self.fittingpopups_ui[listindex].comboBox_paddingMode_pop.findText(paddingMode, QtCore.Qt.MatchFixedString)
            if index >= 0:
                self.fittingpopups_ui[listindex].comboBox_paddingMode_pop.setCurrentIndex(index)
    
            self.fittingpopups_ui[listindex].checkBox_trainLastNOnly_pop.setChecked(train_last_layers)
            self.fittingpopups_ui[listindex].spinBox_trainLastNOnly_pop.setValue(train_last_layers_n)
            self.fittingpopups_ui[listindex].checkBox_trainDenseOnly_pop.setChecked(train_dense_layers)
            self.fittingpopups_ui[listindex].checkBox_dropout_pop.setChecked(dropout_expert_on)
            do_text = [str(do_i) for do_i in dropout_expert]
            self.fittingpopups_ui[listindex].lineEdit_dropout_pop.setText((', '.join(do_text)))
            self.fittingpopups_ui[listindex].checkBox_lossW.setChecked(lossW_expert_on)
            self.fittingpopups_ui[listindex].pushButton_lossW.setEnabled(lossW_expert_on)
            self.fittingpopups_ui[listindex].lineEdit_lossW.setText(str(lossW_expert))
    
            if channels==1:
                channel_text = "Grayscale"
            elif channels==3:
                channel_text = "RGB"
            self.fittingpopups_ui[listindex].comboBox_colorMode_pop.addItems([channel_text])
    
            ###############Continue with training data:augmentation############
            #Rotating could create edge effects. Avoid this by making crop a bit larger for now
            #Worst case would be a 45degree rotation:
            cropsize2 = np.sqrt(crop**2+crop**2)
            cropsize2 = np.ceil(cropsize2 / 2.) * 2 #round to the next even number
        
            #Dictionary defining affine image augmentation options:
            aug_paras = {"v_flip":v_flip,"h_flip":h_flip,"rotation":rotation,"width_shift":width_shift,"height_shift":height_shift,"zoom":zoom,"shear":shear}
                         
            Histories,Index,Saved,Stopwatch,LearningRate = [],[],[],[],[]
            if collection==True:
               HISTORIES = [ [] for model in model_keras]
               SAVED = [ [] for model in model_keras]
    
            counter = 0
            saving_failed = False #when saving fails, this becomes true and the user will be informed at the end of training

            #Save the initial values (Epoch 1)
            update_para_dict()
            
            #Dictionary for records in metrics
            model_metrics_records = {}
            model_metrics_records["accuracy"] = 0 #accuracy  starts at zero and approaches 1 during training         
            model_metrics_records["val_accuracy"] = 0 #accuracy  starts at zero and approaches 1 during training         
            model_metrics_records["loss"] = 9E20 ##loss starts very high and approaches 0 during training         
            model_metrics_records["val_loss"] = 9E20 ##loss starts very high and approaches 0 during training         
            for key in model_keras.metrics_names:
                if 'precision' in key or 'recall' in key or 'auc' in key:
                    model_metrics_records[key] = 0 #those metrics start at zero and approach 1         
                    model_metrics_records["val_"+key] = 0 #those metrics start at zero and approach 1         
    
            gen_train_refresh = False
            time_start = time.perf_counter()
            t1 = time.perf_counter() #Initialize a timer; this is used to save the meta file every few seconds
            t2 =  time.perf_counter() #Initialize a timer; this is used update the fitting parameters
            while counter < nr_epochs:#nr_epochs: #resample nr_epochs times
                #Only keep fitting if the respective window is open:
                isVisible = self.fittingpopups[listindex].isVisible()
                if isVisible:                    
                    ############Keras image augmentation#####################
                    #Start the first iteration:                
                    X_train,y_train,xtra_train = [],[],[]
                    t3 = time.perf_counter()
                    for i in range(len(SelectedFiles_train)):
                        if len(DATA)==0 or gen_train_refresh:
                            #Replace true means that individual cells could occur several times
                            gen_train = aid_img.gen_crop_img(cropsize2,rtdc_path_train[i],nr_events_epoch_train[i],random_images=shuffle_train[i],replace=True,zoom_factor=zoom_factors_train[i],zoom_order=zoom_order,color_mode=self.get_color_mode(),padding_mode=paddingMode,xtra_in=xtra_in) 
                            gen_train_refresh = False
                        else:
                            gen_train = aid_img.gen_crop_img_ram(DATA,rtdc_path_train[i],nr_events_epoch_train[i],random_images=shuffle_train[i],replace=True,xtra_in=xtra_in) #Replace true means that individual cells could occur several times
                            if self.actionVerbose.isChecked():
                                print("Loaded data from RAM")
                        data_ = next(gen_train)
                        X_train.append(data_[0])
                        y_train.append(np.repeat(indices_train[i],X_train[-1].shape[0]))
                        if xtra_in==True:
                            xtra_train.append(data_[2])
                        del data_
                        
                    X_train = np.concatenate(X_train)
                    X_train = X_train.astype(np.uint8)
                    y_train = np.concatenate(y_train)
                    if xtra_in==True:
                        print("Retrieve Xtra Data...")
                        xtra_train = np.concatenate(xtra_train)
                    
                    t4 = time.perf_counter()
                    if verbose == 1:
                        print("Time to load data (from .rtdc or RAM) and crop="+str(t4-t3))
                    
                    if len(X_train.shape)==4:
                        channels=3
                    elif len(X_train.shape)==3:
                        channels=1
                    else:
                        print("Invalid data dimension:" +str(X_train.shape))
                    if channels==1:
                        #Add the "channels" dimension
                        X_train = np.expand_dims(X_train,3)
    
                    t3 = time.perf_counter()
                    #Some parallellization: use nr_threads (number of CPUs)
                    nr_threads = 1 #Somehow for MNIST and CIFAR, processing always took longer for nr_threads>1 . I tried nr_threads=2,4,8,16,24
                    if nr_threads == 1:
                        X_batch = aid_img.affine_augm(X_train,v_flip,h_flip,rotation,width_shift,height_shift,zoom,shear) #Affine image augmentation
                        y_batch = np.copy(y_train)
                    else:
                        #Divde data in 4 batches
                        X_train = np.array_split(X_train,nr_threads)
                        y_train = np.array_split(y_train,nr_threads)
    
                        self.X_batch = [False] * nr_threads
                        self.y_batch = [False] * nr_threads
                        self.counter_aug = 0
                        self.Workers_augm = []
                        
                        def imgaug_worker(aug_paras,progress_callback,history_callback):
                            i = aug_paras["i"]
                            self.X_batch[i] = aid_img.affine_augm(aug_paras["X_train"],v_flip,h_flip,rotation,width_shift,height_shift,zoom,shear)
                            self.y_batch[i] = aug_paras["y_train"]
                            self.counter_aug+=1
    
                        t3_a = time.perf_counter()
                        for i in range(nr_threads):
                            aug_paras_ = copy.deepcopy(aug_paras)
                            aug_paras_["i"] = i
                            aug_paras_["X_train"]=X_train[i]#augparas contains rotation and so on. X_train and y_train are overwritten in each iteration (for each worker new X_train)
                            aug_paras_["y_train"]=y_train[i]
                            
                            self.Workers_augm.append(Worker(imgaug_worker,aug_paras_))                            
                            self.threadpool.start(self.Workers_augm[i])
                            
                        while self.counter_aug < nr_threads:
                            time.sleep(0.01)#Wait 0.1s, then check the len again
                        t3_b = time.perf_counter()
                        if verbose == 1:
                            print("Time to perform affine augmentation_internal ="+str(t3_b-t3_a))
    
                        X_batch = np.concatenate(self.X_batch)
                        y_batch = np.concatenate(self.y_batch)
       
                    Y_batch = to_categorical(y_batch, nr_classes)# * 2 - 1
                    t4 = time.perf_counter()
                    if verbose == 1:
                        print("Time to perform affine augmentation ="+str(t4-t3))
                            
                    t3 = time.perf_counter()            
                    #Now do the final cropping to the actual size that was set by user
                    dim = X_batch.shape
                    if dim[2]!=crop:
                        remove = int(dim[2]/2.0 - crop/2.0)
                        X_batch = X_batch[:,remove:remove+crop,remove:remove+crop,:] #crop to crop x crop pixels #TensorFlow
                    t4 = time.perf_counter()
    #                    if verbose == 1:
    #                        print("Time to crop to final size="+str(t4-t3))
    
                    X_batch_orig = np.copy(X_batch) #save into new array and do some iterations with varying noise/brightness
                    #reuse this X_batch_orig a few times since this augmentation was costly
                    keras_iter_counter = 0
                    while keras_iter_counter < keras_refresh_nr_epochs and counter < nr_epochs:
                        keras_iter_counter+=1
                        #if t2-t1>5: #check for changed settings every 5 seconds
                        if self.actionVerbose.isChecked()==True:
                            verbose = 1
                        else:
                            verbose = 0                            
                                                                        
                        #Another while loop if the user wants to reuse the keras-augmented data
                        #several times and only apply brightness augmentation:
                        brightness_iter_counter = 0
                        while brightness_iter_counter < brightness_refresh_nr_epochs and counter < nr_epochs:
                            #In each iteration, start with non-augmented data
                            X_batch = np.copy(X_batch_orig)#copy from X_batch_orig, X_batch will be altered without altering X_batch_orig            
                            X_batch = X_batch.astype(np.uint8)                            
                            
                            #########X_batch = X_batch.astype(float)########## No float yet :) !!!
                            
                            brightness_iter_counter += 1
                            if self.actionVerbose.isChecked()==True:
                                verbose = 1
                            else:
                                verbose = 0                            
    
                            if self.fittingpopups_ui[listindex].checkBox_ApplyNextEpoch.isChecked():
                                nr_epochs = int(self.fittingpopups_ui[listindex].spinBox_NrEpochs.value())
                                #Keras stuff
                                keras_refresh_nr_epochs = int(self.fittingpopups_ui[listindex].spinBox_RefreshAfterEpochs_pop.value())                                
                                h_flip = bool(self.fittingpopups_ui[listindex].checkBox_HorizFlip_pop.isChecked())
                                v_flip = bool(self.fittingpopups_ui[listindex].checkBox_VertFlip_pop.isChecked())
                                rotation = float(self.fittingpopups_ui[listindex].lineEdit_Rotation_pop.text())
                                width_shift = float(self.fittingpopups_ui[listindex].lineEdit_widthShift_pop.text())
                                height_shift = float(self.fittingpopups_ui[listindex].lineEdit_heightShift_pop.text())
                                zoom = float(self.fittingpopups_ui[listindex].lineEdit_zoomRange_pop.text())
                                shear = float(self.fittingpopups_ui[listindex].lineEdit_shearRange_pop.text())
                                #Brightness stuff
                                brightness_refresh_nr_epochs = int(self.fittingpopups_ui[listindex].spinBox_RefreshAfterNrEpochs_pop.value())
                                brightness_add_lower = float(self.fittingpopups_ui[listindex].spinBox_PlusLower_pop.value())
                                brightness_add_upper = float(self.fittingpopups_ui[listindex].spinBox_PlusUpper_pop.value())
                                brightness_mult_lower = float(self.fittingpopups_ui[listindex].doubleSpinBox_MultLower_pop.value())
                                brightness_mult_upper = float(self.fittingpopups_ui[listindex].doubleSpinBox_MultUpper_pop.value())
                                gaussnoise_mean = float(self.fittingpopups_ui[listindex].doubleSpinBox_GaussianNoiseMean_pop.value())
                                gaussnoise_scale = float(self.fittingpopups_ui[listindex].doubleSpinBox_GaussianNoiseScale_pop.value())
    
                                contrast_on = bool(self.fittingpopups_ui[listindex].checkBox_contrast_pop.isChecked())
                                contrast_lower = float(self.fittingpopups_ui[listindex].doubleSpinBox_contrastLower_pop.value())
                                contrast_higher = float(self.fittingpopups_ui[listindex].doubleSpinBox_contrastHigher_pop.value())
                                saturation_on = bool(self.fittingpopups_ui[listindex].checkBox_saturation_pop.isChecked())
                                saturation_lower = float(self.fittingpopups_ui[listindex].doubleSpinBox_saturationLower_pop.value())
                                saturation_higher = float(self.fittingpopups_ui[listindex].doubleSpinBox_saturationHigher_pop.value())
                                hue_on = bool(self.fittingpopups_ui[listindex].checkBox_hue_pop.isChecked())
                                hue_delta = float(self.fittingpopups_ui[listindex].doubleSpinBox_hueDelta_pop.value())
    
                                avgBlur_on = bool(self.fittingpopups_ui[listindex].checkBox_avgBlur_pop.isChecked())        
                                avgBlur_min = int(self.fittingpopups_ui[listindex].spinBox_avgBlurMin_pop.value())
                                avgBlur_max = int(self.fittingpopups_ui[listindex].spinBox_avgBlurMax_pop.value())
                    
                                gaussBlur_on = bool(self.fittingpopups_ui[listindex].checkBox_gaussBlur_pop.isChecked())        
                                gaussBlur_min = int(self.fittingpopups_ui[listindex].spinBox_gaussBlurMin_pop.value())
                                gaussBlur_max = int(self.fittingpopups_ui[listindex].spinBox_gaussBlurMax_pop.value())
                    
                                motionBlur_on = bool(self.fittingpopups_ui[listindex].checkBox_motionBlur_pop.isChecked())        
                                motionBlur_kernel = str(self.fittingpopups_ui[listindex].lineEdit_motionBlurKernel_pop.text())
                                motionBlur_angle = str(self.fittingpopups_ui[listindex].lineEdit_motionBlurAngle_pop.text())
                                
                                motionBlur_kernel = tuple(ast.literal_eval(motionBlur_kernel)) #translate string in the lineEdits to a tuple
                                motionBlur_angle = tuple(ast.literal_eval(motionBlur_angle)) #translate string in the lineEdits to a tuple
    
                                #Expert mode stuff
                                expert_mode = bool(self.fittingpopups_ui[listindex].groupBox_expertMode_pop.isChecked())
                                batchSize_expert = int(self.fittingpopups_ui[listindex].spinBox_batchSize.value())
                                epochs_expert = int(self.fittingpopups_ui[listindex].spinBox_epochs.value())
                                
                                learning_rate_expert_on = bool(self.fittingpopups_ui[listindex].groupBox_learningRate_pop.isChecked())
                                learning_rate_const_on = bool(self.fittingpopups_ui[listindex].radioButton_LrConst.isChecked())
                                learning_rate_const = float(self.fittingpopups_ui[listindex].doubleSpinBox_learningRate.value())
                
                                learning_rate_cycLR_on = bool(self.fittingpopups_ui[listindex].radioButton_LrCycl.isChecked())
                                try:
                                    cycLrMin = float(self.fittingpopups_ui[listindex].lineEdit_cycLrMin.text())
                                    cycLrMax = float(self.fittingpopups_ui[listindex].lineEdit_cycLrMax.text())
                                except:
                                    cycLrMin = []
                                    cycLrMax = []
                                cycLrMethod = str(self.fittingpopups_ui[listindex].comboBox_cycLrMethod.currentText())
                                clr_settings = self.fittingpopups_ui[listindex].clr_settings.copy() #Get a copy of the current optimizer_settings. .copy prevents that changes in the UI have immediate effect
                                cycLrStepSize = aid_dl.get_cyclStepSize(SelectedFiles,clr_settings["step_size"],batchSize_expert)
                                cycLrGamma = clr_settings["gamma"]

                                learning_rate_expo_on = bool(self.fittingpopups_ui[listindex].radioButton_LrExpo.isChecked())
                                expDecInitLr = float(self.fittingpopups_ui[listindex].doubleSpinBox_expDecInitLr.value())
                                expDecSteps = int(self.fittingpopups_ui[listindex].spinBox_expDecSteps.value())
                                expDecRate = float(self.fittingpopups_ui[listindex].doubleSpinBox_expDecRate.value())

                                loss_expert_on = bool(self.fittingpopups_ui[listindex].checkBox_expt_loss_pop.isChecked())
                                loss_expert = str(self.fittingpopups_ui[listindex].comboBox_expt_loss_pop.currentText())
                                optimizer_expert_on = bool(self.fittingpopups_ui[listindex].checkBox_optimizer_pop.isChecked())
                                optimizer_expert = str(self.fittingpopups_ui[listindex].comboBox_optimizer.currentText())
                                optimizer_settings = self.fittingpopups_ui[listindex].optimizer_settings.copy() #Get a copy of the current optimizer_settings. .copy prevents that changes in the UI have immediate effect
                                paddingMode_ = str(self.fittingpopups_ui[listindex].comboBox_paddingMode_pop.currentText())
                                print("paddingMode_:"+str(paddingMode_))
                                if paddingMode_ != paddingMode:
                                    print("Changed the padding mode!")
                                    gen_train_refresh = True#otherwise changing paddingMode will not have any effect
                                    paddingMode = paddingMode_
                                    
                                train_last_layers = bool(self.fittingpopups_ui[listindex].checkBox_trainLastNOnly_pop.isChecked())             
                                train_last_layers_n = int(self.fittingpopups_ui[listindex].spinBox_trainLastNOnly_pop.value())              
                                train_dense_layers = bool(self.fittingpopups_ui[listindex].checkBox_trainDenseOnly_pop.isChecked())             
                                dropout_expert_on = bool(self.fittingpopups_ui[listindex].checkBox_dropout_pop.isChecked())             
                                try:
                                    dropout_expert = str(self.fittingpopups_ui[listindex].lineEdit_dropout_pop.text()) #due to the validator, there are no squ.brackets
                                    dropout_expert = "["+dropout_expert+"]"
                                    dropout_expert = ast.literal_eval(dropout_expert)
                                except:
                                    dropout_expert = []
                                lossW_expert_on = bool(self.fittingpopups_ui[listindex].checkBox_lossW.isChecked())             
                                lossW_expert = str(self.fittingpopups_ui[listindex].lineEdit_lossW.text())             
                                class_weight = self.get_class_weight(self.fittingpopups_ui[listindex].SelectedFiles,lossW_expert) #

                                print("Updating parameter file (meta.xlsx)!")
                                update_para_dict()
    
                                #Changes in expert mode can affect the model: apply changes now:
                                if expert_mode==True:
                                    if collection==False: #Expert mode is currently not supported for Collections
                                        expert_mode_before = True
    
                                        #Apply changes to the trainable states:
                                        if train_last_layers==True:#Train only the last n layers
                                            if verbose:
                                                print("Train only the last "+str(train_last_layers_n)+ " layer(s)")
                                            trainable_new = (len(trainable_original)-train_last_layers_n)*[False]+train_last_layers_n*[True]
                                            #Change the trainability states. Model compilation is done inside model_change_trainability
                                            summary = aid_dl.model_change_trainability(model_keras,trainable_new,model_metrics_t,nr_classes,loss_expert,optimizer_settings,learning_rate_const)
                                            if model_keras_p!=None:#if this is NOT None, there exists a parallel model, which also needs to be re-compiled
                                                model_metrics_t = aid_dl.get_metrics_tensors(self.get_metrics(),nr_classes)
                                                aid_dl.model_compile(model_keras_p,loss_expert,optimizer_settings,learning_rate_const,model_metrics_t,nr_classes)
                                                print("Recompiled parallel model due to train_last_layers==True")
                                            text1 = "Expert mode: Request for custom trainability states: train only the last "+str(train_last_layers_n)+ " layer(s)\n"
                                            #text2 = "\n--------------------\n"
                                            self.fittingpopups_ui[listindex].textBrowser_FittingInfo.append(text1+summary)
                                        if train_dense_layers==True:#Train only dense layers
                                            if verbose:
                                                print("Train only dense layers")
                                            layer_dense_ind = ["Dense" in x for x in layer_names]
                                            layer_dense_ind = np.where(np.array(layer_dense_ind)==True)[0] #at which indices are dropout layers?
                                            #create a list of trainable states
                                            trainable_new = len(trainable_original)*[False]
                                            for index in layer_dense_ind:
                                                trainable_new[index] = True
                                            #Change the trainability states. Model compilation is done inside model_change_trainability
                                            summary = aid_dl.model_change_trainability(model_keras,trainable_new,model_metrics_t,nr_classes,loss_expert,optimizer_settings,learning_rate_const)                 
                                            if model_keras_p!=None:#if this is NOT None, there exists a parallel model, which also needs to be re-compiled
                                                model_metrics_t = aid_dl.get_metrics_tensors(self.get_metrics(),nr_classes)
                                                aid_dl.model_compile(model_keras_p,loss_expert,optimizer_settings,learning_rate_const,model_metrics_t,nr_classes)
                                                print("Recompiled parallel model due to train_dense_layers==True")
                                            text1 = "Expert mode: Request for custom trainability states: train only dense layer(s)\n"
                                            #text2 = "\n--------------------\n"
                                            self.fittingpopups_ui[listindex].textBrowser_FittingInfo.append(text1+summary)
    
                                        if dropout_expert_on==True:
                                            #The user apparently want to change the dropout rates
                                            do_list = aid_dl.get_dropout(model_keras)#Get a list of dropout values of the current model
                                            #Compare the dropout values in the model to the dropout values requested by user
                                            if len(dropout_expert)==1:#if the user gave a float
                                                dropout_expert_list = len(do_list)*dropout_expert #convert to list
                                            elif len(dropout_expert)>1:
                                                dropout_expert_list = dropout_expert
                                                if not len(dropout_expert_list)==len(do_list):
                                                    text = "Issue with dropout: you defined "+str(len(dropout_expert_list))+" dropout rates, but model has "+str(len(do_list))+" dropout layers"
                                                    self.fittingpopups_ui[listindex].textBrowser_FittingInfo.append(text)
                                            else:
                                                text = "Could not understand user input at Expert->Dropout"
                                                self.fittingpopups_ui[listindex].textBrowser_FittingInfo.append(text)
                                                dropout_expert_list = []
    
                                            if len(dropout_expert_list)>0 and do_list!=dropout_expert_list:#if the dropout rates of the current model is not equal to the required do_list from user...
                                                #Change dropout. Model .compile happens inside change_dropout function
                                                do_changed = aid_dl.change_dropout(model_keras,dropout_expert_list,model_metrics_t,nr_classes,loss_expert,optimizer_settings,learning_rate_const)
                                                if model_keras_p!=None:#if model_keras_p is NOT None, there exists a parallel model, which also needs to be re-compiled
                                                    model_metrics_t = aid_dl.get_metrics_tensors(self.get_metrics(),nr_classes)
                                                    aid_dl.model_compile(model_keras_p,loss_expert,optimizer_settings,learning_rate_const,model_metrics_t,nr_classes)
                                                    print("Recompiled parallel model due to changed dropout. I'm not sure if this works already!")
    
                                                if do_changed==1:
                                                    text_do = "Dropout rate(s) in model was/were changed to: "+str(dropout_expert_list)
                                                else:
                                                    text_do = "Dropout rate(s) in model was/were not changed"
                                            else:
                                                text_do = "Dropout rate(s) in model was/were not changed"
                                            if verbose:
                                                print(text_do)
                                            self.fittingpopups_ui[listindex].textBrowser_FittingInfo.append(text_do)
                                        if learning_rate_expert_on==True:
                                            #get the current lr_dict
                                            lr_dict_now = aid_dl.get_lr_dict(learning_rate_const_on,learning_rate_const,
                                                                               learning_rate_cycLR_on,cycLrMin,cycLrMax,
                                                                               cycLrMethod,cycLrStepSize,
                                                                               learning_rate_expo_on,
                                                                               expDecInitLr,expDecSteps,expDecRate,cycLrGamma)
                                            if not lr_dict_now.equals(lr_dict_original):#in case the dataframes dont equal...
                                                #generate a new callback
                                                callback_lr = aid_dl.get_lr_callback(learning_rate_const_on,learning_rate_const,
                                                                                   learning_rate_cycLR_on,cycLrMin,cycLrMax,
                                                                                   cycLrMethod,cycLrStepSize,
                                                                                   learning_rate_expo_on,
                                                                                   expDecInitLr,expDecSteps,expDecRate,cycLrGamma)
                                                #update lr_dict_original
                                                lr_dict_original = lr_dict_now.copy()
                                        else:
                                            callback_lr = None

                                        if optimizer_expert_on==True:
                                            optimizer_settings_now = self.fittingpopups_ui[listindex].optimizer_settings.copy()
                                            if not optimizer_settings_now == optimizer_settings:#in case the dataframes dont equal...
                                                #grab these new optimizer values
                                                optimizer_settings = optimizer_settings_now.copy()
                                            
                                ############################Invert 'expert' settings#########################
                                if expert_mode==False and expert_mode_before==True: #if the expert mode was selected before, change the parameters back to original vlaues
                                    if verbose:
                                        print("Expert mode was used before and settings are now inverted")
    
                                    #Re-set trainable states back to original state                                    
                                    if verbose:
                                        print("Change 'trainable' layers back to original state")
                                    summary = aid_dl.model_change_trainability(model_keras,trainable_original,model_metrics,nr_classes,loss_expert,optimizer_settings,learning_rate_const)                 
                                    if model_keras_p!=None:#if model_keras_p is NOT None, there exists a parallel model, which also needs to be re-compiled
                                        model_metrics_t = aid_dl.get_metrics_tensors(self.get_metrics(),nr_classes)
                                        aid_dl.model_compile(model_keras_p,loss_expert,optimizer_settings,learning_rate_const,model_metrics_t,nr_classes)
                                        print("Recompiled parallel model to change 'trainable' layers back to original state")
    
                                    text1 = "Expert mode turns off: Request for orignal trainability states:\n"
                                    #text2 = "\n--------------------\n"
                                    self.fittingpopups_ui[listindex].textBrowser_FittingInfo.append(text1+summary)
                                    if verbose:
                                        print("Change dropout rates in dropout layers back to original values")
                                    
                                    callback_lr = None#remove learning rate callback
                                    if verbose:
                                        print("Set learning rate callback to None")
                                    
                                    if len(do_list_original)>0:
                                        do_changed = aid_dl.change_dropout(model_keras,do_list_original,model_metrics_t,nr_classes,loss_expert,optimizer_settings,learning_rate_const)
                                        if model_keras_p!=None:#if model_keras_p is NOT None, there exists a parallel model, which also needs to be re-compiled
                                            model_metrics_t = aid_dl.get_metrics_tensors(self.get_metrics(),nr_classes)
                                            aid_dl.model_compile(model_keras_p,loss_expert,optimizer_settings,learning_rate_const,model_metrics_t,nr_classes)
                                            print("Recompiled parallel model to change dropout values back to original state. I'm not sure if this works!")
    
                                        if do_changed==1:
                                            text_do = "Dropout rate(s) in model was/were changed to original values: "+str(do_list_original)
                                        else:
                                            text_do = "Dropout rate(s) in model was/were not changed"                                        
                                        self.fittingpopups_ui[listindex].textBrowser_FittingInfo.append(text_do+"\n")
    
    
                                text_updates = ""
                                #Compare current lr and the lr on expert tab:
                                if collection==False:
                                    lr_current = model_keras.optimizer.get_config()["learning_rate"]
                                else:
                                    lr_current = model_keras[0].optimizer.get_config()["learning_rate"]
    
                                lr_diff = learning_rate_const-lr_current
                                if  abs(lr_diff) > 1e-6:
                                    if collection==False:
                                        K.set_value(model_keras.optimizer.lr, learning_rate_const)
                                    else:
                                        K.set_value(model_keras[0].optimizer.lr, learning_rate_const)
    
                                    text_updates +=  "Changed the learning rate to "+ str(learning_rate_const)+"\n"
                                
                                recompile = False
                                #Compare current optimizer and the optimizer on expert tab:
                                if collection==False:
                                    optimizer_current = aid_dl.get_optimizer_name(model_keras).lower()#get the current optimizer of the model
                                else:
                                    optimizer_current = aid_dl.get_optimizer_name(model_keras[0]).lower()#get the current optimizer of the model
    
                                if optimizer_current!=optimizer_expert.lower():#if the current model has a different optimizer
                                    recompile = True
                                    text_updates+="Changed the optimizer to "+optimizer_expert+"\n"
    
                                #Compare current loss function and the loss-function on expert tab:
                                if collection==False:
                                    loss_ = model_keras.loss
                                else:
                                    loss_ = model_keras[0].loss
                                if loss_!=loss_expert:
                                    recompile = True
                                    model_metrics_records["loss"] = 9E20 #Reset the record for loss because new loss function could converge to a different min. value
                                    model_metrics_records["val_loss"] = 9E20 #Reset the record for loss because new loss function could converge to a different min. value
                                    text_updates+="Changed the loss function to "+loss_expert+"\n"
    
                                if recompile==True and collection==False:
                                    print("Recompiling...")
                                    model_metrics_t = aid_dl.get_metrics_tensors(self.get_metrics(),nr_classes)
                                    aid_dl.model_compile(model_keras,loss_expert,optimizer_settings,learning_rate_const,model_metrics_t,nr_classes)
                                    if model_keras_p!=None:#if model_keras_p is NOT None, there exists a parallel model, which also needs to be re-compiled
                                        model_metrics_t = aid_dl.get_metrics_tensors(self.get_metrics(),nr_classes)
                                        aid_dl.model_compile(model_keras_p,loss_expert,optimizer_settings,learning_rate_const,model_metrics_t,nr_classes)
                                        print("Recompiled parallel model to change optimizer, loss and learninig rate.")
    
                                elif recompile==True and collection==True:
                                    if model_keras_p!=None:#if model_keras_p is NOT None, there exists a parallel model, which also needs to be re-compiled
                                        print("Altering learning rate is not suported for collections (yet)")
                                        return
                                    print("Recompiling...")
                                    for m in model_keras:
                                        model_metrics_t = aid_dl.get_metrics_tensors(self.get_metrics(),nr_classes)
                                        aid_dl.model_compile(m,loss_expert,optimizer_settings,learning_rate_const,model_metrics_t,nr_classes)
    
                                self.fittingpopups_ui[listindex].textBrowser_FittingInfo.append(text_updates)
    
                                #self.model_keras = model_keras #overwrite the model in self
                                self.fittingpopups_ui[listindex].checkBox_ApplyNextEpoch.setChecked(False)
    
    
                            ##########Contrast/Saturation/Hue augmentation#########
                            #is there any of contrast/saturation/hue augmentation to do?
                            X_batch = X_batch.astype(np.uint8)
                            if contrast_on:
                                t_con_aug_1 = time.perf_counter()
                                X_batch = aid_img.contrast_augm_cv2(X_batch,contrast_lower,contrast_higher) #this function is almost 15 times faster than random_contrast from tf!
                                t_con_aug_2 = time.perf_counter()
                                if verbose == 1:
                                    print("Time to augment contrast="+str(t_con_aug_2-t_con_aug_1))
    
                            if saturation_on or hue_on:
                                t_sat_aug_1 = time.perf_counter()
                                X_batch = aid_img.satur_hue_augm_cv2(X_batch.astype(np.uint8),saturation_on,saturation_lower,saturation_higher,hue_on,hue_delta) #Gray and RGB; both values >0!
                                t_sat_aug_2 = time.perf_counter()
                                if verbose == 1:
                                    print("Time to augment saturation/hue="+str(t_sat_aug_2-t_sat_aug_1))
    
                            ##########Average/Gauss/Motion blurring#########
                            #is there any of blurring to do?
                            
                            if avgBlur_on:
                                t_avgBlur_1 = time.perf_counter()
                                X_batch = aid_img.avg_blur_cv2(X_batch,avgBlur_min,avgBlur_max)
                                t_avgBlur_2 = time.perf_counter()
                                if verbose == 1:
                                    print("Time to perform average blurring="+str(t_avgBlur_2-t_avgBlur_1))
    
                            if gaussBlur_on:
                                t_gaussBlur_1 = time.perf_counter()
                                X_batch = aid_img.gauss_blur_cv(X_batch,gaussBlur_min,gaussBlur_max)
                                t_gaussBlur_2 = time.perf_counter()
                                if verbose == 1:
                                    print("Time to perform gaussian blurring="+str(t_gaussBlur_2-t_gaussBlur_1))
    
                            if motionBlur_on:
                                t_motionBlur_1 = time.perf_counter()
                                X_batch = aid_img.motion_blur_cv(X_batch,motionBlur_kernel,motionBlur_angle)
                                t_motionBlur_2 = time.perf_counter()
                                if verbose == 1:
                                    print("Time to perform motion blurring="+str(t_motionBlur_2-t_motionBlur_1))
    
                            ##########Brightness noise#########
                            t3 = time.perf_counter()
                            X_batch = aid_img.brightn_noise_augm_cv2(X_batch,brightness_add_lower,brightness_add_upper,brightness_mult_lower,brightness_mult_upper,gaussnoise_mean,gaussnoise_scale)
                            t4 = time.perf_counter()
                            if verbose == 1:
                                print("Time to augment brightness="+str(t4-t3))
    
                            t3 = time.perf_counter()
                            if norm == "StdScaling using mean and std of all training data":
                                X_batch = aid_img.image_normalization(X_batch,norm,mean_trainingdata,std_trainingdata)
                            else:
                                X_batch = aid_img.image_normalization(X_batch,norm)
                            t4 = time.perf_counter()
                            if verbose == 1:
                                print("Time to apply normalization="+str(t4-t3))
                            
                            #Fitting can be paused
                            while str(self.fittingpopups_ui[listindex].pushButton_Pause_pop.text())=="":
                                time.sleep(2) #wait 2 seconds and then check the text on the button again
    
                            if verbose == 1: 
                                print("X_batch.shape")
                                print(X_batch.shape)

                            if xtra_in==True:
                                print("Add Xtra Data to X_batch")
                                X_batch = [X_batch,xtra_train]

                            #generate a list of callbacks, get empty list if callback_lr is none
                            callbacks = []
                            if callback_lr!=None:
                                callbacks.append(callback_lr)

                            ###################################################
                            ###############Actual fitting######################
                            ###################################################
    
                            if collection==False:
                                if model_keras_p == None:
                                    history = model_keras.fit(X_batch, Y_batch, batch_size=batchSize_expert, epochs=epochs_expert,verbose=verbose, validation_data=(X_valid, Y_valid),class_weight=class_weight,callbacks=callbacks)
                                elif model_keras_p != None:
                                    history = model_keras_p.fit(X_batch, Y_batch, batch_size=batchSize_expert, epochs=epochs_expert,verbose=verbose, validation_data=(X_valid, Y_valid),class_weight=class_weight,callbacks=callbacks)
                                
                                Histories.append(history.history)
                                Stopwatch.append(time.perf_counter()-time_start)
                                learningrate = K.get_value(history.model.optimizer.lr)
                                LearningRate.append(learningrate)

                                #Check if any metric broke a record
                                record_broken = False #initially, assume there is no new record
                                for key in history.history.keys():
                                    value = history.history[key][-1]
                                    record = model_metrics_records[key]
                                    if 'val_accuracy' in key or 'val_precision' in key or 'val_recall' in key or 'val_auc' in key:
                                        #These metrics should go up (towards 1)
                                        if value>record:
                                            model_metrics_records[key] = value
                                            record_broken = True
                                            print(key+" broke record -> Model will be saved" )
    
                                    elif 'val_loss' in key:
                                        #This metric should go down (towards 0)
                                        if value<record:
                                            model_metrics_records[key] = value
                                            record_broken = True
                                            print(key+" broke record -> Model will be saved")
                                                #self.fittingpopups_ui[listindex].textBrowser_FittingInfo.append(text)
    
                                if record_broken:#if any record was broken...
                                    if deviceSelected=="Multi-GPU":#in case of Multi-GPU...
                                        #In case of multi-GPU, first copy the weights of the parallel model to the normal model
                                        model_keras.set_weights(model_keras_p.layers[-2].get_weights())
                                    #Save the model
                                    text = "Save model to following directory: \n"+os.path.dirname(new_modelname)
                                    print(text)

                                    if os.path.exists(os.path.dirname(new_modelname)):
                                        model_keras.save(new_modelname.split(".model")[0]+"_"+str(counter)+".model",save_format='h5')
                                        text = "Record was broken -> saved model"
                                        print(text)
                                        self.fittingpopups_ui[listindex].textBrowser_FittingInfo.append(text)

                                    else:#in case the folder does not exist (anymore), create a folder in temp
                                        #what is the foldername of the model?
                                        text = "Saving failed. Create folder in temp"
                                        print(text)
                                        self.fittingpopups_ui[listindex].textBrowser_FittingInfo.append(text)

                                        saving_failed = True
                                        temp_path = aid_bin.create_temp_folder()#create a temp folder if it does not already exist

                                        text = "Your temp. folder is here: "+str(temp_path)
                                        print(text)
                                        self.fittingpopups_ui[listindex].textBrowser_FittingInfo.append(text)

                                        parentfolder = aid_bin.splitall(new_modelname)[-2]
                                        fname = os.path.split(new_modelname)[-1]

                                        #create that folder in temp if it not exists already
                                        if not os.path.exists(os.path.join(temp_path,parentfolder)):
                                            text = "Create folder in temp:\n"+os.path.join(temp_path,parentfolder)
                                            print(text)
                                            self.fittingpopups_ui[listindex].textBrowser_FittingInfo.append(text)
                                            os.mkdir(os.path.join(temp_path,parentfolder))

                                        #change the new_modelname to a path in temp
                                        new_modelname = os.path.join(temp_path,parentfolder,fname)

                                        #inform user!
                                        text = "Could not find original folder. Files are now saved to "+new_modelname
                                        text = "<span style=\' color: red;\'>" +text+"</span>"
                                        self.fittingpopups_ui[listindex].textBrowser_FittingInfo.append(text)
                                        text = "<span style=\' color: black;\'>" +""+"</span>"
                                        self.fittingpopups_ui[listindex].textBrowser_FittingInfo.append(text)

                                        #Save the  model
                                        model_keras.save(new_modelname.split(".model")[0]+"_"+str(counter)+".model",save_format='h5')
                                        text = "Model saved successfully to temp"
                                        print(text)
                                        self.fittingpopups_ui[listindex].textBrowser_FittingInfo.append(text)

                                        #Also update the excel writer!
                                        writer = pd.ExcelWriter(new_modelname.split(".model")[0]+'_meta.xlsx', engine='openpyxl')
                                        self.fittingpopups_ui[listindex].writer = writer
                                        pd.DataFrame().to_excel(writer,sheet_name='UsedData') #initialize empty Sheet
                                        SelectedFiles_df.to_excel(writer,sheet_name='UsedData')
                                        DataOverview_df.to_excel(writer,sheet_name='DataOverview') #write data overview to separate sheet            
                                        pd.DataFrame().to_excel(writer,sheet_name='Parameters') #initialize empty Sheet
                                        pd.DataFrame().to_excel(writer,sheet_name='History') #initialize empty Sheet

                                    Saved.append(1)
                                
                                #Also save the model upon user-request  
                                elif bool(self.fittingpopups_ui[listindex].checkBox_saveEpoch_pop.isChecked())==True:
                                    if deviceSelected=="Multi-GPU":#in case of Multi-GPU...
                                        #In case of multi-GPU, first copy the weights of the parallel model to the normal model
                                        model_keras.set_weights(model_keras_p.layers[-2].get_weights())
                                    model_keras.save(new_modelname.split(".model")[0]+"_"+str(counter)+".model",save_format='h5')
                                    Saved.append(1)
                                    self.fittingpopups_ui[listindex].checkBox_saveEpoch_pop.setChecked(False)
                                else:
                                    Saved.append(0)
    
                            elif collection==True:
                                for i in range(len(model_keras)):
                                    #Expert-settings return automatically to default values when Expert-mode is unchecked
                                    history = model_keras[i].fit(X_batch, Y_batch, batch_size=batchSize_expert, epochs=epochs_expert,verbose=verbose, validation_data=(X_valid, Y_valid),class_weight=class_weight,callbacks=callbacks)
                                    HISTORIES[i].append(history.history)
                                    learningrate = K.get_value(history.model.optimizer.lr)

                                    print("model_keras_path[i]")
                                    print(model_keras_path[i])
    
                                    #Check if any metric broke a record
                                    record_broken = False #initially, assume there is no new record
                                                                        
                                    for key in history.history.keys():
                                        value = history.history[key][-1]
                                        record = model_metrics_records[key]
                                        if 'val_accuracy' in key or 'val_precision' in key or 'val_recall' in key or 'val_auc' in key:
                                            #These metrics should go up (towards 1)
                                            if value>record:
                                                model_metrics_records[key] = value
                                                record_broken = True
                                                text = key+" broke record -> Model will be saved"
                                                print(text)
                                                self.fittingpopups_ui[listindex].textBrowser_FittingInfo.append(text)

                                                #one could 'break' here, but I want to update all records
                                        elif 'val_loss' in key:
                                            #This metric should go down (towards 0)
                                            if value<record:
                                                model_metrics_records[key] = value
                                                record_broken = True
                                                text = key+" broke record -> Model will be saved"
                                                print(text)
                                                self.fittingpopups_ui[listindex].textBrowser_FittingInfo.append(text)
                                    
                                    #For collections of models:
                                    if record_broken:
                                        #Save the model
                                        model_keras[i].save(model_keras_path[i].split(".model")[0]+"_"+str(counter)+".model")
                                        SAVED[i].append(1)
                                    elif bool(self.fittingpopups_ui[listindex].checkBox_saveEpoch_pop.isChecked())==True:
                                        model_keras[i].save(model_keras_path[i].split(".model")[0]+"_"+str(counter)+".model")
                                        SAVED[i].append(1)
                                        self.fittingpopups_ui[listindex].checkBox_saveEpoch_pop.setChecked(False)
                                    else:
                                        SAVED[i].append(0)
    
    
                            callback_progessbar = float(counter)/nr_epochs
                            progress_callback.emit(100.0*callback_progessbar)
                            history_emit = history.history
                            history_emit["LearningRate"] = [learningrate]
                            history_callback.emit(history_emit)
                            Index.append(counter)
                            
                            t2 =  time.perf_counter()
                            
                            if collection==False:
                                if counter==0:
                                    #If this runs the first time, create the file with header                                    
                                    DF1 = [[ h[h_i][-1] for h_i in h] for h in Histories] #if nb_epoch in .fit() is >1, only save the last history item, beacuse this would a model that could be saved
                                    DF1 = np.r_[DF1]
                                    DF1 = pd.DataFrame( DF1,columns=Histories[0].keys() )
                                    
                                    DF1["Saved"] = Saved
                                    DF1["Time"] = Stopwatch
                                    DF1["LearningRate"] = LearningRate
                                    DF1.index = Index
    
                                    #If this runs the first time, create the file with header
                                    if os.path.isfile(new_modelname.split(".model")[0]+'_meta.xlsx'):
                                        os.chmod(new_modelname.split(".model")[0]+'_meta.xlsx', S_IREAD|S_IRGRP|S_IROTH|S_IWRITE|S_IWGRP|S_IWOTH) #read/write
                                    DF1.to_excel(writer,sheet_name='History')
                                    writer.save()
                                    os.chmod(new_modelname.split(".model")[0]+'_meta.xlsx', S_IREAD|S_IRGRP|S_IROTH)
                                    
                                    meta_saving_t = int(self.fittingpopups_ui[listindex].spinBox_saveMetaEvery.value())
                                    text = "meta.xlsx was saved (automatic saving every "+str(meta_saving_t)+"s)"
                                    print(text)
                                    self.fittingpopups_ui[listindex].textBrowser_FittingInfo.append(text)

                                    #self.fittingpopups_ui[listindex].backup.append({"DF1":DF1})
                                    Index,Histories,Saved,Stopwatch,LearningRate = [],[],[],[],[]#reset the lists
                                    
                                #Get a sensible frequency for saving the dataframe (every 20s)
                                elif t2-t1>int(self.fittingpopups_ui[listindex].spinBox_saveMetaEvery.value()):                                   
                                #elif counter%50==0:  #otherwise save the history to excel after each n epochs
                                    DF1 = [[ h[h_i][-1] for h_i in h] for h in Histories] #if nb_epoch in .fit() is >1, only save the last history item, beacuse this would a model that could be saved
                                    DF1 = np.r_[DF1]
                                    DF1 = pd.DataFrame( DF1,columns=Histories[0].keys() )
                                    DF1["Saved"] = Saved
                                    DF1["Time"] = Stopwatch
                                    DF1["LearningRate"] = LearningRate
                                    DF1.index = Index

                                    #Saving
                                    if os.path.exists(os.path.dirname(new_modelname)):#check if folder is (still) available
                                        if os.path.isfile(new_modelname.split(".model")[0]+'_meta.xlsx'):
                                            os.chmod(new_modelname.split(".model")[0]+'_meta.xlsx', S_IREAD|S_IRGRP|S_IROTH|S_IWRITE|S_IWGRP|S_IWOTH) #make read/write
                                        DF1.to_excel(writer,sheet_name='History', startrow=writer.sheets['History'].max_row,header= False)
                                        writer.save()
                                        os.chmod(new_modelname.split(".model")[0]+'_meta.xlsx', S_IREAD|S_IRGRP|S_IROTH)  #make read only
                                        
                                        meta_saving_t = int(self.fittingpopups_ui[listindex].spinBox_saveMetaEvery.value())
                                        text = "meta.xlsx was saved (automatic saving every "+str(meta_saving_t)+"s to directory:\n)"+new_modelname
                                        print(text)
                                        self.fittingpopups_ui[listindex].textBrowser_FittingInfo.append(text)
                                        
                                        Index,Histories,Saved,Stopwatch,LearningRate = [],[],[],[],[]#reset the lists
                                        t1 = time.perf_counter()
                                    else:#If folder not available, create a folder in temp
                                        text = "Failed to save meta.xlsx. -> Create folder in temp\n"
                                        saving_failed = True
                                        temp_path = aid_bin.create_temp_folder()#create a temp folder if it does not already exist
                                        text += "Your temp folder is here: "+str(temp_path)+"\n"
                                        folder = os.path.split(new_modelname)[-2]
                                        folder = os.path.split(folder)[-1]
                                        fname = os.path.split(new_modelname)[-1]
                                        #create that folder in temp if it does'nt exist already
                                        if not os.path.exists(os.path.join(temp_path,folder)):
                                            os.mkdir(os.path.join(temp_path,folder))
                                            text +="Created directory in temp:\n"+os.path.join(temp_path,folder)

                                        print(text)
                                        #change the new_modelname to a path in temp
                                        new_modelname = os.path.join(temp_path,folder,fname)

                                        #inform user!
                                        text = "Could not find original folder. Files are now saved to "+new_modelname
                                        text = "<span style=\' color: red;\'>" +text+"</span>"#put red text to the infobox
                                        self.fittingpopups_ui[listindex].textBrowser_FittingInfo.append(text)
                                        text = "<span style=\' color: black;\'>" +""+"</span>"#reset textcolor to black
                                        self.fittingpopups_ui[listindex].textBrowser_FittingInfo.append(text)


                                        #update the excel writer
                                        writer = pd.ExcelWriter(new_modelname.split(".model")[0]+'_meta.xlsx', engine='openpyxl')
                                        self.fittingpopups_ui[listindex].writer = writer
                                        pd.DataFrame().to_excel(writer,sheet_name='UsedData') #initialize empty Sheet
                                        SelectedFiles_df.to_excel(writer,sheet_name='UsedData')
                                        DataOverview_df.to_excel(writer,sheet_name='DataOverview') #write data overview to separate sheet            
                                        pd.DataFrame().to_excel(writer,sheet_name='Parameters') #initialize empty Sheet
                                        pd.DataFrame().to_excel(writer,sheet_name='History') #initialize empty Sheet

                                        if os.path.isfile(new_modelname.split(".model")[0]+'_meta.xlsx'):
                                            print("There is already such a file...AID will add new data to it. Please check if this is OK")
                                            os.chmod(new_modelname.split(".model")[0]+'_meta.xlsx', S_IREAD|S_IRGRP|S_IROTH|S_IWRITE|S_IWGRP|S_IWOTH) #read/write
                                        DF1.to_excel(writer,sheet_name='History')
                                        writer.save()
                                        os.chmod(new_modelname.split(".model")[0]+'_meta.xlsx', S_IREAD|S_IRGRP|S_IROTH)
                                        print("meta.xlsx was saved")
                                        Index,Histories,Saved,Stopwatch,LearningRate = [],[],[],[],[]#reset the lists
                                        
                                        
                            if collection==True:
                                if counter==0:
                                    for i in range(len(HISTORIES)):
                                        Histories = HISTORIES[i]
                                        Saved = SAVED[i]
                                        #If this runs the first time, create the file with header
                                        DF1 = [[ h[h_i][-1] for h_i in h] for h in Histories] #if nb_epoch in .fit() is >1, only save the last history item, beacuse this would a model that could be saved
                                        DF1 = np.r_[DF1]
                                        DF1 = pd.DataFrame( DF1,columns=Histories[0].keys() )
                                        DF1["Saved"] = Saved
                                        DF1.index = Index
                                        HISTORIES[i] = []#reset the Histories list
                                        SAVED[i] = []
                                        #If this runs the first time, create the file with header
                                        if os.path.isfile(model_keras_path[i].split(".model")[0]+'_meta.xlsx'):
                                            os.chmod(model_keras_path[i].split(".model")[0]+'_meta.xlsx', S_IREAD|S_IRGRP|S_IROTH|S_IWRITE|S_IWGRP|S_IWOTH) #read/write
                                        DF1.to_excel(Writers[i],sheet_name='History')
                                        Writers[i].save()
                                        os.chmod(model_keras_path[i].split(".model")[0]+'_meta.xlsx', S_IREAD|S_IRGRP|S_IROTH)
                                        print("meta.xlsx was saved")
                                    Index = []#reset the Index list
                                    
                                #Get a sensible frequency for saving the dataframe (every 20s)
                                elif t2-t1>int(self.fittingpopups_ui[listindex].spinBox_saveMetaEvery.value()):                                    
                                    for i in range(len(HISTORIES)):
                                        Histories = HISTORIES[i]
                                        Saved = SAVED[i]
                                        DF1 = [[ h[h_i][-1] for h_i in h] for h in Histories] #if nb_epoch in .fit() is >1, only save the last history item, beacuse this would a model that could be saved
                                        DF1 = np.r_[DF1]
                                        DF1 = pd.DataFrame( DF1,columns=Histories[0].keys() )
                                        DF1["Saved"] = Saved
                                        DF1.index = Index
                                        HISTORIES[i] = []#reset the Histories list
                                        SAVED[i] = []
                                        #Saving
                                        #TODO: save to temp, if harddisk not available to prevent crash.
                                        if os.path.isfile(model_keras_path[i].split(".model")[0]+'_meta.xlsx'):
                                            os.chmod(model_keras_path[i].split(".model")[0]+'_meta.xlsx', S_IREAD|S_IRGRP|S_IROTH|S_IWRITE|S_IWGRP|S_IWOTH) #make read/write
                                        DF1.to_excel(Writers[i],sheet_name='History', startrow=Writers[i].sheets['History'].max_row,header= False)
                                        Writers[i].save()
                                        os.chmod(model_keras_path[i].split(".model")[0]+'_meta.xlsx', S_IREAD|S_IRGRP|S_IROTH)  #make read only
                                        print("meta.xlsx was saved")
                                        t1 = time.perf_counter()
                                    Index = []#reset the Index list
    
                            counter+=1
                        
            progress_callback.emit(100.0)
            
            #If the original storing locating became inaccessible (folder name changed, HD unplugged...)
            #the models and meta are saved to temp folder. Inform the user!!!
            if saving_failed==True:
                path_orig = str(self.fittingpopups_ui[listindex].lineEdit_modelname_pop.text())
                text = "<html><head/><body><p>Original path:<br>"+path_orig+\
                "<br>became inaccessible during training! Files were then saved to:<br>"+\
                new_modelname.split(".model")[0]+"<br>To bring both parts back together\
                , you have manually open the meta files (excel) and copy;paste each sheet. \
                Sorry for the inconvenience.<br>If that happens often, you may contact \
                the main developer and ask him to improve that.</p></body></html>"
                
                text = "<span style=\' font-weight:600; color: red;\'>" +text+"</span>"#put red text to the infobox
                self.fittingpopups_ui[listindex].textBrowser_FittingInfo.append(text)
                print('\a')#make a noise
                self.fittingpopups_ui[listindex].textBrowser_FittingInfo.setStyleSheet("background-color: yellow;")
                self.fittingpopups_ui[listindex].textBrowser_FittingInfo.moveCursor(QtGui.QTextCursor.End)
                
            if collection==False:
                if len(Histories)>0: #if the list for History files is not empty, process it!
                    DF1 = [[ h[h_i][-1] for h_i in h] for h in Histories] #if nb_epoch in .fit() is >1, only save the last history item, beacuse this would a model that could be saved
                    DF1 = np.r_[DF1]
                    DF1 = pd.DataFrame( DF1,columns=Histories[0].keys() )
                    DF1["Saved"] = Saved
                    DF1["Time"] = Stopwatch
                    DF1["LearningRate"] = LearningRate
                    DF1.index = Index
                    Index = []#reset the Index list
                    Histories = []#reset the Histories list
                    Saved = []
                    #does such a file exist already? append! 
                    if not os.path.isfile(new_modelname.split(".model")[0]+'_meta.xlsx'):
                       DF1.to_excel(writer,sheet_name='History')
                    else: # else it exists so append without writing the header
                       DF1.to_excel(writer,sheet_name='History', startrow=writer.sheets['History'].max_row,header= False)
                if os.path.isfile(new_modelname.split(".model")[0]+'_meta.xlsx'):
                    os.chmod(new_modelname.split(".model")[0]+'_meta.xlsx', S_IREAD|S_IRGRP|S_IROTH|S_IWRITE|S_IWGRP|S_IWOTH) #make read/write
                writer.save()
                writer.close()
    
            if collection==True:
                for i in range(len(HISTORIES)):
                    Histories = HISTORIES[i]
                    Saved = SAVED[i]
                    if len(Histories)>0: #if the list for History files is not empty, process it!
                        DF1 = [[ h[h_i][-1] for h_i in h] for h in Histories] #if nb_epoch in .fit() is >1, only save the last history item, beacuse this would a model that could be saved
                        DF1 = np.r_[DF1]
                        DF1 = pd.DataFrame( DF1,columns=Histories[0].keys() )
                        DF1["Saved"] = Saved
                        DF1.index = Index
                        HISTORIES[i] = []#reset the Histories list
                        SAVED[i] = []
                        #does such a file exist already? append! 
                        if not os.path.isfile(model_keras_path[i].split(".model")[0]+'_meta.xlsx'):
                           DF1.to_excel(Writers[i],sheet_name='History')
                        else: # else it exists so append without writing the header
                           DF1.to_excel(writer,sheet_name='History', startrow=writer.sheets['History'].max_row,header= False)
                    if os.path.isfile(model_keras_path[i].split(".model")[0]+'_meta.xlsx'):
                        os.chmod(model_keras_path[i].split(".model")[0]+'_meta.xlsx', S_IREAD|S_IRGRP|S_IROTH|S_IWRITE|S_IWGRP|S_IWOTH) #make read/write
                    Writers[i].save()
                    Writers[i].close()
                    
                Index = []#reset the Index list
                
                
                
            sess.close()
    #        try:
    #            aid_dl.reset_keras(model_keras)
    #        except:
    #            pass



    def action_fit_model(self):
        #Take the initialized model
        #Unfortunately, in TensorFlow it is not possile to pass a model from
        #one thread to another. Therefore I have to load and save the models each time :(
        model_keras = self.model_keras
        if type(model_keras)==tuple:
            collection=True
        else:
            collection=False
            
        #Check if there was a model initialized:
        new_modelname = str(self.lineEdit_modelname.text())
        
        if len(new_modelname)==0:
           msg = QtWidgets.QMessageBox()
           msg.setIcon(QtWidgets.QMessageBox.Information)       
           msg.setText("Please define a path/filename for the model to be fitted!")
           msg.setWindowTitle("Model path/ filename missing!")
           msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
           msg.exec_()
           return

        if model_keras==None:#in case the model got deleted in another task
            self.action_initialize_model(duties="initialize_train")

            print("Had to re-run action_initialize_model!")
            model_keras = self.model_keras
            self.model_keras = None#delete this copy
            
            if model_keras==None:
#                msg = QtWidgets.QMessageBox()
#                msg.setIcon(QtWidgets.QMessageBox.Information)       
#                msg.setText("Model could not be initialized")
#                msg.setWindowTitle("Error")
#                msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
#                msg.exec_()
                return
            if not model_keras==None:
                msg = QtWidgets.QMessageBox()
                msg.setIcon(QtWidgets.QMessageBox.Information)       
                msg.setText("Model is now initialized for you, Please check Model summary window below if everything is correct and then press Fit again!")
                msg.setWindowTitle("No initilized model found!")
                msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                msg.exec_()
                return

        #There should be at least two outputs (index 0 and 1)
        if collection==False:
            #model_config = model_keras.get_config()#["layers"] 
            nr_classes = int(model_keras.output.shape.dims[1])

        if collection==True:
            #model_config = model_keras[1][0].get_config()#["layers"] 
            nr_classes = int(model_keras[1][0].output.shape.dims[1])

        if nr_classes<2:
           msg = QtWidgets.QMessageBox()
           msg.setIcon(QtWidgets.QMessageBox.Information)       
           msg.setText("Please define at least two classes")
           msg.setWindowTitle("Not enough classes")
           msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
           msg.exec_()
           return
        
        if collection==False:
            #define a variable on self which allows the fit_model_worker to load this model and fit
            #(sorry, this is necessary since TensorFlow does not support passing models between threads)
            self.model_keras_path = new_modelname.split(".model")[0]+"_0.model"
            #save a first version of the .model
            model_keras.save(self.model_keras_path,save_format='h5')
            #Delete the variable to save RAM
            model_keras = None #Since this uses TensorFlow, I have to reload the model action_fit_model_worker anyway

        if collection==True:
            #define a variable on self which allows the fit_model_worker to load this model and fit
            #(sorry, this is necessary since TensorFlow does not support passing models between threads)
            self.model_keras_path = [new_modelname.split(".model")[0]+"_"+model_keras[0][i]+".model" for i in range(len(model_keras[0]))]
            for i in range(len(self.model_keras_path)):
                #save a first version of the .model
                model_keras[1][i].save(self.model_keras_path[i])

            #Delete the variable to save RAM
            model_keras = None #Since this uses TensorFlow, I have to reload the model action_fit_model_worker anyway
        #Check that Data is on RAM
        DATA_len = len(self.ram) #this returns the len of a dictionary. The dictionary is supposed to contain the training/validation data; otherwise the data is read from .rtdc data directly (SLOW unless you have ultra-good SSD)

        def popup_data_to_ram(button):
            yes_or_no = button.text()
            if yes_or_no == "&Yes":
                print("Moving data to ram")
                self.actionDataToRamNow_function()
            elif yes_or_no == "&No":
                pass
            
        if DATA_len==0:
           msg = QtWidgets.QMessageBox()
           msg.setIcon(QtWidgets.QMessageBox.Information)       
           msg.setText("Would you like transfer the Data to RAM now?\n(Currently the data is not in RAM and would be read from .rtdc, which slows down fitting dramatically unless you have a super-fast SSD.)")
           msg.setWindowTitle("Data to RAM now?")
           msg.setStandardButtons(QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)
           msg.buttonClicked.connect(popup_data_to_ram)
           msg.exec_()
            
        ###################Popup Window####################################
        self.fittingpopups.append(MyPopup())
        ui = aid_frontend.Fitting_Ui()
        ui.setupUi(self.fittingpopups[-1]) #append the ui to the last element on the list
        self.fittingpopups_ui.append(ui)
        # Increase the popupcounter by one; this will help to coordinate the data flow between main ui and popup
        self.popupcounter += 1
        listindex=self.popupcounter-1
        
        ##############################Define functions#########################
        self.fittingpopups_ui[listindex].pushButton_UpdatePlot_pop.clicked.connect(lambda: self.update_historyplot_pop(listindex))
        self.fittingpopups_ui[listindex].pushButton_Stop_pop.clicked.connect(lambda: self.stop_fitting_pop(listindex))
        self.fittingpopups_ui[listindex].pushButton_Pause_pop.clicked.connect(lambda: self.pause_fitting_pop(listindex))
        self.fittingpopups_ui[listindex].pushButton_saveTextWindow_pop.clicked.connect(lambda: self.saveTextWindow_pop(listindex))
        self.fittingpopups_ui[listindex].pushButton_clearTextWindow_pop.clicked.connect(lambda: self.clearTextWindow_pop(listindex))
        self.fittingpopups_ui[listindex].pushButton_showModelSumm_pop.clicked.connect(lambda: self.showModelSumm_pop(listindex))
        self.fittingpopups_ui[listindex].pushButton_saveModelSumm_pop.clicked.connect(lambda: self.saveModelSumm_pop(listindex))
        #Expert mode functions
        #self.fittingpopups_ui[listindex].checkBox_pTr_pop.toggled.connect(lambda on_or_off: self.partialtrainability_activated_pop(on_or_off,listindex))
        self.fittingpopups_ui[listindex].pushButton_lossW.clicked.connect(lambda: self.lossWeights_popup(listindex))
        self.fittingpopups_ui[listindex].checkBox_lossW.clicked.connect(lambda on_or_off: self.lossWeights_activated(on_or_off,listindex))

        self.fittingpopups_ui[listindex].Form.setWindowTitle(os.path.split(new_modelname)[1])
        self.fittingpopups_ui[listindex].progressBar_Fitting_pop.setValue(0) #set the progress bar to zero
        self.fittingpopups_ui[listindex].pushButton_ShowExamleImgs_pop.clicked.connect(lambda: self.action_show_example_imgs_pop(listindex))
        self.fittingpopups_ui[listindex].tableWidget_HistoryInfo_pop.doubleClicked.connect(lambda item: self.tableWidget_HistoryInfo_pop_dclick(item,listindex))
        #Cyclical learning rate extra settings
        self.fittingpopups_ui[listindex].pushButton_cycLrPopup.clicked.connect(lambda: self.popup_clr_settings(listindex))
        self.fittingpopups_ui[listindex].comboBox_optimizer.currentTextChanged.connect(lambda: self.expert_optimizer_changed(optimizer_text=self.fittingpopups_ui[listindex].comboBox_optimizer.currentText(),listindex=listindex))
        self.fittingpopups_ui[listindex].pushButton_LR_plot.clicked.connect(lambda: self.popup_lr_plot(listindex))

        self.fittingpopups_ui[listindex].doubleSpinBox_learningRate.valueChanged.connect(lambda: self.expert_lr_changed(value=self.fittingpopups_ui[listindex].doubleSpinBox_learningRate.value(),optimizer_text=self.fittingpopups_ui[listindex].comboBox_optimizer.currentText(),listindex=listindex))
        self.fittingpopups_ui[listindex].doubleSpinBox_expDecInitLr.valueChanged.connect(lambda: self.expert_lr_changed(value=self.fittingpopups_ui[listindex].doubleSpinBox_learningRate.value(),optimizer_text=self.fittingpopups_ui[listindex].comboBox_optimizer.currentText(),listindex=listindex))

        self.fittingpopups_ui[listindex].pushButton_optimizer_pop.clicked.connect(lambda: self.optimizer_change_settings_popup(listindex))



        worker = Worker(self.action_fit_model_worker)
        #Get a signal from the worker to update the progressbar
        worker.signals.progress.connect(self.fittingpopups_ui[listindex].progressBar_Fitting_pop.setValue)
        
        #Define a func which prints information during fitting to textbrowser
        #And furthermore provide option to do real-time plotting
        def real_time_info(dic):
            self.fittingpopups_ui[listindex].Histories.append(dic) #append to a list. Will be used for plotting in the "Update plot" function
            OtherMetrics_keys = self.fittingpopups_ui[listindex].RealTime_OtherMetrics.keys()
            #Append to lists for real-time plotting
            self.fittingpopups_ui[listindex].RealTime_Acc.append(dic["accuracy"][0])
            self.fittingpopups_ui[listindex].RealTime_ValAcc.append(dic["val_accuracy"][0])
            self.fittingpopups_ui[listindex].RealTime_Loss.append(dic["loss"][0])
            self.fittingpopups_ui[listindex].RealTime_ValLoss.append(dic["val_loss"][0])

            keys = list(dic.keys())            
            #sort keys alphabetically
            keys_ = [l.lower() for l in keys]
            ind_sort = np.argsort(keys_)
            keys = list(np.array(keys)[ind_sort])
            #First keys should always be acc,loss,val_acc,val_loss -in this order
            keys_first = ["accuracy","loss","val_accuracy","val_loss"]
            for i in range(len(keys_first)):
                if keys_first[i] in keys:
                    ind = np.where(np.array(keys)==keys_first[i])[0][0]
                    if ind!=i:
                        del keys[ind]
                        keys.insert(i,keys_first[i])
    
            for key in keys:
                if "precision" in key or "auc" in key or "recall" in key or "LearningRate" in key:
                    if not key in OtherMetrics_keys: #if this key is missing in self.fittingpopups_ui[listindex].RealTime_OtherMetrics attach it!
                        self.fittingpopups_ui[listindex].RealTime_OtherMetrics[key] = []
                    self.fittingpopups_ui[listindex].RealTime_OtherMetrics[key].append(dic[key])
            dic_text = [("{} {}".format(item, np.round(amount[0],4))) for item, amount in dic.items()]
            text = "Epoch "+str(self.fittingpopups_ui[listindex].epoch_counter)+"\n"+" ".join(dic_text)
            self.fittingpopups_ui[listindex].textBrowser_FittingInfo.append(text)
            self.fittingpopups_ui[listindex].epoch_counter+=1
            if self.fittingpopups_ui[listindex].epoch_counter==1:

                #for each key, put a checkbox on the tableWidget_HistoryInfo_pop
                rowPosition = self.fittingpopups_ui[listindex].tableWidget_HistoryInfo_pop.rowCount()
                self.fittingpopups_ui[listindex].tableWidget_HistoryInfo_pop.insertRow(rowPosition)
                self.fittingpopups_ui[listindex].tableWidget_HistoryInfo_pop.setColumnCount(len(keys))

                for columnPosition in range(len(keys)):#(2,4):
                    key = keys[columnPosition]
                    #for each item, also create 2 checkboxes (train/valid)
                    item = QtWidgets.QTableWidgetItem(str(key))#("item {0} {1}".format(rowNumber, columnNumber))
                    item.setBackground(QtGui.QColor(self.colorsQt[columnPosition]))
                    item.setFlags( QtCore.Qt.ItemIsUserCheckable | QtCore.Qt.ItemIsEnabled  )
                    item.setCheckState(QtCore.Qt.Unchecked)
                    self.fittingpopups_ui[listindex].tableWidget_HistoryInfo_pop.setItem(rowPosition, columnPosition, item)
            self.fittingpopups_ui[listindex].tableWidget_HistoryInfo_pop.resizeColumnsToContents()
            self.fittingpopups_ui[listindex].tableWidget_HistoryInfo_pop.resizeRowsToContents()


            ########################Real-time plotting#########################
            if self.fittingpopups_ui[listindex].checkBox_realTimePlotting_pop.isChecked():
                #get the range for the real time fitting
                if hasattr(self.fittingpopups_ui[listindex], 'historyscatters'):#if update plot was hit before
                    x = range(len(self.fittingpopups_ui[listindex].Histories))
                    realTimeEpochs = self.fittingpopups_ui[listindex].spinBox_realTimeEpochs.value()
                    if len(x)>realTimeEpochs:
                        x = x[-realTimeEpochs:]
                    #is any metric checked on the table?
                    colcount = int(self.fittingpopups_ui[listindex].tableWidget_HistoryInfo_pop.columnCount())
                    #Collect items that are checked
                    selected_items,Colors = [],[]
                    for colposition in range(colcount):  
                        #is it checked?
                        cb = self.fittingpopups_ui[listindex].tableWidget_HistoryInfo_pop.item(0, colposition)
                        if not cb==None:
                            if cb.checkState() == QtCore.Qt.Checked:
                                selected_items.append(str(cb.text()))
                                Colors.append(cb.background())
                    
                    for i in range(len(self.fittingpopups_ui[listindex].historyscatters)): #iterate over all available plots
                        key = list(self.fittingpopups_ui[listindex].historyscatters.keys())[i]
                        if key in selected_items:
                            if key=="accuracy":
                                y = np.array(self.fittingpopups_ui[listindex].RealTime_Acc).astype(float)
                            elif key=="val_accuracy":
                                y = np.array(self.fittingpopups_ui[listindex].RealTime_ValAcc).astype(float)
                            elif key=="loss":
                                y = np.array(self.fittingpopups_ui[listindex].RealTime_Loss).astype(float)
                            elif key=="val_loss":
                                y = np.array(self.fittingpopups_ui[listindex].RealTime_ValLoss).astype(float)
                            elif "precision" in key or "auc" in key or "recall" in key or "LearningRate" in key:
                               y = np.array(self.fittingpopups_ui[listindex].RealTime_OtherMetrics[key]).astype(float).reshape(-1,)
                            else:
                                return
                            #Only show the last 250 epochs
                            if y.shape[0]>realTimeEpochs:
                                y = y[-realTimeEpochs:]
                            if y.shape[0]==len(x):
                                self.fittingpopups_ui[listindex].historyscatters[key].setData(x, y)#,pen=None,symbol='o',symbolPen=None,symbolBrush=brush,clear=False)
                            else:
                                print("x and y are not the same size! Omitted plotting. I will try again to plot after the next epoch.")

                        pg.QtGui.QApplication.processEvents()

        self.fittingpopups_ui[listindex].epoch_counter = 0
        #self.fittingpopups_ui[listindex].backup = [] #backup of the meta information -> in case the original folder is not accessible anymore
        worker.signals.history.connect(real_time_info)
        
        #Finally start the worker!
        self.threadpool.start(worker)
        self.fittingpopups[listindex].show()















    def action_lr_finder(self):
        #lr_find
        model_keras = self.model_keras
        if type(model_keras)==tuple:
            collection=True
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Warning)       
            msg.setText("LR screening is not supported for Collections of models. Please select single model")
            msg.setWindowTitle("LR screening not supported for Collections!")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return
        else:
            collection=False
            
        #Check if there was a model initialized:
        new_modelname = str(self.lineEdit_modelname.text())
        
        if len(new_modelname)==0:
           msg = QtWidgets.QMessageBox()
           msg.setIcon(QtWidgets.QMessageBox.Information)       
           msg.setText("Please define a path/filename for the model to be fitted!")
           msg.setWindowTitle("Model path/ filename missing!")
           msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
           msg.exec_()
           return

        if model_keras==None:#in case the model got deleted in another task
            self.action_initialize_model(duties="initialize_train")
            print("Had to re-run action_initialize_model!")
            model_keras = self.model_keras
            self.model_keras = None#delete this copy
            
            if model_keras==None:
                return
            if not model_keras==None:
                msg = QtWidgets.QMessageBox()
                msg.setIcon(QtWidgets.QMessageBox.Information)       
                msg.setText("Model is now initialized for you, Please check Model summary window below if everything is correct and then press Fit again!")
                msg.setWindowTitle("No initilized model found!")
                msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                msg.exec_()
                return

        nr_classes = int(model_keras.output.shape.dims[1])

        if nr_classes<2:
           msg = QtWidgets.QMessageBox()
           msg.setIcon(QtWidgets.QMessageBox.Information)       
           msg.setText("Please define at least two classes")
           msg.setWindowTitle("Not enough classes")
           msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
           msg.exec_()
           return
        
        #define a variable on self which allows the fit_model_worker to load this model and fit
        #(sorry, this is necessary since TensorFlow does not support passing models between threads)
        self.model_keras_path = new_modelname.split(".model")[0]+"_0.model"
        #save a first version of the .model
        model_keras.save(self.model_keras_path,save_format='h5')
        #Delete the variable to save RAM
        model_keras = None #Since this uses TensorFlow, I have to reload the model action_fit_model_worker anyway

        #Check that Data is on RAM
        DATA_len = len(self.ram) #this returns the len of a dictionary. The dictionary is supposed to contain the training/validation data; otherwise the data is read from .rtdc data directly (SLOW unless you have ultra-good SSD)

        def popup_data_to_ram(button):
            yes_or_no = button.text()
            if yes_or_no == "&Yes":
                print("Moving data to ram")
                self.actionDataToRamNow_function()
            elif yes_or_no == "&No":
                pass
            
        if DATA_len==0:
           msg = QtWidgets.QMessageBox()
           msg.setIcon(QtWidgets.QMessageBox.Information)       
           msg.setText("Would you like transfer the Data to RAM now?\n(Currently the data is not in RAM and would be read from .rtdc, which slows down fitting dramatically unless you have a super-fast SSD.)")
           msg.setWindowTitle("Data to RAM now?")
           msg.setStandardButtons(QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)
           msg.buttonClicked.connect(popup_data_to_ram)
           msg.exec_()

        worker = Worker(self.action_lr_finder_worker)
        #Get a signal from the worker to update the progressbar
        worker.signals.progress.connect(print)
        worker.signals.history.connect(print)
        #Finally start the worker!
        self.threadpool.start(worker)


    def action_lr_finder_worker(self,progress_callback,history_callback):        
        if self.radioButton_cpu.isChecked():
            gpu_used = False
            deviceSelected = str(self.comboBox_cpu.currentText())
        elif self.radioButton_gpu.isChecked():
            gpu_used = True
            deviceSelected = str(self.comboBox_gpu.currentText())
        gpu_memory = float(self.doubleSpinBox_memory.value())

        #Retrieve more Multi-GPU Options from Menubar:
        cpu_merge = bool(self.actioncpu_merge.isEnabled())
        cpu_relocation = bool(self.actioncpu_relocation.isEnabled())
        cpu_weight_merge = bool(self.actioncpu_weightmerge.isEnabled())    

        #Create config (define which device to use)
        config_gpu = aid_dl.get_config(cpu_nr,gpu_nr,deviceSelected,gpu_memory)
        
        with tf.compat.v1.Session(graph = tf.Graph(), config=config_gpu) as sess:                   
            #get an index of the fitting popup
            #listindex = self.popupcounter-1
            #Get user-specified filename for the new model
            model_keras_path = self.model_keras_path
            
            if type(model_keras_path)==list:
                collection = True
                msg = QtWidgets.QMessageBox()
                msg.setIcon(QtWidgets.QMessageBox.Warning)       
                msg.setText("LR screening is currently not supported for Collections of models. Please use single model")
                msg.setWindowTitle("LR screening not supported for Collections")
                msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                msg.exec_()                        
                return
    
            else:
                collection = False
                #Baustelle
                if deviceSelected=="Multi-GPU" and cpu_weight_merge==True:
                    with tf.device("/cpu:0"):
                        strategy = tf.distribute.MirroredStrategy()
                        with strategy.scope(): 
                            model_keras = load_model(model_keras_path,custom_objects=aid_dl.get_custom_metrics()) 
                else:
                    model_keras = load_model(model_keras_path,custom_objects=aid_dl.get_custom_metrics())
                
            #Initialize a variable for the parallel model
            model_keras_p = None
    
            #Multi-GPU
            if deviceSelected=="Multi-GPU":
                if collection==False:
                    print("Adjusting the model for Multi-GPU")
                    model_keras_p = model_keras#multi_gpu_model(model_keras, gpus=gpu_nr, cpu_merge=cpu_merge, cpu_relocation=cpu_relocation)#indicate the numbers of gpus that you have
                    if self.radioButton_LoadContinueModel.isChecked():#calling multi_gpu_model resets the weights. Hence, they need to be put in place again
                        model_keras_p.layers[-2].set_weights(model_keras.get_weights())
                elif collection==True:
                    print("Collection & Multi-GPU is not supported yet")
                    return
    
            ##############Main function after hitting FIT MODEL####################
            if self.radioButton_LoadRestartModel.isChecked():
                load_modelname = str(self.lineEdit_LoadModelPath.text())
            elif self.radioButton_LoadContinueModel.isChecked():
                load_modelname = str(self.lineEdit_LoadModelPath.text())
            elif self.radioButton_NewModel.isChecked():
                load_modelname = "" #No model is loaded
    
            if collection==False:    
                #model_config = model_keras.get_config()#["layers"] 
                nr_classes = int(model_keras.output.shape.dims[1])
            if collection==True:
                #model_config = model_keras.get_config()#["layers"] 
                nr_classes = int(model_keras[0].output.shape.dims[1])
            
            #Metrics to be displayed during fitting (real-time)
            model_metrics = self.get_metrics()
    
            #Compile model
            if  deviceSelected=="Single-GPU":
                model_keras.compile(loss='categorical_crossentropy',optimizer='adam',metrics=aid_dl.get_metrics_tensors(model_metrics,nr_classes))#dont specify loss and optimizer yet...expert stuff will follow and model will be recompiled
            elif deviceSelected=="Multi-GPU":
                model_keras_p.compile(loss='categorical_crossentropy',optimizer='adam',metrics=aid_dl.get_metrics_tensors(model_metrics,nr_classes))#dont specify loss and optimizer yet...expert stuff will follow and model will be recompiled

            #Collect all information about the fitting routine that was user
            #defined
            if self.actionVerbose.isChecked()==True:
                verbose = 1
            else:
                verbose = 0
    
            trainable_original, layer_names = self.trainable_original, self.layer_names
                
            crop = int(self.spinBox_imagecrop.value())      
            norm = str(self.comboBox_Normalization.currentText())
    
            nr_epochs = int(self.spinBox_NrEpochs.value())
            h_flip = bool(self.checkBox_HorizFlip.isChecked())
            v_flip = bool(self.checkBox_VertFlip.isChecked())
            rotation = float(self.lineEdit_Rotation.text())
     
            width_shift = float(self.lineEdit_widthShift.text())
            height_shift = float(self.lineEdit_heightShift.text())
            zoom = float(self.lineEdit_zoomRange.text())
            shear = float(self.lineEdit_shearRange.text())
            
            brightness_add_lower = float(self.spinBox_PlusLower.value())
            brightness_add_upper = float(self.spinBox_PlusUpper.value())
            brightness_mult_lower = float(self.doubleSpinBox_MultLower.value())
            brightness_mult_upper = float(self.doubleSpinBox_MultUpper.value())
            gaussnoise_mean = float(self.doubleSpinBox_GaussianNoiseMean.value())
            gaussnoise_scale = float(self.doubleSpinBox_GaussianNoiseScale.value())
    
            contrast_on = bool(self.checkBox_contrast.isChecked())        
            contrast_lower = float(self.doubleSpinBox_contrastLower.value())
            contrast_higher = float(self.doubleSpinBox_contrastHigher.value())
            saturation_on = bool(self.checkBox_saturation.isChecked())        
            saturation_lower = float(self.doubleSpinBox_saturationLower.value())
            saturation_higher = float(self.doubleSpinBox_saturationHigher.value())
            hue_on = bool(self.checkBox_hue.isChecked())        
            hue_delta = float(self.doubleSpinBox_hueDelta.value())
    
            avgBlur_on = bool(self.checkBox_avgBlur.isChecked())        
            avgBlur_min = int(self.spinBox_avgBlurMin.value())
            avgBlur_max = int(self.spinBox_avgBlurMax.value())
            gaussBlur_on = bool(self.checkBox_gaussBlur.isChecked())        
            gaussBlur_min = int(self.spinBox_gaussBlurMin.value())
            gaussBlur_max = int(self.spinBox_gaussBlurMax.value())
            motionBlur_on = bool(self.checkBox_motionBlur.isChecked())        
            motionBlur_kernel = str(self.lineEdit_motionBlurKernel.text())
            motionBlur_angle = str(self.lineEdit_motionBlurAngle.text())
            motionBlur_kernel = tuple(ast.literal_eval(motionBlur_kernel)) #translate string in the lineEdits to a tuple
            motionBlur_angle = tuple(ast.literal_eval(motionBlur_angle)) #translate string in the lineEdits to a tuple
    
            if collection==False:
                expert_mode = bool(self.groupBox_expertMode.isChecked())
            elif collection==True:
                expert_mode = self.groupBox_expertMode.setChecked(False)
                print("Expert mode was switched off. Not implemented yet for collections")
                expert_mode = False
    
            learning_rate_const = float(self.doubleSpinBox_learningRate.value())
            loss_expert = str(self.comboBox_expt_loss.currentText()).lower()
            optimizer_expert = str(self.comboBox_optimizer.currentText()).lower()
            optimizer_settings = self.optimizer_settings.copy()
            paddingMode = str(self.comboBox_paddingMode.currentText())#.lower()
    
            train_last_layers = bool(self.checkBox_trainLastNOnly.isChecked())             
            train_last_layers_n = int(self.spinBox_trainLastNOnly.value())              
            train_dense_layers = bool(self.checkBox_trainDenseOnly.isChecked())             
            dropout_expert_on = bool(self.checkBox_dropout.isChecked())             
            try:
                dropout_expert = str(self.lineEdit_dropout.text()) #due to the validator, there are no squ.brackets
                dropout_expert = "["+dropout_expert+"]"
                dropout_expert = ast.literal_eval(dropout_expert)        
            except:
                dropout_expert = []
            lossW_expert = str(self.lineEdit_lossW.text())
    
            #To get the class weights (loss), the SelectedFiles are required 
            SelectedFiles = self.items_clicked_no_rtdc_ds()
            #Check if xtra_data should be used for training
            xtra_in = [s["xtra_in"] for s in SelectedFiles]
            if len(set(xtra_in))==1:
                xtra_in = list(set(xtra_in))[0]
            elif len(set(xtra_in))>1:# False and True is present. Not supported
                print("Xtra data is used only for some files. Xtra data needs to be used either by all or by none!")
                return

            #Get the class weights. This function runs now the first time in the fitting routine. 
            #It is possible that the user chose Custom weights and then changed the classes. Hence first check if 
            #there is a weight for each class available.
            class_weight = self.get_class_weight(SelectedFiles,lossW_expert,custom_check_classes=True)
            if type(class_weight)==list:
                #There has been a mismatch between the classes described in class_weight and the classes available in SelectedFiles!
                lossW_expert = class_weight[0] #overwrite 
                class_weight = class_weight[1]
                print(class_weight)
                print("There has been a mismatch between the classes described in \
                      Loss weights and the classes available in the selected files! \
                      Hence, the Loss weights are set to Balanced")
                
                
            ###############################Expert Mode values##################
            if expert_mode==True:
                #Some settings only need to be changed once, after user clicked apply at next epoch

                #Apply the changes to trainable states:
                if train_last_layers==True:#Train only the last n layers
                    print("Train only the last "+str(train_last_layers_n)+ " layer(s)")
                    trainable_new = (len(trainable_original)-train_last_layers_n)*[False]+train_last_layers_n*[True]
                    summary = aid_dl.model_change_trainability(model_keras,trainable_new,model_metrics,nr_classes,loss_expert,optimizer_settings,learning_rate_const)
                    if model_keras_p!=None:#if this is NOT None, there exists a parallel model, which also needs to be re-compiled
                        model_metrics_t = aid_dl.get_metrics_tensors(self.get_metrics(),nr_classes)
                        aid_dl.model_compile(model_keras_p,loss_expert,optimizer_settings,learning_rate_const,model_metrics_t,nr_classes)
                        print("Recompiled parallel model for train_last_layers==True")
                    text1 = "Expert mode: Request for custom trainability states: train only the last "+str(train_last_layers_n)+ " layer(s)\n"
                    #text2 = "\n--------------------\n"
                    print(text1+summary)
                if train_dense_layers==True:#Train only dense layers
                    print("Train only dense layers")
                    layer_dense_ind = ["Dense" in x for x in layer_names]
                    layer_dense_ind = np.where(np.array(layer_dense_ind)==True)[0] #at which indices are dropout layers?
                    #create a list of trainable states
                    trainable_new = len(trainable_original)*[False]
                    for index in layer_dense_ind:
                        trainable_new[index] = True
                    summary = aid_dl.model_change_trainability(model_keras,trainable_new,model_metrics,nr_classes,loss_expert,optimizer_settings,learning_rate_const)                  
                    if model_keras_p!=None:#if this is NOT None, there exists a parallel model, which also needs to be re-compiled
                        model_metrics_t = aid_dl.get_metrics_tensors(self.get_metrics(),nr_classes)
                        aid_dl.model_compile(model_keras_p,loss_expert,optimizer_settings,learning_rate_const,model_metrics_t,nr_classes)
                        print("Recompiled parallel model for train_dense_layers==True")
                    text1 = "Expert mode: Request for custom trainability states: train only dense layer(s)\n"
                    #text2 = "\n--------------------\n"
                    print(text1+summary)
    
                if dropout_expert_on==True:
                    #The user apparently want to change the dropout rates
                    do_list = aid_dl.get_dropout(model_keras)#Get a list of dropout values of the current model
                    #Compare the dropout values in the model to the dropout values requested by user
                    if len(dropout_expert)==1:#if the user gave a single float
                        dropout_expert_list = len(do_list)*dropout_expert #convert to list
                    elif len(dropout_expert)>1:
                        dropout_expert_list = dropout_expert
                        if not len(dropout_expert_list)==len(do_list):
                            text = "Issue with dropout: you defined "+str(len(dropout_expert_list))+" dropout rates, but model has "+str(len(do_list))+" dropout layers"
                            print(text)
                    else:
                        text = "Could not understand user input at Expert->Dropout"
                        print(text)
                        dropout_expert_list = []
                    if len(dropout_expert_list)>0 and do_list!=dropout_expert_list:#if the dropout rates of the current model is not equal to the required do_list from user...
                        do_changed = aid_dl.change_dropout(model_keras,dropout_expert_list,model_metrics_t,nr_classes,loss_expert,optimizer_settings,learning_rate_const)
                        if model_keras_p!=None:#if this is NOT None, there exists a parallel model, which also needs to be re-compiled
                            model_metrics_t = aid_dl.get_metrics_tensors(self.get_metrics(),nr_classes)
                            aid_dl.model_compile(model_keras_p,loss_expert,optimizer_settings,learning_rate_const,model_metrics_t,nr_classes)
                            print("Recompiled parallel model to change dropout. I'm not sure if this works already!")
                        if do_changed==1:
                            text_do = "Dropout rate(s) in model was/were changed to: "+str(dropout_expert_list)
                        else:
                            text_do = "Dropout rate(s) in model was/were not changed"
                    else:
                        text_do = "Dropout rate(s) in model was/were not changed"
                    print(text_do)    
    
            text_updates = ""
            #Compare current lr and the lr on expert tab:
            if collection == False:
                lr_current = model_keras.optimizer.get_config()["learning_rate"]
            else:
                lr_current = model_keras[0].optimizer.get_config()["learning_rate"]
    
            lr_diff = learning_rate_const-lr_current
            if  abs(lr_diff) > 1e-6:
                if collection == False:
                    K.set_value(model_keras.optimizer.lr, learning_rate_const)
                if collection == True:
                    for m in model_keras:
                        K.set_value(m.optimizer.lr, learning_rate_const)
                text_updates +=  "Changed the learning rate to "+ str(learning_rate_const)+"\n"
            recompile = False
            #Compare current optimizer and the optimizer on expert tab:
            if collection==False:
                optimizer_current = aid_dl.get_optimizer_name(model_keras).lower()#get the current optimizer of the model
            if collection==True:
                optimizer_current = aid_dl.get_optimizer_name(model_keras[0]).lower()#get the current optimizer of the model
    
            if optimizer_current!=optimizer_expert.lower():#if the current model has a different optimizer
                recompile = True
                text_updates+="Changed the optimizer to "+optimizer_expert+"\n"
    
            #Compare current loss function and the loss-function on expert tab:
            if collection==False:
                if model_keras.loss!=loss_expert:
                    recompile = True
                    text_updates+="Changed the loss function to "+loss_expert+"\n"
            if collection==True:
                if model_keras[0].loss!=loss_expert:
                    recompile = True
                    text_updates+="Changed the loss function to "+loss_expert+"\n"
    
    
            if recompile==True:
                print("Recompiling...")
                if collection==False:
                    model_metrics_t = aid_dl.get_metrics_tensors(self.get_metrics(),nr_classes)
                    aid_dl.model_compile(model_keras,loss_expert,optimizer_settings,learning_rate_const,model_metrics_t,nr_classes)
                if collection==True:
                    for m in model_keras[1]:
                        model_metrics_t = aid_dl.get_metrics_tensors(self.get_metrics(),nr_classes)
                        aid_dl.model_compile(m, loss_expert, optimizer_settings, learning_rate_const,model_metrics_t, nr_classes)
                if model_keras_p!=None:#if this is NOT None, there exists a parallel model, which also needs to be re-compiled
                    model_metrics_t = aid_dl.get_metrics_tensors(self.get_metrics(),nr_classes)
                    aid_dl.model_compile(model_keras_p,loss_expert,optimizer_settings,learning_rate_const,model_metrics_t,nr_classes)
                    print("Recompiled parallel model to adjust learning rate, loss, optimizer")
    
            print(text_updates)
        
            ######################Load the Training Data################################
            ind = [selectedfile["TrainOrValid"] == "Train" for selectedfile in SelectedFiles]
            ind = np.where(np.array(ind)==True)[0]
            SelectedFiles_train = np.array(SelectedFiles)[ind]
            SelectedFiles_train = list(SelectedFiles_train)
            indices_train = [selectedfile["class"] for selectedfile in SelectedFiles_train]
            nr_events_epoch_train = [selectedfile["nr_events_epoch"] for selectedfile in SelectedFiles_train]
            rtdc_path_train = [selectedfile["rtdc_path"] for selectedfile in SelectedFiles_train]
            zoom_factors_train = [selectedfile["zoom_factor"] for selectedfile in SelectedFiles_train]
            #zoom_order = [self.actionOrder0.isChecked(),self.actionOrder1.isChecked(),self.actionOrder2.isChecked(),self.actionOrder3.isChecked(),self.actionOrder4.isChecked(),self.actionOrder5.isChecked()]
            #zoom_order = int(np.where(np.array(zoom_order)==True)[0])
            zoom_order = int(self.comboBox_zoomOrder.currentIndex()) #the combobox-index is already the zoom order
            shuffle_train = [selectedfile["shuffle"] for selectedfile in SelectedFiles_train]
            xtra_in = set([selectedfile["xtra_in"] for selectedfile in SelectedFiles_train])   
            if len(xtra_in)>1:# False and True is present. Not supported
                print("Xtra data is used only for some files. Xtra data needs to be used either by all or by none!")
                return
            xtra_in = list(xtra_in)[0]#this is either True or False
            
            if verbose==1:
                print("Length of DATA (in RAM) = "+str(len(self.ram)))

            #If the scaling method is "divide by mean and std of the whole training set":
            if norm == "StdScaling using mean and std of all training data":
                mean_trainingdata,std_trainingdata = [],[]
                for i in range(len(SelectedFiles_train)):
                    if len(self.ram)==0: #Here, the entire training set needs to be used! Not only random images!
                        #Replace=true: means individual cells could occur several times
                        gen_train = aid_img.gen_crop_img(crop,rtdc_path_train[i],random_images=False,zoom_factor=zoom_factors_train[i],zoom_order=zoom_order,color_mode=self.get_color_mode(),padding_mode=paddingMode) 
                    else:
                        gen_train = aid_img.gen_crop_img_ram(self.ram,rtdc_path_train[i],random_images=False) #Replace true means that individual cells could occur several times
                        if self.actionVerbose.isChecked():
                            print("Loaded data from RAM")
                        
                    images = next(gen_train)[0]
                    mean_trainingdata.append(np.mean(images))
                    std_trainingdata.append(np.std(images))
                mean_trainingdata = np.mean(np.array(mean_trainingdata))
                std_trainingdata = np.mean(np.array(std_trainingdata))
                
                if np.allclose(std_trainingdata,0):
                    std_trainingdata = 0.0001
    
                    msg = QtWidgets.QMessageBox()
                    msg.setIcon(QtWidgets.QMessageBox.Information)       
                    text = "<html><head/><body><p>The standard deviation of your training data is zero! This would lead to division by zero. To avoid this, I will divide by 0.0001 instead.</p></body></html>"
                    msg.setText(text) 
                    msg.setWindowTitle("Std. is zero")
                    msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                    msg.exec_()
    
    
            ######################Load the Validation Data################################
            ind = [selectedfile["TrainOrValid"] == "Valid" for selectedfile in SelectedFiles]
            ind = np.where(np.array(ind)==True)[0]
            SelectedFiles_valid = np.array(SelectedFiles)[ind]
            SelectedFiles_valid = list(SelectedFiles_valid)
            indices_valid = [selectedfile["class"] for selectedfile in SelectedFiles_valid]
            nr_events_epoch_valid = [selectedfile["nr_events_epoch"] for selectedfile in SelectedFiles_valid]
            rtdc_path_valid = [selectedfile["rtdc_path"] for selectedfile in SelectedFiles_valid]
            zoom_factors_valid = [selectedfile["zoom_factor"] for selectedfile in SelectedFiles_valid]
            #zoom_order = [self.actionOrder0.isChecked(),self.actionOrder1.isChecked(),self.actionOrder2.isChecked(),self.actionOrder3.isChecked(),self.actionOrder4.isChecked(),self.actionOrder5.isChecked()]
            #zoom_order = int(np.where(np.array(zoom_order)==True)[0])
            zoom_order = int(self.comboBox_zoomOrder.currentIndex()) #the combobox-index is already the zoom order            
            shuffle_valid = [selectedfile["shuffle"] for selectedfile in SelectedFiles_valid]
            xtra_in = set([selectedfile["xtra_in"] for selectedfile in SelectedFiles_valid])   
            if len(xtra_in)>1:# False and True is present. Not supported
                print("Xtra data is used only for some files. Xtra data needs to be used either by all or by none!")
                return
            xtra_in = list(xtra_in)[0]#this is either True or False

            ############Cropping#####################
            percDataV = float(self.popup_lrfinder_ui.doubleSpinBox_percDataV.value())
            percDataV = percDataV/100.0
            
            X_valid,y_valid,Indices,xtra_valid = [],[],[],[]
            for i in range(len(SelectedFiles_valid)):
                if len(self.ram)==0:#if there is no data available on ram
                    #replace=true means individual cells could occur several times
                    gen_valid = aid_img.gen_crop_img(crop,rtdc_path_valid[i],int(np.rint(percDataV*nr_events_epoch_valid[i])),random_images=shuffle_valid[i],replace=True,zoom_factor=zoom_factors_valid[i],zoom_order=zoom_order,color_mode=self.get_color_mode(),padding_mode=paddingMode,xtra_in=xtra_in)
                else:#get a similar generator, using the ram-data
                    gen_valid = aid_img.gen_crop_img_ram(self.ram,rtdc_path_valid[i],int(np.rint(percDataV*nr_events_epoch_valid[i])),random_images=shuffle_valid[i],replace=True,xtra_in=xtra_in) #Replace true means that individual cells could occur several times
                    if self.actionVerbose.isChecked():
                        print("Loaded data from RAM")
                generator_cropped_out = next(gen_valid)
                X_valid.append(generator_cropped_out[0])
                #y_valid.append(np.repeat(indices_valid[i],nr_events_epoch_valid[i]))
                y_valid.append(np.repeat(indices_valid[i],X_valid[-1].shape[0]))
                Indices.append(generator_cropped_out[1])
                xtra_valid.append(generator_cropped_out[2])
                del generator_cropped_out
                
            X_valid = np.concatenate(X_valid)
            y_valid = np.concatenate(y_valid)
            Y_valid = to_categorical(y_valid, nr_classes)# * 2 - 1
            xtra_valid = np.concatenate(xtra_valid)
    
            if len(X_valid.shape)==4:
                channels=3
            elif len(X_valid.shape)==3:
                channels=1
            else:
                print("Invalid data dimension:" +str(X_valid.shape))
            if channels==1:
                #Add the "channels" dimension
                X_valid = np.expand_dims(X_valid,3)
    
            if norm == "StdScaling using mean and std of all training data":
                X_valid = aid_img.image_normalization(X_valid,norm,mean_trainingdata,std_trainingdata)
            else:
                X_valid = aid_img.image_normalization(X_valid,norm)
    
            #Validation data can be cropped to final size already since no augmentation
            #will happen on this data set
            dim_val = X_valid.shape
            print("Current dim. of validation set (pixels x pixels) = "+str(dim_val[2]))
            if dim_val[2]!=crop:
                print("Change dim. (pixels x pixels) of validation set to = "+str(crop))
                remove = int(dim_val[2]/2.0 - crop/2.0)
                X_valid = X_valid[:,remove:remove+crop,remove:remove+crop,:] #crop to crop x crop pixels #TensorFlow
            
            if xtra_in==True:
                print("Add Xtra Data to X_valid")
                X_valid = [X_valid,xtra_valid]
    
    
    
            ###################Load training data####################
            #####################and perform#########################
            ##################Image augmentation#####################

            #Rotating could create edge effects. Avoid this by making crop a bit larger for now
            #Worst case would be a 45degree rotation:
            cropsize2 = np.sqrt(crop**2+crop**2)
            cropsize2 = np.ceil(cropsize2 / 2.) * 2 #round to the next even number

            #Should only a certain percentage of the numbers given in the table be sampled? 
            percDataT = float(self.popup_lrfinder_ui.doubleSpinBox_percDataT.value())
            percDataT = percDataT/100.0

            X_train,y_train,xtra_train = [],[],[]
            t3 = time.perf_counter()
            for i in range(len(SelectedFiles_train)):
                if len(self.ram)==0:
                    #Replace true means that individual cells could occur several times
                    gen_train = aid_img.gen_crop_img(cropsize2,rtdc_path_train[i],int(np.rint(percDataT*nr_events_epoch_train[i])),random_images=shuffle_train[i],replace=True,zoom_factor=zoom_factors_train[i],zoom_order=zoom_order,color_mode=self.get_color_mode(),padding_mode=paddingMode,xtra_in=xtra_in) 
                else:
                    gen_train = aid_img.gen_crop_img_ram(self.ram,rtdc_path_train[i],int(np.rint(percDataT*nr_events_epoch_train[i])),random_images=shuffle_train[i],replace=True,xtra_in=xtra_in) #Replace true means that individual cells could occur several times
                    if self.actionVerbose.isChecked():
                        print("Loaded data from RAM")
                data_ = next(gen_train)
                X_train.append(data_[0])
                y_train.append(np.repeat(indices_train[i],X_train[-1].shape[0]))
                if xtra_in==True:
                    xtra_train.append(data_[2])
                del data_
                
            X_train = np.concatenate(X_train)
            X_train = X_train.astype(np.uint8)
            y_train = np.concatenate(y_train)
            if xtra_in==True:
                print("Retrieve Xtra Data...")
                xtra_train = np.concatenate(xtra_train)
            
            t4 = time.perf_counter()
            if verbose == 1:
                print("Time to load data (from .rtdc or RAM) and crop="+str(t4-t3))
            
            if len(X_train.shape)==4:
                channels=3
            elif len(X_train.shape)==3:
                channels=1
            else:
                print("Invalid data dimension:" +str(X_train.shape))
            if channels==1:
                #Add the "channels" dimension
                X_train = np.expand_dims(X_train,3)

            t3 = time.perf_counter()
            
            #Affine augmentation
            X_train = aid_img.affine_augm(X_train,v_flip,h_flip,rotation,width_shift,height_shift,zoom,shear) #Affine image augmentation
            y_train = np.copy(y_train)
   
            Y_train = to_categorical(y_train, nr_classes)# * 2 - 1
            t4 = time.perf_counter()
            if verbose == 1:
                print("Time to perform affine augmentation ="+str(t4-t3))
                    
            t3 = time.perf_counter()            
            #Now do the final cropping to the actual size that was set by user
            dim = X_train.shape
            if dim[2]!=crop:
                remove = int(dim[2]/2.0 - crop/2.0)
                X_train = X_train[:,remove:remove+crop,remove:remove+crop,:] #crop to crop x crop pixels #TensorFlow
            t4 = time.perf_counter()

            #X_train = np.copy(X_train) #save into new array and do some iterations with varying noise/brightness
            #reuse this X_batch_orig a few times since this augmentation was costly

            if self.actionVerbose.isChecked()==True:
                verbose = 1
            else:
                verbose = 0                            
                                                            
            #In each iteration, start with non-augmented data
            #X_batch = np.copy(X_batch_orig)#copy from X_batch_orig, X_batch will be altered without altering X_batch_orig            
            #X_train = X_train.astype(np.uint8)                            
            
            ##########Contrast/Saturation/Hue augmentation#########
            #is there any of contrast/saturation/hue augmentation to do?
            X_train = X_train.astype(np.uint8)

            if contrast_on:
                t_con_aug_1 = time.perf_counter()
                X_train = aid_img.contrast_augm_cv2(X_train,contrast_lower,contrast_higher) #this function is almost 15 times faster than random_contrast from tf!
                t_con_aug_2 = time.perf_counter()
                if verbose == 1:
                    print("Time to augment contrast="+str(t_con_aug_2-t_con_aug_1))

            if saturation_on or hue_on:
                t_sat_aug_1 = time.perf_counter()
                X_train = aid_img.satur_hue_augm_cv2(X_train.astype(np.uint8),saturation_on,saturation_lower,saturation_higher,hue_on,hue_delta) #Gray and RGB; both values >0!
                t_sat_aug_2 = time.perf_counter()
                if verbose == 1:
                    print("Time to augment saturation/hue="+str(t_sat_aug_2-t_sat_aug_1))

            ##########Average/Gauss/Motion blurring#########
            #is there any of blurring to do?
            if avgBlur_on:
                t_avgBlur_1 = time.perf_counter()
                X_train = aid_img.avg_blur_cv2(X_train,avgBlur_min,avgBlur_max)
                t_avgBlur_2 = time.perf_counter()
                if verbose == 1:
                    print("Time to perform average blurring="+str(t_avgBlur_2-t_avgBlur_1))

            if gaussBlur_on:
                t_gaussBlur_1 = time.perf_counter()
                X_train = aid_img.gauss_blur_cv(X_train,gaussBlur_min,gaussBlur_max)
                t_gaussBlur_2 = time.perf_counter()
                if verbose == 1:
                    print("Time to perform gaussian blurring="+str(t_gaussBlur_2-t_gaussBlur_1))

            if motionBlur_on:
                t_motionBlur_1 = time.perf_counter()
                X_train = aid_img.motion_blur_cv(X_train,motionBlur_kernel,motionBlur_angle)
                t_motionBlur_2 = time.perf_counter()
                if verbose == 1:
                    print("Time to perform motion blurring="+str(t_motionBlur_2-t_motionBlur_1))

            ##########Brightness noise#########
            t3 = time.perf_counter()
            X_train = aid_img.brightn_noise_augm_cv2(X_train,brightness_add_lower,brightness_add_upper,brightness_mult_lower,brightness_mult_upper,gaussnoise_mean,gaussnoise_scale)
            t4 = time.perf_counter()
            if verbose == 1:
                print("Time to augment brightness="+str(t4-t3))

            t3 = time.perf_counter()
            if norm == "StdScaling using mean and std of all training data":
                X_train = aid_img.image_normalization(X_train,norm,mean_trainingdata,std_trainingdata)
            else:
                X_train = aid_img.image_normalization(X_train,norm)
            t4 = time.perf_counter()
            if verbose == 1:
                print("Time to apply normalization="+str(t4-t3))
            
            if verbose == 1: 
                print("X_train.shape")
                print(X_train.shape)

            if xtra_in==True:
                print("Add Xtra Data to X_train")
                X_train = [X_train,xtra_train]

            ###################################################
            ###############Actual fitting######################
            ###################################################

            batch_size = int(self.popup_lrfinder_ui.spinBox_batchSize.value())
            stepsPerEpoch = int(self.popup_lrfinder_ui.spinBox_stepsPerEpoch.value())
            epochs = int(self.popup_lrfinder_ui.spinBox_epochs.value())
            start_lr = float(self.popup_lrfinder_ui.lineEdit_startLr.text())
            stop_lr = float(self.popup_lrfinder_ui.lineEdit_stopLr.text())
            valMetrics = bool(self.popup_lrfinder_ui.checkBox_valMetrics.isChecked())

            ####################lr_find algorithm####################
            if model_keras_p == None:
                lrf = aid_dl.LearningRateFinder(model_keras)
            elif model_keras_p != None:
                lrf = aid_dl.LearningRateFinder(model_keras_p)
            if valMetrics==True:
                lrf.find([X_train,Y_train],[X_valid,Y_valid],start_lr,stop_lr,stepsPerEpoch=stepsPerEpoch,batchSize=batch_size,epochs=epochs)
            else:
                lrf.find([X_train,Y_train],None,start_lr,stop_lr,stepsPerEpoch=stepsPerEpoch,batchSize=batch_size,epochs=epochs)
            
            skipBegin,skipEnd = 10,1
            self.learning_rates = lrf.lrs[skipBegin:-skipEnd]
            self.losses_or = lrf.losses_or[skipBegin:-skipEnd]
            self.losses_sm = lrf.losses_sm[skipBegin:-skipEnd]
            self.accs_or = lrf.accs_or[skipBegin:-skipEnd]
            self.accs_sm = lrf.accs_sm[skipBegin:-skipEnd]
            
            self.val_losses_sm = lrf.val_losses_sm[skipBegin:-skipEnd]
            self.val_losses_or = lrf.val_losses_or[skipBegin:-skipEnd]
            self.val_accs_sm = lrf.val_accs_sm[skipBegin:-skipEnd]
            self.val_accs_or = lrf.val_accs_or[skipBegin:-skipEnd]

            # Enable the groupboxes
            self.popup_lrfinder_ui.groupBox_singleLr.setEnabled(True)
            self.popup_lrfinder_ui.groupBox_LrRange.setEnabled(True)
            
            self.update_lrfind_plot()
            
    def update_lrfind_plot(self):
        if not hasattr(self, 'learning_rates'):
            return

        metric = str(self.popup_lrfinder_ui.comboBox_metric.currentText())
        color = self.popup_lrfinder_ui.pushButton_color.palette().button().color()
        width = int(self.popup_lrfinder_ui.spinBox_lineWidth.value())
        color = list(color.getRgb())
        color = tuple(color)                
        pencolor = pg.mkPen(color, width=width)
        smooth = bool(self.popup_lrfinder_ui.checkBox_smooth.isChecked())

        try:# try to empty the plot
            self.popup_lrfinder_ui.lr_plot.clear()
            #self.popup_lrfinder_ui.lr_plot.removeItem(self.lr_line)
        except:
            pass

        if metric=="Loss" and smooth==True:
            self.y_values = self.losses_sm
        elif metric=="Loss" and smooth==False:
            self.y_values = self.losses_or            
        elif metric=="Loss 1st derivative" and smooth==True:
            self.y_values = np.diff(self.losses_sm,n=1)
        elif metric=="Loss 1st derivative" and smooth==False:
            self.y_values = np.diff(self.losses_or,n=1)
        elif metric=="Accuracy" and smooth==True:
            self.y_values = self.accs_sm
        elif metric=="Accuracy" and smooth==False:
            self.y_values = self.accs_or
        elif metric=="Accuracy 1st derivative" and smooth==True:
            self.y_values = np.diff(self.accs_sm,n=1)
        elif metric=="Accuracy 1st derivative" and smooth==False:
            self.y_values = np.diff(self.accs_or,n=1)

        elif metric=="Val. loss" and smooth==True:
            self.y_values = self.val_losses_sm
        elif metric=="Val. loss" and smooth==False:
            self.y_values = self.val_losses_or
        elif metric=="Val. loss 1st derivative" and smooth==True:
            self.y_values = np.diff(self.val_losses_sm,n=1)
        elif metric=="Val. loss 1st derivative" and smooth==False:
            self.y_values = np.diff(self.val_losses_or,n=1)
        elif metric=="Val. accuracy" and smooth==True:
            self.y_values = self.val_accs_sm
        elif metric=="Val. accuracy" and smooth==False:
            self.y_values = self.val_accs_or
        elif metric=="Val. accuracy 1st derivative" and smooth==True:
            self.y_values = np.diff(self.val_accs_sm,n=1)
        elif metric=="Val. accuracy 1st derivative" and smooth==False:
            self.y_values = np.diff(self.val_accs_or,n=1)
        else:
            print("The combination of "+str(metric)+" and smooth="+str(smooth)+" is not supported!")
        
        if len(self.learning_rates)==len(self.y_values):
            self.lr_line = pg.PlotCurveItem(x=np.log10(self.learning_rates), y=self.y_values,pen=pencolor,name=metric)
        elif len(self.learning_rates)-1==len(self.y_values):
            self.lr_line = pg.PlotCurveItem(x=np.log10(self.learning_rates)[1:], y=self.y_values,pen=pencolor,name=metric)
        else:
            print("No data available. Probably, validation metrics were not computed. Please click Run again.")
            return
        self.popup_lrfinder_ui.lr_plot.addItem(self.lr_line)            

        #In case the groupBox_singleLr is already checked, carry out the function:
        if self.popup_lrfinder_ui.groupBox_singleLr.isChecked():
            self.get_lr_single(on_or_off=True)

        #In case the groupBox_LrRange is already checked, carry out the function:
        if self.popup_lrfinder_ui.groupBox_LrRange.isChecked():
            self.get_lr_range(on_or_off=True)


    def get_lr_single(self,on_or_off):
        if on_or_off==True: #bool(self.popup_lrfinder_ui.groupBox_LrRange.isChecked()):
            ind = np.argmin(self.y_values)#find location of loss-minimum
            mini_x = self.learning_rates[ind]
            mini_x = np.log10(mini_x)
            pen = pg.mkPen(color="w")
            self.lr_single = pg.InfiniteLine(pos=mini_x, angle=90, pen=pen, movable=True)
            self.popup_lrfinder_ui.lr_plot.addItem(self.lr_single)
            
            def position_changed():
                #where did the user drag the region_linfit to?
                new_position = 10**(self.lr_single.value())
                self.popup_lrfinder_ui.lineEdit_singleLr.setText(str(new_position))

            self.lr_single.sigPositionChangeFinished.connect(position_changed)

        if on_or_off==False: #user unchecked the groupbox->remove the InfiniteLine if possible
            try:
                self.popup_lrfinder_ui.lr_plot.removeItem(self.lr_single)
            except:
                pass
        
    
    def get_lr_range(self,on_or_off):
        #print(on_or_off)
        #start_lr = float(self.popup_lrfinder_ui.lineEdit_startLr.text())
        #stop_lr = float(self.popup_lrfinder_ui.lineEdit_stopLr.text())
        if on_or_off==True: #bool(self.popup_lrfinder_ui.groupBox_LrRange.isChecked()):
            start_x = 0.00001
            start_x = np.log10(start_x)
            ind = np.argmin(self.y_values)#find location of loss-minimum
            end_x = self.learning_rates[ind]
            end_x = np.log10(end_x)
            self.lr_region = pg.LinearRegionItem([start_x, end_x], movable=True)
            self.popup_lrfinder_ui.lr_plot.addItem(self.lr_region)

            def region_changed():
                #where did the user drag the region_linfit to?
                new_region = self.lr_region.getRegion()
                new_region_left = 10**(new_region[0])
                new_region_right = 10**(new_region[1])
                self.popup_lrfinder_ui.lineEdit_LrMin.setText(str(new_region_left))
                self.popup_lrfinder_ui.lineEdit_LrMax.setText(str(new_region_right))

            self.lr_region.sigRegionChangeFinished.connect(region_changed)

        if on_or_off==False: #bool(self.popup_lrfinder_ui.groupBox_LrRange.isChecked()):
            try:
                self.popup_lrfinder_ui.lr_plot.removeItem(self.lr_region)
            except:
                pass









        
    def action_show_example_imgs(self): #this function is only for the main window
        if self.actionVerbose.isChecked()==True:
            verbose = 1
        else:
            verbose = 0
        #Get state of the comboboxes!
        tr_or_valid = str(self.comboBox_ShowTrainOrValid.currentText())
        w_or_wo_augm = str(self.comboBox_ShowWOrWoAug.currentText())

        #most of it should be similar to action_fit_model_worker
        #Used files go to a separate sheet on the MetaFile.xlsx
        SelectedFiles = self.items_clicked_no_rtdc_ds()
        #Collect all information about the fitting routine that was user defined
        crop = int(self.spinBox_imagecrop.value())          
        norm = str(self.comboBox_Normalization.currentText())
        h_flip = bool(self.checkBox_HorizFlip.isChecked())
        v_flip = bool(self.checkBox_VertFlip.isChecked())
        rotation = float(self.lineEdit_Rotation.text())
        width_shift = float(self.lineEdit_widthShift.text())
        height_shift = float(self.lineEdit_heightShift.text())
        zoom = float(self.lineEdit_zoomRange.text())
        shear = float(self.lineEdit_shearRange.text())
        brightness_add_lower = float(self.spinBox_PlusLower.value())
        brightness_add_upper = float(self.spinBox_PlusUpper.value())
        brightness_mult_lower = float(self.doubleSpinBox_MultLower.value())
        brightness_mult_upper = float(self.doubleSpinBox_MultUpper.value())
        gaussnoise_mean = float(self.doubleSpinBox_GaussianNoiseMean.value())
        gaussnoise_scale = float(self.doubleSpinBox_GaussianNoiseScale.value())

        contrast_on = bool(self.checkBox_contrast.isChecked())        
        contrast_lower = float(self.doubleSpinBox_contrastLower.value())
        contrast_higher = float(self.doubleSpinBox_contrastHigher.value())
        saturation_on = bool(self.checkBox_saturation.isChecked())        
        saturation_lower = float(self.doubleSpinBox_saturationLower.value())
        saturation_higher = float(self.doubleSpinBox_saturationHigher.value())
        hue_on = bool(self.checkBox_hue.isChecked())        
        hue_delta = float(self.doubleSpinBox_hueDelta.value())

        avgBlur_on = bool(self.checkBox_avgBlur.isChecked())        
        avgBlur_min = int(self.spinBox_avgBlurMin.value())
        avgBlur_max = int(self.spinBox_avgBlurMax.value())

        gaussBlur_on = bool(self.checkBox_gaussBlur.isChecked())        
        gaussBlur_min = int(self.spinBox_gaussBlurMin.value())
        gaussBlur_max = int(self.spinBox_gaussBlurMax.value())

        motionBlur_on = bool(self.checkBox_motionBlur.isChecked())        
        motionBlur_kernel = str(self.lineEdit_motionBlurKernel.text())
        motionBlur_angle = str(self.lineEdit_motionBlurAngle.text())
        
        motionBlur_kernel = tuple(ast.literal_eval(motionBlur_kernel)) #translate string in the lineEdits to a tuple
        motionBlur_angle = tuple(ast.literal_eval(motionBlur_angle)) #translate string in the lineEdits to a tuple

        paddingMode = str(self.comboBox_paddingMode.currentText())#.lower()

        #which index is requested by user:?
        req_index = int(self.spinBox_ShowIndex.value())
        if tr_or_valid=='Training':
            ######################Load the Training Data################################
            ind = [selectedfile["TrainOrValid"] == "Train" for selectedfile in SelectedFiles]
        elif tr_or_valid=='Validation':
            ind = [selectedfile["TrainOrValid"] == "Valid" for selectedfile in SelectedFiles]
        ind = np.where(np.array(ind)==True)[0]
        SelectedFiles = np.array(SelectedFiles)[ind]
        SelectedFiles = list(SelectedFiles)
        indices = [selectedfile["class"] for selectedfile in SelectedFiles]
        ind = np.where(np.array(indices)==req_index)[0]
        if len(ind)<1:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("There is no data for this class available")
            msg.setWindowTitle("Class not available")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return
            
        indices = list(np.array(indices)[ind])
        SelectedFiles = list(np.array(SelectedFiles)[ind])
        nr_events_epoch = len(indices)*[10] #[selectedfile["nr_events_epoch"] for selectedfile in SelectedFiles]
        rtdc_path = [selectedfile["rtdc_path"] for selectedfile in SelectedFiles]
        zoom_factors = [selectedfile["zoom_factor"] for selectedfile in SelectedFiles]
        #zoom_order = [self.actionOrder0.isChecked(),self.actionOrder1.isChecked(),self.actionOrder2.isChecked(),self.actionOrder3.isChecked(),self.actionOrder4.isChecked(),self.actionOrder5.isChecked()]
        #zoom_order = int(np.where(np.array(zoom_order)==True)[0])
        zoom_order = int(self.comboBox_zoomOrder.currentIndex()) #the combobox-index is already the zoom order
        shuffle = [selectedfile["shuffle"] for selectedfile in SelectedFiles]
        #If the scaling method is "divide by mean and std of the whole training set":
        if norm == "StdScaling using mean and std of all training data":
            mean_trainingdata,std_trainingdata = [],[]
            for i in range(len(SelectedFiles)):
                if not self.actionDataToRam.isChecked():
                    #Replace true means that individual cells could occur several times
                    gen = aid_img.gen_crop_img(crop,rtdc_path[i],random_images=False,zoom_factor=zoom_factors[i],zoom_order=zoom_order,color_mode=self.get_color_mode(),padding_mode=paddingMode) 
                else:
                    if len(self.ram)==0:
                        #Replace true means that individual cells could occur several times
                        gen = aid_img.gen_crop_img(crop,rtdc_path[i],random_images=False,zoom_factor=zoom_factors[i],zoom_order=zoom_order,color_mode=self.get_color_mode(),padding_mode=paddingMode)
                    else:    
                        gen = aid_img.gen_crop_img_ram(self.ram,rtdc_path[i],random_images=False) #Replace true means that individual cells could occur several times
                        if self.actionVerbose.isChecked():
                            print("Loaded data from RAM")

                images = next(gen)[0]
                mean_trainingdata.append(np.mean(images))
                std_trainingdata.append(np.std(images))
            mean_trainingdata = np.mean(np.array(mean_trainingdata))
            std_trainingdata = np.mean(np.array(std_trainingdata))
            if np.allclose(std_trainingdata,0):
                std_trainingdata = 0.0001
                print("std_trainingdata was zero and is now set to 0.0001 to avoid div. by zero!")
            if self.actionVerbose.isChecked():
                print("Used all training data to get mean and std for normalization")

        if w_or_wo_augm=='With Augmentation':
            ###############Continue with training data:augmentation############
            #Rotating could create edge effects. Avoid this by making crop a bit larger for now
            #Worst case would be a 45degree rotation:
            cropsize2 = np.sqrt(crop**2+crop**2)
            cropsize2 = np.ceil(cropsize2 / 2.) * 2 #round to the next even number
            
            ############Cropping and image augmentation#####################
            #Start the first iteration:                
            X,y = [],[]
            for i in range(len(SelectedFiles)):
                if not self.actionDataToRam.isChecked():
                    #Replace true means that individual cells could occur several times
                    gen = aid_img.gen_crop_img(cropsize2,rtdc_path[i],10,random_images=True,replace=True,zoom_factor=zoom_factors[i],zoom_order=zoom_order,color_mode=self.get_color_mode(),padding_mode=paddingMode) 
                else:
                    if len(self.ram)==0:
                        #Replace true means that individual cells could occur several times
                        gen = aid_img.gen_crop_img(cropsize2,rtdc_path[i],10,random_images=True,replace=True,zoom_factor=zoom_factors[i],zoom_order=zoom_order,color_mode=self.get_color_mode(),padding_mode=paddingMode) 
                    else:   
                        gen = aid_img.gen_crop_img_ram(self.ram,rtdc_path[i],10,random_images=True,replace=True) #Replace true means that individual cells could occur several times
                        if self.actionVerbose.isChecked():
                            print("Loaded data from RAM")
                try: #When all cells are at the border of the image, the generator will be empty. Avoid program crash by try, except
                    X.append(next(gen)[0])
                except StopIteration:
                    print("All events at border of image and discarded")
                    return
                y.append(np.repeat(indices[i],X[-1].shape[0]))

            X = np.concatenate(X)
            X = X.astype(np.uint8) #make sure we stay in uint8
            y = np.concatenate(y)
        
            if len(X.shape)==4:
                channels=3
            elif len(X.shape)==3:
                channels=1
                X = np.expand_dims(X,3)#Add the "channels" dimension
            else:
                print("Invalid data dimension:" +str(X.shape))
            
            X_batch, y_batch = aid_img.affine_augm(X,v_flip,h_flip,rotation,width_shift,height_shift,zoom,shear), y #Affine image augmentation
            X_batch = X_batch.astype(np.uint8) #make sure we stay in uint8

            #Now do the final cropping to the actual size that was set by user
            dim = X_batch.shape
            if dim[2]!=crop:
                remove = int(dim[2]/2.0 - crop/2.0)
                #X_batch = X_batch[:,:,remove:-remove,remove:-remove] #crop to crop x crop pixels #Theano
                X_batch = X_batch[:,remove:remove+crop,remove:remove+crop,:] #crop to crop x crop pixels #TensorFlow

            ##########Contrast/Saturation/Hue augmentation#########
            #is there any of contrast/saturation/hue augmentation to do?
            if contrast_on:
                X_batch = aid_img.contrast_augm_cv2(X_batch,contrast_lower,contrast_higher) #this function is almost 15 times faster than random_contrast from tf!
            if saturation_on or hue_on:
                X_batch = aid_img.satur_hue_augm_cv2(X_batch.astype(np.uint8),saturation_on,saturation_lower,saturation_higher,hue_on,hue_delta)

            ##########Average/Gauss/Motion blurring#########
            #is there any of blurring to do?
            if avgBlur_on:
                X_batch = aid_img.avg_blur_cv2(X_batch,avgBlur_min,avgBlur_max)
            if gaussBlur_on:
                X_batch = aid_img.gauss_blur_cv(X_batch,gaussBlur_min,gaussBlur_max)
            if motionBlur_on:
                X_batch = aid_img.motion_blur_cv(X_batch,motionBlur_kernel,motionBlur_angle)

            X_batch = aid_img.brightn_noise_augm_cv2(X_batch,brightness_add_lower,brightness_add_upper,brightness_mult_lower,brightness_mult_upper,gaussnoise_mean,gaussnoise_scale)

            if norm == "StdScaling using mean and std of all training data":
                X_batch = aid_img.image_normalization(X_batch,norm,mean_trainingdata,std_trainingdata)
            else:
                X_batch = aid_img.image_normalization(X_batch,norm)
            
            X = X_batch
            if verbose: print("Shape of the shown images is:"+str(X.shape))
            
        elif w_or_wo_augm=='Original image':
            ############Cropping#####################
            X,y = [],[]
            for i in range(len(SelectedFiles)):
                if not self.actionDataToRam.isChecked():
                    #Replace true means that individual cells could occur several times
                    gen = aid_img.gen_crop_img(crop,rtdc_path[i],10,random_images=True,replace=True,zoom_factor=zoom_factors[i],zoom_order=zoom_order,color_mode=self.get_color_mode(),padding_mode=paddingMode) 
                else:
                    if len(self.ram)==0:
                        #Replace true means that individual cells could occur several times
                        gen = aid_img.gen_crop_img(crop,rtdc_path[i],10,random_images=True,replace=True,zoom_factor=zoom_factors[i],zoom_order=zoom_order,color_mode=self.get_color_mode(),padding_mode=paddingMode) 
                    else:                        
                        gen = aid_img.gen_crop_img_ram(self.ram,rtdc_path[i],10,random_images=True,replace=True) #Replace true means that individual cells could occur several times
                        if self.actionVerbose.isChecked():
                            print("Loaded data from RAM")
                try:
                    X.append(next(gen)[0])
                except:
                    return
                y.append(np.repeat(indices[i],X[-1].shape[0]))
            X = np.concatenate(X)
            y = np.concatenate(y)

            if len(X.shape)==4:
                channels=3
            elif len(X.shape)==3:
                channels=1
                X = np.expand_dims(X,3) #Add the "channels" dimension
            else:
                print("Invalid data dimension: " +str(X.shape))
                
        if norm == "StdScaling using mean and std of all training data":
            X = aid_img.image_normalization(X,norm,mean_trainingdata,std_trainingdata)
        else:
            X = aid_img.image_normalization(X,norm)
        
        if verbose: print("Shape of the shown images is: "+str(X.shape))
                
        #Is there already anything shown on the widget?
        children = self.widget_ViewImages.findChildren(QtWidgets.QGridLayout)
        if len(children)>0: #if there is something, delete it!
            for i in reversed(range(self.gridLayout_ViewImages.count())):
                widgetToRemove = self.gridLayout_ViewImages.itemAt(i).widget()
                widgetToRemove.setParent(None)
                widgetToRemove.deleteLater()
        else: #else, create a Gridlayout to put the images
            self.gridLayout_ViewImages = QtWidgets.QGridLayout(self.widget_ViewImages)

        for i in range(5):
            if channels==1:
                img = X[i,:,:,0] #TensorFlow 
            if channels==3:
                img = X[i,:,:,:] #TensorFlow 
            
            #Stretch pixel value to full 8bit range (0-255); only for display
            img = img-np.min(img)
            fac = np.max(img)
            img = (img/fac)*255.0
            img = img.astype(np.uint8)
            if channels==1:
                height, width = img.shape
            if channels==3:
                height, width, _ = img.shape
            
#            qi=QtGui.QImage(img_zoom.data, width, height,width, QtGui.QImage.Format_Indexed8)
#            self.label_image_show = QtWidgets.QLabel(self.widget_ViewImages)
#            self.label_image_show.setPixmap(QtGui.QPixmap.fromImage(qi))
#            self.gridLayout_ViewImages.addWidget(self.label_image_show, 1,i)
#            self.label_image_show.show()
            #Use pygtgraph instead, in order to allow for exporting images
            self.image_show = pg.ImageView(self.widget_ViewImages)
            self.image_show.show()
            if verbose: print("Shape of zoomed image: "+str(img.shape))
            if channels==1:
                self.image_show.setImage(img.T,autoRange=False)
            if channels==3:
                self.image_show.setImage(np.swapaxes(img,0,1),autoRange=False)
                
            self.image_show.ui.histogram.hide()
            self.image_show.ui.roiBtn.hide()
            self.image_show.ui.menuBtn.hide()
            self.gridLayout_ViewImages.addWidget(self.image_show, 1,i)
        self.widget_ViewImages.show()

        
        
    def tableWidget_HistoryInfo_pop_dclick(self,item,listindex):
        if item is not None:
            tableitem = self.fittingpopups_ui[listindex].tableWidget_HistoryInfo_pop.item(item.row(), item.column())
            if str(tableitem.text())!="Show saved only":
                color = QtGui.QColorDialog.getColor()
                if color.getRgb()==(0, 0, 0, 255):#no black!
                    return
                else:
                    tableitem.setBackground(color)
                #self.update_historyplot_pop(listindex)
       
        
        
        
        
    def action_show_example_imgs_pop(self,listindex): #this function is only for the main window
        #Get state of the comboboxes!
        tr_or_valid = str(self.fittingpopups_ui[listindex].comboBox_ShowTrainOrValid_pop.currentText())
        w_or_wo_augm = str(self.fittingpopups_ui[listindex].comboBox_ShowWOrWoAug_pop.currentText())

        #most of it should be similar to action_fit_model_worker
        #Used files go to a separate sheet on the MetaFile.xlsx
        SelectedFiles = self.items_clicked_no_rtdc_ds()
        #Collect all information about the fitting routine that was user defined
        crop = int(self.fittingpopups_ui[listindex].spinBox_imagecrop_pop.value())          
        norm = str(self.fittingpopups_ui[listindex].comboBox_Normalization_pop.currentText())
        h_flip = bool(self.fittingpopups_ui[listindex].checkBox_HorizFlip_pop.isChecked())
        v_flip = bool(self.fittingpopups_ui[listindex].checkBox_VertFlip_pop.isChecked())
        rotation = float(self.fittingpopups_ui[listindex].lineEdit_Rotation_pop.text())
        width_shift = float(self.fittingpopups_ui[listindex].lineEdit_widthShift_pop.text())
        height_shift = float(self.fittingpopups_ui[listindex].lineEdit_heightShift_pop.text())
        zoom = float(self.fittingpopups_ui[listindex].lineEdit_zoomRange_pop.text())
        shear = float(self.fittingpopups_ui[listindex].lineEdit_shearRange_pop.text())
        brightness_add_lower = float(self.fittingpopups_ui[listindex].spinBox_PlusLower_pop.value())
        brightness_add_upper = float(self.fittingpopups_ui[listindex].spinBox_PlusUpper_pop.value())
        brightness_mult_lower = float(self.fittingpopups_ui[listindex].doubleSpinBox_MultLower_pop.value())
        brightness_mult_upper = float(self.fittingpopups_ui[listindex].doubleSpinBox_MultUpper_pop.value())
        gaussnoise_mean = float(self.fittingpopups_ui[listindex].doubleSpinBox_GaussianNoiseMean_pop.value())
        gaussnoise_scale = float(self.fittingpopups_ui[listindex].doubleSpinBox_GaussianNoiseScale_pop.value())
        
        contrast_on = bool(self.fittingpopups_ui[listindex].checkBox_contrast_pop.isChecked())
        contrast_lower = float(self.fittingpopups_ui[listindex].doubleSpinBox_contrastLower_pop.value())
        contrast_higher = float(self.fittingpopups_ui[listindex].doubleSpinBox_contrastHigher_pop.value())
        saturation_on = bool(self.fittingpopups_ui[listindex].checkBox_saturation_pop.isChecked())
        saturation_lower = float(self.fittingpopups_ui[listindex].doubleSpinBox_saturationLower_pop.value())
        saturation_higher = float(self.fittingpopups_ui[listindex].doubleSpinBox_saturationHigher_pop.value())
        hue_on = bool(self.fittingpopups_ui[listindex].checkBox_hue_pop.isChecked())
        hue_delta = float(self.fittingpopups_ui[listindex].doubleSpinBox_hueDelta_pop.value())

        avgBlur_on = bool(self.fittingpopups_ui[listindex].checkBox_avgBlur_pop.isChecked())        
        avgBlur_min = int(self.fittingpopups_ui[listindex].spinBox_avgBlurMin_pop.value())
        avgBlur_max = int(self.fittingpopups_ui[listindex].spinBox_avgBlurMax_pop.value())
    
        gaussBlur_on = bool(self.fittingpopups_ui[listindex].checkBox_gaussBlur_pop.isChecked())        
        gaussBlur_min = int(self.fittingpopups_ui[listindex].spinBox_gaussBlurMin_pop.value())
        gaussBlur_max = int(self.fittingpopups_ui[listindex].spinBox_gaussBlurMax_pop.value())
    
        motionBlur_on = bool(self.fittingpopups_ui[listindex].checkBox_motionBlur_pop.isChecked())        
        motionBlur_kernel = str(self.fittingpopups_ui[listindex].lineEdit_motionBlurKernel_pop.text())
        motionBlur_angle = str(self.fittingpopups_ui[listindex].lineEdit_motionBlurAngle_pop.text())
        
        motionBlur_kernel = tuple(ast.literal_eval(motionBlur_kernel)) #translate string in the lineEdits to a tuple
        motionBlur_angle = tuple(ast.literal_eval(motionBlur_angle)) #translate string in the lineEdits to a tuple

        paddingMode = str(self.fittingpopups_ui[listindex].comboBox_paddingMode_pop.currentText()).lower()

        #which index is requested by user:?
        req_index = int(self.fittingpopups_ui[listindex].spinBox_ShowIndex_pop.value())
        if tr_or_valid=='Training':
            ######################Load the Training Data################################
            ind = [selectedfile["TrainOrValid"] == "Train" for selectedfile in SelectedFiles]
        elif tr_or_valid=='Validation':
            ind = [selectedfile["TrainOrValid"] == "Valid" for selectedfile in SelectedFiles]
        ind = np.where(np.array(ind)==True)[0]
        SelectedFiles = np.array(SelectedFiles)[ind]
        SelectedFiles = list(SelectedFiles)
        indices = [selectedfile["class"] for selectedfile in SelectedFiles]
        ind = np.where(np.array(indices)==req_index)[0]
        if len(ind)<1:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("There is no data for this class available")
            msg.setWindowTitle("Class not available")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return
            
        indices = list(np.array(indices)[ind])
        SelectedFiles = list(np.array(SelectedFiles)[ind])
        nr_events_epoch = len(indices)*[10] #[selectedfile["nr_events_epoch"] for selectedfile in SelectedFiles]
        rtdc_path = [selectedfile["rtdc_path"] for selectedfile in SelectedFiles]
        zoom_factors = [selectedfile["zoom_factor"] for selectedfile in SelectedFiles]
        #zoom_order = [self.actionOrder0.isChecked(),self.actionOrder1.isChecked(),self.actionOrder2.isChecked(),self.actionOrder3.isChecked(),self.actionOrder4.isChecked(),self.actionOrder5.isChecked()]
        #zoom_order = int(np.where(np.array(zoom_order)==True)[0])
        zoom_order = int(self.comboBox_zoomOrder.currentIndex()) #the combobox-index is already the zoom order
        
        shuffle = [selectedfile["shuffle"] for selectedfile in SelectedFiles]
        xtra_in = set([selectedfile["xtra_in"] for selectedfile in SelectedFiles])   
        if len(xtra_in)>1:# False and True is present. Not supported
            print("Xtra data is used only for some files. Xtra data needs to be used either by all or by none!")
            return
        xtra_in = list(xtra_in)[0]#this is either True or False
        
        #If the scaling method is "divide by mean and std of the whole training set":
        if norm == "StdScaling using mean and std of all training data":
            mean_trainingdata,std_trainingdata = [],[]
            for i in range(len(SelectedFiles)):
                if not self.actionDataToRam.isChecked():
                    gen = aid_img.gen_crop_img(crop,rtdc_path[i],random_images=False,zoom_factor=zoom_factors[i],zoom_order=zoom_order,color_mode=self.get_color_mode(),padding_mode=paddingMode,xtra_in=xtra_in) #Replace true means that individual cells could occur several times
                else:
                    if len(self.ram)==0:
                        gen = aid_img.gen_crop_img(crop,rtdc_path[i],random_images=False,zoom_factor=zoom_factors[i],zoom_order=zoom_order,color_mode=self.get_color_mode(),padding_mode=paddingMode,xtra_in=xtra_in) #Replace true means that individual cells could occur several times
                    else:    
                        gen = aid_img.gen_crop_img_ram(self.ram,rtdc_path[i],random_images=False,xtra_in=xtra_in) #Replace true means that individual cells could occur several times
                        if self.actionVerbose.isChecked():
                            print("Loaded data from RAM")

                images = next(gen)[0]
                mean_trainingdata.append(np.mean(images))
                std_trainingdata.append(np.std(images))
            mean_trainingdata = np.mean(np.array(mean_trainingdata))
            std_trainingdata = np.mean(np.array(std_trainingdata))
            if np.allclose(std_trainingdata,0):
                std_trainingdata = 0.0001
                print("std_trainingdata turned out to be zero. I set it to 0.0001, to avoid division by zero!")
            if self.actionVerbose.isChecked():
                print("Used all training data to get mean and std for normalization")

        if w_or_wo_augm=='With Augmentation':
            ###############Continue with training data:augmentation############
            #Rotating could create edge effects. Avoid this by making crop a bit larger for now
            #Worst case would be a 45degree rotation:
            cropsize2 = np.sqrt(crop**2+crop**2)
            cropsize2 = np.ceil(cropsize2 / 2.) * 2 #round to the next even number

            ############Get cropped images with image augmentation#####################
            #Start the first iteration:                
            X,y = [],[]
            for i in range(len(SelectedFiles)):
                if not self.actionDataToRam.isChecked():
                    #Replace true means that individual cells could occur several times
                    gen = aid_img.gen_crop_img(cropsize2,rtdc_path[i],10,random_images=shuffle[i],replace=True,zoom_factor=zoom_factors[i],zoom_order=zoom_order,color_mode=self.get_color_mode(),padding_mode=paddingMode)
                else:
                    if len(self.ram)==0:
                        #Replace true means that individual cells could occur several times
                        gen = aid_img.gen_crop_img(cropsize2,rtdc_path[i],10,random_images=shuffle[i],replace=True,zoom_factor=zoom_factors[i],zoom_order=zoom_order,color_mode=self.get_color_mode(),padding_mode=paddingMode)
                    else:   
                        gen = aid_img.gen_crop_img_ram(self.ram,rtdc_path[i],10,random_images=shuffle[i],replace=True) #Replace true means that individual cells could occur several times
                        if self.actionVerbose.isChecked():
                            print("Loaded data from RAM")
                    
                X.append(next(gen)[0])
                #y.append(np.repeat(indices[i],nr_events_epoch[i]))
                y.append(np.repeat(indices[i],X[-1].shape[0]))

            X = np.concatenate(X)
            y = np.concatenate(y)
            
            if len(X.shape)==4:
                channels=3
            elif len(X.shape)==3:
                channels=1
            else:
                print("Invalid data dimension:" +str(X.shape))
            if channels==1:
                #Add the "channels" dimension
                X = np.expand_dims(X,3)

            X_batch, y_batch = aid_img.affine_augm(X,v_flip,h_flip,rotation,width_shift,height_shift,zoom,shear), y #Affine image augmentation
            X_batch = X_batch.astype(np.uint8) #make sure we stay in uint8

            #Now do the final cropping to the actual size that was set by user
            dim = X_batch.shape
            if dim[2]!=crop:
                remove = int(dim[2]/2.0 - crop/2.0)
                #X_batch = X_batch[:,:,remove:-remove,remove:-remove] #crop to crop x crop pixels #Theano
                X_batch = X_batch[:,remove:remove+crop,remove:remove+crop,:] #crop to crop x crop pixels #TensorFlow

            ##########Contrast/Saturation/Hue augmentation#########
            #is there any of contrast/saturation/hue augmentation to do?
            if contrast_on:
                X_batch = aid_img.contrast_augm_cv2(X_batch,contrast_lower,contrast_higher) #this function is almost 15 times faster than random_contrast from tf!
            if saturation_on or hue_on:
                X_batch = aid_img.satur_hue_augm_cv2(X_batch.astype(np.uint8),saturation_on,saturation_lower,saturation_higher,hue_on,hue_delta)
           
            ##########Average/Gauss/Motion blurring#########
            #is there any of blurring to do?
            if avgBlur_on:
                X_batch = aid_img.avg_blur_cv2(X_batch,avgBlur_min,avgBlur_max)
            if gaussBlur_on:
                X_batch = aid_img.gauss_blur_cv(X_batch,gaussBlur_min,gaussBlur_max)
            if motionBlur_on:
                X_batch = aid_img.motion_blur_cv(X_batch,motionBlur_kernel,motionBlur_angle)

            X_batch = aid_img.brightn_noise_augm_cv2(X_batch,brightness_add_lower,brightness_add_upper,brightness_mult_lower,brightness_mult_upper,gaussnoise_mean,gaussnoise_scale)
            
            if norm == "StdScaling using mean and std of all training data":
                X_batch = aid_img.image_normalization(X_batch,norm,mean_trainingdata,std_trainingdata)
            else:
                X_batch = aid_img.image_normalization(X_batch,norm)
            X = X_batch
            
        elif w_or_wo_augm=='Original image':
            ############Cropping#####################
            X,y = [],[]
            for i in range(len(SelectedFiles)):
                if not self.actionDataToRam.isChecked():
                    #Replace true means that individual cells could occur several times
                    gen = aid_img.gen_crop_img(crop,rtdc_path[i],10,random_images=shuffle[i],replace=True,zoom_factor=zoom_factors[i],zoom_order=zoom_order,color_mode=self.get_color_mode(),padding_mode=paddingMode) 
                else:
                    if len(self.ram)==0:
                        #Replace true means that individual cells could occur several times
                        gen = aid_img.gen_crop_img(crop,rtdc_path[i],10,random_images=shuffle[i],replace=True,zoom_factor=zoom_factors[i],zoom_order=zoom_order,color_mode=self.get_color_mode(),padding_mode=paddingMode) 
                    else:                        
                        gen = aid_img.gen_crop_img_ram(self.ram,rtdc_path[i],10,random_images=shuffle[i],replace=True) #Replace true means that individual cells could occur several times
                        if self.actionVerbose.isChecked():
                            print("Loaded data from RAM")

                X.append(next(gen)[0])
                #y.append(np.repeat(indices[i],nr_events_epoch[i]))
                y.append(np.repeat(indices[i],X[-1].shape[0]))

            X = np.concatenate(X)
            y = np.concatenate(y)
        
            if len(X.shape)==4:
                channels = 3
            elif len(X.shape)==3:
                channels = 1
                X = np.expand_dims(X,3)#Add the "channels" dimension
            else:
                print("Invalid data dimension:" +str(X.shape))
                
            if norm == "StdScaling using mean and std of all training data":
                X = aid_img.image_normalization(X,norm,mean_trainingdata,std_trainingdata)
            else:
                X = aid_img.image_normalization(X,norm)
                
        #Is there already anything shown on the widget?
        children = self.fittingpopups_ui[listindex].widget_ViewImages_pop.findChildren(QtWidgets.QGridLayout)
        if len(children)>0: #if there is something, delete it!
            for i in reversed(range(self.fittingpopups_ui[listindex].gridLayout_ViewImages_pop.count())):
                widgetToRemove = self.fittingpopups_ui[listindex].gridLayout_ViewImages_pop.itemAt(i).widget()
                widgetToRemove.setParent(None)
                widgetToRemove.deleteLater()
        else: #else, create a Gridlayout to put the images
            self.fittingpopups_ui[listindex].gridLayout_ViewImages_pop = QtWidgets.QGridLayout(self.fittingpopups_ui[listindex].widget_ViewImages_pop)

        for i in range(5):
            if channels==1:
                img = X[i,:,:,0]
            if channels==3:
                img = X[i,:,:,:]
            
            #Normalize image to full 8bit range (from 0 to 255)
            img = img-np.min(img)
            fac = np.max(img)
            img = (img/fac)*255.0
            img = img.astype(np.uint8)

#            height, width = img_zoom.shape
#            qi=QtGui.QImage(img_zoom.data, width, height,width, QtGui.QImage.Format_Indexed8)
#            self.label_image_show = QtWidgets.QLabel(self.widget_ViewImages)
#            self.label_image_show.setPixmap(QtGui.QPixmap.fromImage(qi))
#            self.gridLayout_ViewImages_pop.addWidget(self.label_image_show, 1,i)
#            self.label_image_show.show()
            #Use pygtgraph instead, in order to allow for exporting images
            self.fittingpopups_ui[listindex].image_show_pop = pg.ImageView(self.fittingpopups_ui[listindex].widget_ViewImages_pop)
            self.fittingpopups_ui[listindex].image_show_pop.show()
            
            if channels==1:
                self.fittingpopups_ui[listindex].image_show_pop.setImage(img.T,autoRange=False)
            if channels==3:
                self.fittingpopups_ui[listindex].image_show_pop.setImage(np.swapaxes(img,0,1),autoRange=False)
                
            self.fittingpopups_ui[listindex].image_show_pop.ui.histogram.hide()
            self.fittingpopups_ui[listindex].image_show_pop.ui.roiBtn.hide()
            self.fittingpopups_ui[listindex].image_show_pop.ui.menuBtn.hide()
            self.fittingpopups_ui[listindex].gridLayout_ViewImages_pop.addWidget(self.fittingpopups_ui[listindex].image_show_pop, 1,i)
        self.fittingpopups_ui[listindex].widget_ViewImages_pop.show()

    def get_color_mode(self):
        if str(self.comboBox_GrayOrRGB.currentText())=="Grayscale":
            return "Grayscale"
        elif str(self.comboBox_GrayOrRGB.currentText())=="RGB":
            return "RGB"
        else:
            return None
        
    def checkBox_rollingMedian_statechange(self,item):#used in frontend
        self.horizontalSlider_rollmedi.setEnabled(item)
        
    def update_historyplot(self):
        #After loading a history, there are checkboxes available. Check, if user checked some:
        colcount = self.tableWidget_HistoryItems.columnCount()
        #Collect items that are checked
        selected_items = []
        Colors = []
        for colposition in range(colcount):  
            #get checkbox item and; is it checked?
            cb = self.tableWidget_HistoryItems.item(0, colposition)
            if not cb==None:
                if cb.checkState() == QtCore.Qt.Checked:
                    selected_items.append(str(cb.text()))
                    Colors.append(cb.background())
                 
        #Get a list of the color from the background of the table items
        DF1 = self.loaded_history

        #Clear the plot        
        self.widget_Scatterplot.clear()
            
        #Add plot        
        self.plt1 = self.widget_Scatterplot.addPlot()
        self.plt1.showGrid(x=True,y=True)
        self.plt1.addLegend()
        self.plt1.setLabel('bottom', 'Epoch', units='')
        
        self.plot_rollmedis = [] #list for plots of rolling medians
        
        if "Show saved only" in selected_items:
            #nr_of_selected_items = len(selected_items)-1
            #get the "Saved" column from DF1
            saved = DF1["Saved"]
            saved = np.where(np.array(saved==1))[0]
#        else:
#            nr_of_selected_items = len(selected_items)
            
        self.Colors = Colors
        scatter_x,scatter_y = [],[]
        for i in range(len(selected_items)):
            key = selected_items[i]
            if key!="Show saved only":
                df = DF1[key]  
                epochs = range(len(df))
                win = int(self.horizontalSlider_rollmedi.value())
                rollmedi = df.rolling(window=win).median()
                
                if "Show saved only" in selected_items:
                    df = np.array(df)[saved]
                    epochs = np.array(epochs)[saved]
                    rollmedi = pd.DataFrame(df).rolling(window=win).median()

                scatter_x.append(epochs)
                scatter_y.append(df)
                color = self.Colors[i]
                pen_rollmedi = list(color.color().getRgb())
                pen_rollmedi = pg.mkColor(pen_rollmedi)
                pen_rollmedi = pg.mkPen(color=pen_rollmedi,width=6)
                color = list(color.color().getRgb())
                color[-1] = int(0.6*color[-1])
                color = tuple(color)                
                pencolor = pg.mkColor(color)
                brush = pg.mkBrush(color=pencolor)
                self.plt1.plot(epochs, df,pen=None,symbol='o',symbolPen=None,symbolBrush=brush,name=key,clear=False)
                if bool(self.checkBox_rollingMedian.isChecked()):#Should a rolling median be plotted?
                    try:
                        rollmedi = np.array(rollmedi).reshape(rollmedi.shape[0])
                        rm = self.plt1.plot(np.array(epochs), rollmedi,pen=pen_rollmedi,clear=False)
                        self.plot_rollmedis.append(rm)
                    except Exception as e:
                        #There is an issue for the rolling median plotting!
                        msg = QtWidgets.QMessageBox()
                        msg.setIcon(QtWidgets.QMessageBox.Warning)       
                        msg.setText(str(e)+"\n->There are likely too few points to have a rolling median with such a window size ("+str(round(win))+")")
                        msg.setWindowTitle("Error occured when plotting rolling median:")
                        msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                        msg.exec_()                        
                        
        if len(str(self.lineEdit_LoadHistory.text()))==0:
        #if DF1==None:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("Please load History file first (.meta)")
            msg.setWindowTitle("No History file loaded")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return
            
        if len(scatter_x)==0:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("Please select at least one of " +"\n".join(list(DF1.keys())))
            msg.setWindowTitle("No quantity selected")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return

        #Keep the information as lists available for this function
        self.scatter_x_l, self.scatter_y_l = scatter_x,scatter_y
        if bool(self.checkBox_linearFit.isChecked()):
            #Put a liner region on the plot; cover the last 10% of points
            if np.max(np.concatenate(scatter_x))<12:
                start_x = 0
                end_x = np.max(np.concatenate(scatter_x))+1
            else:
                start_x = int(0.9*np.max(np.concatenate(scatter_x)))
                end_x = int(1.0*np.max(np.concatenate(scatter_x)))
            self.region_linfit = pg.LinearRegionItem([start_x, end_x], bounds=[-np.inf,np.inf], movable=True)
            self.plt1.addItem(self.region_linfit)

            def region_changed():
                try: #clear the plot from other fits if there are any
                    if len(self.plot_fits)>0:
                        for i in range(len(self.plot_fits)):
                            self.plt1.legend.removeItem(self.names[i])                                
                            self.plt1.removeItem(self.plot_fits[i])
                except:
                    pass
                #where did the user drag the region_linfit to?
                new_region = self.region_linfit.getRegion()
                #for each curve, do a linear regression
                self.plot_fits,self.names = [], []
                for i in range(len(self.scatter_x_l)):
                    scatter_x_vals = np.array(self.scatter_x_l[i])
                    ind = np.where( (scatter_x_vals<new_region[1]) & (scatter_x_vals>new_region[0]) )
                    scatter_x_vals = scatter_x_vals[ind]
                    scatter_y_vals = np.array(self.scatter_y_l[i])[ind]
                    if len(scatter_x_vals)>1:
                        fit = np.polyfit(scatter_x_vals,scatter_y_vals,1)
                        fit_y = fit[0]*scatter_x_vals+fit[1]    
                        pencolor = pg.mkColor(self.Colors[i].color())
                        pen = pg.mkPen(color=pencolor,width=6)
                        text = 'y='+("{:.2e}".format(fit[0]))+"x + " +("{:.2e}".format(fit[1]))
                        self.names.append(text)
                        self.plot_fits.append(self.plt1.plot(name=text))
                        self.plot_fits[i].setData(scatter_x_vals,fit_y,pen=pen,clear=False,name=text)

            self.region_linfit.sigRegionChangeFinished.connect(region_changed)

        def slider_changed():
            if bool(self.checkBox_rollingMedian.isChecked()):
                #remove other rolling median lines:
                for i in range(len(self.plot_rollmedis)):
                    self.plt1.removeItem(self.plot_rollmedis[i])
                #Start with fresh list 
                self.plot_rollmedis = []
                win = int(self.horizontalSlider_rollmedi.value())
                for i in range(len(self.scatter_x_l)):
                    epochs = np.array(self.scatter_x_l[i])
                    if type(self.scatter_y_l[i]) == pd.core.frame.DataFrame:
                        rollmedi = self.scatter_y_l[i].rolling(window=win).median()
                    else:
                        rollmedi = pd.DataFrame(self.scatter_y_l[i]).rolling(window=win).median()
                    rollmedi = np.array(rollmedi).reshape(rollmedi.shape[0])
                    pencolor = pg.mkColor(self.Colors[i].color())
                    pen_rollmedi = pg.mkPen(color=pencolor,width=6)
                    rm = self.plt1.plot(np.array(epochs), rollmedi,pen=pen_rollmedi,clear=False)
                    self.plot_rollmedis.append(rm)

        self.horizontalSlider_rollmedi.sliderMoved.connect(slider_changed)



        scatter_x = np.concatenate(scatter_x)
        scatter_y = np.concatenate(scatter_y)
        scatter_x_norm = (scatter_x.astype(float))/float(np.max(scatter_x))
        scatter_y_norm = (scatter_y.astype(float))/float(np.max(scatter_y))

        self.model_was_selected_before = False
        def onClick(event):
            #Get all plotting items
            #if len(self.plt1.listDataItems())==nr_of_selected_items+1:
                #delete the last item if the user selected already one:
            if self.model_was_selected_before:
                self.plt1.removeItem(self.plt1.listDataItems()[-1])

            items = self.widget_Scatterplot.scene().items(event.scenePos())
            #get the index of the viewbox
            isviewbox = [type(item)==pg.graphicsItems.ViewBox.ViewBox for item in items]
            index = np.where(np.array(isviewbox)==True)[0]
            vb = np.array(items)[index]
            try: #when user rescaed the vew and clicks somewhere outside, it could appear an IndexError.
                clicked_x =  float(vb[0].mapSceneToView(event.scenePos()).x())
                clicked_y =  float(vb[0].mapSceneToView(event.scenePos()).y())
            except:
                return
            try:
                a1 = (clicked_x)/float(np.max(scatter_x))            
                a2 = (clicked_y)/float(np.max(scatter_y))
            except Exception as e:
                print(str(e))
                return
            #Which is the closest scatter point?
            dist = np.sqrt(( a1-scatter_x_norm )**2 + ( a2-scatter_y_norm )**2)
            index =  np.argmin(dist)
            clicked_x = scatter_x[index]
            clicked_y = scatter_y[index]
            #Update the spinBox
            #self.spinBox_ModelIndex.setValue(int(clicked_x))
            #Modelindex for textBrowser_SelectedModelInfo
            text_index = "\nModelindex: "+str(clicked_x)
            #Indicate the selected model on the scatter plot
            self.plt1.plot([clicked_x], [clicked_y],pen=None,symbol='o',symbolPen='w',clear=False)

            #Get more information about this model
            Modelname = str(self.loaded_para["Modelname"].iloc[0])
            
            path, filename = os.path.split(Modelname)
            filename = filename.split(".model")[0]+"_"+str(clicked_x)+".model" 
            
            path = os.path.join(path,filename)
            if os.path.isfile(path):
                text_path = "\nFile is located in:"+path
            else:
                text_path = "\nFile not found!:"+path+"\nProbably the .model was deleted or not saved"
            text_acc = str(DF1.iloc[clicked_x])
            self.textBrowser_SelectedModelInfo.setText("Loaded model: "+filename+text_index+text_path+"\nPerformance:\n"+text_acc)
            self.model_was_selected_before = True
            self.model_2_convert = path
        self.widget_Scatterplot.scene().sigMouseClicked.connect(onClick)

    def action_load_history(self):
        filename = QtWidgets.QFileDialog.getOpenFileName(self, 'Open meta-data', Default_dict["Path of last model"],"AIDeveloper Meta file (*meta.xlsx)")
        filename = filename[0]
        if not filename.endswith("meta.xlsx"):
            return
        if not os.path.isfile(filename):
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("File not found")
            msg.setWindowTitle("File not found")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return
        self.lineEdit_LoadHistory.setText(filename)
        self.action_plot_history(filename)

    def action_load_history_current(self):
        if self.model_keras_path==None:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("There is no fitting going on")
            msg.setWindowTitle("No current fitting process!")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return
            
        history_path = self.model_keras_path
        if type(history_path)==list:#collection=True
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("Not implemented for collections. Please use 'Load History' button to specify a single .meta file")
            msg.setWindowTitle("Not implemented for collecitons")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return

        filename = history_path.split("_0.model")[0]+"_meta.xlsx"
        
        if not filename.endswith("meta.xlsx"):
            return
        if not os.path.isfile(filename):
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("File not found")
            msg.setWindowTitle("File not found")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return
        self.lineEdit_LoadHistory.setText(filename)
        self.action_plot_history(filename)
        
    def action_plot_history(self,filename):
        #If there is a file, it can happen that fitting is currently going on
        #and with bad luck AID just tries to write to the file. This would cause a crash.
        #Therfore, first try to copy the file to a temporary folder. If that fails,
        #wait 1 seconds and try again
        
        #There needs to be a "temp" folder. If there os none, create it!
        #does temp exist?
        tries = 0 #during fitting, AID sometimes wants to write to the history file. In this case we cant read
        try:
            while tries<15:#try a few times
                try:
                    temp_path = aid_bin.create_temp_folder()#create a temp folder if it does not already exist
                    #Create a  random filename for a temp. file
                    someletters = list("STERNBURGPILS")
                    temporaryfile = np.random.choice(someletters,5,replace=True)
                    temporaryfile = "".join(temporaryfile)+".xlsx"
                    temporaryfile = os.path.join(temp_path,temporaryfile)
                    shutil.copyfile(filename,temporaryfile) #copy the original excel file there
                    dic = pd.read_excel(temporaryfile,sheet_name='History',index_col=0,engine="openpyxl") #open it there
                    self.loaded_history = dic
                    para = pd.read_excel(temporaryfile,sheet_name='Parameters',engine="openpyxl")
                    print(temporaryfile)
                    #delete the tempfile
                    os.remove(temporaryfile)
                    self.loaded_para = para    
                    tries = 16
                except:
                    time.sleep(1.5)
                    tries+=1

        except Exception as e:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Critical)       
            msg.setText(str(e))
            msg.setWindowTitle("Error")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return
            
        #Check if dic exists now
        try:
            keys = list(dic.keys())
        except Exception as e:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Critical)       
            msg.setText(str(e))
            msg.setWindowTitle("Error")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return
        #Remember the path for next time
        Default_dict["Path of last model"] = os.path.split(filename)[0]
        aid_bin.save_aid_settings(Default_dict)
        
        #sort the list alphabetically
        keys_ = [l.lower() for l in keys]
        ind_sort = np.argsort(keys_)
        keys = list(np.array(keys)[ind_sort])

        #First keys should always be acc,loss,val_acc,val_loss -in this order
        keys_first = ["accuracy","loss","val_accuracy","val_loss"]
        for i in range(len(keys_first)):
            if keys_first[i] in keys:
                ind = np.where(np.array(keys)==keys_first[i])[0][0]
                if ind!=i:
                    del keys[ind]
                    keys.insert(i,keys_first[i])
        #Lastly check if there is "Saved" or "Time" present and shift it to the back
        keys_last = ["Saved","Time"]
        for i in range(len(keys_last)):
            if keys_last[i] in keys:
                ind = np.where(np.array(keys)==keys_last[i])[0][0]
                if ind!=len(keys):
                    del keys[ind]
                    keys.append(keys_last[i])


        
        self.tableWidget_HistoryItems.setColumnCount(len(keys)+1) #+1 because of "Show saved only"
        #for each key, put a checkbox on the tableWidget_HistoryInfo_pop
        rowPosition = self.tableWidget_HistoryItems.rowCount()
        if rowPosition==0:
            self.tableWidget_HistoryItems.insertRow(0)
        else:
            rowPosition=0
            
        for columnPosition in range(len(keys)):#(2,4):
            key = keys[columnPosition]
            item = QtWidgets.QTableWidgetItem(str(key))#("item {0} {1}".format(rowNumber, columnNumber))
            item.setBackground(QtGui.QColor(self.colorsQt[columnPosition]))
            item.setFlags( QtCore.Qt.ItemIsUserCheckable | QtCore.Qt.ItemIsEnabled  )
            item.setCheckState(QtCore.Qt.Unchecked)
            self.tableWidget_HistoryItems.setItem(rowPosition, columnPosition, item)

        #One checkbox at the end to switch on/of to show only the models that are saved
        columnPosition = len(keys)
        item = QtWidgets.QTableWidgetItem("Show saved only")#("item {0} {1}".format(rowNumber, columnNumber))
        item.setFlags( QtCore.Qt.ItemIsUserCheckable | QtCore.Qt.ItemIsEnabled  )
        item.setCheckState(QtCore.Qt.Unchecked)
        self.tableWidget_HistoryItems.setItem(rowPosition, columnPosition, item)
        
        self.tableWidget_HistoryItems.resizeColumnsToContents()
        self.tableWidget_HistoryItems.resizeRowsToContents()



    def history_tab_get_model_path(self):#Let user define a model he would like to convert
        #pushButton_LoadModel
        #Open a QFileDialog
        filepath = QtWidgets.QFileDialog.getOpenFileName(self, 'Select a trained model you want to convert', Default_dict["Path of last model"],"Keras Model file (*.model)")
        filepath = filepath[0]
        if os.path.isfile(filepath):
            self.model_2_convert = filepath
            path, filename = os.path.split(filepath)
            try:
                modelindex = filename.split(".model")[0]
                modelindex = int(modelindex.split("_")[-1])
            except:
                modelindex = np.nan
                self.textBrowser_SelectedModelInfo.setText("Error loading model")
                return
            text = "Loaded model: "+filename+"\nModelindex: "+str(modelindex)+"\nFile is located in: "+filepath
            self.textBrowser_SelectedModelInfo.setText(text)
                         

    def history_tab_convertModel(self):
        #Check if there is text in textBrowser_SelectedModelInfo
        path = self.model_2_convert
        try:
            os.path.isfile(path)
        except:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("No file defined!")
            msg.setWindowTitle("No file defined!")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return

        if not os.path.isfile(path):
            #text_path = "\nFile not found!:"+path+"\nProbably the .model was deleted or not saved"
            #self.pushButton_convertModel.setEnabled(False)
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("\nFile not found!:"+path+"\nProbably the .model was deleted or not saved")
            msg.setWindowTitle("File not found")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return

        #If the source format is Keras tensforflow:
        source_format = str(self.combobox_initial_format.currentText())                        
        target_format = str(self.comboBox_convertTo.currentText()) #What is the target format?

        ##TODO: All conversion methods to multiprocessing functions!
        def conversion_successful_msg(text):#Enable the Convert to .nnet button
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText(text)
            msg.setWindowTitle("Successfully converted model!")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()


        ##################Keras TensorFlow -> .nnet############################
        if target_format==".nnet" and source_format=="Keras TensorFlow": 
            ConvertToNnet = 1
            worker = Worker(self.history_tab_convertModel_nnet_worker,ConvertToNnet)
            def get_model_keras_from_worker(dic):
                self.model_keras = dic["model_keras"]
            worker.signals.history.connect(get_model_keras_from_worker)
            def conversion_successful(i):#Enable the Convert to .nnet button
                msg = QtWidgets.QMessageBox()
                msg.setIcon(QtWidgets.QMessageBox.Information)
                text = "Conversion Keras TensorFlow -> .nnet done"
                msg.setText(text)
                msg.setWindowTitle("Successfully converted model!")
                msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                msg.exec_()
                #self.pushButton_convertModel.setEnabled(True) 
            worker.signals.history.connect(conversion_successful)
            self.threadpool.start(worker)

        ##################Keras TensorFlow -> Frozen .pb#######################
        elif target_format=="Frozen TensorFlow .pb" and source_format=="Keras TensorFlow":
            #target filename should be like source +_frozen.pb
            path_new = os.path.splitext(path)[0] + "_frozen.pb"
            aid_dl.convert_kerastf_2_frozen_pb(path,path_new)
            text = "Conversion Keras TensorFlow -> Frozen .pb is done"
            conversion_successful_msg(text)
        ##################Keras TensorFlow -> Optimized .pb####################
        elif target_format=="Optimized TensorFlow .pb" and source_format=="Keras TensorFlow":
            path_new = os.path.splitext(path)[0] + "_optimized.pb"
            aid_dl.convert_kerastf_2_optimized_pb(path,path_new)
            text = "Conversion Keras TensorFlow -> Optimized .pb is done"
            #conversion_successful_msg(text)

        ####################Frozen -> Optimized .pb############################
        elif target_format=="Optimized TensorFlow .pb" and source_format=="Frozen TensorFlow .pb":
            path_new = os.path.splitext(path)[0] + "_optimized.pb"
            aid_dl.convert_frozen_2_optimized_pb(path,path_new)
            text = "Conversion Frozen -> Optimized .pb is done"
            #conversion_successful_msg(text)

        ##################Keras TensorFlow -> ONNX####################
        elif target_format=="ONNX (via tf2onnx)" and source_format=="Keras TensorFlow":
            path_new = os.path.splitext(path)[0] + ".onnx"
            aid_dl.convert_kerastf_2_onnx(path,path_new)
            text = "Conversion Keras TensorFlow -> ONNX (via tf2onnx) is done"
            conversion_successful_msg(text)

        ##################Keras TensorFlow -> ONNX via MMdnn####################
        elif target_format=="ONNX (via MMdnn)" and source_format=="Keras TensorFlow":
            aid_dl.convert_kerastf_2_onnx_mmdnn(path)
            text = "Conversion Keras TensorFlow -> ONNX (via MMdnn) is done"
            conversion_successful_msg(text)

        ##################Keras TensorFlow -> PyTorch Script####################
        elif target_format=="PyTorch Script"  and source_format=="Keras TensorFlow":
            aid_dl.convert_kerastf_2_script(path,"pytorch")
            text = "Conversion Keras TensorFlow -> PyTorch Script is done. You can now use this script and the saved weights to build the model using your PyTorch installation."
            conversion_successful_msg(text)

        ##################Keras TensorFlow -> Caffe Script####################
        elif target_format=="Caffe Script" and source_format=="Keras TensorFlow":
            aid_dl.convert_kerastf_2_script(path,"caffe")
            text = "Conversion Keras TensorFlow -> Caffe Script is done. You can now use this script and the saved weights to build the model using your Caffe installation."
            conversion_successful_msg(text)

        ##################Keras TensorFlow -> CNTK Script####################
        elif target_format=="CNTK Script" and source_format=="Keras TensorFlow":
            aid_dl.convert_kerastf_2_script(path,"cntk")
            text = "Conversion Keras TensorFlow -> CNTK Script is done. You can now use this script and the saved weights to build the model using your CNTK installation."
            conversion_successful_msg(text)

        ##################Keras TensorFlow -> mxnet Script####################
        elif target_format=="MXNet Script" and source_format=="Keras TensorFlow":
            aid_dl.convert_kerastf_2_script(path,"mxnet")
            text = "Conversion Keras TensorFlow -> MXNet Script is done. You can now use this script and the saved weights to build the model using your MXNet installation."
            conversion_successful_msg(text)

        ##################Keras TensorFlow -> onnx Script####################
        elif target_format=="ONNX Script" and source_format=="Keras TensorFlow":
            aid_dl.convert_kerastf_2_script(path,"onnx")
            text = "Conversion Keras TensorFlow -> ONNX Script is done. You can now use this script and the saved weights to build the model using your ONNX installation."
            conversion_successful_msg(text)

        ##################Keras TensorFlow -> TensorFlow Script####################
        elif target_format=="TensorFlow Script" and source_format=="Keras TensorFlow":
            aid_dl.convert_kerastf_2_script(path,"tensorflow")
            text = "Conversion Keras TensorFlow -> TensorFlow Script is done. You can now use this script and the saved weights to build the model using your Tensorflow installation."
            conversion_successful_msg(text)

        ##################Keras TensorFlow -> Keras Script####################
        elif target_format=="Keras Script" and source_format=="Keras TensorFlow":
            aid_dl.convert_kerastf_2_script(path,"keras")
            text = "Conversion Keras TensorFlow -> Keras Script is done. You can now use this script and the saved weights to build the model using your Keras installation."
            conversion_successful_msg(text)

        ##################Keras TensorFlow -> CoreML####################
        elif "CoreML" in target_format and source_format=="Keras TensorFlow":
            aid_dl.convert_kerastf_2_coreml(path)
            text = "Conversion Keras TensorFlow -> CoreML is done."
            conversion_successful_msg(text)

        else:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("Not implemeted (yet)")
            msg.setWindowTitle("Not implemeted (yet)")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
                
        #If that worked without error, save the filepath for next time
        Default_dict["Path of last model"] = os.path.split(path)[0]
        aid_bin.save_aid_settings(Default_dict)


        
    def history_tab_convertModel_nnet_worker(self,ConvertToNnet,progress_callback,history_callback):
        #Define a new session -> Necessary for threading in TensorFlow
        with tf.compat.v1.Session() as sess:
            sess.run(tf.compat.v1.global_variables_initializer()) 
            path = self.model_2_convert
            try:
                model_keras = load_model(path,custom_objects=aid_dl.get_custom_metrics())
            except:
                model_keras = load_model(path)
                
            dic = {"model_keras":model_keras}
            history_callback.emit(dic)
            progress_callback.emit(1)
    
            if ConvertToNnet==1:
                #Since this happened in a thread, TensorFlow cant access it anywhere else
                #Therefore perform Conversion to nnet right away:
                model_config = model_keras.get_config()#["layers"]
                if type(model_config)==dict:
                    model_config = model_config["layers"]#for keras version>2.2.3, there is a change in the output of get_config()
                #Convert model to theano weights format (Only necesary for CNNs)
                for layer in model_keras.layers:
                   if "conv" in layer.__class__.__name__.lower():# in ['Convolution1D', 'Convolution2D']:
                      original_w = K.get_value(layer.W)
                      converted_w = aid_dl.convert_kernel(original_w)
                      K.set_value(layer.W, converted_w)
                
                nnet_path, nnet_filename = os.path.split(self.model_2_convert)
                nnet_filename = nnet_filename.split(".model")[0]+".nnet" 
                out_path = os.path.join(nnet_path,nnet_filename)
                
                #the Input layer seems to be missing for newer versions of keras
                layers = [layer.__class__.__name__ for layer in model_keras.layers]
                if "input" not in layers[0].lower():
                    model_config = model_config[1:]

                aid_dl.dump_to_simple_cpp(model_keras=model_keras,model_config=model_config,output=out_path,verbose=False)
                            
#            sess.close()
#        try:
#            aid_dl.reset_keras()
#        except:
#            print("Could not reset Keras (1)")

        
    def history_tab_ConvertToNnet(self):
        print("Not used")
#        model_keras = self.model_keras
#        model_config = model_keras.get_config()["layers"]
#        #Convert model to theano weights format (Only necesary for CNNs)
#        for layer in model_keras.layers:
#           if layer.__class__.__name__ in ['Convolution1D', 'Convolution2D']:
#              original_w = K.get_value(layer.W)
#              converted_w = convert_kernel(original_w)
#              K.set_value(layer.W, converted_w)
#        
#        nnet_path, nnet_filename = os.path.split(self.model_2_convert)
#        nnet_filename = nnet_filename.split(".model")[0]+".nnet" 
#        out_path = os.path.join(nnet_path,nnet_filename)
#        aid_dl.dump_to_simple_cpp(model_keras=model_keras,model_config=model_config,output=out_path,verbose=False)
#        msg = QtWidgets.QMessageBox()
#        msg.setIcon(QtWidgets.QMessageBox.Information)       
#        msg.setText("Successfully converted model and saved to\n"+out_path)
#        msg.setWindowTitle("Successfully converted model!")
#        msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
#        msg.exec_()
#        self.pushButton_convertModel.setEnabled(False)
#TODO
    def test_nnet(self):
        #I need a function which calls a cpp app that uses the nnet and applies
        #it on a random image.
        #The same image is also used as input the the original .model and 
        #both results are then compared
        print("Not implemented yet")
        print("Placeholder")
        print("Building site")


    def actionDocumentation_function(self):        
        icon = QtGui.QImage(os.path.join(dir_root,"art",Default_dict["Icon theme"],"main_icon_simple_04_256"+icon_suff))
        icon = QtGui.QPixmap(icon).scaledToHeight(32, QtCore.Qt.SmoothTransformation)
        msg = QtWidgets.QMessageBox()
        msg.setIconPixmap(icon)
        text = "Currently, there is no detailed written documentation. AIDeveloper instead makes strong use of tooltips."
        text = "<html><head/><body><p>"+text+"</p></body></html>"
        msg.setText(text)
        msg.setWindowTitle("Documentation")
        msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
        msg.exec_()

    def actionSoftware_function(self):
        if sys.platform == "win32":
            plat = "win"
        elif sys.platform=="darwin":
            plat = "mac"      
        elif sys.platform=="linux":
            plat = "linux"
        else:
            print("Unknown Operating system")
            plat = "Win"
            
        dir_deps = os.path.join(dir_root,"aid_dependencies_"+plat+".txt")#dir to aid_dependencies
        f = open(dir_deps, "r")
        text_modules = f.read()
        f.close()
        icon = QtGui.QImage(os.path.join(dir_root,"art",Default_dict["Icon theme"],"main_icon_simple_04_256"+icon_suff))
        icon = QtGui.QPixmap(icon).scaledToHeight(32, QtCore.Qt.SmoothTransformation)
        msg = QtWidgets.QMessageBox()
        msg.setIconPixmap(icon)
        text = "<html><head/><body><p>AIDeveloper "+str(VERSION)+"<br>"+sys.version+"<br>Click 'Show Details' to retrieve a list of all Python packages used."+"<br>AID_GPU uses CUDA (NVIDIA) to facilitate GPU processing</p></body></html>"
        msg.setText(text)
        msg.setDetailedText(text_modules)
        msg.setWindowTitle("Software")
        msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
        msg.exec_()
        
    def actionAbout_function(self):
        icon = QtGui.QImage(os.path.join(dir_root,"art",Default_dict["Icon theme"],"main_icon_simple_04_256"+icon_suff))
        icon = QtGui.QPixmap(icon).scaledToHeight(32, QtCore.Qt.SmoothTransformation)
        msg = QtWidgets.QMessageBox()
        msg.setIconPixmap(icon)
        text = "AIDeveloper is written and maintained by Maik Herbig. Use maik.herbig@tu-dresden.de to contact the main developer if you find bugs or if you wish a particular feature. Icon theme 2 was mainly designed and created by Konrad Wauer."
        text = "<html><head/><body><p>"+text+"</p></body></html>"
        msg.setText(text)
        msg.setWindowTitle("About")
        msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
        msg.exec_()

    def actionLoadSession_function(self):
        #This function should allow to select and load a metafile and
        #Put the GUI the the corresponing state (place the stuff in the table, click Train/Valid)
        filename = QtWidgets.QFileDialog.getOpenFileName(self, 'Open meta-data', Default_dict["Path of last model"],"AIDeveloper Meta or session file (*meta.xlsx *session.xlsx)")
        filename = filename[0]
        if len(filename)==0:
            return
        xlsx = pd.ExcelFile(filename, engine='openpyxl')
        UsedData = pd.read_excel(xlsx,sheet_name="UsedData",engine="openpyxl")
        Files = list(UsedData["rtdc_path"])

        file_exists = [os.path.exists(url) for url in Files]
        ind_true = np.where(np.array(file_exists)==True)[0]
        UsedData_true = UsedData.iloc[ind_true]
        
        Files_true = list(UsedData_true["rtdc_path"]) #select the indices that are valid

        #Add stuff to table_dragdrop
        rowPosition = int(self.table_dragdrop.rowCount())
        self.dataDropped(Files_true)

        #update the index,  train/valid checkbox and shuffle checkbox
        for i in range(len(Files_true)):
            #set the index (celltype)
            try:
                index = int(np.array(UsedData_true["class"])[i])
            except:
                index = int(np.array(UsedData_true["index"])[i])
                print("You are using an old version of AIDeveloper. Consider upgrading")

            self.table_dragdrop.cellWidget(rowPosition+i, 1).setValue(index)
            #is it checked for train or valid?
            trorvalid = str(np.array(UsedData_true["TrainOrValid"])[i])
            if trorvalid=="Train":
                self.table_dragdrop.item(rowPosition+i, 2).setCheckState(QtCore.Qt.Checked)
            elif trorvalid=="Valid":
                self.table_dragdrop.item(rowPosition+i, 3).setCheckState(QtCore.Qt.Checked)

            #how many cells/epoch during training or validation?
            try:
                nr_events_epoch = str(np.array(UsedData_true["nr_events_epoch"])[i])  
            except:
                nr_events_epoch = str(np.array(UsedData_true["nr_cells_epoch"])[i])  

            self.table_dragdrop.item(rowPosition+i, 6).setText(nr_events_epoch)
            #Shuffle or not?
            shuffle = bool(np.array(UsedData_true["shuffle"])[i])

            if shuffle==False:
                self.table_dragdrop.item(rowPosition+i, 8).setCheckState(QtCore.Qt.Unchecked)
                #Set Cells/Epoch to not editable
                item = self.table_dragdrop.item(rowPosition+i, 6)
                item.setFlags(item.flags() &~QtCore.Qt.ItemIsEnabled &~ QtCore.Qt.ItemIsSelectable )
                #item.setFlags(item.flags() |QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable )
            else:
                self.table_dragdrop.item(rowPosition+i, 8).setCheckState(QtCore.Qt.Checked)


            #zoom_factor = float(np.array(UsedData_true["zoom_factor"])[i])
            zoom_factor = str(np.array(UsedData_true["zoom_factor"])[i])  
            self.table_dragdrop.item(rowPosition+i, 9).setText(zoom_factor)

      
        #Now take care of missing data
        #Take care of missing files (they might have been moved to a different location)
        ind_false = np.where(np.array(file_exists)==False)[0]  
        #Files_false = list(UsedData_false["rtdc_path"]) #select the indices that are valid
        if len(ind_false)>0:
            UsedData_false = UsedData.iloc[ind_false]
            Files_false = list(UsedData_false["rtdc_path"]) #select the indices that are valid
            self.dataDropped(Files_false)

            self.user_selected_path = None
            #Create popup that informs user that there is missing data and let him specify a location
            #to search for the missing files
            def add_missing_files():
                filename = QtWidgets.QFileDialog.getExistingDirectory(self, 'Select directory', Default_dict["Path of last model"])
                user_selected_path = filename
                if len(filename)==0:
                    msg = QtWidgets.QMessageBox()
                    msg.setIcon(QtWidgets.QMessageBox.Information)       
                    msg.setText("Invalid directory")
                    msg.setWindowTitle("Invalid directory")
                    msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                    msg.exec_()
                    return
                #get the hashes 
                hashes = list(np.array(UsedData_false["hash"])[ind_false])
                paths = list(np.array(UsedData_false["rtdc_path"])[ind_false])
                    
                paths_new,info = aid_bin.find_files(user_selected_path,paths,hashes)
                text = ('\n'.join([str(a) +"\t"+ b for a,b in zip(paths_new,info)]))
                self.textBrowser_Info_pop2.setText(text)

                #Add stuff to table_dragdrop
                rowPosition = int(self.table_dragdrop.rowCount())
                self.dataDropped(paths_new)
                
                for i in range(len(paths_new)):
                    #set the index (celltype)
                    try:
                        index = int(np.array(UsedData_false["class"])[i])
                    except:
                        index = int(np.array(UsedData_false["index"])[i])
                        print("You are using an old version of AIDeveloper. Consider upgrading")

                    self.table_dragdrop.cellWidget(rowPosition+i, 1).setValue(index)
                    #is it checked for train or valid?
                    trorvalid = str(np.array(UsedData_false["TrainOrValid"])[i])
                    if trorvalid=="Train":
                        self.table_dragdrop.item(rowPosition+i, 2).setCheckState(QtCore.Qt.Checked)
                    elif trorvalid=="Valid":
                        self.table_dragdrop.item(rowPosition+i, 3).setCheckState(QtCore.Qt.Checked)
                    #how many cells/epoch during training or validation?
                    nr_events_epoch = str(np.array(UsedData_false["nr_events_epoch"])[i]) 

                    #how many cells/epoch during training or validation?
                    try:
                        nr_events_epoch = str(np.array(UsedData_false["nr_events_epoch"])[i])  
                    except:
                        nr_events_epoch = str(np.array(UsedData_false["nr_cells_epoch"])[i])  
                        print("You are using an old version of AIDeveloper. Consider upgrading")

                    self.table_dragdrop.item(rowPosition+i, 6).setText(nr_events_epoch)
                    #Shuffle or not?
                    shuffle = bool(np.array(UsedData_false["shuffle"])[i])
                    if shuffle==False:
                        self.table_dragdrop.item(rowPosition+i, 8).setCheckState(QtCore.Qt.Unchecked)
                        #Set Cells/Epoch to not editable
                        item = self.table_dragdrop.item(rowPosition+i, 6)
                        item.setFlags(item.flags() &~QtCore.Qt.ItemIsEnabled &~ QtCore.Qt.ItemIsSelectable )
                        #item.setFlags(item.flags() |QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable )
                    else:
                        self.table_dragdrop.item(rowPosition+i, 8).setCheckState(QtCore.Qt.Checked)
                    #zoom_factor = float(np.array(UsedData_false["zoom_factor"])[i])
                    zoom_factor = str(np.array(UsedData_false["zoom_factor"])[i])  
                    self.table_dragdrop.item(rowPosition+i, 9).setText(zoom_factor)

            self.w_pop2 = MyPopup()
            self.gridLayout_w_pop2 = QtWidgets.QGridLayout(self.w_pop2)
            self.gridLayout_w_pop2.setObjectName("gridLayout_w_pop2")
            self.verticalLayout_w_pop2 = QtWidgets.QVBoxLayout()
            self.verticalLayout_w_pop2.setObjectName("verticalLayout_w_pop2")
            self.horizontalLayout_w_pop2 = QtWidgets.QHBoxLayout()
            self.horizontalLayout_w_pop2.setObjectName("horizontalLayout_w_pop2")
            self.pushButton_Close_pop2 = QtWidgets.QPushButton(self.centralwidget)
            self.pushButton_Close_pop2.setObjectName("pushButton_Close_pop2")
            self.pushButton_Close_pop2.clicked.connect(self.w_pop2.close)
            self.horizontalLayout_w_pop2.addWidget(self.pushButton_Close_pop2)
            self.pushButton_Search_pop2 = QtWidgets.QPushButton(self.centralwidget)
            self.pushButton_Search_pop2.clicked.connect(add_missing_files)
            self.pushButton_Search_pop2.setObjectName("pushButton_Search")
            self.horizontalLayout_w_pop2.addWidget(self.pushButton_Search_pop2)
            self.verticalLayout_w_pop2.addLayout(self.horizontalLayout_w_pop2)
            self.textBrowser_Info_pop2 = QtWidgets.QTextBrowser(self.centralwidget)
            self.textBrowser_Info_pop2.setObjectName("textBrowser_Info_pop2")
            self.verticalLayout_w_pop2.addWidget(self.textBrowser_Info_pop2)
            self.gridLayout_w_pop2.addLayout(self.verticalLayout_w_pop2, 0, 0, 1, 1)
            self.w_pop2.setWindowTitle("There are missing files. Do you want to search for them?")
            self.pushButton_Close_pop2.setText("No")
            self.pushButton_Search_pop2.setText("Define folder to search files")
            self.w_pop2.show()

        
        #Ask user if only data, or the full set of parameters should be loaded
        msg = QtWidgets.QMessageBox()
        msg.setIcon(QtWidgets.QMessageBox.Question)
        msg.setText(tooltips["msg_loadSession"])
        msg.setWindowTitle("Load only data table all parameters?")
        msg.setStandardButtons(QtGui.QMessageBox.Yes | QtGui.QMessageBox.Save)# | QtGui.QMessageBox.Cancel)
        dataonly = msg.button(QtGui.QMessageBox.Yes)
        dataonly.setText('Data table only')
        allparams = msg.button(QtGui.QMessageBox.Save)
        allparams.setText('Data and all parameters')
#        cancel = msg.button(QtGui.QMessageBox.Cancel)
#        cancel.setText('Cancel')
        msg.exec_()        
 
        #Only update the data table.
        if msg.clickedButton()==dataonly: #show image and heatmap overlay
            pass
        #Load the parameters
        elif msg.clickedButton()==allparams: #show image and heatmap overlay
            Parameters = pd.read_excel(xlsx,sheet_name="Parameters",engine="openpyxl")
            aid_frontend.load_hyper_params(self,Parameters)
#        if msg.clickedButton()==cancel: #show image and heatmap overlay
#            return


        #If all this run without error, save the path.
        Default_dict["Path of last model"] = os.path.split(filename)[0]
        aid_bin.save_aid_settings(Default_dict)
        
        #Update the overview-box
        if self.groupBox_DataOverview.isChecked()==True:
            self.dataOverviewOn()        

    def actionSaveSession_function(self):
        filename = QtWidgets.QFileDialog.getSaveFileName(self, 'Save session', Default_dict["Path of last model"],"AIDeveloper Session file (*_session.xlsx)")
        filename = filename[0]
        path, fname = os.path.split(filename)
                
        if len(fname)==0:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("No valid filename was chosen.")
            msg.setWindowTitle("No valid filename was chosen")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return
        if fname.endswith(".xlsx"):
            fname = fname.split(".xlsx")[0]            
        if fname.endswith("_session"):
            fname = fname.split("_session")[0]            
        if fname.endswith("_meta"):
            fname = fname.split("_meta")[0]
        if fname.endswith(".model"):
            fname = fname.split(".model")[0]
        if fname.endswith(".arch"):
            fname = fname.split(".arch")[0]

        #add the suffix _session.xlsx
        if not fname.endswith("_session.xlsx"):
            fname = fname +"_session.xlsx"
        filename = os.path.join(path,fname)    
        
        writer = pd.ExcelWriter(filename, engine='openpyxl')
        #Used files go to a separate sheet on the -session.xlsx
        SelectedFiles = self.items_clicked()
        SelectedFiles_df = pd.DataFrame(SelectedFiles)
        pd.DataFrame().to_excel(writer,sheet_name='UsedData') #initialize empty Sheet
        SelectedFiles_df.to_excel(writer,sheet_name='UsedData')
        DataOverview_df = self.get_dataOverview()
        DataOverview_df.to_excel(writer,sheet_name='DataOverview') #write data overview to separate sheet            

        #Get all hyper parameters
        Para_dict = pd.DataFrame()
        Para_dict["AIDeveloper_Version"]=VERSION,
        Para_dict["model_zoo_version"]=model_zoo_version,
        try:
            Para_dict["OS"]=platform.platform(),
            Para_dict["CPU"]=platform.processor(),
        except:
            Para_dict["OS"]="Unknown",
            Para_dict["CPU"]="Unknown",
        Para_dict = aid_frontend.get_hyper_params(Para_dict,self)
        Para_dict.to_excel(writer,sheet_name='Parameters')
        
        writer.save()
        writer.close()

        msg = QtWidgets.QMessageBox()
        msg.setIcon(QtWidgets.QMessageBox.Information)       
        msg.setText("Successfully saved as "+filename)
        msg.setWindowTitle("Successfully saved")
        msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
        msg.exec_()
        
        #If all that run without issue, remember the path for next time
        Default_dict["Path of last model"] = os.path.split(filename)[0]        
        aid_bin.save_aid_settings(Default_dict)
        
    def actionClearList_function(self):
        #Remove all items from dragdrop table
        while (self.table_dragdrop.rowCount() > 0):
            self.table_dragdrop.removeRow(0)
        #reset ram
        self.ram = dict()

        #Remove all items from comboBox_chooseRtdcFile
        self.comboBox_chooseRtdcFile.clear()
        self.comboBox_selectData.clear()
        if self.groupBox_DataOverview.isChecked()==True:
            self.dataOverviewOn()

    def actionRemoveSelected_function(self):
        #Which rows are highlighted?
        rows_selected = np.array([index.row() for index in self.table_dragdrop.selectedIndexes()])
        for row in (rows_selected):
            self.table_dragdrop.removeRow(row)
            self.comboBox_chooseRtdcFile.removeItem(row)
            self.comboBox_selectData.removeItem(row)
            #if there are rows below this row, they will move up one step:
            ind = np.where(np.array(rows_selected)>row)[0]
            rows_selected[ind] -= 1

    def actionSaveToPng_function(self):
        #Which table items are selected?
        rows_selected = np.array([index.row() for index in self.table_dragdrop.selectedIndexes()])
        if len(rows_selected)==0:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Warning)       
            msg.setText("Please first select rows in the table!")
            msg.setWindowTitle("No rows selected")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return
        #Ask user to which folder the images should be written:
        filename = QtWidgets.QFileDialog.getSaveFileName(self, 'Save to .png/.jpg', Default_dict["Path of last model"],"Image file format (*.png *.jpg *.bmp *.eps *.gif *.ico *.icns)")
        filename = filename[0]
        if len(filename)==0:
            return
        filename_X, file_extension = os.path.splitext(filename)#divide into path and file_extension if possible
        #Check if the chosen file_extension is valid
        if not file_extension in [".png",".jpg",".bmp",".eps",".gif",".ico",".icns"]:
            print("Invalid file extension detected. Will use .png instead.")
            file_extension = ".png"
            
        #Check the chosen export-options
        if bool(self.actionExport_Original.isChecked())==True:
            print("Export original images")
            save_cropped = False
        elif bool(self.actionExport_Cropped.isChecked())==True:
            print("Export cropped images")
            save_cropped = True
        elif bool(self.actionExport_Off.isChecked())==True:
            print("Exporting is turned off")
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("Plase choose a different Export-option in ->Options->Export")
            msg.setWindowTitle("Export is turned off!")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return

        if save_cropped==True:
            #Collect information for image processing
            cropsize = self.spinBox_imagecrop.value()
            color_mode = str(self.comboBox_loadedRGBorGray.currentText())
            #zoom_methods = [self.actionOrder0.isChecked(),self.actionOrder1.isChecked(),self.actionOrder2.isChecked(),self.actionOrder3.isChecked(),self.actionOrder4.isChecked(),self.actionOrder5.isChecked()]
            #zoom_order = np.where(np.array(zoom_methods)==True)[0]
            zoom_order = int(self.comboBox_zoomOrder.currentIndex()) #the combobox-index is already the zoom order

            index = 0
            for row in (rows_selected):
                #get the corresponding rtdc_path
                rtdc_path = str(self.table_dragdrop.cellWidget(row, 0).text())
                nr_events = None #no number needed as we take all images (replace=False in gen_crop_img)
                zoom_factor = float(self.table_dragdrop.item(row, 9).text())            
                gen = aid_img.gen_crop_img(cropsize,rtdc_path,nr_events=nr_events,replace=False,random_images=False,zoom_factor=zoom_factor,zoom_order=zoom_order,color_mode=color_mode,padding_mode='constant')
                images = next(gen)[0]
                #Save the images data to .png/.jpeg...
                for img in images:
                    img = PIL.Image.fromarray(img)
                    img.save(filename_X+"_"+str(index)+file_extension)
                    index+=1

        if save_cropped==False:#save the original images without pre-processing
            index = 0
            for row in (rows_selected):
                rtdc_path = str(self.table_dragdrop.cellWidget(row, 0).text())
                failed,rtdc_ds = aid_bin.load_rtdc(rtdc_path)
                if failed:
                    msg = QtWidgets.QMessageBox()
                    msg.setIcon(QtWidgets.QMessageBox.Critical)       
                    msg.setText(str(rtdc_ds))
                    msg.setWindowTitle("Error occurred during loading file")
                    msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                    msg.exec_()
                    return
    
                images = rtdc_ds["events"]["image"] #get the images
                #Save the images data to .png/.jpeg...
                for img in images:
                    img = PIL.Image.fromarray(img)
                    img.save(filename_X+"_"+str(index)+file_extension)
                    index+=1
                    
        #If all that run without issue, remember the path for next time
        Default_dict["Path of last model"] = os.path.split(filename)[0]
        aid_bin.save_aid_settings(Default_dict)

    def actionRemoveSelectedPeaks_function(self):
        #Which rows are highlighted?
        rows_selected = np.array([index.row() for index in self.tableWidget_showSelectedPeaks.selectedIndexes()])
        #delete each row only once :)
        rows_selected = np.unique(rows_selected)
        for row in (rows_selected):
            self.tableWidget_showSelectedPeaks.removeRow(row)
            #if there are rows below this row, they will move up one step:
            ind = np.where(np.array(rows_selected)>row)[0]
            rows_selected[ind] -=1
        #Update the widget_showSelectedPeaks
        self.update_peak_plot()


    def actionRemoveAllPeaks_function(self):
        #Remove all items from tableWidget_showSelectedPeaks
        while (self.tableWidget_showSelectedPeaks.rowCount() > 0):
            self.tableWidget_showSelectedPeaks.removeRow(0)

    def actionDataToRamNow_function(self):
        self.statusbar.showMessage("Moving data to RAM")
        #check that the nr. of classes are equal to the model out put
        SelectedFiles = self.items_clicked()
        color_mode = self.get_color_mode()
        zoom_factors = [selectedfile["zoom_factor"] for selectedfile in SelectedFiles]
        #zoom_order = [self.actionOrder0.isChecked(),self.actionOrder1.isChecked(),self.actionOrder2.isChecked(),self.actionOrder3.isChecked(),self.actionOrder4.isChecked(),self.actionOrder5.isChecked()]
        #zoom_order = int(np.where(np.array(zoom_order)==True)[0])
        zoom_order = int(self.comboBox_zoomOrder.currentIndex()) #the combobox-index is already the zoom order

        #Get the user-defined cropping size
        crop = int(self.spinBox_imagecrop.value())          
        #Make the cropsize a bit larger since the images will later be rotated
        cropsize2 = np.sqrt(crop**2+crop**2)
        cropsize2 = np.ceil(cropsize2 / 2.) * 2 #round to the next even number
        
        dic = aid_img.crop_imgs_to_ram(list(SelectedFiles),crop,zoom_factors=zoom_factors,zoom_order=zoom_order,color_mode=color_mode)
        self.ram = dic 
        
        msg = QtWidgets.QMessageBox()
        msg.setIcon(QtWidgets.QMessageBox.Information)       
        msg.setText("Successfully moved data to RAM")
        msg.setWindowTitle("Moved Data to RAM")
        msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
        msg.exec_()

        self.statusbar.showMessage("")







    ###########################################################################
    ###########################################################################
    ###########################################################################
    ###########################################################################
    #######################Functions for Assess model tab######################
    def assessmodel_tab_load_model(self):
        #Get the requested model-name from the chosen metafile
        #Open a QFileDialog
        filename = QtWidgets.QFileDialog.getOpenFileName(self, 'Select a trained model you want to assess', Default_dict["Path of last model"],"Keras Model file (*.model)")
        filename = filename[0]        
        if os.path.isfile(filename):
            #Put this path on the Assess Model tab
            self.lineEdit_LoadModel_2.setText(filename)
            #Save the path to a variable that is then used by history_tab_convertModel_nnet_worker
            self.load_model_path = filename
            
            #Get the modelindex
            path,filename = os.path.split(filename)
            modelindex = filename.split(".model")[0]
            modelindex = int(modelindex.split("_")[-1])
            #Update the modelindex on the Assess Model tab
            self.spinBox_ModelIndex_2.setValue(int(modelindex))

            model_full_h5 = h5py.File(self.load_model_path, 'r')
            model_config = model_full_h5.attrs['model_config']
            model_full_h5.close() #close the hdf5                
            model_config = json.loads(model_config)
            
            in_dim, out_dim = aid_dl.model_in_out_dim(model_config,"config")  
            
            self.spinBox_Crop_2.setValue(int(in_dim[-2]))
            self.spinBox_OutClasses_2.setValue(int(out_dim))
            print("input dimension:"+str(in_dim))
            #Adjust the Color mode in the UI:
            channels = in_dim[-1] #TensorFlow: channels in last dimension
            if channels==1:
                #Set the combobox on Assess model tab to Grayscale; just info for user
                index = self.comboBox_loadedRGBorGray.findText("Grayscale", QtCore.Qt.MatchFixedString)
                if index >= 0:
                    self.comboBox_loadedRGBorGray.setCurrentIndex(index)                                        
                #Check the currently set color_mode. This is important since images are loaded accordingly
                if self.get_color_mode()!="Grayscale":
                    #when model needs Grayscale, set the color mode in comboBox_GrayOrRGB to that
                    index = self.comboBox_GrayOrRGB.findText("Grayscale", QtCore.Qt.MatchFixedString)
                    if index >= 0:
                        self.comboBox_GrayOrRGB.setCurrentIndex(index)                                                            
                    self.statusbar.showMessage("Color Mode set to Grayscale",5000)
            
            elif channels==3:
                #Set the combobox on Assess model tab to Grayscale; just info for user
                index = self.comboBox_loadedRGBorGray.findText("RGB", QtCore.Qt.MatchFixedString)
                if index >= 0:
                    self.comboBox_loadedRGBorGray.setCurrentIndex(index)                                        
                #Check the currently set color_mode. This is important since images are loaded accordingly
                if self.get_color_mode()!="RGB":
                    #when model needs RGB, set the color mode in comboBox_GrayOrRGB to that
                    index = self.comboBox_GrayOrRGB.findText("RGB", QtCore.Qt.MatchFixedString)
                    if index >= 0:
                        self.comboBox_GrayOrRGB.setCurrentIndex(index)
                    self.statusbar.showMessage("Color Mode set to RGB",5000)
            else:
                msg = QtWidgets.QMessageBox()
                msg.setIcon(QtWidgets.QMessageBox.Information)       
                msg.setText("Channel dimensions of model ("+str(channels)+" channels) is not supported. Only 1 or 3 channels are allowed.")
                msg.setWindowTitle("Unsupported channel dimension")
                msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                msg.exec_()
                return

            modelindex = int(self.spinBox_ModelIndex_2.value())
            path,fname = os.path.split(self.load_model_path)    
            fname = fname.split(str(modelindex)+".model")[0]+"meta.xlsx"
            metafile_path = os.path.join(path,fname)

            try:
                img_processing_settings = aid_img.load_model_meta(metafile_path)
                self.img_processing_settings = img_processing_settings
                model_type = str(img_processing_settings["model_type"].values[0])                
                normalization_method = str(img_processing_settings["normalization_method"].values[0])
                
                index = self.comboBox_Normalization_2.findText(normalization_method, QtCore.Qt.MatchFixedString)
                if index >= 0:
                    self.comboBox_Normalization_2.setCurrentIndex(index)
                else:
                    msg = QtWidgets.QMessageBox()
                    msg.setIcon(QtWidgets.QMessageBox.Information)       
                    msg.setText("Unkown normalization method found in .meta file")
                    msg.setWindowTitle("Unkown normalization method")
                    msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                    msg.exec_()
                    return
                    
                self.lineEdit_ModelSelection_2.setText(model_type)

            except: #there is not such a file, or the file cannot be opened
                #Ask the user to choose the normalization method
                self.lineEdit_ModelSelection_2.setText("Unknown")
                self.comboBox_Normalization_2.setEnabled(True)
                msg = QtWidgets.QMessageBox()
                msg.setIcon(QtWidgets.QMessageBox.Information)       
                msg.setText("Meta file not found/ Could not be read. Please specify the normalization method manually (dropdown menu)")
                msg.setWindowTitle("No .meta available")
                msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                msg.exec_()
            
            #If all that run without issue, remember the path for next time
            Default_dict["Path of last model"] = os.path.split(self.load_model_path)[0]
            aid_bin.save_aid_settings(Default_dict)

        else:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("File not found!:\nProbably the .model was deleted or not saved")
            msg.setWindowTitle("File not found")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return
        
    
    def inference_time_worker(self,progress_callback,history_callback):
        if self.radioButton_cpu.isChecked():
            deviceSelected = str(self.comboBox_cpu.currentText())
        elif self.radioButton_gpu.isChecked():
            deviceSelected = str(self.comboBox_gpu.currentText())
        gpu_memory = float(self.doubleSpinBox_memory.value())

        #Create config (define which device to use)
        config_gpu = aid_dl.get_config(cpu_nr,gpu_nr,deviceSelected,gpu_memory)

        #Retrieve more Multi-GPU Options from Menubar:
        cpu_merge = bool(self.actioncpu_merge.isEnabled())
        cpu_relocation = bool(self.actioncpu_relocation.isEnabled())
        cpu_weight_merge = bool(self.actioncpu_weightmerge.isEnabled())    


        #Initiate a fresh session
        with tf.compat.v1.Session(graph = tf.Graph(), config=config_gpu) as sess:
            sess.run(tf.compat.v1.global_variables_initializer())
            #Baustelle
            if deviceSelected=="Multi-GPU" and cpu_weight_merge==True:
                strategy = tf.distribute.MirroredStrategy()
                with tf.device("/cpu:0"):
                    with strategy.scope():
                        model_keras = load_model(self.load_model_path,custom_objects=aid_dl.get_custom_metrics()) 
            else:
                model_keras = load_model(self.load_model_path,custom_objects=aid_dl.get_custom_metrics())
        
            #Multi-GPU
            if deviceSelected=="Multi-GPU":
                print("Adjusting the model for Multi-GPU")
                model_keras = model_keras#multi_gpu_model(model_keras, gpus=gpu_nr, cpu_merge=cpu_merge, cpu_relocation=cpu_relocation)#indicate the numbers of gpus that you have
            
            #Get the model input dimensions
            in_dim, _ = aid_dl.model_in_out_dim(model_keras,"model")
            in_dim = np.array(in_dim)
            ind = np.where(in_dim==None)
            in_dim[ind] = 1
            nr_imgs = self.spinBox_inftime_nr_images.value()
            nr_imgs = int(np.round(float(nr_imgs)/10.0))
            
            #Warm up by predicting a single image
            image = (np.random.randint(0,255,size=in_dim)).astype(np.float32)/255.0
            model_keras.predict(image) # warm up
    
            Times = []
            for k in range(10):
                image = (np.random.randint(0,255,size=in_dim)).astype(np.float32)/255.0
                t1 = time.perf_counter()
                for i in range(nr_imgs):#predict 50 times 20 images
                    model_keras.predict(image)
                t2 = time.perf_counter()
                dt = (t2-t1)/(nr_imgs) #divide by nr_imgs to get time [s] per image
                dt = dt*1000.0 #multiply by 1000 to change to ms range
                dic = {"outp":str(round(dt,3))+"ms"}
                history_callback.emit(dic)
                Times.append(dt)   
        
        
        #Send out the Times
        text = " [ms] Mean: "+str(round(np.mean(Times),3))+"; "+"Median: "+str(round(np.median(Times),3))+"; "+"Min: "+str(round(np.min(Times),3))+"; "+"Max: "+str(round(np.max(Times),3))
        dic = {"outp":text}
        history_callback.emit(dic)
        progress_callback.emit(1) #when finished return one
        self.threadpool_single_queue = 0 #reset the thread-counter

    def inference_time(self):
        if self.radioButton_cpu.isChecked():
            deviceSelected = str(self.comboBox_cpu.currentText())
        elif self.radioButton_gpu.isChecked():
            deviceSelected = str(self.comboBox_gpu.currentText())

        #Inform user that certain config is used for inference
        msg = QtWidgets.QMessageBox()
        msg.setIcon(QtWidgets.QMessageBox.Information)       
        msg.setText("Will use "+deviceSelected+" for inference. To change bewtween CPU and GPU, use the options on the Build-Tab")
        msg.setWindowTitle("CPU used for inference")
        msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
        msg.exec_()

        #Take the model path from the GUI
        self.load_model_path = str(self.lineEdit_LoadModel_2.text())
        if len(self.load_model_path)==0:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("Please define a model path first")
            msg.setWindowTitle("No model path found")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return
        #Increase the thread-counter by one; only after finishing the thread, it will be reset to 0
        self.threadpool_single_queue += 1
        if self.threadpool_single_queue == 1:
            worker = Worker(self.inference_time_worker)
            def get_dt_from_worker(dic):
                outp = dic["outp"]
                self.lineEdit_InferenceTime.setText(outp)
            worker.signals.history.connect(get_dt_from_worker)    
            self.threadpool_single.start(worker)


    def update_check_worker(self,progress_callback,history_callback):
        #Retrieve information from GitHub
        dic = aid_bin.check_for_updates(VERSION)
        #dic = {"Errors":None,"latest_release":latest_release,"latest_release_url":url,"changelog":changelog}
        history_callback.emit(dic)
        progress_callback.emit(1) #when finished return one
        self.threadpool_single_queue = 0 #reset the thread-counter

    def actionUpdate_check_function(self):
        #Increase the thread-counter by one; only after finishing the thread, it will be reset to 0
        self.threadpool_single_queue += 1
        if self.threadpool_single_queue == 1:
            worker = Worker(self.update_check_worker)
            def get_info_from_worker(dic):
                #Create a popup window
                self.popup_updates = MyPopup()
                self.popup_updates_ui = aid_frontend.Ui_Updates()
                self.popup_updates_ui.setupUi(self.popup_updates) #open a popup

                if dic["Errors"]!=None:#if there is an errror (no internet...)
                    #display the error in the textbrowser
                    text = str(dic["Errors"])
                    
                elif dic["Errors"]==None:#No errors! Nice
                    latest_release = dic["latest_release"]
                    
                    if latest_release=="You are up to date":
                        text = "Your major version of AIDeveloper is up-to-date. Check below if there are updates available for that major version. <br>Example: Your major version of AIDeveloper is 0.2.0, then all updates which start with 0.2.x will be compatible."
                        text = "<html><head/><body><p>"+text+"</p></body></html>"
                    else:
                        text = "There is a new major update available. To download, follow this link:"
                        text = text+"<br>"+"<a href ="+dic["latest_release_url"]+">"+dic["latest_release_url"]+"</a>"
                        text = text+"<br>"+dic["changelog"]
                        text = text+"<br>Major updates need to be downloaded and installed manually. After that, you can install minor updates (which correspond to that major version) using the menu below."
                        text = "<html><head/><body><p>"+text+"</p></body></html>"

                #Fill info text (on top of Update Popup window)
                self.popup_updates_ui.textBrowser_majorVersionInfo.setText(text)
                
                #Fill lineEdit "Your version"
                self.popup_updates_ui.lineEdit_yourVersion.setText(VERSION)
                #Add updates to the comboBox
                self.popup_updates_ui.comboBox_updatesOndevice.addItems(dic["tags_update_ondevice"])
                self.popup_updates_ui.comboBox_updatesOnline.addItems(dic["tags_update_online"])

                self.popup_updates.show()
                self.popup_updates_ui.pushButton_installOndevice.clicked.connect(lambda: self.update_aideveloper("local"))
                self.popup_updates_ui.pushButton_installOnline.clicked.connect(lambda: self.update_aideveloper("github"))
                self.popup_updates_ui.pushButton_findFile.clicked.connect(self.update_addLocalFile)

            worker.signals.history.connect(get_info_from_worker)    
            self.threadpool_single.start(worker)

    def actionTerminology_function(self):
        #show a messagebox with link to terminology github page
        icon = QtGui.QImage(os.path.join(dir_root,"art",Default_dict["Icon theme"],"main_icon_simple_04_256"+icon_suff))
        icon = QtGui.QPixmap(icon).scaledToHeight(32, QtCore.Qt.SmoothTransformation)
        msg = QtWidgets.QMessageBox()
        msg.setIconPixmap(icon)
        text = "To learn more about machine learning/ deep learning specific terminology, please visit:<br>"
        url = "<a href=https://github.com/maikherbig/AIDeveloper/tree/master/Terminology>https://github.com/maikherbig/AIDeveloper/tree/master/Terminology</a>"
        text = "<html><head/><body><p>"+text+url+"</p></body></html>"
        msg.setText(text)
        msg.setWindowTitle("ML/DL Terminology")
        msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
        msg.exec_()

    def update_aideveloper(self,source):

        #retrieve the current text on comboBox_availableUpdates
        if source=="local":            
            item_text = str(self.popup_updates_ui.comboBox_updatesOndevice.currentText())
        elif source=="github":            
            item_text = str(self.popup_updates_ui.comboBox_updatesOnline.currentText())
        #Length of the version name should not be 0
        if len(item_text)==0:
            e = "No update available"
            aid_frontend.message(e)
            return
        
        if source=="local":
            #Complete file path (item_text not enough)
            item_path = "AIDeveloper_"+item_text+".zip"
            item_path = os.path.join(dir_root,item_path)  
        
        elif source=="github":
            if item_text=="Bleeding edge":
                #user want the most recent scripts from GitHub.
                downloadprocess = aid_bin.download_aid_repo()
            else:
                #item_text is a tag of the version. Use tag to download the zip
                downloadprocess = aid_bin.download_aid_update(item_text)
            
            #Check if download was successful
            if downloadprocess["success"]==False:#if the download was not done show message
                message = "Download was not conducted. Probably, the file is already present in:/n"+downloadprocess["path_save"]
                aid_frontend.message(message,msg_type="Warning")
                return
            #Retrieve the path of the zip file (contains the update files)
            item_path = downloadprocess["path_save"]
        
        if not os.path.isfile(item_path):#in case that is no file (zip file not created...)
            e = "Update requires a zip file. Could not find/create such a file!"
            aid_frontend.message(e)
            
        #Perform the update (including backup of current version)
        path_backup = aid_bin.update_from_zip(item_path,VERSION)
        
        #message: Installation successful-> need to restart AID
        msg = "Update successful. Please restart AIDeveloper. A backup of your previous version is stored in:\n"+path_backup
        aid_frontend.message(msg,msg_type="Information")

              

    def update_addLocalFile(self):
        #open a filedialog
        filename = QtWidgets.QFileDialog.getOpenFileName(self, 'Choose update file', dir_root,"AID update file (*.zip)")
        filename = filename[0]
        print(filename)
        #Check if the file is a zip
        if not filename.endswith(".zip"):#file has to be .zip
            text = "Chosen file is not a .zip archive!"
            aid_frontend.message(msg_text=text,msg_type="Error")
        #Check that file exists
        if not os.path.isfile(filename):
            text = "File not found"
            aid_frontend.message(msg_text=text,msg_type="Error")
            return
        
        base,_ = os.path.split(filename)
        #ensure that filename obeys the name convention: "AIDeveloper_"+tag_name+".zip"        
        tag_name = datetime.datetime.now().strftime("%Y%m%d_%H-%M-%S")+"-update"
        save_name = "AIDeveloper_"+tag_name+".zip"
        save_name = os.path.join(dir_root,save_name)
        #copy the file to dir_root
        shutil.copy(filename,save_name)

        #append tag_name to combobox
        self.popup_updates_ui.comboBox_updatesOndevice.addItem(tag_name)
        text = "Update is now availabele via the Dropdown menu on the left ("+tag_name+")."
        text += " The file was copied to:\n"
        text += save_name
        aid_frontend.message(msg_text=text,msg_type="Information")
        
        
    def get_validation_data_from_clicked(self,get_normalized=True):
        #Check, if files were clicked
        SelectedFiles = self.items_clicked_no_rtdc_ds()
        ######################Load the Validation Data################################
        ind = [selectedfile["TrainOrValid"] == "Valid" for selectedfile in SelectedFiles]
        ind = np.where(np.array(ind)==True)[0]

        if len(ind)==0:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("No validation data was selected. Please use tab 'Build' and drag/drop to load data")
            msg.setWindowTitle("No validation data selected")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return 0

        SelectedFiles_valid = np.array(SelectedFiles)[ind]
        SelectedFiles_valid = list(SelectedFiles_valid)
        indices_valid = [selectedfile["class"] for selectedfile in SelectedFiles_valid]
        nr_events_epoch_valid = [selectedfile["nr_events_epoch"] for selectedfile in SelectedFiles_valid]
        rtdc_path_valid = [selectedfile["rtdc_path"] for selectedfile in SelectedFiles_valid]
        zoom_factors_valid = [selectedfile["zoom_factor"] for selectedfile in SelectedFiles_valid]
        #zoom_order = [self.actionOrder0.isChecked(),self.actionOrder1.isChecked(),self.actionOrder2.isChecked(),self.actionOrder3.isChecked(),self.actionOrder4.isChecked(),self.actionOrder5.isChecked()]
        #zoom_order = int(np.where(np.array(zoom_order)==True)[0])
        zoom_order = int(self.comboBox_zoomOrder.currentIndex()) #the combobox-index is already the zoom order

        shuffle_valid = [selectedfile["shuffle"] for selectedfile in SelectedFiles_valid]
        xtra_in = set([selectedfile["xtra_in"] for selectedfile in SelectedFiles])   
        if len(xtra_in)>1:# False and True is present. Not supported
            print("Xtra data is used only for some files. Xtra data needs to be used either by all or by none!")
            return
        xtra_in = list(xtra_in)[0]#this is either True or False
        
        #Read other model properties from the Ui
        norm = self.comboBox_Normalization_2.currentText()
        norm = str(norm)
        #if normalization method needs mean/std of training set, the metafile needs to be loaded:
        if norm == "StdScaling using mean and std of all training data":
            modelindex = int(self.spinBox_ModelIndex_2.value())
            path,fname = os.path.split(self.load_model_path)    
            fname = fname.split(str(modelindex)+".model")[0]+"meta.xlsx"
            metafile_path = os.path.join(path,fname)
            parameters = pd.read_excel(metafile_path,sheet_name='Parameters',engine="openpyxl")
            mean_trainingdata = parameters["Mean of training data used for scaling"]
            std_trainingdata = parameters["Std of training data used for scaling"]
        else:
            mean_trainingdata = None
            std_trainingdata = None
            
        crop = int(self.spinBox_Crop_2.value()) 
        paddingMode = str(self.comboBox_paddingMode.currentText())#.lower()

        #read self.ram to new variable ; DONT clear ram after since multiple assessments can run on the same data.
        DATA = self.ram
        #self.ram = dict() #DONT clear the ram here! 
         
        ############Cropping#####################
        X_valid,y_valid,Indices,Xtra_in = [],[],[],[]
        for i in range(len(SelectedFiles_valid)):
            if not self.actionDataToRam.isChecked():
                #Replace=True means that individual cells could occur several times
                gen_valid = aid_img.gen_crop_img(crop,rtdc_path_valid[i],nr_events_epoch_valid[i],random_images=shuffle_valid[i],replace=True,zoom_factor=zoom_factors_valid[i],zoom_order=zoom_order,color_mode=self.get_color_mode(),padding_mode=paddingMode,xtra_in=xtra_in) 
            else: #get a similar generator, using the ram-data
                if len(DATA)==0:
                    #Replace=True means that individual cells could occur several times
                    gen_valid = aid_img.gen_crop_img(crop,rtdc_path_valid[i],nr_events_epoch_valid[i],random_images=shuffle_valid[i],replace=True,zoom_factor=zoom_factors_valid[i],zoom_order=zoom_order,color_mode=self.get_color_mode(),padding_mode=paddingMode,xtra_in=xtra_in) 
                else:
                    if self.actionVerbose.isChecked():
                        print("Loaded data from RAM")
                    gen_valid = aid_img.gen_crop_img_ram(DATA,rtdc_path_valid[i],nr_events_epoch_valid[i],random_images=shuffle_valid[i],replace=True,xtra_in=xtra_in) #Replace=True means that individual cells could occur several times
            
            gen = next(gen_valid)
            X_valid.append(gen[0])
            y_valid.append(np.repeat(indices_valid[i],X_valid[-1].shape[0]))
            Indices.append(gen[1]) #Cell index to track the event in the data-set(not cell-type!)
            Xtra_in.append(gen[2])
            
        X_valid_orig = [X.astype(np.uint8) for X in X_valid]
        X_valid = np.concatenate(X_valid)
        Xtra_in = np.concatenate(Xtra_in)
#        dim = X_valid.shape
#        if dim[2]!=crop:
#            remove = int(dim[2]/2.0 - crop/2.0)
#            #X_batch = X_batch[:,:,remove:-remove,remove:-remove] #crop to crop x crop pixels #Theano
#            X_valid = X_valid[:,remove:-remove,remove:-remove] #crop to crop x crop pixels #TensorFlow

        print("X_valid has following dimension:")
        print(X_valid.shape)
        
        y_valid = np.concatenate(y_valid)

        if len(np.array(X_valid).shape)<3:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("Discarded all events because too far at border of image (check zooming/cropping settings!)")
            msg.setWindowTitle("Empty dataset!")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()       
            return 0

        if get_normalized == True:
            if norm == "StdScaling using mean and std of all training data":
                X_valid = aid_img.image_normalization(X_valid,norm,mean_trainingdata,std_trainingdata)
            else:
                X_valid = aid_img.image_normalization(X_valid,norm)
        else:
            X_valid = None
        dic = {"SelectedFiles_valid":SelectedFiles_valid,"nr_events_epoch_valid":nr_events_epoch_valid,"rtdc_path_valid":rtdc_path_valid,"X_valid_orig":X_valid_orig,"X_valid":X_valid,"y_valid":y_valid,"Indices":Indices,"Xtra_in":Xtra_in}
        self.ValidationSet = dic
        return 1
        
    def export_valid_to_rtdc(self):        
        if not type(self.ValidationSet) is type(None): #If ValidationSet is not none, there has been a ValidationSet loaded already
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("Re-used validation data (from RAM) loaded earlier. If that is not good, please check and uncheck a file on 'Build' tab. This will delete the validation data from RAM")
            msg.setWindowTitle("Re-Used data")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            worked = 1
        else: #Otherwise get the validation data from the stuff that is clicked on 'Build'-Tab
            worked = self.get_validation_data_from_clicked() #after that, self.ValidationSet will exist
        if worked==0:
            return
        
        rtdc_path_valid = self.ValidationSet["rtdc_path_valid"]
        X_valid = []
        X_valid.append(self.ValidationSet["X_valid"][:,:,:,0])

        X_valid_orig = self.ValidationSet["X_valid_orig"]
        Xtra_in = self.ValidationSet["Xtra_in"]
        
        Indices = self.ValidationSet["Indices"]
        y_valid = self.ValidationSet["y_valid"]
        
        #Get a filename from the user for the new file
        filename = QtWidgets.QFileDialog.getSaveFileName(self, 'Save to rtdc', Default_dict["Path of last model"],"rtdc file (*.rtdc)")
        filename = filename[0]
        if len(filename)==0:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("No valid filename was chosen.")
            msg.setWindowTitle("No valid filename was chosen")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return
           
        #add the suffix _Valid_Data.avi or _Valid_Labels.npy
        if not filename.endswith(".rtdc"):
            filename = filename +".rtdc"
        filename_X = filename.split(".rtdc")[0]+"_Valid_Data.rtdc"
        filename_y = filename.split(".rtdc")[0]+"_Valid_Labels.txt"

        if bool(self.actionExport_Original.isChecked())==True:
            print("Export original images")
            save_cropped = False
        elif bool(self.actionExport_Cropped.isChecked())==True:
            print("Export cropped images")
            save_cropped = True
        elif bool(self.actionExport_Off.isChecked())==True:
            print("Exporting is turned off")
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("You could choose a different Exporting option in ->Option->Export")
            msg.setWindowTitle("Export is turned off!")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return

        aid_bin.write_rtdc(filename_X,rtdc_path_valid,X_valid_orig,Indices,cropped=save_cropped,color_mode=self.get_color_mode(),xtra_in=Xtra_in)
        np.savetxt(filename_y,y_valid.astype(int),fmt='%i')
        
        #If all that run without issue, remember the path for next time
        Default_dict["Path of last model"] = os.path.split(filename)[0]     
        aid_bin.save_aid_settings(Default_dict)
        
    def import_valid_from_rtdc(self):
        filename = QtWidgets.QFileDialog.getOpenFileName(self, 'Open Valid_Data.rtdc', Default_dict["Path of last model"],".rtdc file (*_Valid_Data.rtdc)")
        filename = filename[0]
        rtdc_path = filename

        if len(filename)==0:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("No valid filename was chosen.")
            msg.setWindowTitle("No valid filename was chosen")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return

        #Load the corresponding labels
        filename_labels = filename.split("Data.rtdc")[0]+"Labels.txt"
        if not os.path.isfile(filename_labels):
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("No corresponding _Labels.npy file found! Expected it here: "+filename_labels)
            msg.setWindowTitle("No Labels found")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return
        y_valid = np.loadtxt(filename_labels).astype(int)
        #Inform user (statusbar message)
        self.statusbar.showMessage("Loaded labels from "+filename_labels,5000)

        #Read images from .rtdc file
        failed,rtdc_ds = aid_bin.load_rtdc(rtdc_path)
        if failed:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Critical)       
            msg.setText(str(rtdc_ds))
            msg.setWindowTitle("Error occurred during loading file")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return
        
        #Load meta file
        #filename_meta = filename.split("Valid_Data.rtdc")[0]+"meta.xlsx"

        #Make the Image dimensions matching the requirements of the model
        model_in = int(self.spinBox_Crop_2.value())
        model_out = int(self.spinBox_OutClasses_2.value())
        color_mode = str(self.comboBox_loadedRGBorGray.currentText())
        
#        if color_mode=='RGB': #User wants RGB images
#            target_channels = 3
#        if color_mode=='Grayscale': # User want to have Grayscale
#            target_channels = 1

        if model_in==1 and model_out==1:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("Please first define a model. The validation data will then be cropped according to the required model-input size")
            msg.setWindowTitle("No model defined")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return
        
        x_valid = np.array(rtdc_ds["events"]["image"])
        #dim = x_valid.shape[1]
        #channels = x_valid.shape[-1]

        #Get further image processing settings from self.
        zoom_factor = float(self.img_processing_settings["zoom_factor"].values[0])
        zoom_interpol_method = str(self.img_processing_settings["zoom_interpol_method"].values[0])
        padding_mode = str(self.img_processing_settings["padding_mode"].values[0])
        
        #normalization_method = str(self.img_processing_settings["normalization_method"].values[0])
        norm = self.comboBox_Normalization_2.currentText()
        norm = str(norm)
        mean_trainingdata = self.img_processing_settings["mean_trainingdata"].values[0]
        std_trainingdata = self.img_processing_settings["std_trainingdata"].values[0]

        gen_valid = aid_img.gen_crop_img(cropsize=model_in,rtdc_path=rtdc_path,random_images=False,zoom_factor=zoom_factor,zoom_order=zoom_interpol_method,color_mode=color_mode,padding_mode=padding_mode,xtra_in=False)
        x_valid,index,xtra_valid = next(gen_valid)
        #When object is too far at side of image, the frame is dropped.
        #Consider this for y_valid
        y_valid = y_valid[index]
       

        if not model_in==x_valid.shape[-2]:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("Model input dimension ("+str(model_in)+"x"+str(model_in)+"pix) and validation data dimension ("+str(x_valid.shape)+") do not match")
            msg.setWindowTitle("Wrong image dimension")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()

        #Normalize the images
        X_valid_orig = np.copy(x_valid) #copy the cropped but non-normalized images
        if norm == "StdScaling using mean and std of all training data":
            X_valid = aid_img.image_normalization(x_valid,norm,mean_trainingdata,std_trainingdata)
        else:
            X_valid = aid_img.image_normalization(x_valid,norm)
        
        Indices = np.array(range(X_valid.shape[0])) #those are just indices to identify single cells in the file ->not cell-type indices!
        SelectedFiles_valid = None #[].append(rtdc_path)#
        nr_events_epoch_valid = None

        rtdc_h5 = h5py.File(rtdc_path, 'r')
        try:
            Xtra_in = np.array(rtdc_h5["xtra_in"])[index]
        except:
            Xtra_in = []
        rtdc_h5.close() #close the hdf5 
                       

        dic = {"SelectedFiles_valid":SelectedFiles_valid,"nr_events_epoch_valid":nr_events_epoch_valid,"rtdc_path_valid":[rtdc_path],"X_valid_orig":[X_valid_orig],"X_valid":X_valid,"y_valid":y_valid,"Indices":[Indices],"Xtra_in":Xtra_in}
        self.ValidationSet = dic

        self.statusbar.showMessage("Validation data loaded to RAM",5000)
        
        #Update the table
        #Prepare a table in tableWidget_Info
        self.tableWidget_Info_2.setColumnCount(0)#Reset table
        self.tableWidget_Info_2.setRowCount(0)#Reset table
        self.tableWidget_Info_2.setColumnCount(4) #Four columns

        nr_ind = len(set(y_valid)) #number of different labels ("indices")
        nr_rows = nr_ind
        self.tableWidget_Info_2.setRowCount(nr_rows)
        #Wich selected file has the most features?
        header_labels = ["Class","Nr of cells","Clr","Name"]
        self.tableWidget_Info_2.setHorizontalHeaderLabels(header_labels) 
        header = self.tableWidget_Info_2.horizontalHeader()
        for i in range(4):
            header.setResizeMode(i, QtWidgets.QHeaderView.ResizeToContents)        
        
        rowPosition = 0      
        #Total nr of cells for each index
        for index in np.unique(y_valid):
            #put the index in column nr. 0
            item = QtWidgets.QTableWidgetItem()
            item.setFlags(item.flags() &~QtCore.Qt.ItemIsEnabled &~ QtCore.Qt.ItemIsSelectable )
            item.setData(QtCore.Qt.EditRole,str(index))
            self.tableWidget_Info_2.setItem(rowPosition, 0, item)
            #Get the validation files of that index
            ind = np.where(y_valid==index)[0]
            nr_events_epoch = len(ind)
            item = QtWidgets.QTableWidgetItem()
            item.setFlags(item.flags() &~QtCore.Qt.ItemIsEnabled &~ QtCore.Qt.ItemIsSelectable )
            #item.setFlags(QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable)
            item.setData(QtCore.Qt.EditRole, str(np.sum(nr_events_epoch)))
            self.tableWidget_Info_2.setItem(rowPosition, 1, item)
            
            #Column for color
            item = QtWidgets.QTableWidgetItem()
            item.setFlags(item.flags() &~QtCore.Qt.ItemIsEnabled &~ QtCore.Qt.ItemIsSelectable )
            #item.setFlags(QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable)
            item.setData(QtCore.Qt.EditRole, "")
            item.setBackground(QtGui.QColor(self.colorsQt[index]))            
            self.tableWidget_Info_2.setItem(rowPosition, 2, item)

            #Column for User specified name
            item = QtWidgets.QTableWidgetItem()
            #item.setFlags(QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable)
            item.setData(QtCore.Qt.EditRole,str(index))
            self.tableWidget_Info_2.setItem(rowPosition, 3, item)
           
            rowPosition += 1
        self.tableWidget_Info_2.resizeColumnsToContents()            
        self.tableWidget_Info_2.resizeRowsToContents()


                
        
    def cm_interaction(self,item):
        """
        Grab validation data of particular class, load the scores (model.predict)
        and save images to .rtdc, or show them (users decision)
        first, "assess_model_plotting" has the be carried out
        """
        true_label = item.row()
        predicted_label = item.column()

        #If there is X_valid and y_valid on RAM, use it!
        if not type(self.ValidationSet) is type(None): #If X_valid is not none, there has been X_valid loaded already
            self.statusbar.showMessage("Re-used validation data (from RAM) loaded earlier. If that is not good, please check and uncheck a file on 'Build' tab. This will delete the validation data from RAM",2000)
        else: #Otherwise get the validation data from the stuff that is clicked on 'Build'-Tab
            self.get_validation_data_from_clicked() #after that, self.ValidationSet will exist
            self.statusbar.showMessage("Loaded data corresponding to the clicked files on 'Build'-tab",2000)

        rtdc_path_valid = self.ValidationSet["rtdc_path_valid"]
        X_valid_orig = self.ValidationSet["X_valid_orig"] #cropped but non-normalized images
                
        Indices = self.ValidationSet["Indices"]
        y_valid = self.ValidationSet["y_valid"] #load the validation labels to a new variable      
        dic = self.Metrics #gives {"scores":scores,"pred":pred,"conc_target_cell":conc_target_cell,"enrichment":enrichment,"yield_":yield_}
        if len(dic)==0:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("Data was altered. Please run 'Update Plots' again")
            msg.setWindowTitle("Data has changed")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return

        pred = dic["pred"]
        #get the length of each Index-list, 
        lengths = [len(l) for l in Indices]
        starts = np.cumsum(lengths)
        ToSave, y_valid_list, Indices_ = [],[],[] #list; store images remaining to indiv. .rtdc set in there 
        starts = np.array([0]+list(starts))
        for i in range(len(lengths)):
            y_val = y_valid[starts[i]:starts[i+1]]
            pred_ = pred[starts[i]:starts[i+1]]
            #update the indx to prepare for next iteration
            #indx = lengths[i]
            ind = np.where( (y_val==true_label) & (pred_==predicted_label) )[0] #select true_label cells and which of them are clasified as predicted_label
            #Grab the corresponding images
            ToSave.append(X_valid_orig[i][ind,:,:]) #get non-normalized X_valid to new variable
            #X_valid_.append(X_valid[i][ind,:,:]) #get normalized/cropped images ready to run through the model
            y_valid_list.append(y_val[ind])
            Indices_.append(Indices[i][ind]) #get non-normalized X_valid to new variable

        total_number_of_chosen_cells = [len(a) for a in y_valid_list]
        total_number_of_chosen_cells = np.sum(np.array(total_number_of_chosen_cells))

        msg = QtWidgets.QMessageBox()
        msg.setIcon(QtWidgets.QMessageBox.Question)
        text = "<html><head/><body><p>Show images/heatmap or save to .rtdc/.png/.jpg?</p></body></html>"
        msg.setText(text)
        msg.setWindowTitle("Show or save?")
        msg.setStandardButtons(QtGui.QMessageBox.Yes | QtGui.QMessageBox.Save | QtGui.QMessageBox.Cancel)
        show = msg.button(QtGui.QMessageBox.Yes)
        show.setText('Show image/heatmap')
#        show = msg.button(QtGui.QMessageBox.YesToAll)
#        show.setText('Show image/heatmap')
        save_png = msg.button(QtGui.QMessageBox.Save)
        save_png.setText('Save to .rtdc/.png/.jpg...')
        cancel = msg.button(QtGui.QMessageBox.Cancel)
        cancel.setText('Cancel')
        msg.exec_()        
        

        #View image and heatmap overlay (Grad-CAM)
        if msg.clickedButton()==show: #show image and heatmap overlay
            if total_number_of_chosen_cells==0:
                return
            #Get the images that were passed through the model for prediction
            X_valid = self.ValidationSet["X_valid"] #cropped but non-normalized images
            ind = np.where( (y_valid==true_label) & (pred==predicted_label) )[0] #select true_label cells and which of them are clasified as predicted_label
            X_valid_ = X_valid[ind]            

            #Popup window to show images and settings
            self.popup_gradcam = QtGui.QDialog()
            self.popup_gradcam_ui = aid_frontend.popup_cm_interaction()
            self.popup_gradcam_ui.setupUi(self.popup_gradcam) #open a popup to show images and options
            #self.popup_imgRes.setWindowModality(QtCore.Qt.WindowModal)
            #self.popup_gradcam.setWindowModality(QtCore.Qt.ApplicationModal)
            
            #Fill Model info
            self.popup_gradcam_ui.lineEdit_loadModel.setText(self.load_model_path)
            in_dim = int(self.spinBox_Crop_2.value()) #grab value from Assess Tab
            self.popup_gradcam_ui.spinBox_Crop_inpImgSize.setValue(in_dim)#insert value into popup
            out_dim = int(self.spinBox_OutClasses_2.value()) #grab value from Assess Tab
            self.popup_gradcam_ui.spinBox_outpSize.setValue(out_dim) #insert value into popup
            self.popup_gradcam_ui.spinBox_gradCAM_targetClass.setMaximum(out_dim-1)

            #For the grad_cam the name of the final conv-layer needs to be selected
            convlayers = [layer.name for layer in self.model_keras.layers if len(layer.output_shape)==4]
            convlayers = convlayers[::-1] #reverse list
            self.popup_gradcam_ui.comboBox_gradCAM_targetLayer.addItems(convlayers)

            #Connect buttons to functions
            self.popup_gradcam_ui.pushButton_update.clicked.connect(lambda: self.popup_cm_show_update(ToSave,X_valid_))
            self.popup_gradcam_ui.pushButton_reset.clicked.connect(self.popup_cm_reset)
            self.popup_gradcam_ui.pushButton_showSummary.clicked.connect(self.popup_show_model_summary)
            self.popup_gradcam_ui.pushButton_toTensorB.clicked.connect(self.popup_to_tensorboard)

            #Get the original image
            img_display = np.concatenate(ToSave)

            img_display = np.r_[img_display]
            img_display = img_display.swapaxes(1,2)
            img_display = np.append(img_display,img_display[-1:],axis=0)

            self.popup_gradcam_ui.widget_image.setImage(img_display)
            self.popup_gradcam.show()


        #For .rtdc/.png... saving
        elif msg.clickedButton()==save_png: #Save to .rtdc/.png/.jpg/...
            if total_number_of_chosen_cells==0:
                return
            sumlen = np.sum(np.array([len(l) for l in ToSave]))
            self.statusbar.showMessage("Nr. of target cells above threshold = "+str(sumlen),2000)

            filename = QtWidgets.QFileDialog.getSaveFileName(self, 'Save to .rtdc/.png/.jpg', Default_dict["Path of last model"],"File format (*.rtdc *.png *.jpg *.bmp *.eps *.gif *.ico *.icns)")
            filename = filename[0]
            if len(filename)==0:
                return
            filename_X, file_extension = os.path.splitext(filename)#divide into path and file_extension if possible
            #Check if chosen file_extension is valid
            if not file_extension in [".rtdc",".png",".jpg",".bmp",".eps",".gif",".ico",".icns"]:
                print("Invalid file extension detected. Will use .png instead.")
                file_extension = ".png"
            
            if file_extension==".rtdc":#user wants to save to .rtdc
                #add the suffix _Valid_Data.rtdc or _Valid_Labels.txt
                if not filename.endswith(".rtdc"):
                    filename = filename +".rtdc"
                filename_X = filename.split(".rtdc")[0]+"_Valid_Data.rtdc"
                filename_y = filename.split(".rtdc")[0]+"_Valid_Labels.txt"
        
                #Save the labels
                y_valid_list = np.concatenate(y_valid_list)
                #Save the .rtdc data (images and all other stuff)
                #Should cropped or original be saved?
                if bool(self.actionExport_Original.isChecked())==True:
                    print("Export original images")
                    save_cropped = False
                if bool(self.actionExport_Cropped.isChecked())==True:
                    print("Export cropped images")
                    save_cropped = True
                elif bool(self.actionExport_Off.isChecked())==True:
                    print("Exporting is turned off")
                    msg = QtWidgets.QMessageBox()
                    msg.setIcon(QtWidgets.QMessageBox.Information)       
                    msg.setText("You may want to choose a different exporting option in ->Options->Export")
                    msg.setWindowTitle("Export is turned off!")
                    msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                    msg.exec_()
                    return
        
                np.savetxt(filename_y,y_valid_list.astype(int),fmt='%i')      
                aid_bin.write_rtdc(filename_X,rtdc_path_valid,ToSave,Indices_,cropped=save_cropped,color_mode=self.get_color_mode())
    
                #If all that run without issue, remember the path for next time
                Default_dict["Path of last model"] = os.path.split(filename)[0]
                aid_bin.save_aid_settings(Default_dict)
                
                
            else: #some image file format was chosen
                #Should cropped or original be saved?
                if bool(self.actionExport_Original.isChecked())==True:
                    print("Export original images")
                    save_cropped = False
                if bool(self.actionExport_Cropped.isChecked())==True:
                    print("Export cropped images")
                    save_cropped = True
                elif bool(self.actionExport_Off.isChecked())==True:
                    print("Exporting is turned off")
                    msg = QtWidgets.QMessageBox()
                    msg.setIcon(QtWidgets.QMessageBox.Information)       
                    msg.setText("You may want to choose a different exporting option in ->Options->Export")
                    msg.setWindowTitle("Export is turned off!")
                    msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                    msg.exec_()
                    return
    
                #Save the images data to .png/.jpeg...
                index = 0
                for imgs in ToSave:
                    for img in imgs:
                        img = PIL.Image.fromarray(img)
                        img.save(filename_X+"_"+str(index)+file_extension)
                        index+=1
    
                #If all that run without issue, remember the path for next time
                Default_dict["Path of last model"] = os.path.split(filename)[0]
                aid_bin.save_aid_settings(Default_dict)



    def popup_cm_show_update(self,ToSave,X_valid_):
        #ui_item = self.popup_gradcam_ui
        #grab information from the popup window
        show_image = bool(self.popup_gradcam_ui.groupBox_image_Settings.isChecked())
        show_gradCAM = bool(self.popup_gradcam_ui.groupBox_gradCAM_Settings.isChecked())
        alpha_1 = float(self.popup_gradcam_ui.doubleSpinBox_image_alpha.value())
        alpha_2 = float(self.popup_gradcam_ui.doubleSpinBox_gradCAM_alpha.value())
        layer_name = str(self.popup_gradcam_ui.comboBox_gradCAM_targetLayer.currentText()) #self.model_keras exists after assess_model_plotting was carried out
        class_ = int(self.popup_gradcam_ui.spinBox_gradCAM_targetClass.value())
        colormap = str(self.popup_gradcam_ui.comboBox_gradCAM_colorMap.currentText()) #self.model_keras exists after assess_model_plotting was carried out
        colormap = "COLORMAP_"+colormap
        colormap = getattr(cv2, colormap)
        currentindex = self.popup_gradcam_ui.widget_image.currentIndex
        
        if show_image and not show_gradCAM:
            #Get the original image for display
            img_display = np.concatenate(ToSave)
            img_display = np.r_[img_display]
            img_display = img_display.swapaxes(1,2)
            img_display = np.append(img_display,img_display[-1:],axis=0)
        
        if show_gradCAM:#grad-Cam is on
            img_display = np.concatenate(ToSave)
            
            #compare model input dim and dom of provided data
            in_model = self.model_keras.input.shape.as_list()[1:]
            in_data = list(X_valid_.shape[1:])

            channels_model = in_model[-1]
            
            if not in_data==in_model:
                msg = QtWidgets.QMessageBox()
                msg.setIcon(QtWidgets.QMessageBox.Information) 
                msg = "Model input dimension ("+str(in_model)+") not equal to dim. of input data ("+str(in_data)+")"
                msg.setText(msg)
                msg.setWindowTitle("Input dimension error")
                msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                msg.exec_()
                return

            img2 = aid_dl.grad_cam(self.load_model_path, X_valid_, class_, layer_name)#Carry out grad-cam
            img2 = [cv2.applyColorMap(cam_, colormap) for cam_ in img2]#create colormap returns BGR image!
            img2 = [cv2.cvtColor(cam_, cv2.COLOR_BGR2RGB) for cam_ in img2]#convert to RGB
            
            #in case img_display is grayscale, mimick rgb image by stacking 
            
            if channels_model==1:
                print("Triple stacking grayscale channel")
                img_display = [np.stack((img_display_,)*3, axis=-1) for img_display_ in img_display]

            #add heatmap to image, make sure alpha_1=0 if show_image=False
            img_display = [cv2.addWeighted(img_display[i], alpha_1, img2[i], alpha_2, 0) for i in range(X_valid_.shape[0])]
            #ToDo: this only works for RGB images. Adjust expression to work or grayscale and RGB  
            
            img_display = np.r_[img_display]
            img_display = img_display.swapaxes(1,2)
            img_display = np.append(img_display,img_display[-1:],axis=0)

        self.popup_gradcam_ui.widget_image.setImage(img_display)
        self.popup_gradcam_ui.widget_image.setCurrentIndex(currentindex)
        self.popup_gradcam.show()

    def popup_cm_reset(self):
        self.popup_gradcam_ui.groupBox_image_Settings.setChecked(True)
        self.popup_gradcam_ui.groupBox_gradCAM_Settings.setChecked(False)
        #self.popup_gradcam_ui.doubleSpinBox_image_alpha.setValue(1)
        self.popup_gradcam_ui.comboBox_gradCAM_targetLayer.setCurrentIndex(0)
        #self.popup_gradcam_ui.comboBox_gradCAM_colorMap.setCurrentIndex(0)
        self.popup_gradcam_ui.spinBox_gradCAM_targetClass.setValue(0)



    def popup_show_model_summary(self):
        #textbrowser popup        
        self.popup_modelsummary = MyPopup()
        self.popup_modelsummary_ui = aid_frontend.popup_cm_modelsummary()
        self.popup_modelsummary_ui.setupUi(self.popup_modelsummary) #open a popup to show images and options

        text5 = "Model summary:\n"
        summary = []
        self.model_keras.summary(print_fn=summary.append)
        summary = "\n".join(summary)
        text = text5+summary
        self.popup_modelsummary_ui.textBrowser_modelsummary.append(text)
        self.popup_modelsummary.show()


    def popup_to_tensorboard(self):
        #Open the model in tensorboard
        #Issue: I cannot stop the process. The appraoch below, which uses a 
        #separate thread for the function does not solve the issue
        self.threadpool_single_queue += 1
        if self.threadpool_single_queue == 1:
            worker = Worker(self.tensorboad_worker)
            def get_pid_from_worker(dic):
                pid = dic["outp"]
                #print("WORKER-PID")
                #print("pid")
                #os.kill(pid,signal.CTRL_C_EVENT)
                #ToDo Find a way to kill that process!
            worker.signals.history.connect(get_pid_from_worker)
            self.threadpool_single.start(worker)        
            #print("PID-Here:")
            #print(os.getpid())
            #time.sleep(2)
            

    def tensorboad_worker(self,progress_callback,history_callback):
        #send the model to tensorboard (webbased application)
        with tf.compat.v1.Session() as sess:
            model_keras = load_model(self.load_model_path,custom_objects=aid_dl.get_custom_metrics())  
            graph = tf.compat.v1.keras.backend.get_session()
            #graph = K.get_session().graph # Get the sessions graph
            #get a folder for that model in temp
            temp_path = aid_bin.create_temp_folder()
            modelname = os.path.split(self.load_model_path)[-1]
            modelname = modelname.split(".model")[0]
            log_dir = os.path.join(temp_path,modelname)
            tf.summary.create_file_writer(log_dir)
            #writer = tf.summary.FileWriter(logdir=log_dir, graph=graph)#write a log
    
            #tb = program.TensorBoard()            
            tb = program.TensorBoard(default.get_plugins(), assets.get_default_assets_zip_provider())
            #tb.configure(argv=[None, '--logdir', log_dir,"--host","127.0.0.1"])
            tb.configure(argv=[None, '--logdir', log_dir,"--host","localhost"])

            url = tb.launch()
            url = os.path.join(url)
            os.system(r"start "+url)
            pid = os.getpid()
            dic = {"outp":pid}
            #print("WORKER1-PID")
            #print(pid)
            history_callback.emit(dic) #return the pid (use it to kill the process)
            self.threadpool_single_queue = 0 #reset the thread-counter
        time.sleep(0.5)
        


    def copy_cm_to_clipboard(self,cm1_or_cm2):
        if cm1_or_cm2==1:
            table = self.tableWidget_CM1
            cols = table.columnCount()
            header = [table.horizontalHeaderItem(col).text() for col in range(cols)]
        elif cm1_or_cm2==2:
            table = self.tableWidget_CM2
            cols = table.columnCount()
            header = [table.horizontalHeaderItem(col).text() for col in range(cols)]
        elif cm1_or_cm2==3: #this is for the classification report table tableWidget_AccPrecSpec
            table = self.tableWidget_AccPrecSpec
            cols = table.columnCount()
            header = list(range(cols))
        
        rows = table.rowCount()

        tmp_df = pd.DataFrame(columns=header,index=range(rows)) 
        for i in range(rows):
            for j in range(cols):
                try:
                    tmp_df.iloc[i, j] = table.item(i, j).text()
                except:
                    tmp_df.iloc[i, j] = np.nan
                    
        tmp_df.to_clipboard()
        if cm1_or_cm2<3:
            self.statusbar.showMessage("Confusion matrix appended to clipboard.",2000)
        if cm1_or_cm2==3:
            self.statusbar.showMessage("Classification report appended to clipboard.",2000)

    def assess_model_plotting(self):
        if self.load_model_path == None:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("Please define a model path first")
            msg.setWindowTitle("No model path found")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return

       #If there is a ValidationSet on RAM-> use it!
        if not type(self.ValidationSet) is type(None): #If ValidationSet is not none, there has been ValidationSet loaded already
            self.statusbar.showMessage("Use validation data (from RAM) loaded earlier. If that is not good, please check and uncheck a file on 'Build' tab. This will delete the validation data from RAM",5000)
        else: #Otherwise get the validation data from the stuff that is clicked on 'Build'-Tab
            self.get_validation_data_from_clicked() #after that, self.ValidationSet will exist

        #Check if input data is available
        if type(self.ValidationSet)==type(None):
            return
        elif type(self.ValidationSet["X_valid"])==type(None):
            return
        #Check the input dimensions:        
        img_dim = self.ValidationSet["X_valid"].shape[-2]
        model_in = int(self.spinBox_Crop_2.value())
        if model_in!=img_dim:
            self.ValidationSet = None
            self.get_validation_data_from_clicked() #after that, self.ValidationSet will exist
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("New model has different input dimensions (image crop). Validation set is re-loaded (like when you clicked on files on build-tab)")
            msg.setWindowTitle("Automatically re-loaded validation set")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()

        y_valid = self.ValidationSet["y_valid"] #load the validation labels to a new variable
        #X_valid = self.X_valid #<--dont do this since it is used only once (.predict) and it would require additional RAM; instad use self.X_valid for .predict
        #Load the model and predict
        with tf.compat.v1.Session() as sess:
            model_keras = load_model(self.load_model_path,custom_objects=aid_dl.get_custom_metrics())  
            self.model_keras = model_keras #useful to get the list of layers for Grad-CAM; also used to show the summary
            in_dim, _ = aid_dl.model_in_out_dim(model_keras,"model")
            if type(in_dim)==list:
                multi_input = True
                in_dim = in_dim[0]#discard the second (xtra input)
            else:
                multi_input = False                
            
            channels_model = in_dim[-1]
            channels_data = self.ValidationSet["X_valid"].shape[-1]
            
            #Compare channel dimensions of loaded model and validation set
            if channels_model==3 and channels_data==1:
                msg = QtWidgets.QMessageBox()
                msg.setIcon(QtWidgets.QMessageBox.Information)
                text = "Model expects 3 channels, but data has 1 channel!"
                text = text+" Will stack available channel three times to generate RGB image."
                msg.setText(text)
                msg.setWindowTitle("Automatic adjustment of image channels")
                msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                msg.exec_()

                #model wants rgb images, but provided data is grayscale->copy and stack 3 times
                self.ValidationSet["X_valid"] = np.stack((self.ValidationSet["X_valid"][:,:,:,0],)*3, axis=-1)
            
            elif channels_model==1 and channels_data==3:
                msg = QtWidgets.QMessageBox()
                msg.setIcon(QtWidgets.QMessageBox.Information)
                text = "Model expects 1 channel, but data has 3 channels!"
                text = text+" Will use the luminosity formula to convert RGB to grayscale."
                msg.setText(text)
                msg.setWindowTitle("Automatic adjustment of image channels")
                msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                msg.exec_()
                
                #model wants grayscale, but provided data is rgb
                self.ValidationSet["X_valid"] = aid_img.rgb_2_gray(self.ValidationSet["X_valid"])
            
            elif channels_model!=channels_data: #Model and validation data have differnt channel dims
                text = "Model expects "+str(int(channels_model))+" channel(s), but data has "+str(int(channels_data))+" channel(s)!"
                msg = QtWidgets.QMessageBox()
                msg.setIcon(QtWidgets.QMessageBox.Information)       
                msg.setText(text)
                msg.setWindowTitle("Model and data channel dimension not equal!")
                msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                msg.exec_()
                return
            
            if multi_input == False:                
                scores = model_keras.predict(self.ValidationSet["X_valid"])
            if multi_input == True:
                print("self.ValidationSet[Xtra_in]")
                print(self.ValidationSet["Xtra_in"])
                scores = model_keras.predict([self.ValidationSet["X_valid"],self.ValidationSet["Xtra_in"]])
        
        #Get settings from the GUI
        threshold = float(self.doubleSpinBox_sortingThresh.value())#threshold probability obove which a cell is sorted
        target_index = int(self.spinBox_indexOfInterest.value())#index of the cell type that should be sorted for
        thresh_on = bool(self.checkBox_SortingThresh.isChecked())

        #Check that the target index alias "Sorting class" is actually a valid class of the model
        out_dim = int(self.spinBox_OutClasses_2.value())
        if not target_index<out_dim:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Warning)
            msg.setText("You set the 'Sorting class' to "+str(target_index)+" which is not a valid class of the loaded model. The model only has the following classes: "+str(range(out_dim)))
            msg.setWindowTitle("Class not available in the model")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return

        dic = aid_bin.metrics_using_threshold(scores,y_valid,threshold,target_index,thresh_on) #returns dic = {"scores":scores,"pred":pred,"conc_target_cell":conc_target_cell,"enrichment":enrichment,"yield_":yield_}
        self.Metrics = dic #write to a variable #     
        
        pred = dic["pred"]
        
        cm = metrics.confusion_matrix(y_valid,pred,labels=range(scores.shape[1]))
        cm_normalized = 100*cm.astype('float') / cm.sum(axis=1)[:, np.newaxis]
        
        #Show the metrics on tableWidget_CM1 and tableWidget_CM2
        #inds_uni = set(list(set(y_valid))+list(set(pred))) #It could be that a cell-index is not present in the validation data, but, the dimension of the scores tells me, how many indices are supposed to appear
        inds_uni = range(scores.shape[1]) #these indices are explained by model
        
        #look in into tableWidget_Info_2 if there are user defined index names
        
        rowCount = self.tableWidget_Info_2.rowCount()
        #Only counts rows with input
        rowCount = sum([self.tableWidget_Info_2.item(row, 0)!=None for row in range(rowCount)])

        try:
            indices_on_table = [int(self.tableWidget_Info_2.item(row, 0).text()) for row in range(rowCount)]
            names_on_table = [str(self.tableWidget_Info_2.item(row, 3).text()) for row in range(rowCount)]
        except Exception as e:
            #There was an error!
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Critical)       
            msg.setText(str(e))
            msg.setWindowTitle("Error")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return
            
        #Check that len(names_on_table) <= len(inds_uni) ->it is impossible that the model for example can predict 2 classes, but there are 3 different classes in the validation set
        if not len(names_on_table) <= len(inds_uni):
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("Model can only predict "+str(len(inds_uni))+" classes, but validation data contains "+str(len(names_on_table))+" classes")
            msg.setWindowTitle("Too many classes in validation set")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            #return
       
        CellNames = []            
        for ind in inds_uni:
            #check if that index is present on table
            where = np.where(np.array(indices_on_table)==ind)[0]
            if len(where)==1:#if there is exaclty one item...
                CellNames.append(np.array(names_on_table)[where]) #append the corresponding user defined name to a list
            else:
                CellNames.append(str(ind))

        header_labels = [i[0] for i in CellNames]#list(inds_uni)]

        #Table for CM1 - Total Nr of cells
        self.tableWidget_CM1.setRowCount(len(inds_uni))
        self.tableWidget_CM1.setColumnCount(len(inds_uni))            
        
        self.tableWidget_CM1.setHorizontalHeaderLabels(header_labels) 
        self.tableWidget_CM1.setVerticalHeaderLabels(header_labels) 
        for i in inds_uni:
            for j in inds_uni:
                rowPosition = i
                colPosition = j
                #Total nr of cells for each index
                #put the index in column nr. 0
                item = QtWidgets.QTableWidgetItem()
                item.setFlags(QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable)
                item.setData(QtCore.Qt.EditRole,str(cm[i,j]))
                self.tableWidget_CM1.setItem(rowPosition, colPosition, item)
        self.tableWidget_CM1.resizeColumnsToContents()            
        self.tableWidget_CM1.resizeRowsToContents()


        #Table for CM2 - Normalized Confusion matrix
        self.tableWidget_CM2.setRowCount(len(inds_uni))
        self.tableWidget_CM2.setColumnCount(len(inds_uni))
        self.tableWidget_CM2.setHorizontalHeaderLabels(header_labels) 
        self.tableWidget_CM2.setVerticalHeaderLabels(header_labels) 
        for i in range(len(inds_uni)):
            for j in range(len(inds_uni)):
                rowPosition = i
                colPosition = j
                #Total nr of cells for each index
                #put the index in column nr. 0
                item = QtWidgets.QTableWidgetItem()
                item.setFlags(QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable)
                item.setData(QtCore.Qt.EditRole,str(cm_normalized[i,j]))
                self.tableWidget_CM2.setItem(rowPosition, colPosition, item)
        self.tableWidget_CM2.resizeColumnsToContents()            
        self.tableWidget_CM2.resizeRowsToContents()
        
        ############Fill tableWidget_AccPrecSpec with information##########          
                
        #Compute more metrics and put them on the table below                    
        nr_target_init = float(len(np.where(y_valid==target_index)[0])) #number of target cells in the initial sample
        conc_init = nr_target_init/float(len(y_valid)) #concentration of the target cells in the initial sample
                    
        acc = metrics.accuracy_score(y_valid,pred)
        
        #Reset the table
        self.tableWidget_AccPrecSpec.setColumnCount(0)#Reset table
        self.tableWidget_AccPrecSpec.setRowCount(0)#Reset table
        nr_cols = np.max([5,len(inds_uni)+1])
        self.tableWidget_AccPrecSpec.setColumnCount(nr_cols) #Five columns
        self.tableWidget_AccPrecSpec.setRowCount(7+len(inds_uni)+2) #Nr. of rows

        #Put lots and lots of Info on tableWidget_AccPrecSpec
        text_conc_init = "Init. conc. of cells from class/name "+header_labels[target_index]
        self.tableWidget_AccPrecSpec.setItem(0 , 0, QtGui.QTableWidgetItem(text_conc_init))
        item = QtWidgets.QTableWidgetItem()
        item.setData(QtCore.Qt.EditRole, float(np.round(100*conc_init,4)))
        self.tableWidget_AccPrecSpec.setItem(0, 1, item)

        text_conc_final = "Final conc. in target region"
        self.tableWidget_AccPrecSpec.setItem(1 , 0, QtGui.QTableWidgetItem(text_conc_final))
        item = QtWidgets.QTableWidgetItem()
        item.setData(QtCore.Qt.EditRole, float(np.round(dic["conc_target_cell"],4)))
        self.tableWidget_AccPrecSpec.setItem(1, 1, item)

        text_enrich = "Enrichment"
        self.tableWidget_AccPrecSpec.setItem(2 , 0, QtGui.QTableWidgetItem(text_enrich))
        item = QtWidgets.QTableWidgetItem()
        item.setData(QtCore.Qt.EditRole, float(np.round(dic["enrichment"],4)))
        self.tableWidget_AccPrecSpec.setItem(2, 1, item)

        text_yield = "Yield"
        self.tableWidget_AccPrecSpec.setItem(3 , 0, QtGui.QTableWidgetItem(text_yield))
        item = QtWidgets.QTableWidgetItem()
        item.setData(QtCore.Qt.EditRole, float(np.round(dic["yield_"],4)))
        self.tableWidget_AccPrecSpec.setItem(3, 1, item)

        text_acc = "Accuracy"#+str(round(acc,4))+"\n"
        self.tableWidget_AccPrecSpec.setItem(4 , 0, QtGui.QTableWidgetItem(text_acc))
        item = QtWidgets.QTableWidgetItem()
        item.setData(QtCore.Qt.EditRole, float(np.round(acc,4)))
        self.tableWidget_AccPrecSpec.setItem(4, 1, item)

        text_classification_report = "Classification Report"#+metrics.classification_report(y_valid, pred, target_names=header_labels)
        self.tableWidget_AccPrecSpec.setItem(5 , 0, QtGui.QTableWidgetItem(text_classification_report))
        class_rep = metrics.classification_report(y_valid, pred,labels=inds_uni, target_names=header_labels,output_dict =True)

        try:
            df = pd.DataFrame(class_rep)
            df = df.T
            ax_left = df.axes[0]
            for row in range(len(ax_left)):
                self.tableWidget_AccPrecSpec.setItem(7+row, 0, QtGui.QTableWidgetItem(str(ax_left[row])))
            ax_up = df.axes[1]
            for col in range(len(ax_up)):
                self.tableWidget_AccPrecSpec.setItem(6, 1+col, QtGui.QTableWidgetItem(str(ax_up[col])))
                
            for row in range(df.shape[0]):
                for col in range(df.shape[1]):
                    val = df.iloc[row,col]
                    val = float(np.round(val,4))
                    item = QtWidgets.QTableWidgetItem()
                    item.setData(QtCore.Qt.EditRole, val)
                    self.tableWidget_AccPrecSpec.setItem(7+row, 1+col, item)
        except Exception as e:
            #There is an issue loading the files!
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Warning)       
            msg.setText(str(e))
            msg.setWindowTitle("Error")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()

        self.tableWidget_AccPrecSpec.resizeColumnsToContents()            
        self.tableWidget_AccPrecSpec.resizeRowsToContents()

        #AFTER the table is resized to the contents, fill in also information 
        #about the used data:
        rowPosition = self.tableWidget_AccPrecSpec.rowCount()
        self.tableWidget_AccPrecSpec.insertRow(rowPosition) #Insert a new row
        self.tableWidget_AccPrecSpec.setItem(rowPosition , 0, QtGui.QTableWidgetItem("Used Files"))
        rowPosition = self.tableWidget_AccPrecSpec.rowCount()
        self.tableWidget_AccPrecSpec.insertRow(rowPosition) #Insert another row!
        self.tableWidget_AccPrecSpec.setItem(rowPosition , 0, QtGui.QTableWidgetItem("File"))
        
        #dic = {"SelectedFiles_valid":SelectedFiles_valid,"nr_events_epoch_valid":nr_events_epoch_valid,"rtdc_path_valid":[rtdc_path],"X_valid_orig":[X_valid_orig],"X_valid":X_valid,"y_valid":y_valid,"Indices":[Indices]}
        
        rtdc_path_valid = self.ValidationSet["rtdc_path_valid"]
        #nr_events_epoch_valid = self.ValidationSet["nr_events_epoch_valid"]
        y_valid = self.ValidationSet["y_valid"] #y_valid is a long array containing the label of all cell (of all clicked files)
        Indices = self.ValidationSet["Indices"] #Index is a list with arrays containing cell-indices (to track events in a data-set)
        
        
        y_valid_uni = np.unique(np.array(y_valid),return_counts=True)
        #set the column count the at least match the amount of different cell-types available
        if self.tableWidget_AccPrecSpec.columnCount() < len(y_valid_uni[0]):
            diff = len(y_valid_uni[0])-self.tableWidget_AccPrecSpec.columnCount()
            for col_ind in range(diff):
                colPosition = self.tableWidget_AccPrecSpec.columnCount()
                self.tableWidget_AccPrecSpec.insertColumn(colPosition) #Insert a new col for each cell-type
        
        #Create a column for each cell-type
        for col_ind in range(len(y_valid_uni[0])):
            #how_many = y_valid_uni[1][col_ind]
            #self.tableWidget_AccPrecSpec.setItem(rowPosition , 1+col_ind, QtGui.QTableWidgetItem(float(how_many)))
            content = "Class "+str(y_valid_uni[0][col_ind])
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.EditRole, content)
            self.tableWidget_AccPrecSpec.setItem(rowPosition , 1+col_ind, item)
                
        loc = 0
        for row in range(len(rtdc_path_valid)):
            rowPosition = self.tableWidget_AccPrecSpec.rowCount()
            self.tableWidget_AccPrecSpec.insertRow(rowPosition) #Insert a new row for each entry
            self.tableWidget_AccPrecSpec.setItem(rowPosition , 0, QtGui.QTableWidgetItem(rtdc_path_valid[row]))
            #y_valid_uni = np.unique(y_valid[row])
            #self.tableWidget_AccPrecSpec.setItem(rowPosition , 1, QtGui.QTableWidgetItem(np.array(y_valid_uni)))
            #self.tableWidget_AccPrecSpec.setItem(rowPosition , 2, QtGui.QTableWidgetItem(float(nr_events_epoch_valid[row])))
            index = Indices[row] #get the array of indices of a single measurement
            y_valid_i = y_valid[loc:loc+len(index)]
            loc = loc+len(index)
            y_valid_i_uni = np.unique(y_valid_i,return_counts=True)
            for col_ind in range(len(y_valid_i_uni[0])):
                #what is the cell-type
                cell_type = int(y_valid_i_uni[0][col_ind])#cell-type index alway starts with 0. Nr. of cells of cell-type 0 remain to column 1
                how_many = y_valid_i_uni[1][col_ind]
                item = QtWidgets.QTableWidgetItem()
                item.setData(QtCore.Qt.EditRole, int(how_many))
                self.tableWidget_AccPrecSpec.setItem(rowPosition , 1+cell_type, item)

        #Draw the probability histogram
        self.probability_histogram()
        #Finally, also update the third plot
        self.thirdplot()
 
    def create_random_table(self):
        print("def create_random_table only useful for development")
#        matrix = np.random.randint(0,100,size=(3,3))
#        self.tableWidget_CM1.setRowCount(matrix.shape[0])
#        self.tableWidget_CM1.setColumnCount(matrix.shape[1])
#
#        for i in range(matrix.shape[0]):
#            for j in range(matrix.shape[1]):
#                item = QtWidgets.QTableWidgetItem()
#                item.setData(QtCore.Qt.EditRole,str(matrix[i,j]))
#                self.tableWidget_CM1.setItem(i, j, item)
#
#        self.tableWidget_CM1.resizeColumnsToContents()            
#        self.tableWidget_CM1.resizeRowsToContents()

      
    def probability_histogram(self): #    def probability_histogram(self):   
        """
        Grab the scores of each class and show it in histogram
        """
        if len(self.Metrics) ==0: #but if not give message and return
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("There are no Metrics determined yet (use ->'Update Plots' first)")
            msg.setWindowTitle("No Metrics found")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return

        dic = self.Metrics #returns dic = {"scores":scores,"pred":pred,"conc_target_cell":conc_target_cell,"enrichment":enrichment,"yield_":yield_}
        scores = dic["scores"]

        #Get the available cell indices (cell-type identifier)
        inds_uni = range(scores.shape[1]) #these indices are explained by model

        threshold = float(self.doubleSpinBox_sortingThresh.value())#threshold probability obove which a cell is sorted
        target_index = int(self.spinBox_indexOfInterest.value())#index of the cell type that should be sorted for
       
        try:
            #What is the probability of cell with y_valid=i that it belongs to class target_index?
            scores_i = []
            y_valid = self.ValidationSet["y_valid"]
            for i in inds_uni:
                ind = np.where(y_valid==i)[0]
                if len(ind)>0: #if there are no cells available, dont append. In this case there will also be no color defined
                    scores_i.append(scores[ind,target_index])
        except Exception as e:
            #There is an issue loading the files!
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Critical)       
            msg.setText(str(e))
            msg.setWindowTitle("Error")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return
            
        rowCount = self.tableWidget_Info_2.rowCount()
        #only count rows with content
        rowCount = sum([self.tableWidget_Info_2.item(row, 0)!=None for row in range(rowCount)])
        names_on_table = [str(self.tableWidget_Info_2.item(row, 3).text()) for row in range(rowCount)]
        index_on_table = [int(self.tableWidget_Info_2.item(row, 0).text()) for row in range(rowCount)]
        #On which row is the target_index?
        ind = np.where(np.array(index_on_table)==target_index)[0]
        if len(ind) == 1:
            target_name = str(np.array(names_on_table)[ind][0])          
        else:
            target_name = str(target_index)
        
        #Get the user-defined colors from table
        colors_on_table = [self.tableWidget_Info_2.item(row, 2).background() for row in range(rowCount)]
        
        #it can be that the table was not updated and there are more scores than table-items
        if len(colors_on_table)!=len(scores_i):
            #update table
            SelectedFiles = self.items_clicked_no_rtdc_ds()
            self.update_data_overview(SelectedFiles)
            self.update_data_overview_2(SelectedFiles)
            
            #update colors on table
            rowCount = self.tableWidget_Info_2.rowCount()
            #only count rows with content
            rowCount = sum([self.tableWidget_Info_2.item(row, 0)!=None for row in range(rowCount)])
            colors_on_table = [self.tableWidget_Info_2.item(row, 2).background() for row in range(rowCount)]

        #Clear the plot        
        self.widget_probHistPlot.clear()
        #Add plot        
        hist = self.widget_probHistPlot.addPlot()
        hist.showGrid(x=True,y=True)
        hist.setLabel('bottom', "p("+target_name+")", units='')
        hist.setLabel('left', "#", units='')
        
        #Get the user defined histogram style from the combobox
        style = str(self.comboBox_probability_histogram.currentText())
        for i in range(len(scores_i)): # I had previously range(len(scores_i)); but this causes an error if there is a cell-type missing in the validation set
            hist_i = hist.plot()
            if len(scores_i[i])>1:#only continue of there multiple events (histogram does not make sense otherwise)
                range_hist = (scores_i[i].min(), scores_i[i].max())
                first_edge, last_edge = np.lib.histograms._get_outer_edges(scores_i[i], range=range_hist)
                try: #numpy 1.15
                    width = np.lib.histograms._hist_bin_selectors['auto'](scores_i[i])
                except:#numpy >1.15
                    width = np.lib.histograms._hist_bin_selectors['auto'](scores_i[i],(np.min(scores_i[i]),np.min(scores_i[i])))
                try:#prevent crash if width=0
                    n_equal_bins = int(np.ceil(np.lib.histograms._unsigned_subtract(last_edge, first_edge) / width))
                except: 
                    n_equal_bins = 1
                    
                if n_equal_bins>1E4: #Who needs more than 10k bins?!:
                    n_equal_bins = int(1E4)
                else:
                    n_equal_bins='auto'
    
                y,x = np.histogram(scores_i[i], bins=n_equal_bins)
                if style=="Style1":
                    pencolor = pg.mkColor(colors_on_table[i].color())
                    pen = pg.mkPen(color=pencolor,width=5)
                    hist_i.setData(x, y, stepMode=True, pen=pen,clear=False)
                elif style=="Style2":
                    pencolor = pg.mkColor(colors_on_table[i].color())
                    pen = pg.mkPen(color=pencolor,width=10)
                    hist_i.setData(x, y, stepMode=True, pen=pen,clear=False)
                elif style=="Style3":
                    color = colors_on_table[i].color()
                    color.setAlpha(0.6*255.0) 
                    pencolor = pg.mkColor(color)
                    brush = pg.mkBrush(color=pencolor)
                    hist_i.setData(x, y, stepMode=True, fillLevel=0, brush=brush,clear=False)
                elif style=="Style4":
                    color = colors_on_table[i].color()
                    color.setAlpha(0.7*255.0) 
                    pencolor = pg.mkColor(color)
                    brush = pg.mkBrush(color=pencolor)
                    hist_i.setData(x, y, stepMode=True, fillLevel=0, brush=brush,clear=False)
                elif style=="Style5":
                    color = colors_on_table[i].color()
                    color.setAlpha(0.8*255.0) 
                    pencolor = pg.mkColor(color)
                    brush = pg.mkBrush(color=pencolor)
                    hist_i.setData(x, y, stepMode=True, fillLevel=0, brush=brush,clear=False)

        #Add a vertical line indicating the threshold
        self.line = pg.InfiniteLine(pos=threshold, angle=90, pen='w', movable=False)
        hist.addItem(self.line)
        hist.setXRange(0, 1, padding=0)
        
    def thirdplot(self):
        target_index =self.spinBox_indexOfInterest.value()
        cb_text = self.comboBox_3rdPlot.currentText()
        
        if cb_text=='None':
            return
        if cb_text=='ROC-AUC':
            #Check if self.Metrics are available
            if len(self.Metrics) == 0:
                self.assess_model_plotting() #run this function to create self.Metrics
                dic = self.Metrics
            else: #If no Metrics available yet...
                dic = self.Metrics
            if len(dic)==0:
                return
            #Get the ValidationSet
            y_valid = self.ValidationSet["y_valid"] #load the validation labels to a new variable      
            scores = dic["scores"]
            
            inds_uni = list(range(scores.shape[1])) #these indices are explained by model
            #ROC-curve is only available for binary problems:
            Y_valid = to_categorical(y_valid,num_classes=len(inds_uni))
            
            # Compute ROC curve and ROC area for each class
            fpr,tpr,roc_auc = dict(),dict(),dict()
            for i in range(len(inds_uni)):                
                fpr[i], tpr[i], _ = metrics.roc_curve(Y_valid[:, i], scores[:, i])
                roc_auc[i] = metrics.auc(fpr[i], tpr[i])
            
            # Compute micro-average ROC curve and ROC area
            fpr["micro"], tpr["micro"], _ = metrics.roc_curve(Y_valid.ravel(), scores.ravel())
            roc_auc["micro"] = metrics.auc(fpr["micro"], tpr["micro"])

            #Get the user-defined colors from table
            rowCount = self.tableWidget_Info_2.rowCount()
            #only count rows with content
            rowCount = sum([self.tableWidget_Info_2.item(row, 0)!=None for row in range(rowCount)])

            colors_on_table = [self.tableWidget_Info_2.item(row, 2).background() for row in range(rowCount)]

            #Clear the plot        
            self.widget_3rdPlot.clear()
            #Add plot        
            hist = self.widget_3rdPlot.addPlot()
            hist.showGrid(x=True,y=True)
            hist.addLegend()
            hist.setLabel('bottom', "False Positive Rate", units='')
            hist.setLabel('left', "True Positive Rate", units='')
            for i, color in zip(range(len(inds_uni)), colors_on_table):
                text = 'Class '+str(i)+', AUC='+str(round(roc_auc[i],2))
                hist.plot(fpr[i], tpr[i], pen=None,symbol='o',symbolPen=None,symbolBrush=color,name=text,clear=False)
                clr = color.color()
                hist.plot(fpr[i],tpr[i],pen=clr)
            hist.setXRange(0, 1, padding=0)

        if cb_text=='Precision-Recall':
            #Check if self.Metrics are available
            if len(self.Metrics) == 0: 
                self.assess_model_plotting() #run this function to create self.Metrics
                dic = self.Metrics
            else: #Otherwise, there are Metrics available already :) Use them
                dic = self.Metrics
            if len(dic)==0:
                return
              
            #Get the ValidationSet
            y_valid = self.ValidationSet["y_valid"] #load the validation labels to a new variable      
            scores = dic["scores"]#[:,target_index]
            
            inds_uni = list(range(scores.shape[1])) #these indices are explained by model
            #ROC-curve is only available for binary problems:
            Y_valid = to_categorical(y_valid,num_classes=len(inds_uni))
            
            # Compute Precision Recall curve and P-R area for each class
            precision,recall,precision_recall_auc = dict(),dict(),dict()
            for i in range(len(inds_uni)):                
                precision[i], recall[i], _ = metrics.precision_recall_curve(Y_valid[:, i], scores[:, i])
                precision_recall_auc[i] = metrics.auc(recall[i], precision[i])
            
            # Compute micro-average ROC curve and ROC area
            precision["micro"], recall["micro"], _ = metrics.roc_curve(Y_valid.ravel(), scores.ravel())
            precision_recall_auc["micro"] = metrics.auc(recall["micro"],precision["micro"])

            #Get the user-defined colors from table
            rowCount = self.tableWidget_Info_2.rowCount()
            #only count rows with content
            rowCount = sum([self.tableWidget_Info_2.item(row, 0)!=None for row in range(rowCount)])

            colors_on_table = [self.tableWidget_Info_2.item(row, 2).background() for row in range(rowCount)]

            #Clear the plot        
            self.widget_3rdPlot.clear()
            #Add plot        
            hist = self.widget_3rdPlot.addPlot()
            hist.showGrid(x=True,y=True)
            hist.addLegend()
            hist.setLabel('bottom', "Recall", units='')
            hist.setLabel('left', "Precision", units='')
            for i, color in zip(range(len(inds_uni)), colors_on_table):
                text = 'Class '+str(i)+', AUC='+str(round(precision_recall_auc[i],2))
                hist.plot(recall[i],precision[i], pen=None,symbol='o',symbolPen=None,symbolBrush=color,name=text,clear=False)
                clr = color.color()
                hist.plot(recall[i],precision[i],pen=clr)
            hist.setXRange(0, 1, padding=0)

        if cb_text=='Enrichment vs. Threshold' or cb_text=='Yield vs. Threshold' or cb_text=='Conc. vs. Threshold':
            #Check if self.Metrics are available
            if len(self.Metrics) == 0: #if not,
                self.assess_model_plotting() #run this function to create self.Metrics
                dic = self.Metrics
            else: #If Metrics are already available, use it. Load it
                dic = self.Metrics
            if len(dic)==0:
                return
            
            scores = dic["scores"]
            y_valid = self.ValidationSet["y_valid"] #load the validation labels to a new variable

            #The dic only contains metrics for a single threshold, which is not enough
            #call aid_bin.metrics_using_threshold with a range of thresholds:
            #(it might makes sense to evaluate this for each possible target_index. Now only perform the measurement for the user defined target index)
            Dics,Threshs = [],[]
            for thresh in np.linspace(0,1,25):
                dic_ = aid_bin.metrics_using_threshold(scores,y_valid,thresh,target_index) #returns dic = {"scores":scores,"pred":pred,"conc_target_cell":conc_target_cell,"enrichment":enrichment,"yield_":yield_}                    
                Dics.append(dic_)
                Threshs.append(thresh)
            #Collect information in arrays
            enrichment_ = np.array([d["enrichment"] for d in Dics])
            yield__ = np.array([d["yield_"] for d in Dics])
            conc_target_cell = np.array([d["conc_target_cell"] for d in Dics])
            Threshs = np.array(Threshs)

            rowCount = self.tableWidget_Info_2.rowCount()
            #only count rows with content
            rowCount = sum([self.tableWidget_Info_2.item(row, 0)!=None for row in range(rowCount)])

            colors_on_table = [self.tableWidget_Info_2.item(row, 2).background() for row in range(rowCount)]

            #Clear the plot        
            self.widget_3rdPlot.clear()
            #Add plot        
            hist = self.widget_3rdPlot.addPlot()
            hist.showGrid(x=True,y=True)
            hist.addLegend()
            hist.setLabel('bottom', "Threshold", units='')
            color = '#0000ff'
            if cb_text=='Enrichment vs. Threshold':
                hist.setLabel('left', "Enrichment", units='')
                hist.plot(Threshs,enrichment_, pen=None,symbol='o',symbolPen=None,symbolBrush=color,name='',clear=False)
                hist.plot(Threshs,enrichment_,pen=color)
            if cb_text=='Yield vs. Threshold':
                hist.setLabel('left', "Yield", units='')
                hist.plot(Threshs,yield__, pen=None,symbol='o',symbolPen=None,symbolBrush=color,name='',clear=False)
                hist.plot(Threshs,yield__,pen=color)
            if cb_text=='Conc. vs. Threshold':
                hist.setLabel('left', "Conc. of target cell in target region", units='')
                hist.plot(Threshs,conc_target_cell, pen=None,symbol='o',symbolPen=None,symbolBrush=color,name='',clear=False)
                hist.plot(Threshs,conc_target_cell,pen=color)

            hist.setXRange(0, 1, padding=0)
            #Add indicator for the currently used threshold
            threshold = float(self.doubleSpinBox_sortingThresh.value())#threshold probability obove which a cell is sorted
            self.line = pg.InfiniteLine(pos=threshold, angle=90, pen='w', movable=False)
            hist.addItem(self.line)



    def classify(self):
        #Very similar function to "Update Plot". But here, no graphs are produced
        #Resulting scores/predictions etc are simply stored to excel file
        #This function does NOT take labels.

        #Check if a model was defined
        if self.load_model_path == None:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("Please define a model path first")
            msg.setWindowTitle("No model path found")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return

        AvailableFiles = self.items_available()
        rtdc_paths = [file_["rtdc_path"] for file_ in AvailableFiles] 
        #Classify all datasets or just one?
        Files,FileIndex = [],[]
        if self.radioButton_selectAll.isChecked():
            Files = rtdc_paths
            FileIndex = list(range(len(Files)))
            if len(Files)==0:
                msg = QtWidgets.QMessageBox()
                msg.setIcon(QtWidgets.QMessageBox.Information)       
                msg.setText("There are no files on the 'Build'-Tab")
                msg.setWindowTitle("No files found")
                msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                msg.exec_()
                return
        elif self.radioButton_selectDataSet.isChecked():
            rtdc_path = self.comboBox_selectData.currentText()
            Files.append(rtdc_path)
            #get the index of this file on the table
            FileIndex = [int(self.comboBox_selectData.currentIndex())]
            #FileIndex = list(np.where(np.array(rtdc_path)==np.array(rtdc_paths))[0])
        else:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("Please use the radiobuttons (left) to indicate if all or only a selected file should be classified.")
            msg.setWindowTitle("No file(s) specified")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return
          
        print("Chosen file(s):")
        print(Files)

        #what input size is required by loaded model?
        crop = int(self.spinBox_Crop_2.value())
        norm = str(self.comboBox_Normalization_2.currentText())
        paddingMode = str(self.comboBox_paddingMode.currentText())

        color_mode = self.get_color_mode()

        zoom_factors = [selectedfile["zoom_factor"] for selectedfile in AvailableFiles]
        #zoom_order = [self.actionOrder0.isChecked(),self.actionOrder1.isChecked(),self.actionOrder2.isChecked(),self.actionOrder3.isChecked(),self.actionOrder4.isChecked(),self.actionOrder5.isChecked()]
        #zoom_order = int(np.where(np.array(zoom_order)==True)[0])
        zoom_order = int(self.comboBox_zoomOrder.currentIndex()) #the combobox-index is already the zoom order

        xtra_in = set([selectedfile["xtra_in"] for selectedfile in AvailableFiles])   
        if len(xtra_in)>1:# False and True is present. Not supported
            print("Xtra data is used only for some files. Xtra data needs to be used either by all or by none!")
            return
        xtra_in = list(xtra_in)[0]#this is either True or False

        #if normalization method needs mean/std of training set, the metafile needs to be loaded:
        if norm == "StdScaling using mean and std of all training data":
            modelindex = int(self.spinBox_ModelIndex_2.value())
            path,fname = os.path.split(self.load_model_path)    
            fname = fname.split(str(modelindex)+".model")[0]+"meta.xlsx"
            metafile_path = os.path.join(path,fname)
            parameters = pd.read_excel(metafile_path,sheet_name='Parameters',engine="openpyxl")
            mean_trainingdata = parameters["Mean of training data used for scaling"]
            std_trainingdata = parameters["Std of training data used for scaling"]
        else:
            mean_trainingdata = None
            std_trainingdata = None
        
        #Create config (define which device to use)
        if self.radioButton_cpu.isChecked():
            deviceSelected = str(self.comboBox_cpu.currentText())
        elif self.radioButton_gpu.isChecked():
            deviceSelected = str(self.comboBox_gpu.currentText())
        gpu_memory = float(self.doubleSpinBox_memory.value())
        config_gpu = aid_dl.get_config(cpu_nr,gpu_nr,deviceSelected,gpu_memory)
                    
        with tf.compat.v1.Session(graph = tf.Graph(), config=config_gpu) as sess:     
            model_keras = load_model(self.load_model_path,custom_objects=aid_dl.get_custom_metrics())
            in_dim, _ = aid_dl.model_in_out_dim(model_keras,"model")
            #Get the color mode of the model
            channels_model = in_dim[-1]
            if channels_model==1:
                color_mode='Grayscale'
            elif channels_model==3:
                color_mode='RGB'
            else:
                print("Invalid number of channels. AID only supports grayscale (1 channel) and RGB (3 channels) images.")
    
    
            #Get the user-set export option (Excel or to 'userdef0' in .rtdc file)
            export_option = str(self.comboBox_scoresOrPrediction.currentText())
        
            if export_option == "Add predictions to .rtdc file (userdef0)" or export_option=="Add pred&scores to .rtdc file (userdef0 to 9)":
            #Users sometimes need to have Donor-ID (Parent foldername) added to the .rtdc file
            #Ask the user: Do you want to get a specific fixed addon to filename, OR do you want to have the parent-foldername added?
                msg = QtWidgets.QMessageBox()
                msg.setIcon(QtWidgets.QMessageBox.Question)
                text = "Do you want to get a specific fixed addon to filename, <b>or do you want to have the parent-foldername added for each file individually?"
                text = "<html><head/><body><p>"+text+"</p></body></html>"
                msg.setText(text)
                msg.setWindowTitle("Filename-addon for created files")
                msg.addButton(QtGui.QPushButton('Specific fixed addon...'), QtGui.QMessageBox.YesRole)
                msg.addButton(QtGui.QPushButton('Parent foldername'), QtGui.QMessageBox.NoRole)
                msg.addButton(QtGui.QPushButton('Cancel'), QtGui.QMessageBox.RejectRole)
                retval = msg.exec_()
    
                if retval==0: 
                    #Get some user input:
                    fname_addon, ok = QtWidgets.QInputDialog.getText(self, 'Specific fixed addon...', 'Enter filname addon:')
                    if ok:
                        fname_addon = str(fname_addon)
                    else:
                        return
                elif retval==1:
                    fname_addon = "Action:GetParentFolderName!"
                else:
                    return
    
            #Iterate over all Files
            for iterable in range(len(Files)):#rtdc_path in Files:
                print("Files:"+str(Files))

                print("iterable:"+str(iterable))
                rtdc_path = Files[iterable]
                print("rtdc_path:"+str(rtdc_path))
                print("FileIndex:"+str(FileIndex))
                print("zoom_factors:"+str(zoom_factors))

                f_index = FileIndex[iterable]
                zoom_factor = zoom_factors[f_index]
                
                #get all images, cropped correcetly
                gen_train = aid_img.gen_crop_img(crop,rtdc_path,replace=True,random_images=False,zoom_factor=zoom_factor,zoom_order=zoom_order,color_mode=color_mode,padding_mode=paddingMode,xtra_in=xtra_in)
                x_train,index,xtra_train = next(gen_train) #x_train-images of all cells, index-original index of all cells           
                
                if norm == "StdScaling using mean and std of all training data":
                    x_train = aid_img.image_normalization(x_train,norm,mean_trainingdata,std_trainingdata)
                else:
                    x_train = aid_img.image_normalization(x_train,norm)
                            
                #Check the input dimensions:
                img_dim = x_train.shape[-2]
                model_in = int(self.spinBox_Crop_2.value())
                if model_in!=img_dim:
                    msg = QtWidgets.QMessageBox()
                    msg.setIcon(QtWidgets.QMessageBox.Information)       
                    msg.setText("New model has different input dimensions (image crop). Validation set is re-loaded (clicked files on build-tab)")
                    msg.setWindowTitle("Input dimensions not fitting")
                    msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                    msg.exec_()
    
                scores = model_keras.predict(x_train)
                scores_normal = np.copy(scores)
                pred_normal = np.argmax(scores_normal,axis=1)
                
                #Get settings from the GUI
                threshold = float(self.doubleSpinBox_sortingThresh.value())#threshold probability obove which a cell is sorted
                target_index = int(self.spinBox_indexOfInterest.value())#index of the cell type that should be sorted for
    
                #Use argmax for prediction (threshold can only be applied to one index)
                pred_normal = np.argmax(scores,axis=1)
    
                #First: check the scores_in_function of the sorting index and adjust them using the threshold
                pred_thresh = np.array([1 if p>threshold else 0 for p in scores[:,target_index]])
                #replace the corresponding column in the scores_in_function
                scores[:,target_index] = pred_thresh
                #Determine the prediction again, considering the threshold for the target index
                pred_thresh = np.argmax(scores,axis=1)
                
                normal_or_thresh = bool(self.checkBox_SortingThresh.isChecked())
                if normal_or_thresh==True: #(if its true means its normal means p=0.5)
                    prediction_to_rtdc_ds = pred_normal
                if normal_or_thresh==False: #(if its false means its thresholded for some class)
                    prediction_to_rtdc_ds = pred_thresh
                
                if export_option == "Scores and predictions to Excel sheet":
                    info = np.array([[self.load_model_path],[rtdc_path],[target_index],[threshold]]).T
                    info = pd.DataFrame(info,columns=["load_model_path","rtdc_path","target_class","threshold"])
                    #Combine all information in nice excel sheet
                    filename = rtdc_path.split(".rtdc")[0]+"_Prediction.xlsx"
                    writer = pd.ExcelWriter(filename, engine='openpyxl')
                    #Used files go to a separate sheet on the -session.xlsx
                    pd.DataFrame().to_excel(writer,sheet_name='Info') #initialize empty Sheet
                    info.to_excel(writer,sheet_name='Info')
    
                    pd.DataFrame().to_excel(writer,sheet_name='Scores_normal') #initialize empty Sheet
                    pd.DataFrame(scores_normal).to_excel(writer,sheet_name='Scores_normal')
    
                    pd.DataFrame().to_excel(writer,sheet_name='Prediction_normal') #initialize empty Sheet
                    pd.DataFrame(pred_normal).to_excel(writer,sheet_name='Prediction_normal')
    
                    pd.DataFrame().to_excel(writer,sheet_name='Scores_thresholded') #initialize empty Sheet
                    pd.DataFrame(scores).to_excel(writer,sheet_name='Scores_thresholded')
    
                    pd.DataFrame().to_excel(writer,sheet_name='Prediction_thresholded') #initialize empty Sheet
                    pd.DataFrame(pred_thresh).to_excel(writer,sheet_name='Prediction_thresholded')
    
                    writer.save()
                    writer.close()
    
                if export_option == "Add predictions to .rtdc file (userdef0)" or export_option==         "Add pred&scores to .rtdc file (userdef0 to 9)":           
                    #Get initial dataset
                    failed,rtdc_ds = aid_bin.load_rtdc(rtdc_path)
                    if failed:
                        msg = QtWidgets.QMessageBox()
                        msg.setIcon(QtWidgets.QMessageBox.Critical)       
                        msg.setText(str(rtdc_ds))
                        msg.setWindowTitle("Error occurred during loading file")
                        msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                        msg.exec_()
                        return
                    
                    failed,rtdc_ds = aid_bin.load_rtdc(rtdc_path)
                    rtdc_ds_len = rtdc_ds["events"]["image"].shape[0] #this way is actually faster than asking any other feature for its len :)
                    prediction_fillnan = np.full([rtdc_ds_len], np.nan)#put initially np.nan for all cells
    
                    if export_option == "Add pred&scores to .rtdc file (userdef0 to 9)":
                        classes = scores_normal.shape[1]
                        if classes>9:
                            classes = 9#set the max number of classes to 9. It cannot saved more to .rtdc due to limitation of userdef
                        scores_fillnan = np.full([rtdc_ds_len,classes], np.nan)
    
                    #Make sure the predictions get again to the same length as the initial data set
                    #Fill array with corresponding predictions
                    for i in range(len(prediction_to_rtdc_ds)):
                        indx = index[i]
                        prediction_fillnan[indx] = prediction_to_rtdc_ds[i]
                        if export_option == "Add pred&scores to .rtdc file (userdef0 to 9)":
                            #for class_ in range(classes):
                            scores_fillnan[indx,0:classes] = scores_normal[i,0:classes]
    
                    #Get savename
                    path, rtdc_file = os.path.split(rtdc_path)
    
                    if fname_addon!="Action:GetParentFolderName!":#dont get the parentfoldername, instead used user defined addon!
                        savename = rtdc_path.split(".rtdc")[0]
                        savename = savename+"_"+str(fname_addon)+".rtdc"
                        
                    elif fname_addon=="Action:GetParentFolderName!":                        
                        savename = rtdc_path.split(".rtdc")[0]
                        parentfolder = aid_bin.splitall(rtdc_path)[-2]
                        savename = savename+"_"+str(parentfolder)+".rtdc"
                    else:
                        return
                    
                    if not os.path.isfile(savename):#if such a file does not yet exist...
                        savename = savename
                    else:#such a file already exists!!!
                        #Avoid to overwriting an existing file:
                        print("Adding additional number since file exists!")
                        addon = 1
                        while os.path.isfile(savename):
                            savename = savename.split(".rtdc")[0]
                            if addon>1:
                                savename = savename.split("_"+str(addon-1))[0]
                            savename = savename+"_"+str(addon)+".rtdc"
                            addon += 1        
                    
                    print(savename)                    
                    shutil.copy(rtdc_path, savename) #copy original file
                    #append to hdf5 file
                    with h5py.File(savename, mode="a") as h5:
                        if "userdef0" in h5["events"]:
                            del h5["events/userdef0"]
                        h5["events/userdef0"] = prediction_fillnan

                        if export_option == "Add pred&scores to .rtdc file (userdef0 to 9)":
                            #add the scores to userdef1...9
                            userdef_ind = 1
                            for class_ in range(classes):
                                if "userdef"+str(userdef_ind) in h5["events"]:
                                    del h5["events/userdef"+str(userdef_ind)]
                                scores_i = scores_fillnan[:,class_]
                                h5["events/userdef"+str(userdef_ind)] = scores_i
                                userdef_ind += 1
        

    #####################Python Editor/Console#################################
    def pythonInRun(self):
        self.threadpool_single_queue += 1
        if self.threadpool_single_queue == 1:
            worker = Worker(self.pythonInRun_Worker)                          
            self.threadpool.start(worker)

    def pythonInRun_Worker(self,progress_callback,history_callback):
        code = self.plainTextEdit_pythonIn.toPlainText()
        codeOut = io.StringIO()
        out,error = "",""
        # capture output
        sys.stdout = codeOut
        try:
            exec(code,globals())
        except Exception as e:
            error = str(e)
        # restore stdout
        sys.stdout = sys.__stdout__
        out = codeOut.getvalue()
        codeOut.close()
        
        text_out = "Out:\n"+out
        text_error = "Error:\n"+error
        
        #Print both to textBrowser_pythonOut
        self.textBrowser_pythonOut.append(text_out)
        if len(error)>0:
            self.textBrowser_pythonOut.append(text_error)

        self.threadpool_single_queue = 0 #reset thread counter
        
    def pythonInClear(self):
        self.plainTextEdit_pythonIn.clear()
        self.lineEdit_pythonCurrentFile.clear()

    def pythonInSaveAs(self):
        filename = QtWidgets.QFileDialog.getSaveFileName(self, 'Save code', Default_dict["Path of last model"],"Python file (*.py)")
        filename = filename[0]
        if len(filename)==0:
            return
        #add the suffix .csv
        if not filename.endswith(".py"):
            filename = filename +".py"               

        code = self.plainTextEdit_pythonIn.toPlainText()

        myfile = open(filename,'w')#Open the file with writing permission
        myfile.write(code)        
        myfile.close()
        self.lineEdit_pythonCurrentFile.setText(filename)

    def pythonInOpen(self):
        filename = QtWidgets.QFileDialog.getOpenFileName(self, 'Open Python file', Default_dict["Path of last model"],"Python file (*.py)")
        filename = filename[0]
        if not filename.endswith(".py"):
            return
        if not os.path.isfile(filename):
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("File not found")
            msg.setWindowTitle("File not found")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return
        with open(filename, 'r') as myfile:
            data = myfile.read()
        
        self.plainTextEdit_pythonIn.clear()
        self.plainTextEdit_pythonIn.insertPlainText(data)
        self.lineEdit_pythonCurrentFile.setText(filename)

    def pythonOutClear(self):
        self.textBrowser_pythonOut.clear()

    #Show cpu and ram usage on the status bar
    def cpu_ram_worker(self,progress_callback,history_callback):
        while True:
            cpu,ram = psutil.cpu_percent(),psutil.virtual_memory().percent
            #Count the number of visible fittingpopups
            Nr_fittings = [a.isVisible() for a in self.fittingpopups]
            Nr_fittings = int(np.sum(Nr_fittings))
            self.statusbar_cpuRam.setText("CPU: "+str(cpu)+"%  RAM: "+str(ram)+"% Jobs: "+str(Nr_fittings))            
            time.sleep(2)

    def delete_ram(self):
        if self.actionVerbose.isChecked()==True:
            print("Deleting Ram")
        self.ram = dict()
        self.ValidationSet = None
        self.Metrics = dict()

    def quit_app(self):
        sys.exit()

    def closeEvent(self, event):
        sys.exit()


